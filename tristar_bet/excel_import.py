from __future__ import annotations

import math
import re
import warnings
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Any, Iterable

from .models import (
    AdsorptiveProperties,
    FreeSpaceInfo,
    IsothermPoint,
    RunConditions,
    SampleInfo,
    SmpHeader,
    TargetPressureRow,
    TriStarResult,
)


MMHG_TO_KPA = 101.325 / 760.0
KPA_TO_MMHG = 760.0 / 101.325
CM3_STP_PER_MMOL = 22.414
EXCEL_EPOCH = datetime(1899, 12, 30)


class ExcelParseError(ValueError):
    """Raised when an Excel workbook is not a supported BET export."""


@dataclass(frozen=True)
class SheetGrid:
    name: str
    values: list[list[Any]]

    @property
    def nrows(self) -> int:
        return len(self.values)

    @property
    def ncols(self) -> int:
        return max((len(row) for row in self.values), default=0)

    def cell(self, row: int, column: int) -> Any:
        if row < 0 or row >= len(self.values):
            return None
        values = self.values[row]
        if column < 0 or column >= len(values):
            return None
        return values[column]


@dataclass(frozen=True)
class ExcelWorkbook:
    sheets: list[SheetGrid]

    def text_values(self) -> Iterable[str]:
        for sheet in self.sheets:
            for row in sheet.values:
                for value in row:
                    text = _text(value)
                    if text:
                        yield text


@dataclass(frozen=True)
class IsothermTable:
    points: list[IsothermPoint]
    source: str


def load_excel(path: str | Path) -> TriStarResult:
    return OfficialExcelParser().parse(path)


class OfficialExcelParser:
    def parse(self, path: str | Path) -> TriStarResult:
        file_path = Path(path)
        workbook = _read_workbook(file_path)
        labels = _collect_labels(workbook.sheets)
        _add_quantachrome_text_labels(workbook, labels)
        _add_jwgb_labels(workbook, labels)
        instrument = _detect_instrument(workbook)
        sample = self._build_sample(file_path, labels, workbook, instrument)
        run_conditions = self._build_run_conditions(labels)
        free_space = self._build_free_space(labels)
        adsorptive_properties = self._build_adsorptive_properties(labels, run_conditions)

        table = self._parse_bsd_isotherm(workbook)
        if table is None:
            table = self._parse_jwgb_isotherm(workbook)
        if table is None:
            table = self._parse_belmaster_isotherm(workbook)
        if table is None:
            table = self._parse_quantachrome_text_isotherm(workbook, sample)
        if table is None:
            table = self._parse_micromeritics_flex_report_isotherm(workbook, sample)
        if table is None:
            table = self._parse_micromeritics_isotherm(workbook, sample)
        if table is None or not table.points:
            raise ExcelParseError(f"No supported isotherm table found in {file_path}")

        created_raw, created_time = self._parse_started_time(labels)
        completed_raw, completed_time = self._parse_completed_time(labels)
        report_raw, report_time = self._parse_report_time(labels)
        modified_raw, modified_time = _file_modified_timestamp(file_path)
        if not created_time:
            created_raw, created_time = modified_raw, modified_time
        header = SmpHeader(
            file_path=str(file_path.resolve()),
            file_name=file_path.name,
            byte_count=file_path.stat().st_size,
            magic="OFFICIAL_EXCEL",
            version=instrument.get("instrument_software", ""),
            created_raw=created_raw,
            created_time=created_time,
            modified_raw=modified_raw,
            modified_time=modified_time,
            directory_offset=0,
            directory_size=0,
        )

        method_options: dict[str, Any] = {
            **instrument,
            "excel_import_source": table.source,
            "excel_quantity_source": "official_report_quantity_adsorbed_cm3_g_stp",
            "excel_phase_source": "pressure_sequence_peak" if table.source.startswith("micromeritics") else "sheet_branch_marker",
            "excel_elapsed_time_note": "source display text preserved; not used for analysis",
            "target_pressure_table_source": "measured_isotherm_pressure_sequence",
            "sample_saved_time": modified_time,
            "sample_saved_raw": modified_raw,
            "sample_saved_time_source": "Windows file LastWriteTime",
            "excel_missing_method_settings": (
                "Excel reports often omit method settings such as selected thickness curve, "
                "BET/t-plot ranges, smoothing and BJH correction details."
            ),
        }
        if created_time:
            method_options["test_started_time"] = created_time
            method_options["test_started_raw"] = created_raw
        if completed_time:
            method_options["test_completed_time"] = completed_time
            method_options["test_completed_raw"] = completed_raw
        if report_time:
            method_options["excel_report_time"] = report_time
            method_options["excel_report_raw"] = report_raw
        duration_text, duration_seconds = _duration_text(created_raw, completed_raw)
        if duration_text:
            method_options["test_duration_time"] = duration_text
            method_options["test_duration_seconds"] = duration_seconds
            method_options["test_duration_source"] = "Excel Started/Completed fields"
        self._add_report_values(labels, method_options)
        self._add_bsd_report_values(workbook, method_options)
        self._add_jwgb_report_values(workbook, method_options)
        self._add_quantachrome_text_report_values(workbook, method_options)
        self._add_micromeritics_flex_report_values(workbook, method_options)
        self._add_bet_fit_range(workbook, method_options)

        return TriStarResult(
            header=header,
            subsets=[],
            sample=sample,
            run_conditions=run_conditions,
            target_pressure_table=_target_pressure_table_from_isotherm(table.points),
            free_space=free_space,
            po_records=[],
            isotherm=table.points,
            adsorptive_properties=adsorptive_properties,
            log_messages=[],
            sample_tube_strings=[],
            method_options=method_options,
            raw_strings={},
        )

    def _build_sample(
        self,
        file_path: Path,
        labels: dict[str, Any],
        workbook: ExcelWorkbook,
        instrument: dict[str, Any],
    ) -> SampleInfo:
        sample_name = _as_clean_string(_first_label(labels, "sample", "样品名称"))
        if not sample_name and instrument.get("instrument_manufacturer") == "MicrotracBEL":
            sample_name = self._belmaster_sample_name(workbook, labels)
        if not sample_name:
            data_file = _as_clean_string(_first_label(labels, "data file", "file"))
            sample_name = Path(data_file).stem if data_file else file_path.stem
        return SampleInfo(
            sample_name=sample_name,
            operator=_as_clean_string(_first_label(labels, "operator", "测试人员")),
            submitter=_as_clean_string(_first_label(labels, "submitter", "送样人员", "送样单位")),
            bar_code=_as_clean_string(_first_label(labels, "bar code")),
            sample_mass_g=_number(_first_label(labels, "sample mass", "sample weight/g", "样品质量", "脱气后样品质量")),
            sample_density_g_cm3=_number(_first_label(labels, "sample density", "样品密度")),
        )

    @staticmethod
    def _belmaster_sample_name(workbook: ExcelWorkbook, labels: dict[str, Any]) -> str:
        for sheet in workbook.sheets:
            if sheet.name.lower() not in {"summary", "isotherm"}:
                continue
            for row_index in range(min(sheet.nrows, 10)):
                if _label_key(sheet.cell(row_index, 0)):
                    continue
                value = sheet.cell(row_index, 2)
                if not isinstance(value, str):
                    continue
                text = _as_clean_string(value)
                if text and not any(mark in text.lower() for mark in ("data file", "vacuum", ".dat")):
                    return text
        data_file = _as_clean_string(_first_label(labels, "data file"))
        return Path(data_file).stem if data_file else ""

    @staticmethod
    def _build_run_conditions(labels: dict[str, Any]) -> RunConditions:
        raw_adsorptive = _as_clean_string(_first_label(labels, "analysis adsorptive", "adsorptive", "吸附质"))
        adsorptive = _adsorptive_short(raw_adsorptive)
        bath_temperature = _temperature_k(
            _first_label(labels, "analysis bath temp.", "analysis bath temp", "adsorption temperature", "temperature")
        )
        if bath_temperature is None:
            bath_temperature = _adsorptive_temperature_k(raw_adsorptive)
        equilibration = _number(_first_label(labels, "equilibration interval"))
        return RunConditions(
            evacuation_rate_mmHg_s=None,
            unrestricted_evacuate_from_mmHg=None,
            evacuation_time_h=None,
            leak_test_time_s=None,
            equilibration_interval_s=equilibration,
            free_space_equilibration_time_h=None,
            ambient_free_space_entered_cm3=_number(_first_label(labels, "ambient free space")),
            analysis_free_space_entered_cm3=_number(_first_label(labels, "analysis free space")),
            desorption_test_time_s=None,
            po_reference_mmHg=None,
            bath_temperature_K=bath_temperature,
            adsorptive_short=adsorptive,
            adsorptive_name=adsorptive,
        )

    @staticmethod
    def _build_free_space(labels: dict[str, Any]) -> FreeSpaceInfo:
        warm = _number(_first_label(labels, "warm free space"))
        cold = _number(_first_label(labels, "cold free space"))
        generic = _number(_first_label(labels, "free space"))
        analysis = _number(_first_label(labels, "analysis free space"))
        ambient = _number(_first_label(labels, "ambient free space"))
        return FreeSpaceInfo(
            analysis_entered_cm3=analysis,
            ambient_entered_cm3=ambient,
            nonideality_factor=None,
            cold_free_space_cm3=cold,
            warm_free_space_cm3=warm,
            stem_volume_cm3=None,
            vbath_cm3=None,
            vfree_factor_cm3=generic,
            vfree_factor_source="official_excel_report",
        )

    @staticmethod
    def _build_adsorptive_properties(labels: dict[str, Any], run_conditions: RunConditions) -> AdsorptiveProperties | None:
        cross_section = _number(_first_label(labels, "molecular cross-sectional area"))
        density_factor = _number(_first_label(labels, "density conversion factor"))
        psat_kpa = _number(_first_label(labels, "saturation vapor pressure"))
        if not run_conditions.adsorptive_short and cross_section is None and density_factor is None and psat_kpa is None:
            return None
        psat_table = []
        if psat_kpa is not None:
            psat_table.append({"index": 1, "saturation_pressure_mmHg": psat_kpa * KPA_TO_MMHG})
        return AdsorptiveProperties(
            adsorptive=run_conditions.adsorptive_short,
            mnemonic=run_conditions.adsorptive_short,
            max_manifold_pressure_mmHg=None,
            max_manifold_pressure_kPa=None,
            nonideality_factor=None,
            density_conversion_factor=density_factor,
            thermal_transpiration_hard_sphere_A=None,
            thermal_transpiration_hard_sphere_nm=None,
            molecular_cross_sectional_area_nm2=cross_section,
            ui_field_rel101=None,
            psat_table=psat_table,
        )

    def _parse_jwgb_isotherm(self, workbook: ExcelWorkbook) -> IsothermTable | None:
        if not _is_jwgb_workbook(workbook):
            return None
        sheet = next((item for item in workbook.sheets if item.name.lower() == "isotherm"), None)
        if sheet is None:
            return None
        table_position = self._find_jwgb_isotherm_table(sheet)
        if table_position is None:
            return None
        header_row, columns = table_position
        points: list[IsothermPoint] = []
        phase = "adsorption"
        point_index = 1
        for row_index in range(header_row + 1, sheet.nrows):
            relative = _number(sheet.cell(row_index, columns["relative_pressure"]))
            quantity = _number(sheet.cell(row_index, columns["quantity"]))
            if relative is None or quantity is None:
                if points and phase == "adsorption":
                    phase = "desorption"
                    point_index = 1
                continue
            pressure_kpa = _number(sheet.cell(row_index, columns["pressure_kpa"]))
            saturation_kpa = _number(sheet.cell(row_index, columns["saturation_kpa"]))
            elapsed_seconds = None
            elapsed_col = columns.get("elapsed")
            if elapsed_col is not None:
                elapsed_seconds = _elapsed_display_seconds(sheet.cell(row_index, elapsed_col))
            if pressure_kpa is None:
                pressure_kpa = float(relative) * (saturation_kpa or 101.325)
            if saturation_kpa is None or saturation_kpa <= 0.0:
                saturation_kpa = pressure_kpa / float(relative) if relative > 0.0 else 101.325
            row_id = _number(sheet.cell(row_index, columns["id"]))
            point_index_value = int(row_id) if row_id is not None else point_index
            points.append(
                IsothermPoint(
                    index=point_index_value,
                    phase=phase,
                    record_rel_offset=0,
                    absolute_pressure_mmHg=float(pressure_kpa) * KPA_TO_MMHG,
                    relative_pressure=float(relative),
                    raw_internal_cm3_stp=float(quantity),
                    saturation_pressure_mmHg=float(saturation_kpa) * KPA_TO_MMHG,
                    elapsed_seconds=elapsed_seconds,
                    quantity_adsorbed_cm3_g_stp=float(quantity),
                    quantity_adsorbed_mmol_g=float(quantity) / CM3_STP_PER_MMOL,
                )
            )
            point_index += 1
        return IsothermTable(points=points, source=f"jwgb:{sheet.name}") if points else None

    @staticmethod
    def _find_jwgb_isotherm_table(sheet: SheetGrid) -> tuple[int, dict[str, int]] | None:
        for row_index in range(sheet.nrows):
            headers = [_normalize_header(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
            columns: dict[str, int] = {}
            for column_index, header in enumerate(headers):
                compact = header.replace(" ", "").replace("po", "p0")
                if compact == "id":
                    columns["id"] = column_index
                elif compact in {"p(kpa)", "pressure(kpa)"}:
                    columns["pressure_kpa"] = column_index
                elif compact in {"p/p0", "p/po"}:
                    columns["relative_pressure"] = column_index
                elif compact.startswith("v(") and "cm3/gstp" in compact:
                    columns["quantity"] = column_index
                elif compact in {"p0(kpa)", "po(kpa)"}:
                    columns["saturation_kpa"] = column_index
                elif compact == "time":
                    columns["elapsed"] = column_index
            required = {"id", "pressure_kpa", "relative_pressure", "quantity", "saturation_kpa"}
            if required.issubset(columns):
                return row_index, columns
        return None

    def _parse_belmaster_isotherm(self, workbook: ExcelWorkbook) -> IsothermTable | None:
        sheet = next((item for item in workbook.sheets if item.name.lower() == "isotherm"), None)
        if sheet is None:
            return None
        header = None
        for row_index, row in enumerate(sheet.values):
            normalized = [_normalize_header(value) for value in row]
            if "p/p0" in normalized and any("va/cm3" in item or "va/cm" in item for item in normalized):
                header = row_index
                break
        if header is None:
            return None

        phase = "adsorption"
        points: list[IsothermPoint] = []
        point_index = 0
        for row_index in range(header + 1, sheet.nrows):
            first = _as_clean_string(sheet.cell(row_index, 0)).upper()
            if first in {"ADS", "ADSORPTION"}:
                phase = "adsorption"
                continue
            if first in {"DES", "DESORPTION"}:
                phase = "desorption"
                continue
            pe_kpa = _number(sheet.cell(row_index, 1))
            p0_kpa = _number(sheet.cell(row_index, 2))
            relative = _number(sheet.cell(row_index, 3))
            quantity = _number(sheet.cell(row_index, 4))
            if pe_kpa is None or p0_kpa is None or relative is None or quantity is None:
                continue
            points.append(
                IsothermPoint(
                    index=point_index,
                    phase=phase,
                    record_rel_offset=0,
                    absolute_pressure_mmHg=pe_kpa * KPA_TO_MMHG,
                    relative_pressure=relative,
                    raw_internal_cm3_stp=quantity,
                    saturation_pressure_mmHg=p0_kpa * KPA_TO_MMHG,
                    elapsed_seconds=None,
                    quantity_adsorbed_cm3_g_stp=quantity,
                    quantity_adsorbed_mmol_g=quantity / CM3_STP_PER_MMOL,
                )
            )
            point_index += 1
        return IsothermTable(points=points, source=f"belmaster:{sheet.name}") if points else None

    def _parse_quantachrome_text_isotherm(
        self,
        workbook: ExcelWorkbook,
        sample: SampleInfo,
    ) -> IsothermTable | None:
        if not _is_quantachrome_text_workbook(workbook):
            return None
        sheet = next((item for item in workbook.sheets if item.name.lower() == "isotherm"), None)
        if sheet is None:
            return None
        raw_rows = _quantachrome_text_numeric_rows_after_units(sheet, expected_columns=2)
        if len(raw_rows) < 3:
            return None

        max_position = max(range(len(raw_rows)), key=lambda index: raw_rows[index][0])
        points: list[IsothermPoint] = []
        for index, row in enumerate(raw_rows, start=1):
            relative = float(row[0])
            quantity = float(row[1])
            if not (0.0 < relative < 1.1) or quantity < 0.0:
                continue
            phase = "adsorption" if index - 1 <= max_position else "desorption"
            points.append(
                IsothermPoint(
                    index=index,
                    phase=phase,
                    record_rel_offset=0,
                    absolute_pressure_mmHg=relative * 760.0,
                    relative_pressure=relative,
                    raw_internal_cm3_stp=quantity * sample.sample_mass_g if sample.sample_mass_g else quantity,
                    saturation_pressure_mmHg=760.0,
                    elapsed_seconds=None,
                    quantity_adsorbed_cm3_g_stp=quantity,
                    quantity_adsorbed_mmol_g=quantity / CM3_STP_PER_MMOL,
                )
            )
        return IsothermTable(points=points, source=f"quantachrome_text:{sheet.name}") if points else None

    def _parse_bsd_isotherm(self, workbook: ExcelWorkbook) -> IsothermTable | None:
        if not _is_bsd_workbook(workbook):
            return None
        for sheet in workbook.sheets:
            table_position = self._find_bsd_isotherm_table(sheet)
            if table_position is None:
                continue
            header_row, columns = table_position
            raw_rows: list[dict[str, Any]] = []
            invalid_streak = 0
            for row_index in range(header_row + 1, sheet.nrows):
                serial = _number(sheet.cell(row_index, columns["serial"]))
                pressure_pa = _number(sheet.cell(row_index, columns["pressure_pa"]))
                p0_pa = _number(sheet.cell(row_index, columns["p0_pa"]))
                relative = _number(sheet.cell(row_index, columns["relative_pressure"]))
                quantity = _number(sheet.cell(row_index, columns["quantity"]))
                if serial is None or pressure_pa is None or p0_pa is None or relative is None or quantity is None:
                    if raw_rows:
                        invalid_streak += 1
                        if invalid_streak >= 3:
                            break
                    continue
                invalid_streak = 0
                elapsed_seconds = None
                elapsed_col = columns.get("elapsed")
                if elapsed_col is not None:
                    elapsed_seconds = _elapsed_display_seconds(sheet.cell(row_index, elapsed_col))
                raw_rows.append(
                    {
                        "relative": relative,
                        "absolute_mmHg": pressure_pa * KPA_TO_MMHG / 1000.0,
                        "saturation_mmHg": p0_pa * KPA_TO_MMHG / 1000.0,
                        "quantity": quantity,
                        "elapsed_seconds": elapsed_seconds,
                    }
                )
            if not raw_rows:
                continue
            peak_index = max(range(len(raw_rows)), key=lambda index: raw_rows[index]["relative"])
            points = []
            for index, row in enumerate(raw_rows):
                quantity = float(row["quantity"])
                points.append(
                    IsothermPoint(
                        index=index,
                        phase="adsorption" if index <= peak_index else "desorption",
                        record_rel_offset=0,
                        absolute_pressure_mmHg=float(row["absolute_mmHg"]),
                        relative_pressure=float(row["relative"]),
                        raw_internal_cm3_stp=quantity,
                        saturation_pressure_mmHg=float(row["saturation_mmHg"]),
                        elapsed_seconds=row["elapsed_seconds"],
                        quantity_adsorbed_cm3_g_stp=quantity,
                        quantity_adsorbed_mmol_g=quantity / CM3_STP_PER_MMOL,
                    )
                )
            return IsothermTable(points=points, source=f"bsd:{sheet.name}")
        return None

    @staticmethod
    def _find_bsd_isotherm_table(sheet: SheetGrid) -> tuple[int, dict[str, int]] | None:
        for row_index in range(sheet.nrows):
            headers = [_normalize_header(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
            columns: dict[str, int] = {}
            for column_index, header in enumerate(headers):
                compact = header.replace(" ", "")
                if compact == "serial":
                    columns["serial"] = column_index
                elif "p/(pa)" in compact or compact == "p(pa)":
                    columns["pressure_pa"] = column_index
                elif "p0/(pa)" in compact or compact == "p0(pa)":
                    columns["p0_pa"] = column_index
                elif compact in {"p/p0", "p/po"}:
                    columns["relative_pressure"] = column_index
                elif "v/(cm3/g.stp)" in compact or "∑v/(cm3/g.stp)" in compact:
                    columns["quantity"] = column_index
                elif compact in {"t/(h:m)", "t(s)"}:
                    columns["elapsed"] = column_index
            required = {"serial", "pressure_pa", "p0_pa", "relative_pressure", "quantity"}
            if required.issubset(columns):
                return row_index, columns
        return None

    def _parse_micromeritics_flex_report_isotherm(
        self,
        workbook: ExcelWorkbook,
        sample: SampleInfo,
    ) -> IsothermTable | None:
        if not _is_micromeritics_flex_workbook(workbook):
            return None
        sheet = next((item for item in workbook.sheets if item.name == "Isotherm Tabular Report"), None)
        if sheet is None:
            return None
        table_position = self._find_micromeritics_flex_isotherm_table(sheet)
        if table_position is None:
            return None
        header_row, columns = table_position
        saturation_kpa = None
        saturation_col = columns.get("saturation")
        if saturation_col is not None:
            for row_index in range(header_row + 1, min(sheet.nrows, header_row + 8)):
                saturation_kpa = _number(sheet.cell(row_index, saturation_col))
                if saturation_kpa is not None:
                    break

        raw_rows: list[dict[str, float]] = []
        invalid_streak = 0
        for row_index in range(header_row + 1, sheet.nrows):
            relative = _number(sheet.cell(row_index, columns["relative_pressure"]))
            quantity = _number(sheet.cell(row_index, columns["quantity"]))
            row_saturation = _number(sheet.cell(row_index, saturation_col)) if saturation_col is not None else None
            if row_saturation is not None:
                saturation_kpa = row_saturation
            if relative is None or quantity is None:
                if raw_rows:
                    invalid_streak += 1
                    if invalid_streak >= 4:
                        break
                continue
            if not (0.0 < relative < 1.1):
                if raw_rows:
                    break
                continue
            invalid_streak = 0
            absolute_mmHg = relative * saturation_kpa * KPA_TO_MMHG if saturation_kpa is not None else relative * 760.0
            raw_rows.append(
                {
                    "relative": float(relative),
                    "quantity": float(quantity),
                    "absolute_mmHg": float(absolute_mmHg),
                    "saturation_mmHg": float(saturation_kpa * KPA_TO_MMHG) if saturation_kpa is not None else 760.0,
                }
            )
        if not raw_rows:
            return None

        max_relative = max(row["relative"] for row in raw_rows)
        peak_index = max(
            index
            for index, row in enumerate(raw_rows)
            if math.isclose(row["relative"], max_relative, rel_tol=0.0, abs_tol=1e-12)
        )
        points = []
        for index, row in enumerate(raw_rows, start=1):
            quantity = float(row["quantity"])
            points.append(
                IsothermPoint(
                    index=index,
                    phase="adsorption" if index - 1 <= peak_index else "desorption",
                    record_rel_offset=0,
                    absolute_pressure_mmHg=float(row["absolute_mmHg"]),
                    relative_pressure=float(row["relative"]),
                    raw_internal_cm3_stp=quantity * sample.sample_mass_g if sample.sample_mass_g else quantity,
                    saturation_pressure_mmHg=float(row["saturation_mmHg"]),
                    elapsed_seconds=None,
                    quantity_adsorbed_cm3_g_stp=quantity,
                    quantity_adsorbed_mmol_g=quantity / CM3_STP_PER_MMOL,
                )
            )
        return IsothermTable(points=points, source=f"micromeritics_flex_report:{sheet.name}")

    @staticmethod
    def _find_micromeritics_flex_isotherm_table(sheet: SheetGrid) -> tuple[int, dict[str, int]] | None:
        for row_index in range(sheet.nrows):
            headers = [_normalize_header(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
            columns: dict[str, int] = {}
            for column_index, header in enumerate(headers):
                compact = header.replace(" ", "")
                if _is_relative_pressure_header(header):
                    columns["relative_pressure"] = column_index
                elif "quantityadsorbed" in compact and "cm3/gstp" in compact:
                    columns["quantity"] = column_index
                elif "saturationpressure" in compact and "kpa" in compact:
                    columns["saturation"] = column_index
            if {"relative_pressure", "quantity"}.issubset(columns):
                return row_index, columns
        return None

    def _parse_micromeritics_isotherm(self, workbook: ExcelWorkbook, sample: SampleInfo) -> IsothermTable | None:
        for sheet in workbook.sheets:
            table_position = self._find_micromeritics_table(sheet)
            if table_position is None:
                continue
            header_row, rel_col = table_position
            raw_rows: list[dict[str, Any]] = []
            invalid_streak = 0
            for row_index in range(header_row + 1, sheet.nrows):
                relative = _number(sheet.cell(row_index, rel_col))
                absolute = _number(sheet.cell(row_index, rel_col + 1))
                quantity = _number(sheet.cell(row_index, rel_col + 2))
                elapsed_seconds = _elapsed_display_seconds(sheet.cell(row_index, rel_col + 3))
                saturation = _number(sheet.cell(row_index, rel_col + 4))
                if relative is None or absolute is None or quantity is None:
                    if raw_rows:
                        invalid_streak += 1
                        if invalid_streak >= 5:
                            break
                    continue
                invalid_streak = 0
                raw_rows.append(
                    {
                        "relative": relative,
                        "absolute": absolute,
                        "quantity": quantity,
                        "elapsed_seconds": elapsed_seconds,
                        "saturation": saturation,
                    }
                )
            if not raw_rows:
                continue
            peak_index = max(range(len(raw_rows)), key=lambda index: raw_rows[index]["relative"])
            points = []
            for index, row in enumerate(raw_rows):
                quantity = float(row["quantity"])
                points.append(
                    IsothermPoint(
                        index=index,
                        phase="adsorption" if index <= peak_index else "desorption",
                        record_rel_offset=0,
                        absolute_pressure_mmHg=float(row["absolute"]),
                        relative_pressure=float(row["relative"]),
                        raw_internal_cm3_stp=quantity * sample.sample_mass_g if sample.sample_mass_g else quantity,
                        saturation_pressure_mmHg=row["saturation"],
                        elapsed_seconds=row["elapsed_seconds"],
                        quantity_adsorbed_cm3_g_stp=quantity,
                        quantity_adsorbed_mmol_g=quantity / CM3_STP_PER_MMOL,
                    )
                )
            return IsothermTable(points=points, source=f"micromeritics:{sheet.name}")
        return None

    @staticmethod
    def _find_micromeritics_table(sheet: SheetGrid) -> tuple[int, int] | None:
        for row_index in range(sheet.nrows):
            for column_index in range(sheet.ncols - 2):
                current = _normalize_header(sheet.cell(row_index, column_index))
                next_one = _normalize_header(sheet.cell(row_index, column_index + 1))
                next_two = _normalize_header(sheet.cell(row_index, column_index + 2))
                if (
                    "relative pressure" in current
                    and "absolute pressure" in next_one
                    and "quantity adsorbed" in next_two
                ):
                    return row_index, column_index
        return None

    @staticmethod
    def _parse_started_time(labels: dict[str, Any]) -> tuple[int, str]:
        started = _first_label(labels, "started", "开始时间")
        if started:
            return _timestamp_pair(started)
        analysis_date = _first_label(labels, "analysis date")
        analysis_time = _first_label(labels, "analysis time")
        combined = _combine_date_time(analysis_date, analysis_time)
        return _timestamp_pair(combined) if combined else (0, "")

    @staticmethod
    def _parse_completed_time(labels: dict[str, Any]) -> tuple[int, str]:
        return _timestamp_pair(_first_label(labels, "completed", "结束时间"))

    @staticmethod
    def _parse_report_time(labels: dict[str, Any]) -> tuple[int, str]:
        return _timestamp_pair(_first_label(labels, "report time"))

    @staticmethod
    def _add_report_values(labels: dict[str, Any], method_options: dict[str, Any]) -> None:
        mappings = {
            "bet surface area": "excel_bet_surface_area_m2_g",
            "langmuir surface area": "excel_langmuir_surface_area_m2_g",
            "t-plot micropore area": "excel_t_plot_micropore_area_m2_g",
            "t-plot external surface area": "excel_t_plot_external_surface_area_m2_g",
            "t-plot micropore volume": "excel_t_plot_micropore_volume_cm3_g",
        }
        for label, option_key in mappings.items():
            value = _number(_first_label(labels, label))
            if value is not None:
                method_options[option_key] = value
                method_options["official_fit_value_usage"] = "validation_only"

    @staticmethod
    def _add_bet_fit_range(workbook: ExcelWorkbook, method_options: dict[str, Any]) -> None:
        if "stored_bet_pressure_min" in method_options and "stored_bet_pressure_max" in method_options:
            return
        bsd_range = _bsd_bet_pressure_range(workbook)
        if bsd_range is not None:
            p_min, p_max, source = bsd_range
            method_options["stored_bet_pressure_min"] = p_min
            method_options["stored_bet_pressure_max"] = p_max
            method_options["excel_bet_range_source"] = source
            method_options["official_fit_range_usage"] = "validation_only"
            return
        fit_range = _bet_fit_pressure_range(workbook)
        if fit_range is None:
            return
        p_min, p_max, source = fit_range
        method_options["stored_bet_pressure_min"] = p_min
        method_options["stored_bet_pressure_max"] = p_max
        method_options["excel_bet_range_source"] = source
        method_options["official_fit_range_usage"] = "validation_only"

    @staticmethod
    def _add_bsd_report_values(workbook: ExcelWorkbook, method_options: dict[str, Any]) -> None:
        if not _is_bsd_workbook(workbook):
            return
        method_options["bsd_excel_import"] = True
        method_options["bsd_bjh_thickness_method"] = "halsey"
        method_options["bsd_bjh_kelvin_factor_nm"] = 0.954853
        for sheet in workbook.sheets:
            title = sheet.name
            if "BET" in title:
                _update_from_label_pairs(
                    sheet,
                    method_options,
                    {
                        "斜率a": "bsd_bet_slope",
                        "截距b": "bsd_bet_intercept",
                        "相关系数r": "bsd_bet_r",
                        "BET常数C": "bsd_bet_c_constant",
                        "单层吸附量Vm": "bsd_bet_monolayer_capacity_cm3_g_stp",
                        "BET比表面积": "excel_bet_surface_area_m2_g",
                    },
                )
            elif "Langmuir" in title:
                _update_from_label_pairs(
                    sheet,
                    method_options,
                    {
                        "斜率a": "bsd_langmuir_slope",
                        "截距b": "bsd_langmuir_intercept",
                        "相关系数r": "bsd_langmuir_r",
                        "常数C": "bsd_langmuir_c_constant",
                        "单层吸附量Vm": "bsd_langmuir_monolayer_capacity_cm3_g_stp",
                        "Langmuir比表面积": "excel_langmuir_surface_area_m2_g",
                    },
                )
                langmuir_range = _bsd_range_from_label(sheet, "P取点范围/kPa")
                if langmuir_range is not None:
                    method_options["bsd_langmuir_pressure_min_kpa"] = langmuir_range[0]
                    method_options["bsd_langmuir_pressure_max_kpa"] = langmuir_range[1]
                    method_options["official_fit_range_usage"] = "validation_only"
            elif "T-Plot" in title or "T-plot" in title:
                _update_from_label_pairs(
                    sheet,
                    method_options,
                    {
                        "斜率a": "bsd_t_plot_slope",
                        "截距b": "bsd_t_plot_intercept",
                        "相关系数r": "bsd_t_plot_r",
                        "T-Plot微孔容积": "excel_t_plot_micropore_volume_cm3_g",
                        "T-Plot微孔比表面积": "excel_t_plot_micropore_area_m2_g",
                        "T-Plot外比表面积": "excel_t_plot_external_surface_area_m2_g",
                    },
                )
                t_range = _bsd_range_from_label(sheet, "P/P0取点范围")
                if t_range is not None:
                    method_options["bsd_t_plot_pressure_min"] = t_range[0]
                    method_options["bsd_t_plot_pressure_max"] = t_range[1]
                    method_options["official_fit_range_usage"] = "validation_only"
            elif "BJH" in title:
                phase_key = "adsorption" if "吸附" in title else "desorption" if "脱附" in title else "unknown"
                prefix = f"bsd_bjh_{phase_key}"
                _update_from_label_pairs(
                    sheet,
                    method_options,
                    {
                        "BJH累积孔容积": f"{prefix}_pore_volume_cm3_g",
                        "BJH平均孔直径": f"{prefix}_average_diameter_nm",
                        "累计孔面积S": f"{prefix}_pore_area_m2_g",
                        "最可几孔直径": f"{prefix}_mode_diameter_nm",
                        "D10孔直径": f"{prefix}_d10_nm",
                        "D50孔直径": f"{prefix}_d50_nm",
                        "D90孔直径": f"{prefix}_d90_nm",
                        "D99孔直径": f"{prefix}_d99_nm",
                    },
                )
                table_rows = _bsd_bjh_table_rows(sheet, phase_key)
                if table_rows:
                    method_options["official_bjh_table_usage"] = "validation_only"
                    method_options[f"{prefix}_rows"] = table_rows
        if any(
            key in method_options
            for key in (
                "excel_bet_surface_area_m2_g",
                "excel_langmuir_surface_area_m2_g",
                "excel_t_plot_micropore_volume_cm3_g",
                "excel_t_plot_micropore_area_m2_g",
                "excel_t_plot_external_surface_area_m2_g",
            )
        ):
            method_options["official_fit_value_usage"] = "validation_only"

    @staticmethod
    def _add_jwgb_report_values(workbook: ExcelWorkbook, method_options: dict[str, Any]) -> None:
        if not _is_jwgb_workbook(workbook):
            return
        method_options["jwgb_excel_import"] = True
        method_options["vendor_bjh_thickness_method"] = "halsey"
        method_options["vendor_bjh_correction"] = "standard"
        method_options["vendor_bjh_smooth_derivative"] = False
        method_options["jwgb_t_plot_thickness_method"] = "harkins_jura"
        method_options["jwgb_bjh_thickness_method"] = "halsey"
        method_options["jwgb_bjh_kelvin_factor_nm"] = 0.954853
        method_options["official_fit_value_usage"] = "validation_only"
        method_options["official_fit_range_usage"] = "vendor_default_from_excel_fit_tables"
        method_options["use_official_excel_fit_ranges"] = True

        info = _jwgb_info_labels(workbook)
        mapping = {
            "单点BET比表面积在P/Po为0.20000": "jwgb_single_point_bet_surface_area_m2_g",
            "BET Surface Area": "excel_bet_surface_area_m2_g",
            "Langmuir比表面分析设置": "excel_langmuir_surface_area_m2_g",
            "t-Plot Micropore Area": "excel_t_plot_micropore_area_m2_g",
            "t-Plot External Surface Area": "excel_t_plot_external_surface_area_m2_g",
            "BJH Adsorption Cumulative Surface Area": "jwgb_bjh_adsorption_cumulative_area_m2_g",
            "BJH Desorption Cumulative Surface Area": "jwgb_bjh_desorption_cumulative_area_m2_g",
            "吸附总孔体积在P/Po为0.99000": "jwgb_total_pore_volume_cm3_g",
            "t-Plot Micropore Volume": "excel_t_plot_micropore_volume_cm3_g",
            "BJH Adsorption Cumulative Volume": "jwgb_bjh_adsorption_cumulative_volume_cm3_g",
            "BJH Desorption Cumulative Volume": "jwgb_bjh_desorption_cumulative_volume_cm3_g",
            "Average Pore Diameter (4V/A)": "jwgb_average_pore_diameter_4v_a_nm",
            "BJH Adsorption Average Pore Diameter (4V/A)": "jwgb_bjh_adsorption_average_diameter_4v_a_nm",
            "BJH Adsorption Most Frequent Pore Diameter (dV/dD)": "jwgb_bjh_adsorption_mode_diameter_dv_dD_nm",
            "BJH Desorption Average Pore Diameter (4V/A)": "jwgb_bjh_desorption_average_diameter_4v_a_nm",
            "BJH Desorption Most Frequent Pore Diameter (dV/dD)": "jwgb_bjh_desorption_mode_diameter_dv_dD_nm",
        }
        for label, option_key in mapping.items():
            value = _number(info.get(label))
            if value is not None:
                method_options[option_key] = value

        for sheet_name, prefix in (
            ("BET Surface Area", "bet"),
            ("Langmuir Surface Area", "langmuir"),
            ("t-Plot ", "t_plot"),
        ):
            fit_range = _jwgb_fit_pressure_range(workbook, sheet_name)
            if fit_range is None:
                continue
            p_min, p_max, ids = fit_range
            method_options[f"stored_{prefix}_pressure_min"] = p_min
            method_options[f"stored_{prefix}_pressure_max"] = p_max
            method_options[f"excel_{prefix}_range_source"] = f"JWGB {sheet_name} ID rows"
            method_options[f"jwgb_{prefix}_fit_point_ids"] = ids

        for phase in ("adsorption", "desorption"):
            rows = _jwgb_bjh_distribution_rows(workbook, phase)
            if rows:
                method_options["official_bjh_table_usage"] = "validation_only"
                method_options[f"jwgb_bjh_{phase}_rows"] = rows

    @staticmethod
    def _add_quantachrome_text_report_values(workbook: ExcelWorkbook, method_options: dict[str, Any]) -> None:
        if not _is_quantachrome_text_workbook(workbook):
            return
        method_options["quantachrome_excel_import"] = True
        method_options["vendor_bjh_thickness_method"] = "harkins_jura"
        method_options["vendor_bjh_smooth_derivative"] = False
        method_options["quantachrome_bjh_thickness_method"] = "harkins_jura"
        method_options["quantachrome_bjh_table_source"] = "NovaWin text Excel BJH table"
        method_options["official_bjh_table_usage"] = "validation_only"
        for phase in ("adsorption", "desorption"):
            rows = _quantachrome_text_bjh_distribution_rows(workbook, phase)
            if rows:
                method_options[f"quantachrome_bjh_{phase}_rows"] = rows

    @staticmethod
    def _add_micromeritics_flex_report_values(workbook: ExcelWorkbook, method_options: dict[str, Any]) -> None:
        if not _is_micromeritics_flex_workbook(workbook):
            return
        method_options["micromeritics_flex_excel_import"] = True
        method_options["vendor_bjh_thickness_method"] = "harkins_jura"
        method_options["vendor_bjh_correction"] = "faas"
        method_options["vendor_bjh_smooth_derivative"] = False
        method_options["micromeritics_flex_bjh_source"] = "isotherm_recalculation"
        method_options["official_bjh_table_usage"] = "validation_only"

        langmuir_range = _micromeritics_flex_langmuir_range(workbook)
        if langmuir_range is not None:
            method_options["stored_langmuir_pressure_min"] = langmuir_range[0]
            method_options["stored_langmuir_pressure_max"] = langmuir_range[1]
            method_options["excel_langmuir_range_source"] = langmuir_range[2]
            method_options["official_fit_range_usage"] = "validation_only"

        t_plot_range = _micromeritics_flex_t_plot_range(workbook)
        if t_plot_range is not None:
            method_options["stored_t_plot_pressure_min"] = t_plot_range[0]
            method_options["stored_t_plot_pressure_max"] = t_plot_range[1]
            method_options["excel_t_plot_range_source"] = t_plot_range[2]
            method_options["official_fit_range_usage"] = "validation_only"


def _read_workbook(path: Path) -> ExcelWorkbook:
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path)
    raise ExcelParseError(f"Unsupported Excel extension: {path.suffix}")


def _read_xls(path: Path) -> ExcelWorkbook:
    try:
        import xlrd
    except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
        raise ExcelParseError("Reading .xls files requires xlrd. Please run: python -m pip install xlrd") from exc
    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise ExcelParseError(f"Unsupported or corrupt .xls workbook: {path}") from exc
    sheets = []
    for sheet in book.sheets():
        rows = []
        for row_index in range(sheet.nrows):
            values = []
            for column_index in range(sheet.ncols):
                cell = sheet.cell(row_index, column_index)
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    values.append(None)
                elif cell.ctype == xlrd.XL_CELL_DATE:
                    values.append(_xlrd_datetime(cell.value, book.datemode))
                else:
                    values.append(cell.value)
            rows.append(values)
        sheets.append(SheetGrid(name=sheet.name, values=rows))
    if not sheets:
        raise ExcelParseError(f"Workbook has no sheets: {path}")
    return ExcelWorkbook(sheets=sheets)


def _read_xlsx(path: Path) -> ExcelWorkbook:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:  # pragma: no cover - dependency guard
        raise ExcelParseError("Reading .xlsx/.xlsm files requires openpyxl. Please run: python -m pip install openpyxl") from exc
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            book = load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:
        raise ExcelParseError(f"Unsupported or corrupt .xlsx/.xlsm workbook: {path}") from exc
    sheets = []
    for worksheet in book.worksheets:
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(list(row))
        sheets.append(SheetGrid(name=worksheet.title, values=rows))
    if not sheets:
        raise ExcelParseError(f"Workbook has no sheets: {path}")
    return ExcelWorkbook(sheets=sheets)


def _xlrd_datetime(value: float, datemode: int) -> Any:
    try:
        import xlrd

        moment = xlrd.xldate_as_datetime(value, datemode)
    except Exception:
        return value
    if 0 <= value < 1:
        return moment.time().replace(microsecond=0)
    if moment.time() == time(0, 0):
        return moment.date()
    return moment.replace(microsecond=0)


def _collect_labels(sheets: list[SheetGrid]) -> dict[str, Any]:
    labels: dict[str, Any] = {}
    for sheet in sheets:
        for row in sheet.values:
            for column_index, value in enumerate(row):
                label = _label_key(value)
                if not _is_label(label, value):
                    continue
                label_value = _scan_label_value(row, column_index)
                if label_value in (None, ""):
                    continue
                labels.setdefault(label, label_value)
    return labels


def _add_quantachrome_text_labels(workbook: ExcelWorkbook, labels: dict[str, Any]) -> None:
    if not _is_quantachrome_text_workbook(workbook):
        return
    for text in _quantachrome_text_lines(workbook):
        sample = re.search(r"\bSample ID:\s*(.*?)\s+Filename:", text, re.IGNORECASE)
        if sample:
            labels.setdefault("sample", sample.group(1).strip())
        filename = re.search(r"\bFilename:\s*(\S+)", text, re.IGNORECASE)
        if filename:
            labels.setdefault("file", filename.group(1).strip())
        operator = re.search(r"\bOperator:\s*(\S+)", text, re.IGNORECASE)
        if operator:
            labels.setdefault("operator", operator.group(1).strip())
        sample_mass = re.search(r"\bSample weight:\s*([0-9.+\-Ee]+)\s*g\b", text, re.IGNORECASE)
        if sample_mass:
            labels.setdefault("sample mass", sample_mass.group(1))
        completed = re.search(r"\bEnd of run:\s*([0-9/: -]+?)\s+Instrument:", text, re.IGNORECASE)
        if completed:
            labels.setdefault("completed", completed.group(1).strip())
        analysis_gas = re.search(
            r"\bAnalysis gas:\s*(.*?)\s+Bath Temp:\s*([0-9.+\-Ee]+)\s*K\b",
            text,
            re.IGNORECASE,
        )
        if analysis_gas:
            labels.setdefault("analysis adsorptive", analysis_gas.group(1).strip())
            labels.setdefault("analysis bath temp", f"{analysis_gas.group(2)} K")
        properties = re.search(
            r"\bMolec\.\s*Wt\.:\s*([0-9.+\-Ee]+)\s+Cross Section:\s*([0-9.+\-Ee]+).*?"
            r"Liquid Density:\s*([0-9.+\-Ee]+)\s*g/cc",
            text,
            re.IGNORECASE,
        )
        if properties:
            molecular_weight = float(properties.group(1))
            cross_section_angstrom2 = float(properties.group(2))
            liquid_density = float(properties.group(3))
            labels.setdefault("molecular cross-sectional area", cross_section_angstrom2 / 100.0)
            if liquid_density > 0.0:
                density_factor = (molecular_weight / liquid_density) / (CM3_STP_PER_MMOL * 1000.0)
                labels.setdefault("density conversion factor", density_factor)
    if "quantachrome novawin" in "\n".join(_quantachrome_text_lines(workbook)).lower():
        labels.setdefault("adsorptive", labels.get("analysis adsorptive", "Nitrogen"))


def _add_jwgb_labels(workbook: ExcelWorkbook, labels: dict[str, Any]) -> None:
    if not _is_jwgb_workbook(workbook):
        return
    info = _jwgb_info_labels(workbook)
    label_mapping = {
        "送样单位": "submitter",
        "样品名称": "sample",
        "样品编号": "bar code",
        "吸附质": "adsorptive",
        "质量": "sample mass",
        "开始时间": "started",
        "结束时间": "completed",
        "BET Surface Area": "bet surface area",
        "Langmuir比表面分析设置": "langmuir surface area",
        "t-Plot Micropore Area": "t-plot micropore area",
        "t-Plot External Surface Area": "t-plot external surface area",
        "t-Plot Micropore Volume": "t-plot micropore volume",
    }
    for source, target in label_mapping.items():
        value = info.get(source)
        if value not in (None, ""):
            labels.setdefault(target, value)
    if "adsorptive" in labels and "analysis adsorptive" not in labels:
        labels["analysis adsorptive"] = labels["adsorptive"]


def _scan_label_value(row: list[Any], column_index: int) -> Any:
    for offset in range(1, 5):
        index = column_index + offset
        if index >= len(row):
            return None
        candidate = row[index]
        if _is_separator(candidate):
            return None
        if _is_blank(candidate):
            continue
        if _is_label(_label_key(candidate), candidate):
            return None
        return candidate
    return None


KNOWN_LABELS = {
    "adsorptive",
    "adsorption temperature",
    "analysis adsorptive",
    "analysis bath temp",
    "analysis bath temp.",
    "analysis date",
    "analysis free space",
    "analysis time",
    "automatic degas",
    "bar code",
    "cold free space",
    "comment",
    "completed",
    "data file",
    "density conversion factor",
    "equilibration interval",
    "file",
    "free space",
    "langmuir surface area",
    "molecular cross-sectional area",
    "molecular diameter",
    "operator",
    "report time",
    "sample",
    "sample density",
    "sample mass",
    "sample weight/g",
    "saturation vapor pressure",
    "started",
    "submitter",
    "t-plot external surface area",
    "t-plot micropore area",
    "t-plot micropore volume",
    "warm free space",
}


def _is_label(label: str, value: Any) -> bool:
    text = _as_clean_string(value)
    return bool(label) and (text.endswith(":") or label in KNOWN_LABELS)


def _label_key(value: Any) -> str:
    text = _as_clean_string(value).replace("：", ":")
    text = re.sub(r"\s+", " ", text).strip()
    text = text.rstrip(":").strip()
    return text.lower()


def _detect_instrument(workbook: ExcelWorkbook) -> dict[str, Any]:
    texts = list(workbook.text_values())
    joined = "\n".join(texts[:500])
    lower = joined.lower()

    def first_contains(pattern: str) -> str:
        regex = re.compile(pattern, re.IGNORECASE)
        return next((text for text in texts if regex.search(text)), "")

    if "bsd-660" in lower or "贝士德" in joined or "bsd instrument" in lower:
        model = first_contains(r"BSD-660[^\s]*") or "BSD-660"
        version = first_contains(r"V\.\d+(?:\.\d+)+(?:\s+Date\s+\d{2}\.\d{2}\.\d{2})?")
        return {
            "instrument_manufacturer": "BSD",
            "instrument_model": model,
            "instrument_software": version,
        }
    if _is_jwgb_workbook(workbook):
        return {
            "instrument_manufacturer": "JWGB",
            "instrument_model": "JWGB surface area and porosity analyzer",
            "instrument_software": "JWGB Excel export",
        }
    if "microactive for tristar ii plus" in lower:
        software = first_contains(r"MicroActive for TriStar II Plus")
        return {
            "instrument_manufacturer": "Micromeritics",
            "instrument_model": "TriStar II Plus",
            "instrument_software": software,
        }
    if re.search(r"\bflex\s+\d+(?:\.\d+)+", lower) and re.search(r"\b3500\b", lower):
        software = first_contains(r"\bFlex\s+\d+(?:\.\d+)+")
        return {
            "instrument_manufacturer": "Micromeritics",
            "instrument_model": "3Flex 3500",
            "instrument_software": software,
        }
    if "tristar ii 3020" in lower:
        software = first_contains(r"TriStar II 3020")
        return {
            "instrument_manufacturer": "Micromeritics",
            "instrument_model": "TriStar II 3020",
            "instrument_software": software,
        }
    if "asap 2460" in lower:
        software = first_contains(r"ASAP 2460")
        return {
            "instrument_manufacturer": "Micromeritics",
            "instrument_model": "ASAP 2460",
            "instrument_software": software,
        }
    if "asap 2020 plus" in lower:
        software = first_contains(r"ASAP 2020 Plus")
        return {
            "instrument_manufacturer": "Micromeritics",
            "instrument_model": "ASAP 2020 Plus",
            "instrument_software": software,
        }
    if "belmaster" in lower:
        software = first_contains(r"BELMaster")
        return {
            "instrument_manufacturer": "MicrotracBEL",
            "instrument_model": "BELMaster",
            "instrument_software": software,
        }
    if "quantachrome" in lower or "autosorb" in lower or "novawin" in lower or "quadrasorb" in lower:
        version = first_contains(r"version\s+\d+(?:\.\d+)+")
        software = f"NovaWin {version}".strip() if version else first_contains(r"Quantachrome|Autosorb|NovaWin")
        model = "QuadraSorb" if "quadrasorb" in lower else "Autosorb"
        return {
            "instrument_manufacturer": "Quantachrome",
            "instrument_model": model,
            "instrument_software": software,
        }
    return {
        "instrument_manufacturer": "",
        "instrument_model": "",
        "instrument_software": "",
    }


def _first_label(labels: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = labels.get(key.lower())
        if value not in (None, ""):
            return value
    return None


def _target_pressure_table_from_isotherm(isotherm: list[IsothermPoint]) -> list[TargetPressureRow]:
    previous_by_phase: dict[str, float] = {}
    rows = []
    for row_index, point in enumerate(isotherm, start=1):
        pressure = float(point.relative_pressure)
        previous = previous_by_phase.get(point.phase, 0.0)
        rows.append(
            TargetPressureRow(
                row=row_index,
                branch=point.phase,
                starting_pressure_p_po=previous,
                ending_pressure_p_po=pressure,
                pressure_increment_p_po=pressure - previous,
                ending_pressure_rel_offset=0,
            )
        )
        previous_by_phase[point.phase] = pressure
    return rows


def _is_bsd_workbook(workbook: ExcelWorkbook) -> bool:
    for text in workbook.text_values():
        lower = text.lower()
        if "bsd-660" in lower or "bsd instrument" in lower or "贝士德" in text:
            return True
    return False


def _is_jwgb_workbook(workbook: ExcelWorkbook) -> bool:
    names = {sheet.name.strip().lower() for sheet in workbook.sheets}
    required = {
        "info",
        "isotherm",
        "bet surface area",
        "langmuir surface area",
        "t-plot",
        "bjh adsorption",
        "bjh desorption",
    }
    if not required.issubset(names):
        return False
    info = _jwgb_info_labels(workbook)
    return all(key in info for key in ("样品名称", "吸附质", "质量", "BET Surface Area"))


def _jwgb_info_labels(workbook: ExcelWorkbook) -> dict[str, Any]:
    sheet = next((item for item in workbook.sheets if item.name.strip().lower() == "info"), None)
    if sheet is None:
        return {}
    labels: dict[str, Any] = {}
    for row in sheet.values:
        if not row:
            continue
        label = _as_clean_string(row[0])
        if not label:
            continue
        labels[label] = row[1] if len(row) > 1 else None
    return labels


def _jwgb_adsorption_pressure_by_id(workbook: ExcelWorkbook) -> dict[int, float]:
    sheet = next((item for item in workbook.sheets if item.name.strip().lower() == "isotherm"), None)
    if sheet is None:
        return {}
    table = OfficialExcelParser._find_jwgb_isotherm_table(sheet)
    if table is None:
        return {}
    header_row, columns = table
    pressures: dict[int, float] = {}
    phase = "adsorption"
    for row_index in range(header_row + 1, sheet.nrows):
        point_id = _number(sheet.cell(row_index, columns["id"]))
        pressure = _number(sheet.cell(row_index, columns["relative_pressure"]))
        quantity = _number(sheet.cell(row_index, columns["quantity"]))
        if point_id is None or pressure is None or quantity is None:
            if pressures and phase == "adsorption":
                phase = "desorption"
            continue
        if phase == "adsorption" and 0.0 < float(pressure) < 1.0:
            pressures[int(point_id)] = float(pressure)
    return pressures


def _jwgb_fit_pressure_range(workbook: ExcelWorkbook, sheet_name: str) -> tuple[float, float, list[int]] | None:
    sheet = next((item for item in workbook.sheets if item.name.strip().lower() == sheet_name.strip().lower()), None)
    if sheet is None:
        return None
    pressure_by_id = _jwgb_adsorption_pressure_by_id(workbook)
    ids: list[int] = []
    fallback_pressures: list[float] = []
    for row_index in range(1, sheet.nrows):
        point_id = _number(sheet.cell(row_index, 0))
        pressure = _number(sheet.cell(row_index, 1))
        quantity = _number(sheet.cell(row_index, 2))
        if point_id is None or pressure is None or quantity is None:
            continue
        ids.append(int(point_id))
        if 0.0 < float(pressure) < 1.0:
            fallback_pressures.append(float(pressure))
    pressures = [pressure_by_id[point_id] for point_id in ids if point_id in pressure_by_id]
    if len(pressures) < 3:
        pressures = fallback_pressures
    if len(pressures) < 3:
        return None
    return min(pressures), max(pressures), ids


def _jwgb_bjh_distribution_rows(workbook: ExcelWorkbook, phase: str) -> list[dict[str, float]]:
    sheet_name = "BJH Adsorption" if phase == "adsorption" else "BJH Desorption"
    sheet = next((item for item in workbook.sheets if item.name.strip().lower() == sheet_name.lower()), None)
    if sheet is None:
        return []
    rows: list[dict[str, float]] = []
    for row_index in range(1, sheet.nrows):
        point_id = _number(sheet.cell(row_index, 0))
        diameter = _number(sheet.cell(row_index, 2))
        incremental_volume = _number(sheet.cell(row_index, 3))
        cumulative_volume = _number(sheet.cell(row_index, 4))
        differential_per_nm = _number(sheet.cell(row_index, 5))
        differential_log = _number(sheet.cell(row_index, 6))
        if diameter is None or incremental_volume is None or cumulative_volume is None:
            continue
        diameter_range = _diameter_range_from_text(sheet.cell(row_index, 1))
        row: dict[str, float] = {
            "phase": phase,
            "point_id": float(point_id) if point_id is not None else float(len(rows) + 1),
            "pore_diameter_nm": float(diameter),
            "incremental_pore_volume_cm3_g": float(incremental_volume),
            "cumulative_pore_volume_cm3_g": float(cumulative_volume),
        }
        if diameter_range is not None:
            row["pore_diameter_range_high_nm"] = diameter_range[0]
            row["pore_diameter_range_low_nm"] = diameter_range[1]
        if differential_per_nm is not None:
            row["differential_pore_volume_per_nm_cm3_g_nm"] = float(differential_per_nm)
        if differential_log is not None:
            row["differential_pore_volume_cm3_g"] = float(differential_log)
        rows.append(row)
    return rows


def _is_quantachrome_text_workbook(workbook: ExcelWorkbook) -> bool:
    names = {sheet.name.lower() for sheet in workbook.sheets}
    if not {"isotherm", "bjh-adsorption", "bjh-desorption"}.issubset(names):
        return False
    joined = "\n".join(_quantachrome_text_lines(workbook)[:30]).lower()
    return "quantachrome novawin" in joined or "for nova instruments" in joined


def _quantachrome_text_lines(workbook: ExcelWorkbook) -> list[str]:
    lines: list[str] = []
    for sheet in workbook.sheets:
        for row in sheet.values:
            if not row:
                continue
            text = _as_clean_string(row[0])
            if text:
                lines.append(text)
    return lines


def _quantachrome_text_numeric_rows_after_units(
    sheet: SheetGrid,
    *,
    expected_columns: int,
) -> list[list[float]]:
    start_row = None
    for row_index in range(sheet.nrows):
        text = _as_clean_string(sheet.cell(row_index, 0)).strip().lower()
        if text == "cc/g" or text.startswith("nm "):
            start_row = row_index + 1
            break
    if start_row is None:
        return []

    rows: list[list[float]] = []
    for row_index in range(start_row, sheet.nrows):
        text = _as_clean_string(sheet.cell(row_index, 0))
        numbers = _numbers_from_text(text)
        if len(numbers) < expected_columns:
            continue
        rows.append(numbers[:expected_columns])
    return rows


def _quantachrome_text_bjh_distribution_rows(workbook: ExcelWorkbook, phase: str) -> list[dict[str, float]]:
    sheet_name = "BJH-adsorption" if phase == "adsorption" else "BJH-desorption"
    sheet = next((item for item in workbook.sheets if item.name.lower() == sheet_name.lower()), None)
    if sheet is None:
        return []
    table_rows = _quantachrome_text_numeric_rows_after_units(sheet, expected_columns=7)
    rows: list[dict[str, float]] = []
    for values in table_rows:
        diameter, cumulative_volume, cumulative_area, dv_d, ds_d, dv_log, ds_log = values[:7]
        if diameter <= 0.0 or cumulative_volume < 0.0:
            continue
        rows.append(
            {
                "phase": phase,
                "pore_diameter_nm": float(diameter),
                "cumulative_pore_volume_cm3_g": float(cumulative_volume),
                "cumulative_pore_area_m2_g": float(cumulative_area),
                "differential_pore_volume_per_nm_cm3_g_nm": float(dv_d),
                "differential_pore_area_per_nm_m2_g_nm": float(ds_d),
                "differential_pore_volume_cm3_g": float(dv_log),
                "differential_pore_area_m2_g": float(ds_log),
            }
        )
    return rows


def _is_micromeritics_flex_workbook(workbook: ExcelWorkbook) -> bool:
    has_flex = False
    has_3500 = False
    for text in workbook.text_values():
        clean = _as_clean_string(text)
        lower = clean.lower()
        if re.search(r"\bflex\s+\d+(?:\.\d+)+", lower):
            has_flex = True
        if clean == "3500" or "3flex" in lower:
            has_3500 = True
        if has_flex and has_3500:
            return True
    return False


def _micromeritics_flex_saturation_kpa(workbook: ExcelWorkbook) -> float | None:
    sheet = next((item for item in workbook.sheets if item.name == "Isotherm Tabular Report"), None)
    if sheet is None:
        return None
    for row_index in range(sheet.nrows):
        for column_index in range(sheet.ncols):
            header = _normalize_header(sheet.cell(row_index, column_index)).replace(" ", "")
            if "saturationpressure" not in header or "kpa" not in header:
                continue
            for data_row in range(row_index + 1, min(sheet.nrows, row_index + 8)):
                value = _number(sheet.cell(data_row, column_index))
                if value is not None and value > 0.0:
                    return float(value)
    return None


def _micromeritics_flex_langmuir_range(workbook: ExcelWorkbook) -> tuple[float, float, str] | None:
    sheet = next((item for item in workbook.sheets if item.name == "Langmuir Tabular Report"), None)
    if sheet is None:
        return None
    saturation_kpa = _micromeritics_flex_saturation_kpa(workbook) or 101.325
    for row_index in range(sheet.nrows):
        headers = [_normalize_header(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
        pressure_col = None
        quantity_col = None
        for column_index, header in enumerate(headers):
            compact = header.replace(" ", "")
            if "pressure(kpa)" in compact:
                pressure_col = column_index
            elif "quantityadsorbed" in compact:
                quantity_col = column_index
        if pressure_col is None or quantity_col is None:
            continue
        pressures: list[float] = []
        blank_streak = 0
        for data_row in range(row_index + 1, sheet.nrows):
            pressure_kpa = _number(sheet.cell(data_row, pressure_col))
            quantity = _number(sheet.cell(data_row, quantity_col))
            if pressure_kpa is None or quantity is None:
                if pressures:
                    blank_streak += 1
                    if blank_streak >= 2:
                        break
                continue
            relative = float(pressure_kpa) / float(saturation_kpa)
            if 0.0 < relative <= 1.1:
                pressures.append(relative)
                blank_streak = 0
            elif pressures:
                break
        if len(pressures) >= 3:
            return min(pressures), max(pressures), f"Micromeritics Flex Langmuir report:{sheet.name}"
    return None


def _micromeritics_flex_t_plot_range(workbook: ExcelWorkbook) -> tuple[float, float, str] | None:
    sheet = next((item for item in workbook.sheets if item.name == "t-Plot Tabular Report"), None)
    if sheet is None:
        return None
    for row_index in range(sheet.nrows):
        headers = [_normalize_header(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
        rel_col = None
        fitted_col = None
        for column_index, header in enumerate(headers):
            if _is_relative_pressure_header(header):
                rel_col = column_index
            elif header == "fitted":
                fitted_col = column_index
        if rel_col is None or fitted_col is None:
            continue
        pressures: list[float] = []
        for data_row in range(row_index + 1, sheet.nrows):
            pressure = _number(sheet.cell(data_row, rel_col))
            fitted = _as_clean_string(sheet.cell(data_row, fitted_col))
            if pressure is None:
                if pressures:
                    break
                continue
            if fitted == "*":
                pressures.append(float(pressure))
        if len(pressures) >= 3:
            return min(pressures), max(pressures), f"Micromeritics Flex t-Plot fitted rows:{sheet.name}"
    return None


def _micromeritics_flex_bjh_distribution_rows(workbook: ExcelWorkbook, phase: str) -> list[dict[str, float]]:
    phase_title = "adsorption" if phase == "adsorption" else "desorption"
    for sheet in workbook.sheets:
        title_found = False
        for row in sheet.values:
            for value in row:
                text = _as_clean_string(value).lower()
                if f"bjh {phase_title} pore distribution report" in text:
                    title_found = True
                    break
            if title_found:
                break
        if not title_found:
            continue

        header_row = None
        columns: dict[str, int] = {}
        for row_index in range(sheet.nrows):
            headers = [_normalize_header(sheet.cell(row_index, column_index)) for column_index in range(sheet.ncols)]
            for column_index, header in enumerate(headers):
                compact = header.replace(" ", "")
                if "porediameterrange" in compact or "porewidthrange" in compact:
                    columns["range"] = column_index
                elif "averagediameter" in compact or "averagewidth" in compact:
                    columns["average"] = column_index
                elif "incrementalporevolume" in compact:
                    columns["incremental_volume"] = column_index
                elif "cumulativeporevolume" in compact:
                    columns["cumulative_volume"] = column_index
                elif "incrementalporearea" in compact:
                    columns["incremental_area"] = column_index
                elif "cumulativeporearea" in compact:
                    columns["cumulative_area"] = column_index
            if {"range", "average", "incremental_volume", "cumulative_volume"}.issubset(columns):
                header_row = row_index
                break
        if header_row is None:
            continue

        rows: list[dict[str, float]] = []
        blank_streak = 0
        for row_index in range(header_row + 1, sheet.nrows):
            average = _number(sheet.cell(row_index, columns["average"]))
            incremental_volume = _number(sheet.cell(row_index, columns["incremental_volume"]))
            cumulative_volume = _number(sheet.cell(row_index, columns["cumulative_volume"]))
            diameter_range = _diameter_range_from_text(sheet.cell(row_index, columns["range"]))
            if average is None or incremental_volume is None or cumulative_volume is None or diameter_range is None:
                if rows:
                    blank_streak += 1
                    if blank_streak >= 2:
                        break
                continue
            blank_streak = 0
            high, low = diameter_range
            row: dict[str, float] = {
                "phase": phase,
                "pore_diameter_range_high_nm": high,
                "pore_diameter_range_low_nm": low,
                "pore_diameter_nm": float(average),
                "incremental_pore_volume_cm3_g": float(incremental_volume),
                "cumulative_pore_volume_cm3_g": float(cumulative_volume),
            }
            incremental_area_col = columns.get("incremental_area")
            cumulative_area_col = columns.get("cumulative_area")
            incremental_area = _number(sheet.cell(row_index, incremental_area_col)) if incremental_area_col is not None else None
            cumulative_area = _number(sheet.cell(row_index, cumulative_area_col)) if cumulative_area_col is not None else None
            if incremental_area is not None:
                row["incremental_pore_area_m2_g"] = float(incremental_area)
            if cumulative_area is not None:
                row["cumulative_pore_area_m2_g"] = float(cumulative_area)
            rows.append(row)
        if rows:
            return rows
    return []


def _diameter_range_from_text(value: Any) -> tuple[float, float] | None:
    text = _as_clean_string(value)
    numbers = re.findall(r"[-+]?(?:(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?", text)
    if len(numbers) < 2:
        return None
    try:
        left = float(numbers[0].replace(",", ""))
        right = float(numbers[1].replace(",", ""))
    except ValueError:
        return None
    return max(left, right), min(left, right)


def _bsd_bet_pressure_range(workbook: ExcelWorkbook) -> tuple[float, float, str] | None:
    for sheet in workbook.sheets:
        if "BET" not in sheet.name:
            continue
        values = _bsd_numeric_table_column(sheet, "p/p0")
        if len(values) >= 3:
            return min(values), max(values), f"BSD BET data table:{sheet.name}"
        label_range = _bsd_range_from_label(sheet, "P/P0取点范围")
        if label_range is not None:
            return label_range[0], label_range[1], f"BSD BET range label:{sheet.name}"
    return None


def _bsd_numeric_table_column(sheet: SheetGrid, header_text: str) -> list[float]:
    target = _normalize_header(header_text).replace(" ", "")
    for row_index in range(sheet.nrows):
        for column_index in range(sheet.ncols):
            header = _normalize_header(sheet.cell(row_index, column_index)).replace(" ", "")
            if header != target:
                continue
            values: list[float] = []
            blank_streak = 0
            for data_row in range(row_index + 1, sheet.nrows):
                value = _number(sheet.cell(data_row, column_index))
                if value is None:
                    if values:
                        blank_streak += 1
                        if blank_streak >= 2:
                            return values
                    continue
                if 0.0 < value < 1.0:
                    values.append(float(value))
                    blank_streak = 0
                elif values:
                    return values
            if values:
                return values
    return []


def _bsd_range_from_label(sheet: SheetGrid, label_text: str) -> tuple[float, float] | None:
    for row_index, row in enumerate(sheet.values):
        for column_index, value in enumerate(row):
            if label_text not in _as_clean_string(value):
                continue
            for offset in range(1, 7):
                raw = sheet.cell(row_index, column_index + offset)
                parsed = _range_from_text(raw)
                if parsed is not None:
                    return parsed
    return None


def _bsd_bjh_table_rows(sheet: SheetGrid, phase: str) -> list[dict[str, float]]:
    start: tuple[int, int] | None = None
    for row_index in range(sheet.nrows):
        for column_index in range(max(0, sheet.ncols - 8)):
            header = _normalize_header(sheet.cell(row_index, column_index)).replace(" ", "")
            next_header = _as_clean_string(sheet.cell(row_index, column_index + 1)).replace(" ", "")
            next_two = _as_clean_string(sheet.cell(row_index, column_index + 2)).replace(" ", "")
            if header in {"p/p0", "p/po"} and next_header == "dnm" and next_two == "Dnm":
                start = (row_index, column_index)
                break
        if start is not None:
            break
    if start is None:
        return []

    header_row, start_col = start
    raw_rows: list[dict[str, float]] = []
    blank_streak = 0
    for row_index in range(header_row + 1, sheet.nrows):
        values = [_number(sheet.cell(row_index, start_col + offset)) for offset in range(9)]
        if any(value is None for value in values):
            if raw_rows:
                blank_streak += 1
                if blank_streak >= 2:
                    break
            continue
        blank_streak = 0
        (
            relative_pressure,
            kelvin_pore_diameter,
            pore_diameter,
            cumulative_volume,
            cumulative_area,
            differential_volume_per_nm,
            differential_area_per_nm,
            differential_volume,
            differential_area,
        ) = [float(value) for value in values]
        raw_rows.append(
            {
                "phase": phase,
                "relative_pressure": relative_pressure,
                "kelvin_pore_diameter_nm": kelvin_pore_diameter,
                "pore_diameter_nm": pore_diameter,
                "cumulative_pore_volume_cm3_g": cumulative_volume,
                "cumulative_pore_area_m2_g": cumulative_area,
                "differential_pore_volume_per_nm_cm3_g_nm": differential_volume_per_nm,
                "differential_pore_area_per_nm_m2_g_nm": differential_area_per_nm,
                "differential_pore_volume_cm3_g": differential_volume,
                "differential_pore_area_m2_g": differential_area,
            }
        )
    return raw_rows


def _update_from_label_pairs(sheet: SheetGrid, options: dict[str, Any], mapping: dict[str, str]) -> None:
    for row_index, row in enumerate(sheet.values):
        for column_index, value in enumerate(row):
            text = _as_clean_string(value)
            if not text:
                continue
            for label, option_key in mapping.items():
                if label not in text or option_key in options:
                    continue
                candidate = _next_numeric_value(sheet, row_index, column_index)
                if candidate is not None:
                    options[option_key] = candidate


def _next_numeric_value(sheet: SheetGrid, row_index: int, column_index: int) -> float | None:
    for offset in range(1, 8):
        value = _number(sheet.cell(row_index, column_index + offset))
        if value is not None:
            return value
    return None


def _bet_fit_pressure_range(workbook: ExcelWorkbook) -> tuple[float, float, str] | None:
    for sheet in workbook.sheets:
        candidate = _find_bet_transform_table(sheet)
        if candidate is not None:
            return candidate
    return None


def _find_bet_transform_table(sheet: SheetGrid) -> tuple[float, float, str] | None:
    for row_index in range(sheet.nrows):
        for column_index in range(sheet.ncols - 1):
            current = _normalize_header(sheet.cell(row_index, column_index))
            next_one = _normalize_header(sheet.cell(row_index, column_index + 1))
            next_two = _normalize_header(sheet.cell(row_index, column_index + 2))
            if not _is_relative_pressure_header(current):
                continue
            if _is_bet_transform_header(next_one):
                pressures = _numeric_pressure_column(sheet, row_index + 1, column_index)
                if len(pressures) >= 3:
                    return min(pressures), max(pressures), f"BET Surface Area Plot:{sheet.name}"
            if "quantity adsorbed" in next_one and _is_bet_transform_header(next_two):
                pressures = _numeric_pressure_column(sheet, row_index + 1, column_index)
                if len(pressures) >= 3:
                    return min(pressures), max(pressures), f"BET Surface Area Report:{sheet.name}"
    return None


def _numeric_pressure_column(sheet: SheetGrid, start_row: int, column_index: int) -> list[float]:
    values: list[float] = []
    blank_streak = 0
    for row_index in range(start_row, sheet.nrows):
        pressure = _number(sheet.cell(row_index, column_index))
        if pressure is None:
            if values:
                blank_streak += 1
                if blank_streak >= 2:
                    break
            continue
        if 0.0 < pressure < 1.0:
            values.append(float(pressure))
            blank_streak = 0
            continue
        if values:
            break
    return values


def _is_relative_pressure_header(value: str) -> bool:
    compact = value.replace(" ", "").replace("°", "0").replace("po", "p0")
    return "relative pressure" in value or compact in {"p/p0"}


def _is_bet_transform_header(value: str) -> bool:
    compact = value.replace(" ", "").replace("°", "0").replace("po", "p0")
    return "1/[q(p0/p-1)]" in compact or "1/[q(p0/p-1)]" in compact.replace("−", "-")


def _normalize_header(value: Any) -> str:
    return _as_clean_string(value).lower().replace("³", "3").replace("²", "2")


def _as_clean_string(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date) and not isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _text(value: Any) -> str:
    return _as_clean_string(value)


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None
    text = _as_clean_string(value)
    if not text:
        return None
    match = re.search(r"[-+]?(?:(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?", text)
    if not match:
        return None
    try:
        number = float(match.group(0).replace(",", ""))
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def _numbers_from_text(value: Any) -> list[float]:
    text = _as_clean_string(value)
    if not text:
        return []
    numbers: list[float] = []
    for match in re.findall(r"[-+]?(?:(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?", text):
        try:
            number = float(match.replace(",", ""))
        except ValueError:
            continue
        if math.isfinite(number):
            numbers.append(number)
    return numbers


def _range_from_text(value: Any) -> tuple[float, float] | None:
    text = _as_clean_string(value)
    if not text:
        return None
    numbers = re.findall(r"[-+]?(?:(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d*)?|\.\d+)(?:[Ee][-+]?\d+)?", text)
    if len(numbers) < 2:
        return None
    try:
        left = float(numbers[0].replace(",", ""))
        right = float(numbers[1].replace(",", ""))
    except ValueError:
        return None
    return (min(left, right), max(left, right))


def _adsorptive_short(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    first = re.split(r"[,，;；\s]+", text, maxsplit=1)[0].strip()
    return first or text


def _adsorptive_temperature_k(value: str) -> float | None:
    text = value.strip()
    if not text:
        return None
    match = re.search(r"([-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*[kK]\b", text)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            return None
    match = re.search(r"([-+]?(?:\d+(?:\.\d*)?|\.\d+))\s*(?:℃|°C|°c)\b", text)
    if match:
        try:
            return float(match.group(1)) + 273.15
        except ValueError:
            return None
    return None


def _temperature_k(value: Any) -> float | None:
    number = _number(value)
    if number is None:
        return None
    text = _as_clean_string(value).lower()
    if "°c" in text or re.search(r"\bc\b", text):
        return number + 273.15
    if number < 0.0:
        return number + 273.15
    return number


def _elapsed_display_seconds(value: Any) -> int | None:
    text = _as_clean_string(value)
    match = re.match(r"^\s*(\d+):([0-5]\d)(?::([0-5]\d))?\s*$", text)
    if not match:
        return None
    first = int(match.group(1))
    second = int(match.group(2))
    third = int(match.group(3) or 0)
    if match.group(3) is None:
        return first * 60 + second
    return first * 3600 + second * 60 + third


def _timestamp_pair(value: Any) -> tuple[int, str]:
    moment = _as_datetime(value)
    if moment is None:
        return 0, _as_clean_string(value)
    try:
        raw = int(moment.timestamp())
    except (OverflowError, OSError, ValueError):
        raw = 0
    return raw, moment.strftime("%Y-%m-%d %H:%M:%S")


def _as_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value.replace(microsecond=0)
    if isinstance(value, date) and not isinstance(value, datetime):
        return datetime.combine(value, time())
    text = _as_clean_string(value)
    if not text:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M:%S",
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _combine_date_time(date_value: Any, time_value: Any) -> datetime | None:
    date_part = _as_datetime(date_value)
    if date_part is None:
        return None
    if isinstance(time_value, time):
        return datetime.combine(date_part.date(), time_value)
    time_part = _as_datetime(time_value)
    if time_part is not None:
        return datetime.combine(date_part.date(), time_part.time())
    text = _as_clean_string(time_value)
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, fmt).time()
        except ValueError:
            continue
        return datetime.combine(date_part.date(), parsed)
    return date_part


def _duration_text(start_raw: int, end_raw: int) -> tuple[str, int]:
    if not start_raw or not end_raw or end_raw < start_raw:
        return "", 0
    total = int(end_raw - start_raw)
    hours, remainder = divmod(total, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}", total


def _file_modified_timestamp(path: Path) -> tuple[int, str]:
    try:
        raw = int(path.stat().st_mtime)
    except OSError:
        return 0, ""
    try:
        return raw, datetime.fromtimestamp(raw).strftime("%Y-%m-%d %H:%M:%S")
    except (OSError, OverflowError, ValueError):
        return raw, ""


def _is_blank(value: Any) -> bool:
    return value is None or _as_clean_string(value) == ""


def _is_separator(value: Any) -> bool:
    return _as_clean_string(value) == "|"
