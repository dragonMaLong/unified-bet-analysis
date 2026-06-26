from __future__ import annotations

import copy
import os
import math
import sys
import threading
from datetime import datetime
from pathlib import Path
from typing import Iterable

os.environ.setdefault("PYQTGRAPH_QT_LIB", "PyQt5")

import numpy as np
import pyqtgraph as pg
from pyqtgraph.Qt import QtCore, QtGui, QtWidgets

from tristar_bet import BELMasterParseError, ExcelParseError, TriStarParseError, load_file
from tristar_bet.analysis import (
    DEFAULT_THICKNESS_METHOD,
    DFT_DEFAULT_ANALYSIS_TYPE,
    DFT_DEFAULT_GEOMETRY,
    DFT_DEFAULT_MODEL,
    DFT_DEFAULT_REGULARIZATION,
    DFT_REGULARIZATION_VALUES,
    HK_ADSORBENT_PRESETS,
    HK_ADSORPTIVE_PRESETS,
    HK_DEFAULT_ADSORBENT,
    HK_DEFAULT_ADSORPTIVE,
    HK_DEFAULT_GEOMETRY,
    HK_DEFAULT_INTERACTION_PARAMETER_ERG_CM4,
    THICKNESS_METHOD_DEFAULT_PARAMS,
    analysis_bundle,
    automatic_bet_range,
    automatic_langmuir_range,
    automatic_t_plot_pressure_range,
    bet_analysis,
    bjh_pore_distribution,
    calculate_hk_interaction_parameter,
    density_conversion_factor,
    dft_pore_distribution,
    dh_pore_distribution,
    horvath_kawazoe_pore_distribution,
    langmuir_analysis,
    t_plot_analysis,
    t_plot_analysis_by_thickness,
    thickness_nm,
)
from tristar_bet.dft_models import dft_model_options
from tristar_bet.reference_thickness import (
    DEFAULT_REFERENCE_DIR,
    normalize_reference_points,
    read_reference_points,
    write_reference_points,
)
from tristar_bet.update_checker import DEFAULT_UPDATE_REPOSITORY, UpdateInfo, check_for_update
from tristar_bet.updater import UpdateDownloadError, download_update, launch_update_and_exit
from tristar_bet.ui.plots import (
    ACTIVE_LINE_WIDTH,
    BJH_DIFFERENTIAL_LOG,
    BJH_DISPLAY_METRIC_ORDER,
    DEFAULT_COLORS,
    HK_CUMULATIVE_VOLUME,
    HK_DIFFERENTIAL_LINEAR,
    HK_DISPLAY_METRIC_ORDER,
    bjh_display_axis_label,
    bjh_display_metric_label,
    hk_display_axis_label,
    hk_display_metric_label,
    link_sample_curve_hover_plots,
    make_plot,
    normalize_hk_display_metric,
    plot_bet_multi,
    plot_bet_selection,
    plot_bjh_selection,
    plot_dh_distribution_multi,
    plot_dh_distribution_placeholder,
    plot_dft_diagnostics,
    plot_dft_distribution_multi,
    plot_dft_selection,
    plot_hk_distribution_multi,
    plot_hk_distribution_placeholder,
    plot_hk_selection,
    plot_isotherm_multi,
    plot_isotherm_selection,
    plot_bjh_distribution_multi,
    plot_langmuir_points_multi,
    plot_langmuir_selection,
    plot_pore_distribution_placeholder,
    plot_t_plot_points_multi,
    plot_t_plot_selection,
    normalize_bjh_display_metrics,
    replace_bet_fit_line,
    replace_langmuir_fit_line,
    replace_t_plot_fit_line,
)
from tristar_bet.version import __version__


APP_NAME = "BET 综合分析-DragonScience"
APP_VERSION = __version__
APP_ICON_PATH = Path(__file__).resolve().parent.parent / "assets" / "BET-logo.png"
UPDATE_REPOSITORY = DEFAULT_UPDATE_REPOSITORY
AUTO_UPDATE_CHECK_DELAY_MS = 3000
FIT_ANALYSIS_CACHE_LIMIT = 2048
BJH_DISTRIBUTION_CACHE_LIMIT = 1024
Signal = getattr(QtCore, "Signal", None) or getattr(QtCore, "pyqtSignal")


def _qt_enum_int(value) -> int:
    try:
        return int(value)
    except TypeError:
        return int(value.value)


def _make_update_available_icon(size: int = 28) -> QtGui.QIcon:
    pixmap = QtGui.QPixmap(size, size)
    pixmap.fill(QtCore.Qt.transparent)
    painter = QtGui.QPainter(pixmap)
    painter.setRenderHint(QtGui.QPainter.Antialiasing, True)
    scale = size / 28.0

    def s(value: float) -> float:
        return float(value) * scale

    cloud = QtGui.QPainterPath()
    cloud.addEllipse(QtCore.QRectF(s(3.0), s(12.2), s(11.5), s(10.0)))
    cloud.addEllipse(QtCore.QRectF(s(8.2), s(5.2), s(14.8), s(15.0)))
    cloud.addEllipse(QtCore.QRectF(s(17.0), s(10.2), s(9.0), s(9.8)))
    cloud.addRoundedRect(QtCore.QRectF(s(4.2), s(16.2), s(21.5), s(8.0)), s(4.0), s(4.0))
    painter.setPen(QtCore.Qt.NoPen)
    painter.setBrush(QtGui.QColor("#2563eb"))
    painter.drawPath(cloud.simplified())

    arrow = QtGui.QPainterPath()
    arrow.addRoundedRect(QtCore.QRectF(s(13.0), s(10.5), s(3.0), s(8.5)), s(1.4), s(1.4))
    arrow.moveTo(s(8.9), s(17.0))
    arrow.lineTo(s(14.5), s(22.5))
    arrow.lineTo(s(20.1), s(17.0))
    arrow.closeSubpath()
    painter.setBrush(QtGui.QColor("#ffffff"))
    painter.drawPath(arrow)
    painter.end()
    return QtGui.QIcon(pixmap)


class UpdateCheckWorker(QtCore.QObject):
    finished = Signal(object, bool)
    failed = Signal(str, bool)

    def __init__(self, current_version: str, repository: str, manual: bool) -> None:
        super().__init__()
        self.current_version = current_version
        self.repository = repository
        self.manual = manual

    def run(self) -> None:
        try:
            info = check_for_update(self.current_version, repository=self.repository)
        except Exception as exc:
            self.failed.emit(str(exc), self.manual)
            return
        self.finished.emit(info, self.manual)


class UpdateDownloadWorker(QtCore.QObject):
    progress = Signal(int, int)
    finished = Signal(object, str)
    failed = Signal(str)

    def __init__(self, info: UpdateInfo) -> None:
        super().__init__()
        self.info = info

    def run(self) -> None:
        try:
            path = download_update(
                self.info,
                progress_callback=lambda downloaded, total: self.progress.emit(downloaded, total),
            )
        except Exception as exc:
            self.failed.emit(str(exc))
            return
        self.finished.emit(self.info, str(path))


class SelectAllCheckBox(QtWidgets.QCheckBox):
    def nextCheckState(self) -> None:
        if self.checkState() == QtCore.Qt.Checked:
            self.setCheckState(QtCore.Qt.Unchecked)
        else:
            self.setCheckState(QtCore.Qt.Checked)


class SampleTableWidget(QtWidgets.QTableWidget):
    rowMoveRequested = Signal(int, int)
    smpFilesDropped = Signal(list)
    LONG_PRESS_MS = 220
    FROZEN_COLUMN_COUNT = 2

    def __init__(self, *args, **kwargs) -> None:
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)
        self.viewport().setAcceptDrops(True)
        self._syncing_frozen_columns = False
        self._drag_source_row = -1
        self._drag_start_pos = QtCore.QPoint()
        self._drag_timer = QtCore.QElapsedTimer()
        self._dragging_row = False
        self._drop_indicator = QtWidgets.QFrame(self.viewport())
        self._drop_indicator.setFixedHeight(2)
        self._drop_indicator.setStyleSheet("background: #2563eb;")
        self._drop_indicator.hide()
        self._init_frozen_columns()

    def frozen_header(self):
        return self._frozen_table.horizontalHeader()

    def _init_frozen_columns(self) -> None:
        self._frozen_table = QtWidgets.QTableView(self)
        self._frozen_table.setModel(self.model())
        self._frozen_table.setSelectionModel(self.selectionModel())
        self._frozen_table.setAcceptDrops(True)
        self._frozen_table.viewport().setAcceptDrops(True)
        self._frozen_table.setFocusPolicy(QtCore.Qt.NoFocus)
        self._frozen_table.setFrameShape(QtWidgets.QFrame.NoFrame)
        self._frozen_table.setShowGrid(False)
        self._frozen_table.setAlternatingRowColors(False)
        self._frozen_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self._frozen_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self._frozen_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self._frozen_table.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self._frozen_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self._frozen_table.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)
        self._frozen_table.setStyleSheet(
            """
            QTableView {
                border: 0;
                background: #ffffff;
                alternate-background-color: #ffffff;
            }
            QTableView::item:selected {
                background: #e0ecff;
            }
            QTableView::item:focus {
                outline: none;
            }
            QTableView::indicator {
                width: 11px;
                height: 11px;
                border-radius: 6px;
                border: 1px solid #6b7280;
                background: white;
            }
            QTableView::indicator:checked {
                border: 1px solid #2563eb;
                background: #2563eb;
            }
            QHeaderView::section {
                background: #f9fafb;
                border: 0;
                border-right: 1px solid #d1d5db;
                border-bottom: 1px solid #d1d5db;
                color: #374151;
                font-weight: 600;
                padding: 4px 8px 4px 6px;
            }
            """
        )
        self._frozen_table.verticalHeader().hide()
        self._frozen_table.verticalHeader().setDefaultSectionSize(self.verticalHeader().defaultSectionSize())
        frozen_header = self._frozen_table.horizontalHeader()
        frozen_header.setSectionsMovable(False)
        frozen_header.setHighlightSections(False)
        frozen_header.setDefaultAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        frozen_header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        self._frozen_table.viewport().installEventFilter(self)

        for column in range(self.model().columnCount()):
            self._frozen_table.setColumnHidden(column, column >= self.FROZEN_COLUMN_COUNT)

        self.horizontalHeader().sectionResized.connect(self._on_main_section_resized)
        frozen_header.sectionResized.connect(self._on_frozen_section_resized)
        self.verticalHeader().sectionResized.connect(self._on_main_row_resized)
        self.verticalScrollBar().valueChanged.connect(self._frozen_table.verticalScrollBar().setValue)
        self._frozen_table.verticalScrollBar().valueChanged.connect(self.verticalScrollBar().setValue)
        self._frozen_table.show()
        self.sync_frozen_row_heights()
        self._update_frozen_geometry()

    def setRowCount(self, rows: int) -> None:
        super().setRowCount(rows)
        self.sync_frozen_row_heights()

    def sync_frozen_row_heights(self) -> None:
        if not hasattr(self, "_frozen_table"):
            return
        self._frozen_table.verticalHeader().setDefaultSectionSize(self.verticalHeader().defaultSectionSize())
        for row in range(self.rowCount()):
            self._frozen_table.setRowHeight(row, self.rowHeight(row))

    def eventFilter(self, obj, event) -> bool:
        frozen_table = getattr(self, "_frozen_table", None)
        if frozen_table is not None and obj is frozen_table.viewport():
            if event.type() in (QtCore.QEvent.DragEnter, QtCore.QEvent.DragMove):
                if self._accept_file_drag_event(event):
                    return True
            if event.type() == QtCore.QEvent.Drop:
                if self._accept_file_drop_event(event):
                    return True
            if event.type() == QtCore.QEvent.ContextMenu:
                row = self._frozen_table.rowAt(event.pos().y())
                if row >= 0:
                    self.selectRow(row)
                    self.customContextMenuRequested.emit(QtCore.QPoint(0, event.pos().y()))
                    return True
            if event.type() == QtCore.QEvent.MouseButtonPress and event.button() == QtCore.Qt.LeftButton:
                self._begin_row_drag(self._frozen_to_main_viewport_pos(event.pos()))
                return False
            if event.type() == QtCore.QEvent.MouseMove:
                if self._update_row_drag(self._frozen_to_main_viewport_pos(event.pos()), event.buttons()):
                    return True
            if event.type() == QtCore.QEvent.MouseButtonRelease:
                if self._finish_row_drag(self._frozen_to_main_viewport_pos(event.pos()), event.button()):
                    return True
                if event.button() == QtCore.Qt.LeftButton:
                    self._reset_row_drag()
                return False
        return super().eventFilter(obj, event)

    def _frozen_to_main_viewport_pos(self, position: QtCore.QPoint) -> QtCore.QPoint:
        return self.viewport().mapFromGlobal(self._frozen_table.viewport().mapToGlobal(position))

    def scrollTo(self, index, hint=QtWidgets.QAbstractItemView.EnsureVisible) -> None:
        if index.isValid():
            horizontal_value = self.horizontalScrollBar().value()
            super().scrollTo(index, hint)
            if index.column() < self.FROZEN_COLUMN_COUNT or self.selectionBehavior() == QtWidgets.QAbstractItemView.SelectRows:
                self.horizontalScrollBar().setValue(horizontal_value)

    def resizeEvent(self, event) -> None:
        super().resizeEvent(event)
        self._update_frozen_geometry()

    def setColumnWidth(self, column: int, width: int) -> None:
        super().setColumnWidth(column, width)
        if hasattr(self, "_frozen_table") and column < self.FROZEN_COLUMN_COUNT:
            self._frozen_table.setColumnWidth(column, width)
            self._update_frozen_geometry()

    def setRowHeight(self, row: int, height: int) -> None:
        super().setRowHeight(row, height)
        if hasattr(self, "_frozen_table"):
            self._frozen_table.setRowHeight(row, height)

    def _on_main_section_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        if logical_index >= self.FROZEN_COLUMN_COUNT or self._syncing_frozen_columns:
            self._update_frozen_geometry()
            return
        self._syncing_frozen_columns = True
        try:
            self._frozen_table.setColumnWidth(logical_index, new_size)
        finally:
            self._syncing_frozen_columns = False
        self._update_frozen_geometry()

    def _on_frozen_section_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        if logical_index >= self.FROZEN_COLUMN_COUNT or self._syncing_frozen_columns:
            return
        self._syncing_frozen_columns = True
        try:
            super().setColumnWidth(logical_index, new_size)
        finally:
            self._syncing_frozen_columns = False
        self._update_frozen_geometry()

    def _on_main_row_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        self._frozen_table.setRowHeight(logical_index, new_size)

    def _frozen_width(self) -> int:
        return sum(self.columnWidth(column) for column in range(self.FROZEN_COLUMN_COUNT))

    def _update_frozen_geometry(self) -> None:
        if not hasattr(self, "_frozen_table"):
            return
        width = self._frozen_width()
        self._frozen_table.setGeometry(
            self.frameWidth(),
            self.frameWidth(),
            width,
            self.viewport().height() + self.horizontalHeader().height(),
        )
        self._frozen_table.raise_()

    def mousePressEvent(self, event) -> None:
        if event.button() == QtCore.Qt.LeftButton:
            self._begin_row_drag(event.pos())
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event) -> None:
        if not self._update_row_drag(event.pos(), event.buttons()):
            super().mouseMoveEvent(event)
            return
        event.accept()

    def mouseReleaseEvent(self, event) -> None:
        if self._finish_row_drag(event.pos(), event.button()):
            event.accept()
            return
        if event.button() == QtCore.Qt.LeftButton:
            self._reset_row_drag()
        super().mouseReleaseEvent(event)

    def _begin_row_drag(self, position: QtCore.QPoint) -> None:
        row = self.rowAt(position.y())
        if row >= 0:
            self._drag_source_row = row
            self._drag_start_pos = position
            self._drag_timer.start()
            self._dragging_row = False
        else:
            self._reset_row_drag()

    def _update_row_drag(self, position: QtCore.QPoint, buttons) -> bool:
        if not (buttons & QtCore.Qt.LeftButton) or self._drag_source_row < 0:
            return False
        distance = (position - self._drag_start_pos).manhattanLength()
        if not self._dragging_row:
            if self._drag_timer.elapsed() < self.LONG_PRESS_MS or distance < QtWidgets.QApplication.startDragDistance():
                return True
            self._dragging_row = True
            self.setCursor(QtCore.Qt.ClosedHandCursor)
            self._frozen_table.setCursor(QtCore.Qt.ClosedHandCursor)

        insert_row = self._drop_insert_row(position)
        self._show_drop_indicator(insert_row)
        self._auto_scroll(position)
        return True

    def _finish_row_drag(self, position: QtCore.QPoint, button) -> bool:
        if button != QtCore.Qt.LeftButton or not self._dragging_row:
            return False
        source_row = self._drag_source_row
        insert_row = self._drop_insert_row(position)
        self._reset_row_drag()
        if source_row >= 0:
            self.rowMoveRequested.emit(source_row, insert_row)
        return True

    def leaveEvent(self, event) -> None:
        if not self._dragging_row:
            self._drop_indicator.hide()
        super().leaveEvent(event)

    def _reset_row_drag(self) -> None:
        if self._dragging_row:
            self.unsetCursor()
            self._frozen_table.unsetCursor()
        self._drag_source_row = -1
        self._dragging_row = False
        self._drop_indicator.hide()

    def _drop_insert_row(self, position: QtCore.QPoint) -> int:
        row_count = self.rowCount()
        if row_count == 0:
            return 0
        row = self.rowAt(position.y())
        if row < 0:
            return 0 if position.y() < 0 else row_count
        midpoint = self.rowViewportPosition(row) + self.rowHeight(row) / 2
        return row if position.y() < midpoint else row + 1

    def _show_drop_indicator(self, insert_row: int) -> None:
        row_count = self.rowCount()
        if row_count == 0:
            self._drop_indicator.hide()
            return
        if insert_row <= 0:
            y = self.rowViewportPosition(0)
        elif insert_row >= row_count:
            last_row = row_count - 1
            y = self.rowViewportPosition(last_row) + self.rowHeight(last_row)
        else:
            y = self.rowViewportPosition(insert_row)
        self._drop_indicator.setGeometry(0, max(0, int(y) - 1), self.viewport().width(), 2)
        self._drop_indicator.show()
        self._drop_indicator.raise_()

    def _auto_scroll(self, position: QtCore.QPoint) -> None:
        margin = 24
        step = 18
        scroll_bar = self.verticalScrollBar()
        if position.y() < margin:
            scroll_bar.setValue(scroll_bar.value() - step)
        elif position.y() > self.viewport().height() - margin:
            scroll_bar.setValue(scroll_bar.value() + step)

    def dragEnterEvent(self, event) -> None:
        if self._accept_file_drag_event(event):
            return
        super().dragEnterEvent(event)

    def dragMoveEvent(self, event) -> None:
        if self._accept_file_drag_event(event):
            return
        super().dragMoveEvent(event)

    def dropEvent(self, event) -> None:
        if self._accept_file_drop_event(event):
            return
        super().dropEvent(event)

    def _accept_file_drag_event(self, event) -> bool:
        paths = self._smp_paths_from_mime_data(event.mimeData())
        if paths:
            event.acceptProposedAction()
            return True
        return False

    def _accept_file_drop_event(self, event) -> bool:
        paths = self._smp_paths_from_mime_data(event.mimeData())
        if not paths:
            return False
        event.acceptProposedAction()
        self.smpFilesDropped.emit(paths)
        return True

    @staticmethod
    def _smp_paths_from_mime_data(mime_data) -> list[str]:
        if not mime_data.hasUrls():
            return []
        paths = []
        for url in mime_data.urls():
            if not url.isLocalFile():
                continue
            path = Path(url.toLocalFile())
            if path.is_file() and path.suffix.lower() in (".smp", ".dat", ".qps", ".xls", ".xlsx", ".xlsm"):
                paths.append(str(path))
        return paths


class _NoFocusDelegate(QtWidgets.QStyledItemDelegate):
    def initStyleOption(self, option, index) -> None:
        super().initStyleOption(option, index)
        option.state &= ~QtWidgets.QStyle.State_HasFocus
        foreground = index.data(QtCore.Qt.ForegroundRole)
        color = foreground.color() if isinstance(foreground, QtGui.QBrush) else QtGui.QColor("#111827")
        option.palette.setColor(QtGui.QPalette.HighlightedText, color)


class _FrozenColumnsDelegate(_NoFocusDelegate):
    def initStyleOption(self, option, index) -> None:
        super().initStyleOption(option, index)


def _check_state_value(state) -> int:
    return int(getattr(state, "value", state))


VISIBLE_COLUMN = 0
FILE_COLUMN = 1
TEST_TIME_COLUMN = 2
BET_COLUMN = 3
LANGMUIR_COLUMN = 4
T_PLOT_COLUMN = 5
BJH_PORE_VOLUME_COLUMN = 6
PORE_VOLUME_METHOD_BJH = "bjh"
PORE_VOLUME_METHOD_DH = "dh"
PORE_VOLUME_METHOD_HK = "hk"
PORE_VOLUME_METHOD_DFT = "dft"
SUPPORTED_DATA_SUFFIXES = (".smp", ".dat", ".qps", ".xls", ".xlsx", ".xlsm")
BET_DEFAULT_RANGE = (0.05, 0.30)
BET_PLOT_RANGE = (0.0, 1.0)
LANGMUIR_DEFAULT_RANGE = (0.05, 0.30)
LANGMUIR_PLOT_RANGE = (0.0, 1.0)
T_PLOT_DEFAULT_PRESSURE_RANGE = (0.20, 0.50)
T_PLOT_PLOT_RANGE = (0.0, 1.0)
HK_DEFAULT_PRESSURE_RANGE = (0.0, 0.05)
CM3_STP_PER_MMOL = 22.414
SURFACE_AREA_CORRECTION_FACTOR = 1.0
CUSTOM_BET_COLOR = "#2563eb"
T_PLOT_THICKNESS_METHOD_LABELS = {
    "reference": "参比",
    "kjs": "Kruk-Jaroniec-Sayari",
    "halsey": "Halsey",
    "harkins_jura": "Harkins-Jura",
    "broekhoff_de_boer": "Broekhoff-De Boer",
    "carbon_black_stsa": "碳黑STSA",
}
T_PLOT_THICKNESS_PARAM_DEFAULTS = {
    method: {key: value for key, value in params.items() if key != "scale"}
    for method, params in THICKNESS_METHOD_DEFAULT_PARAMS.items()
}
DEFAULT_T_PLOT_THICKNESS_METHOD = DEFAULT_THICKNESS_METHOD
DEFAULT_T_PLOT_THICKNESS_PARAMS = dict(T_PLOT_THICKNESS_PARAM_DEFAULTS[DEFAULT_T_PLOT_THICKNESS_METHOD])
DEFAULT_T_PLOT_SURFACE_AREA_MODE = "BET"
DEFAULT_T_PLOT_SURFACE_AREA_INPUT = 1.0
DEFAULT_T_PLOT_SURFACE_AREA_CORRECTION = 1.0
DEFAULT_BJH_THICKNESS_METHOD = "reference"
DEFAULT_BJH_REFERENCE_FILE = DEFAULT_REFERENCE_DIR / "sio2oh.thk"
DEFAULT_BJH_CORRECTION = "standard"
DEFAULT_BJH_OPEN_PORE_FRACTION = 0.0
DEFAULT_BJH_SMOOTH_DERIVATIVE = True
DEFAULT_BJH_SHOW_ADSORPTION = True
DEFAULT_BJH_SHOW_DESORPTION = False
DEFAULT_BJH_DIFFERENTIAL_MODE = BJH_DIFFERENTIAL_LOG
DEFAULT_BJH_DISPLAY_METRICS = (BJH_DIFFERENTIAL_LOG,)
DEFAULT_BJH_PORE_VOLUME_RANGE = (2.0, 10.0)
DEFAULT_DH_SHOW_ADSORPTION = True
DEFAULT_DH_SHOW_DESORPTION = False
DEFAULT_DH_SMOOTH_DERIVATIVE = True
DEFAULT_DH_THICKNESS_METHOD = "reference"
DEFAULT_DH_DIFFERENTIAL_MODE = BJH_DIFFERENTIAL_LOG
DEFAULT_DH_DISPLAY_METRICS = (BJH_DIFFERENTIAL_LOG,)
DEFAULT_HK_GEOMETRY = HK_DEFAULT_GEOMETRY
DEFAULT_HK_ADSORBENT = HK_DEFAULT_ADSORBENT
DEFAULT_HK_ADSORPTIVE = HK_DEFAULT_ADSORPTIVE
DEFAULT_HK_INTERACTION_PARAMETER_MODE = "input"
DEFAULT_HK_INTERACTION_PARAMETER = HK_DEFAULT_INTERACTION_PARAMETER_ERG_CM4
DEFAULT_HK_CHENG_YANG_CORRECTION = False
DEFAULT_HK_SMOOTH_DERIVATIVE = False
DEFAULT_HK_DISPLAY_METRIC = HK_DIFFERENTIAL_LINEAR
DEFAULT_HK_PORE_VOLUME_RANGE = (0.6, 1.0)
DEFAULT_DFT_ANALYSIS_TYPE = DFT_DEFAULT_ANALYSIS_TYPE
DEFAULT_DFT_GEOMETRY = DFT_DEFAULT_GEOMETRY
DEFAULT_DFT_MODEL = DFT_DEFAULT_MODEL
DEFAULT_DFT_REGULARIZATION = DFT_DEFAULT_REGULARIZATION
DEFAULT_DFT_PORE_VOLUME_RANGE = (0.6, 10.0)
T_PLOT_PANEL_COLLAPSED_WIDTH = 360
T_PLOT_PANEL_EXPANDED_WIDTH = 660
BJH_PANEL_COLLAPSED_WIDTH = 380
BJH_PANEL_EXPANDED_WIDTH = 660
DFT_PANEL_WIDTH = 430
REGION_LINE_COLOR = "#2563eb"
REGION_LINE_HOVER_COLOR = "#dc2626"
REGION_FILL_COLOR = (37, 99, 235, 34)
REGION_FILL_HOVER_COLOR = (37, 99, 235, 48)
BJH_REGION_LINE_COLOR = "#16a34a"
BJH_REGION_LINE_HOVER_COLOR = "#15803d"
BJH_REGION_FILL_COLOR = (22, 163, 74, 34)
BJH_REGION_FILL_HOVER_COLOR = (22, 163, 74, 48)
REFERENCE_INVALID_BACKGROUND = "#fee2e2"
REFERENCE_INVALID_FOREGROUND = "#b91c1c"


def _region_pen(color: str) -> QtGui.QPen:
    pen = pg.mkPen(color, width=3)
    pen.setStyle(QtCore.Qt.DashLine)
    return pen


def _default_t_plot_thickness_params_by_method() -> dict[str, dict[str, float]]:
    return {method: dict(params) for method, params in T_PLOT_THICKNESS_PARAM_DEFAULTS.items()}


def _default_bjh_reference_params() -> dict[str, object]:
    try:
        points = read_reference_points(DEFAULT_BJH_REFERENCE_FILE)
    except OSError:
        points = []
    return {
        "reference_name": DEFAULT_BJH_REFERENCE_FILE.name,
        "reference_path": str(DEFAULT_BJH_REFERENCE_FILE),
        "reference_points": points,
    }


def _default_bjh_thickness_params_by_method() -> dict[str, dict[str, object]]:
    params_by_method = {
        method: dict(params)
        for method, params in T_PLOT_THICKNESS_PARAM_DEFAULTS.items()
    }
    params_by_method["reference"] = _default_bjh_reference_params()
    return params_by_method


def _t_plot_thickness_label(method: str) -> str:
    return T_PLOT_THICKNESS_METHOD_LABELS.get(method, T_PLOT_THICKNESS_METHOD_LABELS[DEFAULT_T_PLOT_THICKNESS_METHOD])


def _float_equal(left: object, right: object, *, tol: float = 1e-9) -> bool:
    try:
        return abs(float(left) - float(right)) <= tol
    except (TypeError, ValueError):
        return False


def _reference_points_equal(left: object, right: object) -> bool:
    left_points = normalize_reference_points(left)
    right_points = normalize_reference_points(right)
    if len(left_points) != len(right_points):
        return False
    return all(
        _float_equal(left_pressure, right_pressure) and _float_equal(left_thickness, right_thickness)
        for (left_pressure, left_thickness), (right_pressure, right_thickness) in zip(left_points, right_points)
    )


def _thickness_params_equal(active: dict[str, object], default: dict[str, object]) -> bool:
    for key, default_value in default.items():
        active_value = active.get(key)
        if key == "reference_points":
            if not _reference_points_equal(active_value, default_value):
                return False
        elif isinstance(default_value, (int, float)):
            if not _float_equal(active_value, default_value):
                return False
        elif active_value != default_value:
            return False
    return True


def _settings_values_equal(active: object, default: object) -> bool:
    if isinstance(default, (int, float)):
        return _float_equal(active, default)
    return active == default


def _settings_mapping_equal(active: dict[str, object], default: dict[str, object]) -> bool:
    if set(active) != set(default):
        return False
    for key, default_value in default.items():
        if not _settings_values_equal(active.get(key), default_value):
            return False
    return True


def _normalize_bjh_display_metrics(value: object) -> list[str]:
    return [_normalize_bjh_display_metric(value)]


def _normalize_bjh_display_metric(value: object) -> str:
    metrics = normalize_bjh_display_metrics(value)
    return metrics[0] if metrics else DEFAULT_BJH_DISPLAY_METRICS[0]


def _default_user_directory() -> Path:
    desktop = Path.home() / "Desktop"
    return desktop if desktop.exists() else Path.home()


class RegularizationSlider(QtWidgets.QWidget):
    valueChanged = Signal(float)
    valueChangeFinished = Signal(float)
    _AXIS_LOG_FACTOR = 5.0

    def __init__(self, values: Iterable[float], value: float, parent=None) -> None:
        super().__init__(parent)
        self._values = sorted({float(v) for v in values})
        self._value = float(value)
        self._dragging = False
        self._editor = QtWidgets.QLineEdit(self)
        validator = QtGui.QDoubleValidator(0.0, 10.0, 5, self._editor)
        validator.setNotation(QtGui.QDoubleValidator.StandardNotation)
        self._editor.setValidator(validator)
        self._editor.setAlignment(QtCore.Qt.AlignCenter)
        self._editor.setToolTip("请输入 0 到 10 之间的数字")
        self._editor.hide()
        self._editor.editingFinished.connect(self._finish_inline_edit)
        self.setMinimumHeight(76)
        self.setMinimumWidth(320)
        self.setMouseTracking(True)
        self.setCursor(QtCore.Qt.PointingHandCursor)

    def value(self) -> float:
        return float(self._value)

    def setValue(self, value: float, *, emit: bool = False) -> None:
        value = self._clamp_value(value)
        if _float_equal(value, self._value, tol=1e-12):
            self._value = value
            self.update()
            return
        self._value = value
        self.update()
        if emit:
            self.valueChanged.emit(self._value)

    def paintEvent(self, event) -> None:
        super().paintEvent(event)
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.Antialiasing)
        left, right = self._track_bounds()
        y = self.height() - 24
        active_x = self._value_to_x(self._value)

        left_button, right_button = self._button_rects()
        self._draw_arrow_button(painter, left_button, "left")
        self._draw_arrow_button(painter, right_button, "right")

        painter.setPen(QtGui.QPen(QtGui.QColor("#6b7280"), 2))
        painter.drawLine(QtCore.QLineF(left, y, right, y))
        painter.setPen(QtGui.QPen(QtGui.QColor(CUSTOM_BET_COLOR), 4))
        painter.drawLine(QtCore.QLineF(left, y, active_x, y))

        tick_pen = QtGui.QPen(QtGui.QColor("#cbd5e1"), 1)
        painter.setPen(tick_pen)
        for index in range(len(self._values)):
            x = self._fraction_to_x(self._preset_fraction(index))
            painter.drawLine(QtCore.QLineF(x, y - 8, x, y + 8))

        knob_radius = 11
        painter.setBrush(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
        painter.setPen(QtGui.QPen(QtGui.QColor("#ffffff"), 3))
        painter.drawEllipse(QtCore.QPointF(active_x, y), knob_radius, knob_radius)

        label = self._format_value(self._value)
        font = painter.font()
        font.setPointSize(9)
        font.setBold(True)
        painter.setFont(font)
        metrics = QtGui.QFontMetrics(font)
        rect = self._label_rect(label, metrics)
        painter.setPen(QtGui.QPen(QtGui.QColor("#d1d5db"), 1))
        painter.setBrush(QtGui.QBrush(QtGui.QColor("#ffffff")))
        painter.drawRoundedRect(rect, 4, 4)
        painter.setPen(QtGui.QPen(QtGui.QColor("#111827"), 1))
        painter.drawText(rect, QtCore.Qt.AlignCenter, label)

    def mousePressEvent(self, event) -> None:
        if event.button() != QtCore.Qt.LeftButton:
            super().mousePressEvent(event)
            return
        left_button, right_button = self._button_rects()
        if left_button.contains(event.pos()):
            self._step_preset(-1)
            self.valueChangeFinished.emit(self._value)
            event.accept()
            return
        if right_button.contains(event.pos()):
            self._step_preset(1)
            self.valueChangeFinished.emit(self._value)
            event.accept()
            return
        if self._current_label_rect().contains(event.pos()):
            self._begin_inline_edit()
            event.accept()
            return
        self._dragging = True
        self.setValue(self._x_to_value(event.pos().x()), emit=True)
        event.accept()

    def mouseMoveEvent(self, event) -> None:
        if self._dragging:
            self.setValue(self._x_to_value(event.pos().x()), emit=True)
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event) -> None:
        if event.button() == QtCore.Qt.LeftButton and self._dragging:
            self._dragging = False
            self.setValue(self._x_to_value(event.pos().x()), emit=True)
            self.valueChangeFinished.emit(self._value)
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def mouseDoubleClickEvent(self, event) -> None:
        if event.button() != QtCore.Qt.LeftButton:
            super().mouseDoubleClickEvent(event)
            return
        left_button, right_button = self._button_rects()
        if left_button.contains(event.pos()):
            self._step_preset(-1)
            self.valueChangeFinished.emit(self._value)
            event.accept()
            return
        if right_button.contains(event.pos()):
            self._step_preset(1)
            self.valueChangeFinished.emit(self._value)
            event.accept()
            return
        if self._current_label_rect().contains(event.pos()):
            self._begin_inline_edit()
        event.accept()

    def _begin_inline_edit(self) -> None:
        rect = self._current_label_rect()
        self._editor.setGeometry(rect)
        self._editor.setText(self._format_value(self._value))
        self._editor.selectAll()
        self._editor.show()
        self._editor.setFocus(QtCore.Qt.MouseFocusReason)

    def _finish_inline_edit(self) -> None:
        if not self._editor.isVisible():
            return
        text = self._editor.text().strip()
        self._editor.hide()
        try:
            value = float(text)
        except ValueError:
            self.update()
            return
        if 0.0 <= value <= 10.0:
            self.setValue(value, emit=True)
            self.valueChangeFinished.emit(self._value)
        else:
            self.update()

    def isDragging(self) -> bool:
        return bool(self._dragging)

    def _track_bounds(self) -> tuple[float, float]:
        return (46.0, max(48.0, float(self.width() - 46)))

    def _button_rects(self) -> tuple[QtCore.QRect, QtCore.QRect]:
        y = self.height() - 39
        return QtCore.QRect(4, y, 30, 30), QtCore.QRect(self.width() - 34, y, 30, 30)

    def _draw_arrow_button(self, painter: QtGui.QPainter, rect: QtCore.QRect, direction: str) -> None:
        painter.setPen(QtGui.QPen(QtGui.QColor("#cbd5e1"), 1))
        painter.setBrush(QtGui.QBrush(QtGui.QColor("#f8fafc")))
        painter.drawRoundedRect(rect, 5, 5)
        center = rect.center()
        if direction == "left":
            points = [
                QtCore.QPointF(center.x() - 4, center.y()),
                QtCore.QPointF(center.x() + 4, center.y() - 6),
                QtCore.QPointF(center.x() + 4, center.y() + 6),
            ]
        else:
            points = [
                QtCore.QPointF(center.x() + 4, center.y()),
                QtCore.QPointF(center.x() - 4, center.y() - 6),
                QtCore.QPointF(center.x() - 4, center.y() + 6),
            ]
        painter.setPen(QtCore.Qt.NoPen)
        painter.setBrush(QtGui.QBrush(QtGui.QColor("#334155")))
        painter.drawPolygon(QtGui.QPolygonF(points))

    def _value_to_x(self, value: float) -> float:
        fraction = self._value_to_fraction(value)
        return self._fraction_to_x(fraction)

    def _fraction_to_x(self, fraction: float) -> float:
        left, right = self._track_bounds()
        return left + (right - left) * max(0.0, min(1.0, float(fraction)))

    def _x_to_value(self, x: float) -> float:
        left, right = self._track_bounds()
        if right <= left:
            return self._value
        fraction = max(0.0, min(1.0, (float(x) - left) / (right - left)))
        return self._fraction_to_value(fraction)

    def _value_to_fraction(self, value: float) -> float:
        value = max(0.0, min(float(value), 10.0))
        values = self._values
        if not values or len(values) == 1:
            return 0.0
        if value <= values[0]:
            return 0.0
        if value >= values[-1]:
            return 1.0
        for index, preset in enumerate(values):
            if abs(value - preset) <= max(1e-12, abs(preset) * 1e-9):
                return self._index_fraction_to_axis_fraction(index / (len(values) - 1))
        for index in range(len(values) - 1):
            low = values[index]
            high = values[index + 1]
            if low <= value <= high:
                if low <= 0.0:
                    local = value / high if high > 0.0 else 0.0
                else:
                    local = (math.log10(value) - math.log10(low)) / (math.log10(high) - math.log10(low))
                index_fraction = (index + max(0.0, min(1.0, local))) / (len(values) - 1)
                return self._index_fraction_to_axis_fraction(index_fraction)
        return 1.0

    def _fraction_to_value(self, fraction: float) -> float:
        fraction = max(0.0, min(1.0, float(fraction)))
        values = self._values
        if not values:
            return self._value
        if len(values) == 1:
            return values[0]
        index_fraction = self._axis_fraction_to_index_fraction(fraction)
        position = index_fraction * (len(values) - 1)
        index = int(math.floor(position))
        if index >= len(values) - 1:
            return values[-1]
        local = position - index
        low = values[index]
        high = values[index + 1]
        if low <= 0.0:
            return low + (high - low) * local
        return 10.0 ** (math.log10(low) + (math.log10(high) - math.log10(low)) * local)

    def _preset_fraction(self, index: int) -> float:
        if len(self._values) <= 1:
            return 0.0
        index_fraction = int(index) / (len(self._values) - 1)
        return self._index_fraction_to_axis_fraction(index_fraction)

    @classmethod
    def _index_fraction_to_axis_fraction(cls, index_fraction: float) -> float:
        index_fraction = max(0.0, min(1.0, float(index_fraction)))
        return math.log1p(cls._AXIS_LOG_FACTOR * index_fraction) / math.log1p(cls._AXIS_LOG_FACTOR)

    @classmethod
    def _axis_fraction_to_index_fraction(cls, axis_fraction: float) -> float:
        axis_fraction = max(0.0, min(1.0, float(axis_fraction)))
        return math.expm1(axis_fraction * math.log1p(cls._AXIS_LOG_FACTOR)) / cls._AXIS_LOG_FACTOR

    def _current_label_rect(self) -> QtCore.QRect:
        font = self.font()
        font.setPointSize(9)
        font.setBold(True)
        return self._label_rect(self._format_value(self._value), QtGui.QFontMetrics(font))

    def _label_rect(self, label: str, metrics: QtGui.QFontMetrics) -> QtCore.QRect:
        active_x = self._value_to_x(self._value)
        text_width = metrics.horizontalAdvance(label) + 14
        rect_x = max(4, min(self.width() - text_width - 4, int(active_x - text_width / 2)))
        return QtCore.QRect(rect_x, 6, text_width, 24)

    def _step_preset(self, direction: int) -> None:
        if not self._values:
            return
        value = self._value
        if direction < 0:
            candidates = [preset for preset in self._values if preset < value - 1e-12]
            next_value = candidates[-1] if candidates else self._values[0]
        else:
            candidates = [preset for preset in self._values if preset > value + 1e-12]
            next_value = candidates[0] if candidates else self._values[-1]
        self.setValue(next_value, emit=True)

    @staticmethod
    def _clamp_value(value: float) -> float:
        try:
            value = float(value)
        except (TypeError, ValueError):
            value = DEFAULT_DFT_REGULARIZATION
        if not np.isfinite(value):
            value = DEFAULT_DFT_REGULARIZATION
        return max(0.0, min(value, 10.0))

    @staticmethod
    def _format_value(value: float) -> str:
        return f"{float(value):.5f}"


class FileImportDialog(QtWidgets.QDialog):
    def __init__(
        self,
        parent=None,
        initial_dir: Path | str | None = None,
        existing_paths: Iterable[str] | None = None,
        available_sort: tuple[int, QtCore.Qt.SortOrder] | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("导入数据文件")
        self.resize(1060, 640)
        self.setMinimumSize(860, 520)
        self.current_directory = Path(initial_dir or Path.cwd())
        self._available_paths: list[Path] = []
        self._selected_paths: list[Path] = [Path(p) for p in (existing_paths or [])]
        self._available_sort_column = int(available_sort[0]) if available_sort is not None else 0
        self._available_sort_order = available_sort[1] if available_sort is not None else QtCore.Qt.AscendingOrder

        folder_label = QtWidgets.QLabel("文件夹")
        self.folder_edit = QtWidgets.QLineEdit(str(self.current_directory))
        self.folder_edit.returnPressed.connect(self._set_directory_from_edit)
        browse_button = QtWidgets.QToolButton()
        browse_button.setText("...")
        browse_button.setToolTip("选择文件夹")
        browse_button.clicked.connect(self._browse_directory)
        refresh_button = QtWidgets.QToolButton()
        refresh_button.setText("刷新")
        refresh_button.clicked.connect(self._scan_directory)

        folder_layout = QtWidgets.QHBoxLayout()
        folder_layout.addWidget(folder_label)
        folder_layout.addWidget(self.folder_edit, 1)
        folder_layout.addWidget(browse_button)
        folder_layout.addWidget(refresh_button)

        self.search_edit = QtWidgets.QLineEdit()
        self.search_edit.setPlaceholderText("按文件名筛选")
        self.search_edit.textChanged.connect(lambda _text: self._populate_tables())

        search_layout = QtWidgets.QHBoxLayout()
        search_layout.addWidget(QtWidgets.QLabel("筛选"))
        search_layout.addWidget(self.search_edit, 1)

        self.available_table = self._make_file_table()
        self.selected_table = self._make_file_table()
        self.available_table.setSortingEnabled(True)
        self.available_table.horizontalHeader().sortIndicatorChanged.connect(self._on_available_sort_changed)
        self.available_table.itemDoubleClicked.connect(lambda _item: self._move_selected_to_right())
        self.selected_table.itemDoubleClicked.connect(lambda _item: self._move_selected_to_left())

        available_box = self._make_group("可导入文件", self.available_table)
        selected_box = self._make_group("待导入文件", self.selected_table)

        self.to_right_button = self._arrow_button(">", "添加选中文件")
        self.to_left_button = self._arrow_button("<", "移回选中文件")
        self.all_right_button = self._arrow_button(">>", "添加全部文件")
        self.all_left_button = self._arrow_button("<<", "全部移回")
        self.to_right_button.clicked.connect(self._move_selected_to_right)
        self.to_left_button.clicked.connect(self._move_selected_to_left)
        self.all_right_button.clicked.connect(self._move_all_to_right)
        self.all_left_button.clicked.connect(self._move_all_to_left)

        move_layout = QtWidgets.QVBoxLayout()
        move_layout.addStretch(1)
        for button in (self.to_right_button, self.to_left_button, self.all_right_button, self.all_left_button):
            move_layout.addWidget(button)
        move_layout.addStretch(1)

        self.move_up_button = self._arrow_button("↑", "上移选中文件")
        self.move_down_button = self._arrow_button("↓", "下移选中文件")
        self.move_up_button.clicked.connect(lambda: self._move_selected_rows(-1))
        self.move_down_button.clicked.connect(lambda: self._move_selected_rows(1))
        order_layout = QtWidgets.QHBoxLayout()
        order_layout.addStretch(1)
        order_layout.addWidget(self.move_up_button)
        order_layout.addWidget(self.move_down_button)

        selected_panel = QtWidgets.QWidget()
        selected_layout = QtWidgets.QVBoxLayout(selected_panel)
        selected_layout.setContentsMargins(0, 0, 0, 0)
        selected_layout.addWidget(selected_box, 1)
        selected_layout.addLayout(order_layout)

        picker_layout = QtWidgets.QHBoxLayout()
        picker_layout.addWidget(available_box, 1)
        picker_layout.addLayout(move_layout)
        picker_layout.addWidget(selected_panel, 1)

        self.count_label = QtWidgets.QLabel("")
        self.import_button = QtWidgets.QPushButton("导入")
        self.import_button.clicked.connect(self.accept)
        cancel_button = QtWidgets.QPushButton("取消")
        cancel_button.clicked.connect(self.reject)

        bottom_layout = QtWidgets.QHBoxLayout()
        bottom_layout.addWidget(self.count_label, 1)
        bottom_layout.addWidget(self.import_button)
        bottom_layout.addWidget(cancel_button)

        main_layout = QtWidgets.QVBoxLayout(self)
        main_layout.setContentsMargins(12, 12, 12, 12)
        main_layout.setSpacing(10)
        main_layout.addLayout(folder_layout)
        main_layout.addLayout(search_layout)
        main_layout.addLayout(picker_layout, 1)
        main_layout.addLayout(bottom_layout)

        for table in (self.available_table, self.selected_table):
            table.itemSelectionChanged.connect(self._update_buttons)
        self._scan_directory()

    def selected_paths(self) -> list[str]:
        return [str(path) for path in self._selected_paths]

    def available_sort(self) -> tuple[int, QtCore.Qt.SortOrder]:
        return (self._available_sort_column, self._available_sort_order)

    @staticmethod
    def _make_group(title: str, table: QtWidgets.QTableWidget) -> QtWidgets.QGroupBox:
        group = QtWidgets.QGroupBox(title)
        layout = QtWidgets.QVBoxLayout(group)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.addWidget(table)
        return group

    @staticmethod
    def _make_file_table() -> QtWidgets.QTableWidget:
        table = QtWidgets.QTableWidget(0, 4)
        table.setHorizontalHeaderLabels(["文件名", "格式", "修改时间", "大小"])
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setAlternatingRowColors(True)
        table.setSortingEnabled(False)
        table.verticalHeader().setVisible(False)
        table.verticalHeader().setDefaultSectionSize(26)
        table.horizontalHeader().setStretchLastSection(False)
        table.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.Stretch)
        table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        return table

    @staticmethod
    def _arrow_button(text: str, tooltip: str) -> QtWidgets.QPushButton:
        button = QtWidgets.QPushButton(text)
        button.setFixedSize(44, 32)
        button.setToolTip(tooltip)
        return button

    def _browse_directory(self) -> None:
        directory = QtWidgets.QFileDialog.getExistingDirectory(self, "选择数据文件夹", str(self.current_directory))
        if not directory:
            return
        self.current_directory = Path(directory)
        self.folder_edit.setText(str(self.current_directory))
        self._scan_directory()

    def _set_directory_from_edit(self) -> None:
        directory = Path(self.folder_edit.text().strip())
        if not directory.is_dir():
            QtWidgets.QMessageBox.warning(self, "文件夹不存在", str(directory))
            self.folder_edit.setText(str(self.current_directory))
            return
        self.current_directory = directory
        self._scan_directory()

    def _scan_directory(self) -> None:
        directory = self.current_directory
        self.folder_edit.setText(str(directory))
        selected = {self._path_key(path) for path in self._selected_paths}
        try:
            files = [
                path
                for path in directory.iterdir()
                if path.is_file() and path.suffix.lower() in SUPPORTED_DATA_SUFFIXES and self._path_key(path) not in selected
            ]
        except OSError as exc:
            QtWidgets.QMessageBox.warning(self, "无法读取文件夹", str(exc))
            files = []
        self._available_paths = sorted(files, key=lambda path: path.name.lower())
        self._populate_tables()

    def _populate_tables(self) -> None:
        query = self.search_edit.text().strip().lower()
        available = [path for path in self._available_paths if query in path.name.lower()]
        self._fill_table(self.available_table, available)
        self._fill_table(self.selected_table, self._selected_paths)
        self._update_buttons()

    def _fill_table(self, table: QtWidgets.QTableWidget, paths: list[Path]) -> None:
        table.setSortingEnabled(False)
        table.setRowCount(0)
        for path in paths:
            row = table.rowCount()
            table.insertRow(row)
            values = [path.name, path.suffix.upper().lstrip("."), self._modified_text(path), self._size_text(path)]
            for column, value in enumerate(values):
                item = QtWidgets.QTableWidgetItem(value)
                item.setData(QtCore.Qt.UserRole, str(path))
                if column in {1, 3}:
                    item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                table.setItem(row, column, item)
        table.setSortingEnabled(table is self.available_table)
        if table is self.available_table:
            table.sortItems(self._available_sort_column, self._available_sort_order)

    def _on_available_sort_changed(self, column: int, order: QtCore.Qt.SortOrder) -> None:
        self._available_sort_column = int(column)
        self._available_sort_order = order

    @staticmethod
    def _modified_text(path: Path) -> str:
        try:
            return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
        except OSError:
            return ""

    @staticmethod
    def _size_text(path: Path) -> str:
        try:
            size = path.stat().st_size
        except OSError:
            return ""
        if size >= 1024 * 1024:
            return f"{size / (1024 * 1024):.1f} MB"
        if size >= 1024:
            return f"{size / 1024:.1f} KB"
        return f"{size} B"

    @staticmethod
    def _path_key(path: Path) -> str:
        try:
            return str(path.resolve()).lower()
        except OSError:
            return str(path).lower()

    def _selected_table_paths(self, table: QtWidgets.QTableWidget) -> list[Path]:
        rows = sorted({index.row() for index in table.selectionModel().selectedRows()})
        paths = []
        for row in rows:
            item = table.item(row, 0)
            if item is not None:
                paths.append(Path(str(item.data(QtCore.Qt.UserRole))))
        return paths

    def _move_selected_to_right(self) -> None:
        paths = self._selected_table_paths(self.available_table)
        self._add_to_selected(paths)

    def _move_all_to_right(self) -> None:
        self._add_to_selected(list(self._available_paths))

    def _add_to_selected(self, paths: list[Path]) -> None:
        if not paths:
            return
        selected_keys = {self._path_key(path) for path in self._selected_paths}
        for path in paths:
            key = self._path_key(path)
            if key not in selected_keys:
                self._selected_paths.append(path)
                selected_keys.add(key)
        moved = {self._path_key(path) for path in paths}
        self._available_paths = [path for path in self._available_paths if self._path_key(path) not in moved]
        self._populate_tables()

    def _move_selected_to_left(self) -> None:
        paths = self._selected_table_paths(self.selected_table)
        self._remove_from_selected(paths)

    def _move_all_to_left(self) -> None:
        self._remove_from_selected(list(self._selected_paths))

    def _remove_from_selected(self, paths: list[Path]) -> None:
        if not paths:
            return
        removed = {self._path_key(path) for path in paths}
        self._selected_paths = [path for path in self._selected_paths if self._path_key(path) not in removed]
        existing = {self._path_key(path) for path in self._available_paths}
        for path in paths:
            if path.exists() and self._path_key(path) not in existing:
                self._available_paths.append(path)
                existing.add(self._path_key(path))
        self._available_paths.sort(key=lambda path: path.name.lower())
        self._populate_tables()

    def _move_selected_rows(self, direction: int) -> None:
        rows = sorted({index.row() for index in self.selected_table.selectionModel().selectedRows()})
        if not rows or (direction < 0 and rows[0] == 0) or (direction > 0 and rows[-1] >= len(self._selected_paths) - 1):
            return
        if direction > 0:
            rows = list(reversed(rows))
        for row in rows:
            target = row + direction
            self._selected_paths[row], self._selected_paths[target] = self._selected_paths[target], self._selected_paths[row]
        selected_after = [row + direction for row in rows]
        self._populate_tables()
        self.selected_table.clearSelection()
        for row in selected_after:
            self.selected_table.selectRow(row)

    def _update_buttons(self) -> None:
        has_available_selection = bool(self.available_table.selectionModel().selectedRows())
        has_selected_selection = bool(self.selected_table.selectionModel().selectedRows())
        self.to_right_button.setEnabled(has_available_selection)
        self.to_left_button.setEnabled(has_selected_selection)
        self.all_right_button.setEnabled(bool(self._available_paths))
        self.all_left_button.setEnabled(bool(self._selected_paths))
        self.move_up_button.setEnabled(has_selected_selection and min(self._selected_rows(self.selected_table), default=0) > 0)
        self.move_down_button.setEnabled(
            has_selected_selection
            and max(self._selected_rows(self.selected_table), default=-1) < len(self._selected_paths) - 1
        )
        self.import_button.setEnabled(bool(self._selected_paths))
        self.count_label.setText(f"可导入 {len(self._available_paths)} 个，待导入 {len(self._selected_paths)} 个")

    @staticmethod
    def _selected_rows(table: QtWidgets.QTableWidget) -> list[int]:
        return sorted({index.row() for index in table.selectionModel().selectedRows()})


class HorvathKawazoePropertiesDialog(QtWidgets.QDialog):
    PROPERTY_FIELDS = (
        ("diameter_nm", "直径:", "nm", "{:.4f}"),
        ("zero_diameter_nm", "零能量处直径", "nm", "{:.4f}"),
        ("polarizability_cm3", "极化", "cm3", "{:.3e}"),
        ("susceptibility_cm3", "磁化率", "cm3", "{:.3e}"),
        ("density_per_cm2", "密度", "molecule/cm2", "{:.3e}"),
    )

    def __init__(
        self,
        parent=None,
        *,
        adsorbent_key: str = DEFAULT_HK_ADSORBENT,
        adsorptive_key: str = DEFAULT_HK_ADSORPTIVE,
        adsorbent_properties: dict[str, object] | None = None,
        adsorptive_properties: dict[str, object] | None = None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle("Horvath-Kawazoe 物理性质")
        self.setMinimumWidth(520)
        self._adsorbent_edits: dict[str, QtWidgets.QLineEdit] = {}
        self._adsorptive_edits: dict[str, QtWidgets.QLineEdit] = {}

        self.adsorbent_combo = QtWidgets.QComboBox()
        for key, props in HK_ADSORBENT_PRESETS.items():
            self.adsorbent_combo.addItem(str(props.get("label", key)), key)
        self.adsorptive_combo = QtWidgets.QComboBox()
        for key, props in HK_ADSORPTIVE_PRESETS.items():
            self.adsorptive_combo.addItem(str(props.get("label", key)), key)

        adsorbent_group = self._make_properties_group("吸附剂", self.adsorbent_combo, self._adsorbent_edits)
        adsorptive_group = self._make_properties_group("吸附物", self.adsorptive_combo, self._adsorptive_edits)

        button_row = QtWidgets.QHBoxLayout()
        ok_button = QtWidgets.QPushButton("OK")
        cancel_button = QtWidgets.QPushButton("结束")
        ok_button.clicked.connect(self.accept)
        cancel_button.clicked.connect(self.reject)
        button_row.addStretch(1)
        button_row.addWidget(ok_button)
        button_row.addSpacing(24)
        button_row.addWidget(cancel_button)
        button_row.addStretch(1)

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(10)
        layout.addWidget(adsorbent_group)
        layout.addWidget(adsorptive_group)
        layout.addLayout(button_row)

        self.adsorbent_combo.currentIndexChanged.connect(lambda _index: self._populate_fields("adsorbent"))
        self.adsorptive_combo.currentIndexChanged.connect(lambda _index: self._populate_fields("adsorptive"))
        self._set_combo_data(self.adsorbent_combo, adsorbent_key)
        self._set_combo_data(self.adsorptive_combo, adsorptive_key)
        self._populate_fields("adsorbent", adsorbent_properties)
        self._populate_fields("adsorptive", adsorptive_properties)

    @classmethod
    def _make_properties_group(
        cls,
        title: str,
        combo: QtWidgets.QComboBox,
        edits: dict[str, QtWidgets.QLineEdit],
    ) -> QtWidgets.QGroupBox:
        group = QtWidgets.QGroupBox(title)
        layout = QtWidgets.QGridLayout(group)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setHorizontalSpacing(8)
        layout.setVerticalSpacing(6)
        layout.addWidget(QtWidgets.QLabel("描述"), 0, 0)
        layout.addWidget(combo, 0, 1, 1, 2)
        validator = QtGui.QDoubleValidator(group)
        validator.setNotation(QtGui.QDoubleValidator.ScientificNotation)
        for row, (key, label, unit, _fmt) in enumerate(cls.PROPERTY_FIELDS, start=1):
            edit = QtWidgets.QLineEdit()
            edit.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
            edit.setValidator(validator)
            edit.setFixedWidth(130)
            edits[key] = edit
            layout.addWidget(QtWidgets.QLabel(label), row, 0)
            layout.addWidget(edit, row, 1)
            layout.addWidget(QtWidgets.QLabel(unit), row, 2)
        layout.setColumnStretch(3, 1)
        return group

    @staticmethod
    def _set_combo_data(combo: QtWidgets.QComboBox, target_key: str) -> None:
        for index in range(combo.count()):
            if str(combo.itemData(index)) == str(target_key):
                combo.setCurrentIndex(index)
                return

    def _populate_fields(self, kind: str, overrides: dict[str, object] | None = None) -> None:
        if kind == "adsorbent":
            combo = self.adsorbent_combo
            presets = HK_ADSORBENT_PRESETS
            edits = self._adsorbent_edits
        else:
            combo = self.adsorptive_combo
            presets = HK_ADSORPTIVE_PRESETS
            edits = self._adsorptive_edits
        key = str(combo.currentData())
        props = dict(presets.get(key, next(iter(presets.values()))))
        if overrides:
            props.update(overrides)
        for field_key, _label, _unit, fmt in self.PROPERTY_FIELDS:
            value = props.get(field_key, 0.0)
            try:
                text = fmt.format(float(value))
            except (TypeError, ValueError):
                text = str(value)
            edits[field_key].setText(text)

    def selected_adsorbent_key(self) -> str:
        return str(self.adsorbent_combo.currentData())

    def selected_adsorptive_key(self) -> str:
        return str(self.adsorptive_combo.currentData())

    def adsorbent_properties(self) -> dict[str, object]:
        return self._read_properties(self._adsorbent_edits, self.adsorbent_combo.currentText())

    def adsorptive_properties(self) -> dict[str, object]:
        return self._read_properties(self._adsorptive_edits, self.adsorptive_combo.currentText())

    def _read_properties(self, edits: dict[str, QtWidgets.QLineEdit], label: str) -> dict[str, object]:
        props: dict[str, object] = {"label": label}
        for field_key, _label, _unit, _fmt in self.PROPERTY_FIELDS:
            try:
                props[field_key] = float(edits[field_key].text())
            except (KeyError, TypeError, ValueError):
                props[field_key] = 0.0
        return props


class MainWindow(QtWidgets.QMainWindow):
    def __init__(self) -> None:
        super().__init__()
        self.results = []
        self.visible_results: list[bool] = []
        self.active_index = -1
        self.sample_colors = list(DEFAULT_COLORS)
        self.sample_items = []
        self.custom_bet_fit_ranges: dict[int, tuple[float, float]] = {}
        self.custom_langmuir_fit_ranges: dict[int, tuple[float, float]] = {}
        self.custom_t_plot_fit_ranges: dict[int, tuple[float, float]] = {}
        self.custom_t_plot_settings: dict[int, dict[str, object]] = {}
        self.custom_bjh_settings: dict[int, dict[str, object]] = {}
        self.custom_dh_settings: dict[int, dict[str, object]] = {}
        self.custom_hk_settings: dict[int, dict[str, object]] = {}
        self.custom_dft_settings: dict[int, dict[str, object]] = {}
        self.bjh_pore_volume_range: tuple[float, float] = DEFAULT_BJH_PORE_VOLUME_RANGE
        self._updating_table = False
        self._updating_sample_checks = False
        self._updating_sample_column_widths = False
        self._sample_column_widths_initialized = False
        self.sample_column_widths: dict[int, int] = {}
        self.test_time_sort_ascending = False
        self.bet_sort_ascending = False
        self.langmuir_sort_ascending = False
        self.t_plot_sort_ascending = False
        self.bjh_pore_sort_ascending = False
        self.region = None
        self._isotherm_selection_items = []
        self.bet_region = None
        self._bet_fit_line = None
        self._bet_selection_item = None
        self._bet_x_range = None
        self._bet_plot_p_range = None
        self.langmuir_region = None
        self._langmuir_fit_line = None
        self._langmuir_selection_item = None
        self._langmuir_x_range = None
        self._langmuir_plot_p_range = None
        self.t_plot_region = None
        self._t_plot_fit_line = None
        self._t_plot_selection_item = None
        self._t_plot_x_range = None
        self._t_plot_p_range = None
        self._fit_analysis_cache: dict[tuple[object, ...], object] = {}
        self.bjh_region = None
        self._bjh_selection_items = []
        self._bjh_distribution_rows_by_key: dict[tuple[int, str], list[dict[str, float]]] = {}
        self._bjh_distribution_cache: dict[tuple[object, ...], list[dict[str, float]]] = {}
        self._bjh_diameter_log_bounds: tuple[float, float] | None = None
        self.dh_region = None
        self._dh_selection_items = []
        self._dh_distribution_rows_by_key: dict[tuple[int, str], list[dict[str, float]]] = {}
        self._dh_distribution_cache: dict[tuple[object, ...], list[dict[str, float]]] = {}
        self._dh_diameter_log_bounds: tuple[float, float] | None = None
        self._hk_distribution_rows_by_key: dict[tuple[int, str], list[dict[str, float]]] = {}
        self._hk_distribution_cache: dict[tuple[object, ...], list[dict[str, float]]] = {}
        self.hk_region = None
        self._hk_selection_items = []
        self._hk_width_log_bounds: tuple[float, float] | None = None
        self.hk_pore_volume_range: tuple[float, float] = DEFAULT_HK_PORE_VOLUME_RANGE
        self._dft_distribution_rows_by_index: dict[int, list[dict[str, float]]] = {}
        self._dft_result_cache: dict[tuple[object, ...], object] = {}
        self.dft_region = None
        self._dft_selection_items = []
        self._dft_width_log_bounds: tuple[float, float] | None = None
        self.dft_pore_volume_range: tuple[float, float] = DEFAULT_DFT_PORE_VOLUME_RANGE
        self._dft_diagnostic_line = None
        self._syncing_t_plot_controls = False
        self.t_plot_thickness_method = DEFAULT_T_PLOT_THICKNESS_METHOD
        self.t_plot_thickness_params_by_method = _default_t_plot_thickness_params_by_method()
        self.t_plot_thickness_params = dict(DEFAULT_T_PLOT_THICKNESS_PARAMS)
        self.t_plot_surface_area_mode = DEFAULT_T_PLOT_SURFACE_AREA_MODE
        self.t_plot_surface_area_input = DEFAULT_T_PLOT_SURFACE_AREA_INPUT
        self.t_plot_surface_area_correction = DEFAULT_T_PLOT_SURFACE_AREA_CORRECTION
        self._syncing_bjh_controls = False
        self.bjh_thickness_method = DEFAULT_BJH_THICKNESS_METHOD
        self.bjh_thickness_params_by_method = _default_bjh_thickness_params_by_method()
        self.bjh_thickness_params = dict(self.bjh_thickness_params_by_method[DEFAULT_BJH_THICKNESS_METHOD])
        self.bjh_correction = DEFAULT_BJH_CORRECTION
        self.bjh_open_pore_fraction = DEFAULT_BJH_OPEN_PORE_FRACTION
        self.bjh_smooth_derivative = DEFAULT_BJH_SMOOTH_DERIVATIVE
        self.bjh_show_adsorption = DEFAULT_BJH_SHOW_ADSORPTION
        self.bjh_show_desorption = DEFAULT_BJH_SHOW_DESORPTION
        self.bjh_differential_mode = DEFAULT_BJH_DIFFERENTIAL_MODE
        self.bjh_display_metrics = list(DEFAULT_BJH_DISPLAY_METRICS)
        self._syncing_dh_controls = False
        self.dh_thickness_method = DEFAULT_DH_THICKNESS_METHOD
        self.dh_thickness_params_by_method = _default_bjh_thickness_params_by_method()
        self.dh_thickness_params = dict(self.dh_thickness_params_by_method[DEFAULT_DH_THICKNESS_METHOD])
        self.dh_smooth_derivative = DEFAULT_DH_SMOOTH_DERIVATIVE
        self.dh_show_adsorption = DEFAULT_DH_SHOW_ADSORPTION
        self.dh_show_desorption = DEFAULT_DH_SHOW_DESORPTION
        self.dh_differential_mode = DEFAULT_DH_DIFFERENTIAL_MODE
        self.dh_display_metrics = list(DEFAULT_DH_DISPLAY_METRICS)
        self._syncing_hk_controls = False
        self.hk_geometry = DEFAULT_HK_GEOMETRY
        self.hk_adsorbent_key = DEFAULT_HK_ADSORBENT
        self.hk_adsorptive_key = DEFAULT_HK_ADSORPTIVE
        self.hk_adsorbent_properties = dict(HK_ADSORBENT_PRESETS[DEFAULT_HK_ADSORBENT])
        self.hk_adsorptive_properties = dict(HK_ADSORPTIVE_PRESETS[DEFAULT_HK_ADSORPTIVE])
        self.hk_interaction_parameter_mode = DEFAULT_HK_INTERACTION_PARAMETER_MODE
        self.hk_interaction_parameter = DEFAULT_HK_INTERACTION_PARAMETER
        self.hk_cheng_yang_correction = DEFAULT_HK_CHENG_YANG_CORRECTION
        self.hk_smooth_derivative = DEFAULT_HK_SMOOTH_DERIVATIVE
        self.hk_display_metric = DEFAULT_HK_DISPLAY_METRIC
        self._syncing_dft_controls = False
        self.dft_analysis_type = DEFAULT_DFT_ANALYSIS_TYPE
        self.dft_geometry = DEFAULT_DFT_GEOMETRY
        self.dft_model = DEFAULT_DFT_MODEL
        self.dft_regularization = DEFAULT_DFT_REGULARIZATION
        self.dft_regularization_apply_all = False
        self._pending_dft_regularization_apply_all_value: float | None = None
        self._dft_regularization_preview_active = False
        self._dft_regularization_refresh_timer = QtCore.QTimer(self)
        self._dft_regularization_refresh_timer.setSingleShot(True)
        self._dft_regularization_refresh_timer.setInterval(40)
        self._dft_regularization_refresh_timer.timeout.connect(self._preview_deferred_dft_regularization_refresh)
        self.reference_tables: dict[str, QtWidgets.QTableWidget] = {}
        self.reference_name_edits: dict[str, QtWidgets.QLineEdit] = {}
        self._syncing_reference_tables = False
        self.region_is_log = False
        self._isotherm_region_custom = False
        self._last_isotherm_region_range: tuple[float, float] | None = None
        self._setting_isotherm_region = False
        self._metrics_pending = False
        self._bet_region_pending = False
        self._langmuir_region_pending = False
        self._t_plot_region_pending = False
        self._bjh_region_pending = False
        self._dh_region_pending = False
        self._hk_region_pending = False
        self._dft_region_pending = False
        self._syncing_region_changes = False
        self._setting_bet_region = False
        self._setting_langmuir_region = False
        self._setting_t_plot_region = False
        self._setting_bjh_region = False
        self._setting_dh_region = False
        self._setting_hk_region = False
        self._setting_dft_region = False
        self._checking_for_updates = False
        self._update_thread: QtCore.QThread | None = None
        self._update_worker: UpdateCheckWorker | None = None
        self._update_download_thread: QtCore.QThread | None = None
        self._update_download_worker: UpdateDownloadWorker | None = None
        self._update_progress_dialog: QtWidgets.QProgressDialog | None = None
        self._available_update_info: UpdateInfo | None = None
        self.settings = QtCore.QSettings("UnifiedBET", "TriStarBetAppZh")
        self.settings.remove("bjh_differential_mode")
        self.settings.remove("bjh_display_metrics")
        self.import_directory = self._read_directory_setting("import_directory")
        self.export_directory = self._read_directory_setting("export_directory")
        self._import_available_sort = (
            int(self.settings.value("import_available_sort_column", 0)),
            QtCore.Qt.SortOrder(
                int(self.settings.value("import_available_sort_order", _qt_enum_int(QtCore.Qt.AscendingOrder)))
            ),
        )

        self.setWindowTitle(f"{APP_NAME} v{APP_VERSION}")
        if APP_ICON_PATH.exists():
            self.setWindowIcon(QtGui.QIcon(str(APP_ICON_PATH)))
        self.resize(1280, 780)

        open_button = QtWidgets.QPushButton("导入文件")
        open_button.clicked.connect(self.open_files)
        export_button = QtWidgets.QPushButton("导出文件")
        export_button.clicked.connect(self.export_xlsx)
        self.update_available_button = QtWidgets.QToolButton()
        self.update_available_button.setIcon(_make_update_available_icon(28))
        self.update_available_button.setIconSize(QtCore.QSize(24, 24))
        self.update_available_button.setFixedSize(32, 32)
        self.update_available_button.setAutoRaise(True)
        self.update_available_button.setCursor(QtCore.Qt.PointingHandCursor)
        self.update_available_button.setToolTip("发现新版本，点击更新")
        self.update_available_button.clicked.connect(self._show_pending_update_dialog)
        self.update_available_button.hide()
        self.update_button = QtWidgets.QPushButton("软件更新")
        self.update_button.setToolTip("联网检查 Gitee/GitHub 更新源中是否有新版")
        self.update_button.clicked.connect(self.check_for_updates)
        for button in (open_button, export_button, self.update_button):
            button.setFixedHeight(32)
            button.setSizePolicy(QtWidgets.QSizePolicy.Fixed, QtWidgets.QSizePolicy.Fixed)
            button.setFixedWidth(96)

        self.select_all_check = SelectAllCheckBox()
        self.select_all_check.setTristate(True)
        self.select_all_check.setCheckState(QtCore.Qt.Checked)
        self.select_all_check.setCursor(QtCore.Qt.PointingHandCursor)
        self.select_all_check.setToolTip("显示或隐藏全部样品")
        self.select_all_check.stateChanged.connect(self.on_select_all_changed)
        self.select_all_check.setStyleSheet(
            """
            QCheckBox::indicator {
                width: 12px;
                height: 12px;
                border-radius: 7px;
                border: 1px solid #6b7280;
                background: white;
            }
            QCheckBox::indicator:checked {
                border: 1px solid #2563eb;
                background: #2563eb;
            }
            QCheckBox::indicator:indeterminate {
                border: 1px solid #2563eb;
                background: #93c5fd;
            }
            """
        )

        self.sample_table = SampleTableWidget(0, 7)
        self.sample_table.setHorizontalHeaderLabels(
            ["", "文件名", "测试时间", "BET(m2/g)", "Langmuir(m2/g)", "t-Plot外比(m2/g)", "选区孔容量(cm3/g)"]
        )
        sample_header = self.sample_table.horizontalHeader()
        sample_header.setVisible(True)
        sample_header.setSectionsMovable(False)
        sample_header.setHighlightSections(False)
        sample_header.setStretchLastSection(False)
        sample_header.setDefaultAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        sample_header.setSectionResizeMode(QtWidgets.QHeaderView.Interactive)
        sample_header.sectionClicked.connect(self.on_sample_header_clicked)
        sample_header.sectionResized.connect(self.on_sample_header_resized)
        self.sample_table.horizontalHeaderItem(TEST_TIME_COLUMN).setToolTip("点击按测试时间排序")
        self.sample_table.horizontalHeaderItem(BET_COLUMN).setToolTip("点击按BET比表面积排序")
        self.sample_table.horizontalHeaderItem(BET_COLUMN).setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.sample_table.horizontalHeaderItem(LANGMUIR_COLUMN).setToolTip("点击按Langmuir比表面积排序")
        self.sample_table.horizontalHeaderItem(LANGMUIR_COLUMN).setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.sample_table.horizontalHeaderItem(T_PLOT_COLUMN).setToolTip("点击按t-Plot外比表面积排序")
        self.sample_table.horizontalHeaderItem(T_PLOT_COLUMN).setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.sample_table.horizontalHeaderItem(BJH_PORE_VOLUME_COLUMN).setToolTip("点击按 BJH 选区孔容量排序")
        self.sample_table.horizontalHeaderItem(BJH_PORE_VOLUME_COLUMN).setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self._update_pore_volume_header()
        self.sample_table.verticalHeader().setVisible(False)
        self.sample_table.verticalHeader().setDefaultSectionSize(28)
        self.sample_table.sync_frozen_row_heights()
        self.sample_table.setShowGrid(False)
        self.sample_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.sample_table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        self.sample_table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        self.sample_table.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.sample_table.setHorizontalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        self.sample_table.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarAsNeeded)
        self.sample_table.setMinimumHeight(70)
        self.sample_table.setColumnWidth(VISIBLE_COLUMN, 30)
        self.sample_table.setColumnWidth(FILE_COLUMN, 170)
        self.sample_table.setColumnWidth(TEST_TIME_COLUMN, 250)
        self.sample_table.setColumnWidth(BET_COLUMN, 120)
        self.sample_table.setColumnWidth(LANGMUIR_COLUMN, 200)
        self.sample_table.setColumnWidth(T_PLOT_COLUMN, 200)
        self.sample_table.setColumnWidth(BJH_PORE_VOLUME_COLUMN, 190)
        self.sample_table.setItemDelegate(_NoFocusDelegate(self.sample_table))
        self.sample_table._frozen_table.setItemDelegate(_FrozenColumnsDelegate(self.sample_table._frozen_table))
        self.sample_table.itemChanged.connect(self.on_sample_item_changed)
        self.sample_table.currentCellChanged.connect(self.on_active_cell_changed)
        self.sample_table.rowMoveRequested.connect(self.move_sample_row)
        self.sample_table.smpFilesDropped.connect(self.append_files)
        self.sample_table.setContextMenuPolicy(QtCore.Qt.CustomContextMenu)
        self.sample_table.customContextMenuRequested.connect(self.show_sample_context_menu)
        self.sample_table.horizontalScrollBar().valueChanged.connect(self._position_header_controls)
        self.sample_table.setStyleSheet(
            """
            QTableWidget {
                border: 1px solid #d1d5db;
                background: #ffffff;
            }
            QTableWidget::item:selected {
                background: #e0ecff;
            }
            QTableWidget::item:focus {
                outline: none;
            }
            QTableWidget::indicator {
                width: 11px;
                height: 11px;
                border-radius: 6px;
                border: 1px solid #6b7280;
                background: white;
            }
            QTableWidget::indicator:checked {
                border: 1px solid #2563eb;
                background: #2563eb;
            }
            QHeaderView::section {
                background: #f9fafb;
                border: 0;
                border-right: 1px solid #d1d5db;
                border-bottom: 1px solid #d1d5db;
                color: #374151;
                font-weight: 600;
                padding: 4px 8px 4px 6px;
            }
            """
        )
        self.select_all_check.setParent(self.sample_table.frozen_header())
        self.select_all_check.show()

        self.metrics_table = self._make_table(["参数", "值"])

        self.isotherm_plot = make_plot("吸附/脱附等温线", "吸附量 (cm3/g STP)", "相对压力 (P/P0)")
        self.bet_plot = make_plot("BET 拟合", "P/[V(P0-P)]", "相对压力 (P/P0)")
        self.langmuir_plot = make_plot("Langmuir 拟合", "(P/P0) / V", "相对压力 (P/P0)")
        self.t_plot = make_plot("t-Plot", "液体体积 (cm3/g)", "统计膜厚 t (nm)")
        self.pore_plot = make_plot(
            "BJH 孔径分布",
            bjh_display_axis_label(self.bjh_display_metrics),
            "孔径 (nm)",
            legend_position="right",
        )

        self.dh_plot = make_plot(
            "Dollimore-Heal 孔径分布",
            bjh_display_axis_label(self.dh_display_metrics),
            "孔径 (nm)",
            legend_position="right",
        )

        self.bet_default_button = QtWidgets.QPushButton("默认")
        self.bet_default_button.setToolTip("按自动 BET 选点算法重新计算")
        self.hk_plot = make_plot(
            "Horvath-Kawazoe 孔径分布",
            hk_display_axis_label(self.hk_display_metric),
            "孔宽 W (nm)",
            legend_position="right",
        )

        self.dft_plot = make_plot(
            "DFT 孔径分布",
            "dV/dlogW (cm3/g)",
            "孔宽 W (nm)",
            legend_position="right",
        )
        self.dft_diagnostic_plot = make_plot(
            "拟合误差 vs 正则化",
            "RMS 拟合误差 (mmol/g)",
            "正则化",
        )
        self.dft_diagnostic_plot.setMinimumWidth(360)
        for plot in (
            self.isotherm_plot,
            self.bet_plot,
            self.langmuir_plot,
            self.t_plot,
            self.pore_plot,
            self.dh_plot,
            self.hk_plot,
            self.dft_plot,
        ):
            setattr(plot, "_sample_curve_selected_callback", self._select_sample_from_curve)
        link_sample_curve_hover_plots(
            self.isotherm_plot,
            self.bet_plot,
            self.langmuir_plot,
            self.t_plot,
            self.pore_plot,
            self.dh_plot,
            self.hk_plot,
            self.dft_plot,
        )

        self.bet_default_button.setMinimumWidth(76)
        self.bet_default_button.clicked.connect(self.reset_bet_fit_to_default)
        bet_controls = QtWidgets.QHBoxLayout()
        bet_controls.setContentsMargins(6, 2, 6, 6)
        bet_controls.addWidget(self.bet_default_button)
        bet_controls.addStretch(1)
        self.bet_tab = QtWidgets.QWidget()
        bet_tab_layout = QtWidgets.QVBoxLayout(self.bet_tab)
        bet_tab_layout.setContentsMargins(0, 0, 0, 0)
        bet_tab_layout.setSpacing(0)
        bet_tab_layout.addWidget(self.bet_plot, 1)
        bet_tab_layout.addLayout(bet_controls)

        self.langmuir_default_button = QtWidgets.QPushButton("默认")
        self.langmuir_default_button.setToolTip("按默认 Langmuir 区间 0.05-0.30 重新计算")
        self.langmuir_default_button.setMinimumWidth(76)
        self.langmuir_default_button.clicked.connect(self.reset_langmuir_fit_to_default)
        langmuir_controls = QtWidgets.QHBoxLayout()
        langmuir_controls.setContentsMargins(6, 2, 6, 6)
        langmuir_controls.addWidget(self.langmuir_default_button)
        langmuir_controls.addStretch(1)
        self.langmuir_tab = QtWidgets.QWidget()
        langmuir_tab_layout = QtWidgets.QVBoxLayout(self.langmuir_tab)
        langmuir_tab_layout.setContentsMargins(0, 0, 0, 0)
        langmuir_tab_layout.setSpacing(0)
        langmuir_tab_layout.addWidget(self.langmuir_plot, 1)
        langmuir_tab_layout.addLayout(langmuir_controls)

        t_plot_plot_panel = QtWidgets.QWidget()
        t_plot_plot_layout = QtWidgets.QVBoxLayout(t_plot_plot_panel)
        t_plot_plot_layout.setContentsMargins(0, 0, 0, 0)
        t_plot_plot_layout.setSpacing(0)
        t_plot_plot_layout.addWidget(self.t_plot, 1)
        self.t_plot_options_panel = self._make_t_plot_options_panel()
        self.t_plot_tab = QtWidgets.QWidget()
        t_plot_tab_layout = QtWidgets.QHBoxLayout(self.t_plot_tab)
        t_plot_tab_layout.setContentsMargins(0, 0, 0, 0)
        t_plot_tab_layout.setSpacing(0)
        t_plot_tab_layout.addWidget(self.t_plot_options_panel)
        t_plot_tab_layout.addWidget(t_plot_plot_panel, 1)

        bjh_plot_panel = QtWidgets.QWidget()
        bjh_plot_layout = QtWidgets.QVBoxLayout(bjh_plot_panel)
        bjh_plot_layout.setContentsMargins(0, 0, 0, 0)
        bjh_plot_layout.setSpacing(0)
        bjh_plot_layout.addWidget(self.pore_plot, 1)
        self.bjh_options_panel = self._make_bjh_options_panel()
        self.bjh_tab = QtWidgets.QWidget()
        bjh_tab_layout = QtWidgets.QHBoxLayout(self.bjh_tab)
        bjh_tab_layout.setContentsMargins(0, 0, 0, 0)
        bjh_tab_layout.setSpacing(0)
        bjh_tab_layout.addWidget(self.bjh_options_panel)
        bjh_tab_layout.addWidget(bjh_plot_panel, 1)

        dh_plot_panel = QtWidgets.QWidget()
        dh_plot_layout = QtWidgets.QVBoxLayout(dh_plot_panel)
        dh_plot_layout.setContentsMargins(0, 0, 0, 0)
        dh_plot_layout.setSpacing(0)
        dh_plot_layout.addWidget(self.dh_plot, 1)
        self.dh_options_panel = self._make_dh_options_panel()
        self.dh_tab = QtWidgets.QWidget()
        dh_tab_layout = QtWidgets.QHBoxLayout(self.dh_tab)
        dh_tab_layout.setContentsMargins(0, 0, 0, 0)
        dh_tab_layout.setSpacing(0)
        dh_tab_layout.addWidget(self.dh_options_panel)
        dh_tab_layout.addWidget(dh_plot_panel, 1)

        hk_plot_panel = QtWidgets.QWidget()
        hk_plot_layout = QtWidgets.QVBoxLayout(hk_plot_panel)
        hk_plot_layout.setContentsMargins(0, 0, 0, 0)
        hk_plot_layout.setSpacing(0)
        hk_plot_layout.addWidget(self.hk_plot, 1)
        self.hk_options_panel = self._make_hk_options_panel()
        self.hk_tab = QtWidgets.QWidget()
        hk_tab_layout = QtWidgets.QHBoxLayout(self.hk_tab)
        hk_tab_layout.setContentsMargins(0, 0, 0, 0)
        hk_tab_layout.setSpacing(0)
        hk_tab_layout.addWidget(self.hk_options_panel)
        hk_tab_layout.addWidget(hk_plot_panel, 1)

        dft_plot_panel = QtWidgets.QWidget()
        dft_plot_layout = QtWidgets.QVBoxLayout(dft_plot_panel)
        dft_plot_layout.setContentsMargins(0, 0, 0, 0)
        dft_plot_layout.setSpacing(0)
        dft_plot_layout.addWidget(self.dft_plot, 1)
        self.dft_options_panel = self._make_dft_options_panel()
        self.dft_tab = QtWidgets.QWidget()
        dft_tab_layout = QtWidgets.QHBoxLayout(self.dft_tab)
        dft_tab_layout.setContentsMargins(0, 0, 0, 0)
        dft_tab_layout.setSpacing(0)
        dft_tab_layout.addWidget(self.dft_options_panel)
        dft_tab_layout.addWidget(dft_plot_panel, 1)

        self.plot_tabs = QtWidgets.QTabWidget()
        self.plot_tabs.addTab(self.bet_tab, "BET")
        self.plot_tabs.addTab(self.langmuir_tab, "Langmuir")
        self.plot_tabs.addTab(self.t_plot_tab, "t-Plot")
        self.plot_tabs.addTab(self.bjh_tab, "BJH")
        self.plot_tabs.addTab(self.dh_tab, "DH")
        self.plot_tabs.addTab(self.hk_tab, "HK")
        self.plot_tabs.addTab(self.dft_tab, "DFT")
        self.plot_tabs.currentChanged.connect(self.on_plot_tab_changed)

        self.isotherm_table = self._make_table(
            [
                "#",
                "阶段",
                "P/P0",
                "压力(mmHg)",
                "吸附量(cm3/g STP)",
                "吸附量(mmol/g)",
                "Po(mmHg)",
                "Elapsed",
            ]
        )
        self.target_table = self._make_table(["行", "阶段", "起始P/P0", "终止P/P0", "步长P/P0", "偏移"])
        self.condition_table = self._make_table(["字段", "值"])
        self.log_table = self._make_table(["来源", "偏移", "文本"])
        self.report_module_table = self._make_table(["SUBSET", "偏移", "文本"])

        self.detail_tabs = QtWidgets.QTabWidget()
        self.detail_tabs.addTab(self.metrics_table, "结果参数")
        self.detail_tabs.addTab(self.condition_table, "样品/条件")
        self.detail_tabs.addTab(self.isotherm_table, "实际等温线")
        self.detail_tabs.addTab(self.target_table, "目标压力表")
        self.detail_tabs.addTab(self.report_module_table, "报告模块")
        self.detail_tabs.addTab(self.log_table, "日志/样品管")

        sample_panel = QtWidgets.QWidget()
        sample_panel_layout = QtWidgets.QVBoxLayout(sample_panel)
        sample_panel_layout.setContentsMargins(0, 0, 0, 0)
        sample_panel_layout.setSpacing(0)
        sample_panel_layout.addWidget(self.sample_table, 1)

        self.left_splitter = QtWidgets.QSplitter(QtCore.Qt.Vertical)
        self.left_splitter.addWidget(sample_panel)
        self.left_splitter.addWidget(self.detail_tabs)
        self.left_splitter.setChildrenCollapsible(False)
        self.left_splitter.setHandleWidth(8)
        self.left_splitter.setSizes([190, 520])
        self.left_splitter.setStyleSheet(
            """
            QSplitter::handle:vertical {
                background: #e5e7eb;
                margin: 2px 0;
            }
            QSplitter::handle:vertical:hover {
                background: #93c5fd;
            }
            """
        )

        side_panel = QtWidgets.QWidget()
        side_layout = QtWidgets.QVBoxLayout(side_panel)
        side_layout.setContentsMargins(6, 6, 6, 6)
        side_layout.setSpacing(6)
        button_row = QtWidgets.QHBoxLayout()
        button_row.setContentsMargins(0, 0, 0, 0)
        button_row.setSpacing(6)
        button_row.addWidget(open_button)
        button_row.addWidget(export_button)
        button_row.addStretch(1)
        button_row.addWidget(self.update_available_button)
        button_row.addWidget(self.update_button)
        side_layout.addLayout(button_row)
        side_layout.addWidget(self.left_splitter, 1)

        self.dft_diagnostic_plot = make_plot(
            "拟合误差 vs 正则化",
            "RMS 拟合误差 (mmol/g)",
            "正则化",
        )
        self.dft_diagnostic_plot.setMinimumWidth(360)
        self.dft_diagnostic_plot.hide()
        bottom_plot_panel = QtWidgets.QWidget()
        bottom_plot_layout = QtWidgets.QHBoxLayout(bottom_plot_panel)
        bottom_plot_layout.setContentsMargins(0, 0, 0, 0)
        bottom_plot_layout.setSpacing(6)
        bottom_plot_layout.addWidget(self.isotherm_plot, 2)
        bottom_plot_layout.addWidget(self.dft_diagnostic_plot, 1)

        right_splitter = QtWidgets.QSplitter(QtCore.Qt.Vertical)
        right_splitter.addWidget(self.plot_tabs)
        right_splitter.addWidget(bottom_plot_panel)
        right_splitter.setChildrenCollapsible(False)
        right_splitter.setHandleWidth(8)
        right_splitter.setStretchFactor(0, 3)
        right_splitter.setStretchFactor(1, 2)
        right_splitter.setSizes([480, 300])
        right_splitter.setStyleSheet(
            """
            QSplitter::handle:vertical {
                background: #e5e7eb;
                margin: 2px 0;
            }
            QSplitter::handle:vertical:hover {
                background: #93c5fd;
            }
            """
        )

        splitter = QtWidgets.QSplitter(QtCore.Qt.Horizontal)
        splitter.addWidget(side_panel)
        splitter.addWidget(right_splitter)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 3)
        splitter.setSizes([390, 890])
        self.setCentralWidget(splitter)

        self.statusBar().showMessage("打开或拖入 SMP、DAT、QPS 或官方 Excel 导出文件")
        self.refresh_all()
        self._sync_select_all_state()
        QtCore.QTimer.singleShot(0, self._position_header_controls)
        QtCore.QTimer.singleShot(AUTO_UPDATE_CHECK_DELAY_MS, self._auto_check_for_updates)

    def _auto_check_for_updates(self) -> None:
        self.check_for_updates(manual=False)

    def check_for_updates(self, _checked: bool = False, *, manual: bool = True) -> None:
        if self._checking_for_updates:
            if manual:
                self.statusBar().showMessage("正在检查软件更新...", 3000)
            return

        self._checking_for_updates = True
        update_button = getattr(self, "update_button", None)
        if update_button is not None:
            update_button.setEnabled(False)
        if manual:
            self.statusBar().showMessage("正在连接更新源检查软件更新...", 3000)

        thread = QtCore.QThread(self)
        worker = UpdateCheckWorker(APP_VERSION, UPDATE_REPOSITORY, manual)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.finished.connect(self._on_update_check_finished)
        worker.failed.connect(self._on_update_check_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.failed.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(self._clear_update_check_worker)
        self._update_thread = thread
        self._update_worker = worker
        thread.start()

    def _on_update_check_finished(self, info: UpdateInfo, manual: bool) -> None:
        self._finish_update_check(manual)
        if not info.update_available:
            self._set_update_available_indicator(None)
            message = f"当前已是最新版本 v{info.current_version}"
            if manual:
                QtWidgets.QMessageBox.information(self, "软件更新", message)
            return

        self._set_update_available_indicator(info)
        if not manual:
            self.statusBar().showMessage(f"发现新版本 v{info.latest_version}，点击软件更新左侧图标即可更新。", 8000)
            return

        self._show_update_available_dialog(info)

    def _show_pending_update_dialog(self) -> None:
        info = self._available_update_info
        if info is None:
            self.check_for_updates(manual=True)
            return
        self._show_update_available_dialog(info)

    def _show_update_available_dialog(self, info: UpdateInfo) -> None:
        title = f"发现新版本 v{info.latest_version}"
        download_hint = f"安装包: {info.asset_name}" if info.asset_name else "安装包: 自动选择"
        release_notes = str(info.release_notes or "").strip()
        notes_hint = f"\n\n本次更新内容:\n{release_notes}" if release_notes else "\n\n本次更新内容:\n暂无更新说明。"
        message = (
            f"当前版本: v{info.current_version}\n"
            f"最新版本: v{info.latest_version}\n\n"
            f"来源: DragonScience\n"
            f"{download_hint}"
            f"{notes_hint}\n\n"
            "是否现在下载并重启到新版本？"
        )
        box = QtWidgets.QMessageBox(self)
        box.setIcon(QtWidgets.QMessageBox.Information)
        box.setWindowTitle("软件更新")
        box.setText(title)
        box.setInformativeText(message)
        update_button = box.addButton("更新", QtWidgets.QMessageBox.AcceptRole)
        box.addButton("稍后", QtWidgets.QMessageBox.RejectRole)
        exec_func = getattr(box, "exec", box.exec_)
        exec_func()
        if box.clickedButton() == update_button:
            self._download_and_install_update(info)

    def _set_update_available_indicator(self, info: UpdateInfo | None, *, enabled: bool = True) -> None:
        self._available_update_info = info
        button = getattr(self, "update_available_button", None)
        if button is None:
            return
        has_update = info is not None and bool(info.update_available)
        button.setVisible(has_update)
        button.setEnabled(bool(enabled))
        if has_update:
            button.setToolTip(f"发现新版本 v{info.latest_version}，点击更新")
        else:
            button.setToolTip("发现新版本，点击更新")

    def _on_update_check_failed(self, message: str, manual: bool) -> None:
        self._finish_update_check(manual)
        if manual:
            QtWidgets.QMessageBox.warning(self, "软件更新检查失败", message)
        else:
            self.statusBar().showMessage(f"自动检查软件更新失败: {message}", 5000)

    def _finish_update_check(self, manual: bool) -> None:
        self._checking_for_updates = False
        update_button = getattr(self, "update_button", None)
        if update_button is not None:
            update_button.setEnabled(True)
        if not manual:
            self.settings.setValue("updates/last_auto_check_date", datetime.now().date().isoformat())

    def _clear_update_check_worker(self) -> None:
        self._update_thread = None
        self._update_worker = None

    def _download_and_install_update(self, info: UpdateInfo) -> None:
        if self._update_download_thread is not None:
            self.statusBar().showMessage("正在下载软件更新...", 3000)
            return
        if not info.download_url:
            self._open_update_page(info)
            return

        self._show_update_progress(info)
        update_button = getattr(self, "update_button", None)
        if update_button is not None:
            update_button.setEnabled(False)
        self._set_update_available_indicator(info, enabled=False)

        thread = QtCore.QThread(self)
        worker = UpdateDownloadWorker(info)
        worker.moveToThread(thread)
        thread.started.connect(worker.run)
        worker.progress.connect(self._on_update_download_progress)
        worker.finished.connect(self._on_update_download_finished)
        worker.failed.connect(self._on_update_download_failed)
        worker.finished.connect(thread.quit)
        worker.failed.connect(thread.quit)
        worker.finished.connect(worker.deleteLater)
        worker.failed.connect(worker.deleteLater)
        thread.finished.connect(thread.deleteLater)
        thread.finished.connect(self._clear_update_download_worker)
        self._update_download_thread = thread
        self._update_download_worker = worker
        thread.start()

    def _show_update_progress(self, info: UpdateInfo) -> None:
        dialog = QtWidgets.QProgressDialog(f"正在下载 v{info.latest_version}...", None, 0, 100, self)
        dialog.setWindowTitle("软件更新")
        dialog.setWindowModality(QtCore.Qt.WindowModal)
        dialog.setAutoClose(False)
        dialog.setAutoReset(False)
        dialog.setMinimumDuration(0)
        dialog.setValue(0)
        dialog.setCancelButton(None)
        dialog.setStyleSheet(
            """
            QProgressBar {
                border: 1px solid #9ca3af;
                border-radius: 4px;
                text-align: center;
                background: #f3f4f6;
            }
            QProgressBar::chunk {
                background: #22c55e;
                border-radius: 3px;
            }
            """
        )
        self._update_progress_dialog = dialog
        dialog.show()

    def _on_update_download_progress(self, downloaded: int, total: int) -> None:
        dialog = self._update_progress_dialog
        if dialog is None:
            return
        if total > 0:
            dialog.setRange(0, 100)
            value = max(0, min(100, int(downloaded * 100 / total)))
            dialog.setValue(value)
            dialog.setLabelText(f"正在下载软件更新... {downloaded / 1024 / 1024:.1f} / {total / 1024 / 1024:.1f} MB")
        else:
            dialog.setRange(0, 0)
            dialog.setLabelText(f"正在下载软件更新... {downloaded / 1024 / 1024:.1f} MB")

    def _on_update_download_finished(self, info: UpdateInfo, path: str) -> None:
        dialog = self._update_progress_dialog
        if dialog is not None:
            dialog.setRange(0, 100)
            dialog.setValue(100)
            dialog.setLabelText("下载完成，正在安装更新...")
        self.statusBar().showMessage(f"已下载 v{info.latest_version}，正在安装更新，请稍后重新打开软件。", 3000)
        QtCore.QTimer.singleShot(800, lambda: self._launch_downloaded_update(path))

    def _on_update_download_failed(self, message: str) -> None:
        dialog = self._update_progress_dialog
        if dialog is not None:
            dialog.close()
            self._update_progress_dialog = None
        update_button = getattr(self, "update_button", None)
        if update_button is not None:
            update_button.setEnabled(True)
        if self._available_update_info is not None:
            self._set_update_available_indicator(self._available_update_info, enabled=True)
        QtWidgets.QMessageBox.warning(self, "软件更新失败", message)

    def _clear_update_download_worker(self) -> None:
        self._update_download_thread = None
        self._update_download_worker = None

    def _launch_downloaded_update(self, path: str) -> None:
        try:
            launch_update_and_exit(Path(path))
        except UpdateDownloadError as exc:
            if self._update_progress_dialog is not None:
                self._update_progress_dialog.close()
                self._update_progress_dialog = None
            QtWidgets.QMessageBox.warning(self, "软件更新失败", str(exc))
            update_button = getattr(self, "update_button", None)
            if update_button is not None:
                update_button.setEnabled(True)
            if self._available_update_info is not None:
                self._set_update_available_indicator(self._available_update_info, enabled=True)
            return
        if self._update_progress_dialog is not None:
            self._update_progress_dialog.close()
            self._update_progress_dialog = None
        killer = threading.Timer(5.0, lambda: os._exit(0))
        killer.daemon = True
        killer.start()
        for widget in QtWidgets.QApplication.topLevelWidgets():
            widget.close()
        QtWidgets.QApplication.exit(0)

    def _open_update_page(self, info: UpdateInfo) -> None:
        url = info.download_url or info.release_url
        if not url:
            QtWidgets.QMessageBox.warning(self, "软件更新", "没有可打开的下载链接。")
            return
        QtGui.QDesktopServices.openUrl(QtCore.QUrl(url))

    def _make_t_plot_options_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        panel.setFixedWidth(T_PLOT_PANEL_COLLAPSED_WIDTH)
        panel_layout = QtWidgets.QVBoxLayout(panel)
        panel_layout.setContentsMargins(6, 6, 6, 6)
        panel_layout.setSpacing(8)

        thickness_group = QtWidgets.QGroupBox("厚度曲线")
        thickness_layout = QtWidgets.QVBoxLayout(thickness_group)
        thickness_layout.setContentsMargins(8, 8, 8, 8)
        thickness_layout.setSpacing(4)
        self.t_plot_method_radios = {}
        self.t_plot_method_group = QtWidgets.QButtonGroup(panel)
        self.t_plot_method_group.setExclusive(True)
        self.t_plot_param_spins = {}
        self.t_plot_formula_buttons = {}
        self.t_plot_formula_widgets = {}
        self.t_plot_formula_expanded = {}
        thickness_methods = [
            ("reference", "参比", True),
            ("kjs", "Kruk-Jaroniec-Sayari E", True),
            ("halsey", "Halsey", True),
            ("harkins_jura", "Harkins and Jura 厚度", True),
            ("broekhoff_de_boer", "Broekhoff-De Boer 厚度", True),
            ("carbon_black_stsa", "碳黑STSA", True),
        ]
        for key, label, enabled in thickness_methods:
            row = self._make_thickness_method_row(key, label, enabled, context="t_plot")
            thickness_layout.addWidget(row)

        surface_group = QtWidgets.QGroupBox("表面积")
        surface_layout = QtWidgets.QVBoxLayout(surface_group)
        surface_layout.setContentsMargins(8, 8, 8, 8)
        surface_layout.setSpacing(6)
        self.surface_area_bet_radio = QtWidgets.QRadioButton("BET")
        self.surface_area_langmuir_radio = QtWidgets.QRadioButton("Langmuir")
        self.surface_area_input_radio = QtWidgets.QRadioButton("输入")
        self.surface_area_bet_radio.setChecked(True)
        self.surface_area_bet_radio.toggled.connect(self._on_t_plot_surface_area_mode_changed)
        self.surface_area_langmuir_radio.toggled.connect(self._on_t_plot_surface_area_mode_changed)
        self.surface_area_input_radio.toggled.connect(self._on_t_plot_surface_area_mode_changed)
        surface_layout.addWidget(self.surface_area_bet_radio)
        surface_layout.addWidget(self.surface_area_langmuir_radio)
        input_row = QtWidgets.QHBoxLayout()
        input_row.addWidget(self.surface_area_input_radio)
        self.surface_area_input_spin = self._make_param_spin(1.0, 0.0, 1000000.0, 3)
        self.surface_area_input_spin.setEnabled(False)
        self.surface_area_input_spin.valueChanged.connect(self._on_t_plot_surface_area_input_changed)
        input_row.addWidget(self.surface_area_input_spin)
        input_row.addWidget(QtWidgets.QLabel("m²/g"))
        input_row.addStretch(1)
        surface_layout.addLayout(input_row)

        correction_group = QtWidgets.QGroupBox("表面积校正因子")
        correction_layout = QtWidgets.QVBoxLayout(correction_group)
        correction_layout.setContentsMargins(8, 8, 8, 8)
        self.surface_area_correction_spin = self._make_param_spin(SURFACE_AREA_CORRECTION_FACTOR, 0.0, 1000.0, 3)
        self.surface_area_correction_spin.valueChanged.connect(self._on_t_plot_surface_area_correction_changed)
        correction_layout.addWidget(self.surface_area_correction_spin)

        self.t_plot_default_button = QtWidgets.QPushButton("默认")
        self.t_plot_default_button.setToolTip("重置当前样品的 t-Plot 厚度曲线、表面积选项和拟合区间")
        self.t_plot_default_button.clicked.connect(self.reset_t_plot_fit_to_default)
        default_button_row = QtWidgets.QHBoxLayout()
        default_button_row.setContentsMargins(0, 0, 0, 0)
        default_button_row.addWidget(self.t_plot_default_button)
        default_button_row.addStretch(1)

        panel_layout.addWidget(thickness_group)
        panel_layout.addWidget(surface_group)
        panel_layout.addWidget(correction_group)
        panel_layout.addLayout(default_button_row)
        panel_layout.addStretch(1)
        self._update_t_plot_options_panel_width(panel)
        return panel

    def _make_bjh_options_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        panel.setFixedWidth(BJH_PANEL_COLLAPSED_WIDTH)
        panel_layout = QtWidgets.QVBoxLayout(panel)
        panel_layout.setContentsMargins(6, 6, 6, 6)
        panel_layout.setSpacing(8)

        thickness_group = QtWidgets.QGroupBox("厚度曲线")
        thickness_layout = QtWidgets.QVBoxLayout(thickness_group)
        thickness_layout.setContentsMargins(8, 8, 8, 8)
        thickness_layout.setSpacing(4)
        self.bjh_method_radios = {}
        self.bjh_method_group = QtWidgets.QButtonGroup(panel)
        self.bjh_method_group.setExclusive(True)
        self.bjh_param_spins = {}
        self.bjh_formula_buttons = {}
        self.bjh_formula_widgets = {}
        self.bjh_formula_expanded = {}
        thickness_methods = [
            ("reference", "参比", True),
            ("kjs", "Kruk-Jaroniec-Sayari E", True),
            ("halsey", "Halsey", True),
            ("harkins_jura", "Harkins and Jura 厚度", True),
            ("broekhoff_de_boer", "Broekhoff-De Boer 厚度", True),
            ("carbon_black_stsa", "碳黑STSA", True),
        ]
        for key, label, enabled in thickness_methods:
            thickness_layout.addWidget(self._make_thickness_method_row(key, label, enabled, context="bjh"))

        correction_group = QtWidgets.QGroupBox("BJH 校正")
        correction_layout = QtWidgets.QVBoxLayout(correction_group)
        correction_layout.setContentsMargins(8, 8, 8, 8)
        correction_layout.setSpacing(6)
        self.bjh_standard_radio = QtWidgets.QRadioButton("标准的")
        self.bjh_kjs_correction_radio = QtWidgets.QRadioButton("Kruk-Jaroniec-Sayari E")
        self.bjh_faas_correction_radio = QtWidgets.QRadioButton("Faas 校正")
        self.bjh_standard_radio.setChecked(True)
        for radio in (self.bjh_standard_radio, self.bjh_kjs_correction_radio, self.bjh_faas_correction_radio):
            radio.toggled.connect(self._on_bjh_option_changed)
            correction_layout.addWidget(radio)

        open_fraction_label = QtWidgets.QLabel("两端开口孔的分数")
        self.bjh_open_fraction_spin = self._make_param_spin(0.0, 0.0, 1.0, 2, width=112)
        self.bjh_open_fraction_spin.setToolTip("暂作为 BJH 参数保留；当前标准 BJH 计算中不改变结果")
        self.bjh_open_fraction_spin.valueChanged.connect(self._on_bjh_option_changed)
        open_fraction_row = QtWidgets.QHBoxLayout()
        open_fraction_row.setContentsMargins(0, 0, 0, 0)
        open_fraction_row.addWidget(self.bjh_open_fraction_spin)
        open_fraction_row.addStretch(1)

        self.bjh_smooth_checkbox = QtWidgets.QCheckBox("平滑的微分")
        self.bjh_smooth_checkbox.setChecked(True)
        self.bjh_smooth_checkbox.stateChanged.connect(self._on_bjh_option_changed)
        self.bjh_adsorption_checkbox = QtWidgets.QCheckBox("BJH 吸附")
        self.bjh_desorption_checkbox = QtWidgets.QCheckBox("BJH 脱附")
        self.bjh_adsorption_checkbox.setChecked(True)
        self.bjh_desorption_checkbox.setChecked(False)
        self.bjh_adsorption_checkbox.stateChanged.connect(self._on_bjh_option_changed)
        self.bjh_desorption_checkbox.stateChanged.connect(self._on_bjh_option_changed)

        display_group = QtWidgets.QGroupBox("显示模式")
        display_layout = QtWidgets.QVBoxLayout(display_group)
        display_layout.setContentsMargins(8, 8, 8, 8)
        display_layout.setSpacing(4)
        self.bjh_display_combo = QtWidgets.QComboBox()
        for metric in BJH_DISPLAY_METRIC_ORDER:
            self.bjh_display_combo.addItem(bjh_display_metric_label(metric), metric)
        self.bjh_display_combo.setToolTip("切换 BJH 孔径分布图的纵轴显示方式")
        self._set_bjh_display_combo()
        self.bjh_display_combo.currentIndexChanged.connect(self._on_bjh_display_mode_changed)
        display_layout.addWidget(self.bjh_display_combo)

        self.bjh_default_button = QtWidgets.QPushButton("默认")
        self.bjh_default_button.setToolTip("重置 BJH 厚度曲线、校正参数和显示分支")
        self.bjh_default_button.clicked.connect(self.reset_bjh_to_default)
        self.bjh_apply_all_button = QtWidgets.QPushButton("全部应用")
        self.bjh_apply_all_button.setToolTip("把当前 BJH 厚度曲线、校正参数、平滑和吸/脱附设置应用到所有样品")
        self.bjh_apply_all_button.clicked.connect(self.apply_bjh_settings_to_all)
        self.bjh_all_default_button = QtWidgets.QPushButton("全部默认")
        self.bjh_all_default_button.setToolTip("让所有样品恢复各自文件或厂商默认的 BJH 设置")
        self.bjh_all_default_button.clicked.connect(self.reset_all_bjh_to_default)
        default_button_row = QtWidgets.QHBoxLayout()
        default_button_row.setContentsMargins(0, 0, 0, 0)
        default_button_row.addWidget(self.bjh_default_button)
        default_button_row.addWidget(self.bjh_apply_all_button)
        default_button_row.addWidget(self.bjh_all_default_button)
        default_button_row.addStretch(1)

        panel_layout.addWidget(thickness_group)
        panel_layout.addWidget(correction_group)
        panel_layout.addWidget(open_fraction_label)
        panel_layout.addLayout(open_fraction_row)
        panel_layout.addWidget(self.bjh_smooth_checkbox)
        panel_layout.addWidget(self.bjh_adsorption_checkbox)
        panel_layout.addWidget(self.bjh_desorption_checkbox)
        panel_layout.addWidget(display_group)
        panel_layout.addLayout(default_button_row)
        panel_layout.addStretch(1)
        self._update_bjh_options_panel_width(panel)
        return panel

    def _make_dh_options_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        panel.setFixedWidth(BJH_PANEL_COLLAPSED_WIDTH)
        panel_layout = QtWidgets.QVBoxLayout(panel)
        panel_layout.setContentsMargins(6, 6, 6, 6)
        panel_layout.setSpacing(8)

        thickness_group = QtWidgets.QGroupBox("厚度曲线")
        thickness_layout = QtWidgets.QVBoxLayout(thickness_group)
        thickness_layout.setContentsMargins(8, 8, 8, 8)
        thickness_layout.setSpacing(4)
        self.dh_method_radios = {}
        self.dh_method_group = QtWidgets.QButtonGroup(panel)
        self.dh_method_group.setExclusive(True)
        self.dh_param_spins = {}
        self.dh_formula_buttons = {}
        self.dh_formula_widgets = {}
        self.dh_formula_expanded = {}
        thickness_methods = [
            ("reference", "参比", True),
            ("kjs", "Kruk-Jaroniec-Sayari E", True),
            ("halsey", "Halsey", True),
            ("harkins_jura", "Harkins and Jura 厚度", True),
            ("broekhoff_de_boer", "Broekhoff-De Boer 厚度", True),
            ("carbon_black_stsa", "碳黑STSA", True),
        ]
        for key, label, enabled in thickness_methods:
            thickness_layout.addWidget(self._make_thickness_method_row(key, label, enabled, context="dh"))

        self.dh_adsorption_checkbox = QtWidgets.QCheckBox("DH 吸附")
        self.dh_desorption_checkbox = QtWidgets.QCheckBox("DH 脱附")
        self.dh_adsorption_checkbox.setChecked(DEFAULT_DH_SHOW_ADSORPTION)
        self.dh_desorption_checkbox.setChecked(DEFAULT_DH_SHOW_DESORPTION)
        self.dh_adsorption_checkbox.stateChanged.connect(self._on_dh_option_changed)
        self.dh_desorption_checkbox.stateChanged.connect(self._on_dh_option_changed)

        self.dh_smooth_checkbox = QtWidgets.QCheckBox("平滑的微分")
        self.dh_smooth_checkbox.setChecked(DEFAULT_DH_SMOOTH_DERIVATIVE)
        self.dh_smooth_checkbox.stateChanged.connect(self._on_dh_option_changed)

        display_group = QtWidgets.QGroupBox("显示模式")
        display_layout = QtWidgets.QVBoxLayout(display_group)
        display_layout.setContentsMargins(8, 8, 8, 8)
        display_layout.setSpacing(4)
        self.dh_display_combo = QtWidgets.QComboBox()
        for metric in BJH_DISPLAY_METRIC_ORDER:
            self.dh_display_combo.addItem(bjh_display_metric_label(metric), metric)
        self.dh_display_combo.setToolTip("切换 DH 孔径分布图的纵轴显示方式")
        self._set_dh_display_combo()
        self.dh_display_combo.currentIndexChanged.connect(self._on_dh_display_mode_changed)
        display_layout.addWidget(self.dh_display_combo)

        self.dh_default_button = QtWidgets.QPushButton("默认")
        self.dh_default_button.setToolTip("重置 DH 显示分支、平滑和显示模式")
        self.dh_default_button.clicked.connect(self.reset_dh_to_default)
        self.dh_apply_all_button = QtWidgets.QPushButton("全部应用")
        self.dh_apply_all_button.setToolTip("把当前 DH 厚度曲线、平滑和吸/脱附设置应用到所有样品")
        self.dh_apply_all_button.clicked.connect(self.apply_dh_settings_to_all)
        self.dh_all_default_button = QtWidgets.QPushButton("全部默认")
        self.dh_all_default_button.setToolTip("让所有样品恢复各自文件或厂商默认的 DH 设置")
        self.dh_all_default_button.clicked.connect(self.reset_all_dh_to_default)
        default_button_row = QtWidgets.QHBoxLayout()
        default_button_row.setContentsMargins(0, 0, 0, 0)
        default_button_row.addWidget(self.dh_default_button)
        default_button_row.addWidget(self.dh_apply_all_button)
        default_button_row.addWidget(self.dh_all_default_button)
        default_button_row.addStretch(1)

        panel_layout.addWidget(thickness_group)
        panel_layout.addWidget(self.dh_smooth_checkbox)
        panel_layout.addWidget(self.dh_adsorption_checkbox)
        panel_layout.addWidget(self.dh_desorption_checkbox)
        panel_layout.addWidget(display_group)
        panel_layout.addLayout(default_button_row)
        panel_layout.addStretch(1)
        return panel

    def _make_hk_options_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        panel.setFixedWidth(BJH_PANEL_COLLAPSED_WIDTH)
        panel_layout = QtWidgets.QVBoxLayout(panel)
        panel_layout.setContentsMargins(6, 6, 6, 6)
        panel_layout.setSpacing(8)

        geometry_group = QtWidgets.QGroupBox("孔型")
        geometry_layout = QtWidgets.QVBoxLayout(geometry_group)
        geometry_layout.setContentsMargins(8, 8, 8, 8)
        geometry_layout.setSpacing(4)
        self.hk_geometry_group = QtWidgets.QButtonGroup(panel)
        self.hk_geometry_group.setExclusive(True)
        self.hk_geometry_radios = {}
        for key, label, enabled in (
            ("slit", "狭缝(原始H-K)", True),
            ("cylinder", "圆柱(Satio-Foley)", True),
            ("sphere", "球形", False),
        ):
            radio = QtWidgets.QRadioButton(label)
            radio.setEnabled(enabled)
            if not enabled:
                radio.setToolTip("球形 Cheng-Yang 公式的单位尺度还需要官方导出数据校验，暂不开放")
            radio.setChecked(key == self.hk_geometry)
            radio.toggled.connect(self._on_hk_option_changed)
            self.hk_geometry_group.addButton(radio)
            self.hk_geometry_radios[key] = radio
            geometry_layout.addWidget(radio)

        interaction_group = QtWidgets.QGroupBox("作用参数")
        interaction_layout = QtWidgets.QVBoxLayout(interaction_group)
        interaction_layout.setContentsMargins(8, 8, 8, 8)
        interaction_layout.setSpacing(6)
        self.hk_calculated_radio = QtWidgets.QRadioButton("计算的")
        self.hk_input_radio = QtWidgets.QRadioButton("输入的")
        self.hk_input_radio.setChecked(True)
        self.hk_calculated_radio.toggled.connect(self._on_hk_option_changed)
        self.hk_input_radio.toggled.connect(self._on_hk_option_changed)
        interaction_layout.addWidget(self.hk_calculated_radio)
        calc_label = QtWidgets.QLabel(f"{calculate_hk_interaction_parameter():.3e} erg·cm^4")
        calc_label.setStyleSheet("color: #6b7280; margin-left: 22px;")
        self.hk_calculated_value_label = calc_label
        interaction_layout.addWidget(calc_label)
        input_row = QtWidgets.QHBoxLayout()
        input_row.addWidget(self.hk_input_radio)
        self.hk_interaction_edit = QtWidgets.QLineEdit(f"{self.hk_interaction_parameter:.3e}")
        self.hk_interaction_edit.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self.hk_interaction_edit.setFixedWidth(112)
        validator = QtGui.QDoubleValidator(self.hk_interaction_edit)
        validator.setNotation(QtGui.QDoubleValidator.ScientificNotation)
        self.hk_interaction_edit.setValidator(validator)
        self.hk_interaction_edit.editingFinished.connect(self._on_hk_option_changed)
        input_row.addWidget(self.hk_interaction_edit)
        input_row.addWidget(QtWidgets.QLabel("erg·cm^4"))
        input_row.addStretch(1)
        interaction_layout.addLayout(input_row)
        self.hk_properties_button = QtWidgets.QPushButton("性质")
        self.hk_properties_button.clicked.connect(self._open_hk_properties_dialog)
        property_row = QtWidgets.QHBoxLayout()
        property_row.addWidget(self.hk_properties_button)
        property_row.addStretch(1)
        interaction_layout.addLayout(property_row)

        self.hk_cheng_yang_checkbox = QtWidgets.QCheckBox("使用Cheng-Yang校正")
        self.hk_cheng_yang_checkbox.setChecked(DEFAULT_HK_CHENG_YANG_CORRECTION)
        self.hk_cheng_yang_checkbox.stateChanged.connect(self._on_hk_option_changed)
        self.hk_smooth_checkbox = QtWidgets.QCheckBox("平滑微分")
        self.hk_smooth_checkbox.setChecked(DEFAULT_HK_SMOOTH_DERIVATIVE)
        self.hk_smooth_checkbox.stateChanged.connect(self._on_hk_option_changed)

        display_group = QtWidgets.QGroupBox("显示模式")
        display_layout = QtWidgets.QVBoxLayout(display_group)
        display_layout.setContentsMargins(8, 8, 8, 8)
        display_layout.setSpacing(4)
        self.hk_display_combo = QtWidgets.QComboBox()
        for metric in HK_DISPLAY_METRIC_ORDER:
            self.hk_display_combo.addItem(hk_display_metric_label(metric), metric)
        self._set_hk_display_combo()
        self.hk_display_combo.currentIndexChanged.connect(self._on_hk_display_mode_changed)
        display_layout.addWidget(self.hk_display_combo)

        self.hk_default_button = QtWidgets.QPushButton("默认")
        self.hk_default_button.setToolTip("重置 HK 孔型、物性、作用参数和平滑选项")
        self.hk_default_button.clicked.connect(self.reset_hk_to_default)
        self.hk_apply_all_button = QtWidgets.QPushButton("全部应用")
        self.hk_apply_all_button.setToolTip("把当前 HK 孔型、物性、作用参数和平滑选项应用到所有样品")
        self.hk_apply_all_button.clicked.connect(self.apply_hk_settings_to_all)
        self.hk_all_default_button = QtWidgets.QPushButton("全部默认")
        self.hk_all_default_button.setToolTip("让所有样品恢复默认 HK 设置")
        self.hk_all_default_button.clicked.connect(self.reset_all_hk_to_default)
        default_button_row = QtWidgets.QHBoxLayout()
        default_button_row.setContentsMargins(0, 0, 0, 0)
        default_button_row.addWidget(self.hk_default_button)
        default_button_row.addWidget(self.hk_apply_all_button)
        default_button_row.addWidget(self.hk_all_default_button)
        default_button_row.addStretch(1)

        panel_layout.addWidget(geometry_group)
        panel_layout.addWidget(interaction_group)
        panel_layout.addWidget(self.hk_cheng_yang_checkbox)
        panel_layout.addWidget(self.hk_smooth_checkbox)
        panel_layout.addWidget(display_group)
        panel_layout.addLayout(default_button_row)
        panel_layout.addStretch(1)
        self._update_hk_interaction_controls()
        return panel

    def _make_dft_options_panel(self) -> QtWidgets.QWidget:
        panel = QtWidgets.QWidget()
        panel.setFixedWidth(DFT_PANEL_WIDTH)
        panel_layout = QtWidgets.QVBoxLayout(panel)
        panel_layout.setContentsMargins(6, 6, 6, 6)
        panel_layout.setSpacing(8)

        method_group = QtWidgets.QGroupBox("模型设置")
        method_layout = QtWidgets.QFormLayout(method_group)
        method_layout.setContentsMargins(8, 8, 8, 8)
        method_layout.setHorizontalSpacing(8)
        method_layout.setVerticalSpacing(6)

        self.dft_type_combo = QtWidgets.QComboBox()
        self.dft_type_combo.addItem("DFT 孔径", "dft_pore")
        self.dft_type_combo.addItem("典型", "typical")
        self.dft_type_combo.currentIndexChanged.connect(self._on_dft_option_changed)
        method_layout.addRow("类型", self.dft_type_combo)

        self.dft_geometry_combo = QtWidgets.QComboBox()
        self.dft_geometry_combo.addItem("狭缝", "slit")
        self.dft_geometry_combo.addItem("圆柱", "cylinder")
        self.dft_geometry_combo.currentIndexChanged.connect(self._on_dft_option_changed)
        method_layout.addRow("结构", self.dft_geometry_combo)

        self.dft_model_combo = QtWidgets.QComboBox()
        for key, label in dft_model_options():
            if key == DEFAULT_DFT_MODEL:
                self.dft_model_combo.addItem(label, key)
        if self.dft_model_combo.count() == 0:
            self.dft_model_combo.addItem("N2 - DFT Model", DEFAULT_DFT_MODEL)
        self.dft_model_combo.currentIndexChanged.connect(self._on_dft_option_changed)
        method_layout.addRow("模型", self.dft_model_combo)

        regularization_group = QtWidgets.QGroupBox("正则化")
        regularization_layout = QtWidgets.QVBoxLayout(regularization_group)
        regularization_layout.setContentsMargins(8, 8, 8, 8)
        regularization_layout.setSpacing(4)
        self.dft_regularization_slider = RegularizationSlider(
            DFT_REGULARIZATION_VALUES,
            self.dft_regularization,
            regularization_group,
        )
        self.dft_regularization_slider.valueChanged.connect(self._on_dft_regularization_changed)
        self.dft_regularization_slider.valueChangeFinished.connect(self._finish_deferred_dft_regularization_refresh)
        regularization_layout.addWidget(self.dft_regularization_slider)
        self.dft_regularization_apply_all_checkbox = QtWidgets.QCheckBox("应用全部")
        self.dft_regularization_apply_all_checkbox.setToolTip("勾选后，拖动正则化会同步到所有样品")
        self.dft_regularization_apply_all_checkbox.toggled.connect(self._on_dft_regularization_apply_all_toggled)
        regularization_layout.addWidget(self.dft_regularization_apply_all_checkbox)

        self.dft_default_button = QtWidgets.QPushButton("默认")
        self.dft_default_button.setToolTip("恢复当前样品的 DFT 默认设置")
        self.dft_default_button.clicked.connect(self.reset_dft_to_default)
        self.dft_apply_all_button = QtWidgets.QPushButton("全部应用")
        self.dft_apply_all_button.setToolTip("把当前 DFT 设置应用到所有样品")
        self.dft_apply_all_button.clicked.connect(self.apply_dft_settings_to_all)
        self.dft_all_default_button = QtWidgets.QPushButton("全部默认")
        self.dft_all_default_button.setToolTip("让所有样品恢复各自默认 DFT 设置")
        self.dft_all_default_button.clicked.connect(self.reset_all_dft_to_default)
        button_row = QtWidgets.QHBoxLayout()
        button_row.setContentsMargins(0, 0, 0, 0)
        button_row.addWidget(self.dft_default_button)
        button_row.addWidget(self.dft_apply_all_button)
        button_row.addWidget(self.dft_all_default_button)
        button_row.addStretch(1)

        self.dft_diagnostic_plot = make_plot(
            "拟合误差 vs 正则化",
            "RMS 拟合误差 (mmol/g)",
            "正则化",
        )
        self.dft_diagnostic_plot.setMinimumHeight(190)
        self.dft_diagnostic_plot.setMaximumHeight(260)

        panel_layout.addWidget(method_group)
        panel_layout.addWidget(regularization_group)
        panel_layout.addLayout(button_row)
        self.dft_diagnostic_plot.setVisible(False)
        panel_layout.addWidget(self.dft_diagnostic_plot)
        panel_layout.addStretch(1)
        return panel

    def _make_thickness_method_row(self, key: str, label: str, enabled: bool, *, context: str = "t_plot") -> QtWidgets.QWidget:
        container = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(container)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(2)

        row_layout = QtWidgets.QHBoxLayout()
        row_layout.setContentsMargins(0, 0, 0, 0)
        radio = QtWidgets.QRadioButton(label)
        radio.setEnabled(enabled)
        radio.setChecked(key == self._thickness_method_for_context(context))
        if context == "bjh":
            radio.toggled.connect(lambda checked, method_key=key: self._on_bjh_thickness_method_changed(method_key, checked))
            self.bjh_method_radios[key] = radio
            self.bjh_method_group.addButton(radio)
        elif context == "dh":
            radio.toggled.connect(lambda checked, method_key=key: self._on_dh_thickness_method_changed(method_key, checked))
            self.dh_method_radios[key] = radio
            self.dh_method_group.addButton(radio)
        else:
            radio.toggled.connect(lambda checked, method_key=key: self._on_t_plot_thickness_method_changed(method_key, checked))
            self.t_plot_method_radios[key] = radio
            self.t_plot_method_group.addButton(radio)
        arrow = QtWidgets.QToolButton()
        arrow.setArrowType(QtCore.Qt.DownArrow)
        arrow.setAutoRaise(True)
        arrow.setFixedSize(22, 22)
        row_layout.addWidget(radio, 1)
        row_layout.addWidget(arrow)

        formula = self._make_t_plot_formula_widget(key, context=context) if enabled else self._make_pending_formula_widget()
        formula.setVisible(False)
        if context == "bjh":
            arrow.clicked.connect(lambda _checked=False, method_key=key: self._toggle_bjh_formula(method_key))
            self.bjh_formula_buttons[key] = arrow
            self.bjh_formula_widgets[key] = formula
            self.bjh_formula_expanded[key] = False
        elif context == "dh":
            arrow.clicked.connect(lambda _checked=False, method_key=key: self._toggle_dh_formula(method_key))
            self.dh_formula_buttons[key] = arrow
            self.dh_formula_widgets[key] = formula
            self.dh_formula_expanded[key] = False
        else:
            arrow.clicked.connect(lambda _checked=False, method_key=key: self._toggle_t_plot_formula(method_key))
            self.t_plot_formula_buttons[key] = arrow
            self.t_plot_formula_widgets[key] = formula
            self.t_plot_formula_expanded[key] = False

        layout.addLayout(row_layout)
        layout.addWidget(formula)
        return container

    def _make_t_plot_formula_widget(self, method_key: str, *, context: str = "t_plot") -> QtWidgets.QWidget:
        if method_key == "reference":
            return self._make_reference_table_widget(context=context)
        if method_key in {"kjs", "harkins_jura"}:
            return self._make_power_log_formula_widget(method_key, context=context)
        if method_key == "halsey":
            return self._make_halsey_formula_widget(context=context)
        if method_key == "broekhoff_de_boer":
            return self._make_broekhoff_de_boer_formula_widget(context=context)
        if method_key == "carbon_black_stsa":
            return self._make_carbon_black_stsa_formula_widget(context=context)
        return self._make_pending_formula_widget()

    def _make_reference_table_widget(self, *, context: str = "t_plot") -> QtWidgets.QWidget:
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        outer_layout = QtWidgets.QVBoxLayout(frame)
        outer_layout.setContentsMargins(8, 8, 8, 8)
        outer_layout.setSpacing(6)

        name_row = QtWidgets.QHBoxLayout()
        name_row.setContentsMargins(0, 0, 0, 0)
        name_row.addWidget(QtWidgets.QLabel("名称"))
        name_edit = QtWidgets.QLineEdit()
        name_edit.setMinimumWidth(220)
        name_edit.editingFinished.connect(lambda ctx=context: self._on_reference_name_changed(ctx))
        name_row.addWidget(name_edit, 1)
        outer_layout.addLayout(name_row)

        body_layout = QtWidgets.QHBoxLayout()
        body_layout.setContentsMargins(0, 0, 0, 0)
        body_layout.setSpacing(10)

        table = QtWidgets.QTableWidget(0, 2)
        table.setHorizontalHeaderLabels(["相对\n压力\n(p/p°)", "厚度\n(nm)"])
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        table.verticalHeader().setDefaultSectionSize(25)
        table.verticalHeader().setMinimumWidth(42)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SingleSelection)
        table.setEditTriggers(
            QtWidgets.QAbstractItemView.DoubleClicked
            | QtWidgets.QAbstractItemView.EditKeyPressed
        )
        table.setFixedHeight(230)
        table.setMinimumWidth(250)
        table.itemChanged.connect(lambda item, ctx=context: self._on_reference_table_item_changed(ctx, item))
        body_layout.addWidget(table, 1)

        button_layout = QtWidgets.QVBoxLayout()
        button_layout.setContentsMargins(0, 0, 0, 0)
        button_layout.setSpacing(8)
        for text, handler in (
            ("插入", self._insert_reference_row),
            ("删除", self._delete_reference_row),
            ("清除", self._clear_reference_table),
            ("附加", self._append_reference_row),
        ):
            button = QtWidgets.QPushButton(text)
            button.setFixedWidth(110)
            button.clicked.connect(lambda _checked=False, ctx=context, fn=handler: fn(ctx))
            button_layout.addWidget(button)
        button_layout.addStretch(1)
        for text, handler in (("打开...", self._open_reference_file), ("另存为...", self._save_reference_file_as)):
            button = QtWidgets.QPushButton(text)
            button.setFixedWidth(110)
            button.clicked.connect(lambda _checked=False, ctx=context, fn=handler: fn(ctx))
            button_layout.addWidget(button)
        body_layout.addLayout(button_layout)
        outer_layout.addLayout(body_layout)

        self.reference_tables[context] = table
        self.reference_name_edits[context] = name_edit
        self._sync_reference_table_for_context(context)
        return frame

    def _reference_params_for_context(self, context: str) -> dict[str, object]:
        if context == "bjh":
            params_by_method = self.bjh_thickness_params_by_method
        elif context == "dh":
            params_by_method = self.dh_thickness_params_by_method
        else:
            params_by_method = self.t_plot_thickness_params_by_method
        params = dict(params_by_method.get("reference") or T_PLOT_THICKNESS_PARAM_DEFAULTS.get("reference", {}))
        params["reference_points"] = normalize_reference_points(params.get("reference_points"))
        return params

    def _set_reference_params_for_context(self, context: str, params: dict[str, object]) -> None:
        params = dict(params)
        params["reference_points"] = normalize_reference_points(params.get("reference_points"))
        if context == "bjh":
            self.bjh_thickness_params_by_method["reference"] = params
            if self.bjh_thickness_method == "reference":
                self.bjh_thickness_params = dict(params)
        elif context == "dh":
            self.dh_thickness_params_by_method["reference"] = params
            if self.dh_thickness_method == "reference":
                self.dh_thickness_params = dict(params)
        else:
            self.t_plot_thickness_params_by_method["reference"] = params
            if self.t_plot_thickness_method == "reference":
                self.t_plot_thickness_params = dict(params)

    def _sync_reference_table_for_context(self, context: str) -> None:
        table = self.reference_tables.get(context)
        name_edit = self.reference_name_edits.get(context)
        if table is None or name_edit is None:
            return
        params = self._reference_params_for_context(context)
        points = normalize_reference_points(params.get("reference_points"))
        self._syncing_reference_tables = True
        try:
            name_edit.setText(str(params.get("reference_name") or params.get("reference_path") or ""))
            table.setRowCount(len(points))
            for row, (pressure, thickness) in enumerate(points):
                pressure_item = QtWidgets.QTableWidgetItem(self._format_reference_pressure(pressure))
                thickness_item = QtWidgets.QTableWidgetItem(self._format_reference_thickness(thickness))
                pressure_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                thickness_item.setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                table.setItem(row, 0, pressure_item)
                table.setItem(row, 1, thickness_item)
                self._style_reference_item(pressure_item, 0)
                self._style_reference_item(thickness_item, 1)
        finally:
            self._syncing_reference_tables = False

    def _on_reference_name_changed(self, context: str) -> None:
        self._sync_reference_params_from_table(context)
        self._save_reference_settings_for_active(context)

    def _on_reference_table_item_changed(self, context: str, item: QtWidgets.QTableWidgetItem) -> None:
        if self._syncing_reference_tables:
            return
        self._style_reference_item(item, item.column())
        self._sync_reference_params_from_table(context)
        self._save_reference_settings_for_active(context)
        self._refresh_after_reference_change(context)

    def _sync_reference_params_from_table(self, context: str) -> dict[str, object]:
        table = self.reference_tables.get(context)
        name_edit = self.reference_name_edits.get(context)
        current = self._reference_params_for_context(context)
        points = []
        if table is not None:
            for row in range(table.rowCount()):
                pressure_item = table.item(row, 0)
                thickness_item = table.item(row, 1)
                if pressure_item is None or thickness_item is None:
                    continue
                try:
                    pressure = float(pressure_item.text())
                    thickness = float(thickness_item.text())
                except ValueError:
                    continue
                points.append((pressure, thickness))
        current["reference_points"] = normalize_reference_points(points)
        if name_edit is not None:
            current["reference_name"] = name_edit.text().strip()
        self._set_reference_params_for_context(context, current)
        return current

    def _save_reference_settings_for_active(self, context: str) -> None:
        if context == "bjh":
            self._save_bjh_settings_for_active()
        elif context == "dh":
            self._save_dh_settings_for_active()
        else:
            self._save_t_plot_settings_for_active()

    def _refresh_after_reference_change(self, context: str) -> None:
        if context == "bjh":
            if self.bjh_thickness_method == "reference":
                self.refresh_bjh_plot()
                self._refresh_all_sample_bjh_pore_cells()
                self.refresh_metrics()
        elif context == "dh":
            if self.dh_thickness_method == "reference":
                self._dh_distribution_cache.clear()
                self.refresh_dh_plot()
                self._refresh_all_sample_bjh_pore_cells()
                self.refresh_metrics()
        elif self.t_plot_thickness_method == "reference":
            active = self.active_result()
            if active is not None and self.plot_tabs.currentWidget() is self.t_plot_tab:
                pressure_range = self._current_pressure_region()
                p_min, p_max = pressure_range if pressure_range else (None, None)
                self._refresh_t_plot_plot(active, p_min, p_max, reset_region=True)
            self._refresh_sample_t_plot_cell(self.active_index)
            self.refresh_metrics()

    def _insert_reference_row(self, context: str) -> None:
        table = self.reference_tables.get(context)
        if table is None:
            return
        row = table.currentRow()
        if row < 0:
            row = table.rowCount()
        insert_at = max(0, min(row, table.rowCount()))
        points = self._reference_points_from_table(context)
        if len(points) >= 2:
            if insert_at <= 0:
                left, right = points[0], points[1]
            elif insert_at >= len(points):
                left, right = points[-2], points[-1]
            else:
                left, right = points[insert_at - 1], points[insert_at]
            new_point = ((left[0] + right[0]) / 2.0, (left[1] + right[1]) / 2.0)
        elif points:
            new_point = (points[0][0] + 0.01, points[0][1] + 0.01)
        else:
            new_point = (0.0, 0.0)
        points.insert(insert_at, new_point)
        self._replace_reference_points(context, points, select_row=insert_at)

    def _delete_reference_row(self, context: str) -> None:
        table = self.reference_tables.get(context)
        if table is None or table.currentRow() < 0:
            return
        row = table.currentRow()
        points = self._reference_points_from_table(context)
        if 0 <= row < len(points):
            points.pop(row)
        self._replace_reference_points(context, points, select_row=min(row, len(points) - 1))

    def _clear_reference_table(self, context: str) -> None:
        reply = QtWidgets.QMessageBox.question(
            self,
            "清除",
            "清除整个表？",
            QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No,
            QtWidgets.QMessageBox.No,
        )
        if reply != QtWidgets.QMessageBox.Yes:
            return
        self._replace_reference_points(context, [], select_row=-1)

    def _append_reference_row(self, context: str) -> None:
        points = self._reference_points_from_table(context)
        if len(points) >= 2:
            previous, last = points[-2], points[-1]
            new_point = (last[0] + (last[0] - previous[0]), last[1] + (last[1] - previous[1]))
        elif points:
            new_point = (points[-1][0] + 0.01, points[-1][1] + 0.01)
        else:
            new_point = (0.0, 0.0)
        points.append(new_point)
        self._replace_reference_points(context, points, select_row=len(points) - 1)

    def _open_reference_file(self, context: str) -> None:
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self,
            "打开厚度参比表",
            str(DEFAULT_REFERENCE_DIR),
            "Thickness files (*.thk *.als);;All files (*.*)",
        )
        if not path:
            return
        try:
            points = read_reference_points(path)
        except OSError as exc:
            QtWidgets.QMessageBox.warning(self, "打开失败", str(exc))
            return
        params = self._reference_params_for_context(context)
        params["reference_name"] = Path(path).name
        params["reference_path"] = path
        params["reference_points"] = points
        self._set_reference_params_for_context(context, params)
        self._sync_reference_table_for_context(context)
        self._save_reference_settings_for_active(context)
        self._refresh_after_reference_change(context)

    def _save_reference_file_as(self, context: str) -> None:
        points = self._reference_points_from_table(context)
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "另存为厚度参比表",
            str(DEFAULT_REFERENCE_DIR / "reference.thk"),
            "Thickness files (*.thk);;ALS files (*.als);;All files (*.*)",
        )
        if not path:
            return
        if Path(path).suffix == "":
            path += ".thk"
        try:
            write_reference_points(path, points)
        except OSError as exc:
            QtWidgets.QMessageBox.warning(self, "保存失败", str(exc))
            return
        params = self._reference_params_for_context(context)
        params["reference_name"] = Path(path).name
        params["reference_path"] = path
        params["reference_points"] = points
        self._set_reference_params_for_context(context, params)
        self._sync_reference_table_for_context(context)
        self._save_reference_settings_for_active(context)

    def _replace_reference_points(self, context: str, points, *, select_row: int = -1) -> None:
        params = self._reference_params_for_context(context)
        params["reference_points"] = normalize_reference_points(points)
        self._set_reference_params_for_context(context, params)
        self._sync_reference_table_for_context(context)
        table = self.reference_tables.get(context)
        if table is not None and 0 <= select_row < table.rowCount():
            table.selectRow(select_row)
            table.scrollToItem(table.item(select_row, 0))
        self._save_reference_settings_for_active(context)
        self._refresh_after_reference_change(context)

    def _reference_points_from_table(self, context: str) -> list[tuple[float, float]]:
        params = self._sync_reference_params_from_table(context)
        return list(normalize_reference_points(params.get("reference_points")))

    def _style_reference_item(self, item: QtWidgets.QTableWidgetItem, column: int) -> None:
        try:
            value = float(item.text())
        except ValueError:
            invalid = True
        else:
            invalid = value > (1.0 if column == 0 else 10.0)
        if invalid:
            item.setBackground(QtGui.QBrush(QtGui.QColor(REFERENCE_INVALID_BACKGROUND)))
            item.setForeground(QtGui.QBrush(QtGui.QColor(REFERENCE_INVALID_FOREGROUND)))
        else:
            item.setBackground(QtGui.QBrush(QtGui.QColor("#ffffff")))
            item.setForeground(QtGui.QBrush(QtGui.QColor("#111827")))

    @staticmethod
    def _format_reference_pressure(value: float) -> str:
        return f"{float(value):.9f}"

    @staticmethod
    def _format_reference_thickness(value: float) -> str:
        return f"{float(value):.5f}"

    def _make_param_spin_for_method(
        self,
        method_key: str,
        param_key: str,
        minimum: float,
        maximum: float,
        decimals: int,
        *,
        width: int | None = None,
        context: str = "t_plot",
    ) -> QtWidgets.QDoubleSpinBox:
        if context == "bjh":
            params_by_method = self.bjh_thickness_params_by_method
        elif context == "dh":
            params_by_method = self.dh_thickness_params_by_method
        else:
            params_by_method = self.t_plot_thickness_params_by_method
        params = params_by_method.get(
            method_key,
            T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS),
        )
        spin = self._make_param_spin(params[param_key], minimum, maximum, decimals, width=width)
        if context == "bjh":
            spin.valueChanged.connect(
                lambda _value, changed_method=method_key: self._on_bjh_thickness_param_changed(changed_method)
            )
            self.bjh_param_spins.setdefault(method_key, {})[param_key] = spin
        elif context == "dh":
            spin.valueChanged.connect(
                lambda _value, changed_method=method_key: self._on_dh_thickness_param_changed(changed_method)
            )
            self.dh_param_spins.setdefault(method_key, {})[param_key] = spin
        else:
            spin.valueChanged.connect(
                lambda _value, changed_method=method_key: self._on_t_plot_thickness_param_changed(changed_method)
            )
            self.t_plot_param_spins.setdefault(method_key, {})[param_key] = spin
        return spin

    def _make_power_log_formula_widget(self, method_key: str, *, context: str = "t_plot") -> QtWidgets.QWidget:
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        outer_layout = QtWidgets.QHBoxLayout(frame)
        outer_layout.setContentsMargins(8, 8, 8, 8)
        outer_layout.setSpacing(5)

        outer_layout.addWidget(QtWidgets.QLabel("t = ("))
        fraction_layout = QtWidgets.QVBoxLayout()
        fraction_layout.setContentsMargins(0, 0, 0, 0)
        fraction_layout.setSpacing(3)

        numerator_spin = self._make_param_spin_for_method(
            method_key, "numerator", -1000000.0, 1000000.0, 4, width=126, context=context
        )
        offset_spin = self._make_param_spin_for_method(
            method_key, "offset", -1000.0, 1000.0, 5, width=126, context=context
        )
        exponent_spin = self._make_param_spin_for_method(
            method_key, "exponent", -10.0, 10.0, 4, width=96, context=context
        )
        if method_key == "harkins_jura" and context == "t_plot":
            self.hj_numerator_spin = numerator_spin
            self.hj_offset_spin = offset_spin
            self.hj_exponent_spin = exponent_spin

        numerator_row = QtWidgets.QHBoxLayout()
        numerator_row.setContentsMargins(0, 0, 0, 0)
        numerator_row.addStretch(1)
        numerator_row.addWidget(numerator_spin)
        numerator_row.addStretch(1)

        fraction_line = QtWidgets.QFrame()
        fraction_line.setFrameShape(QtWidgets.QFrame.HLine)
        fraction_line.setFrameShadow(QtWidgets.QFrame.Plain)
        fraction_line.setLineWidth(2)

        denominator_row = QtWidgets.QHBoxLayout()
        denominator_row.setContentsMargins(0, 0, 0, 0)
        denominator_row.setSpacing(4)
        denominator_row.addStretch(1)
        denominator_row.addWidget(offset_spin)
        denominator_row.addWidget(QtWidgets.QLabel("- log(p/p°)"))
        denominator_row.addStretch(1)

        fraction_layout.addLayout(numerator_row)
        fraction_layout.addWidget(fraction_line)
        fraction_layout.addLayout(denominator_row)
        outer_layout.addLayout(fraction_layout)
        outer_layout.addWidget(QtWidgets.QLabel(")^"))
        outer_layout.addWidget(exponent_spin)
        outer_layout.addStretch(1)
        return frame

    def _make_halsey_formula_widget(self, *, context: str = "t_plot") -> QtWidgets.QWidget:
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        outer_layout = QtWidgets.QHBoxLayout(frame)
        outer_layout.setContentsMargins(8, 8, 8, 8)
        outer_layout.setSpacing(5)

        outer_layout.addWidget(QtWidgets.QLabel("t ="))
        prefactor_spin = self._make_param_spin_for_method(
            "halsey", "prefactor", -1000000.0, 1000000.0, 4, width=112, context=context
        )
        numerator_spin = self._make_param_spin_for_method(
            "halsey", "numerator", -1000000.0, 1000000.0, 4, width=112, context=context
        )
        exponent_spin = self._make_param_spin_for_method(
            "halsey", "exponent", -10.0, 10.0, 4, width=96, context=context
        )
        outer_layout.addWidget(prefactor_spin)
        outer_layout.addWidget(QtWidgets.QLabel("("))

        fraction_layout = QtWidgets.QVBoxLayout()
        fraction_layout.setContentsMargins(0, 0, 0, 0)
        fraction_layout.setSpacing(3)
        numerator_row = QtWidgets.QHBoxLayout()
        numerator_row.setContentsMargins(0, 0, 0, 0)
        numerator_row.addWidget(numerator_spin)
        fraction_line = QtWidgets.QFrame()
        fraction_line.setFrameShape(QtWidgets.QFrame.HLine)
        fraction_line.setFrameShadow(QtWidgets.QFrame.Plain)
        fraction_line.setLineWidth(2)
        denominator_row = QtWidgets.QHBoxLayout()
        denominator_row.setContentsMargins(0, 0, 0, 0)
        denominator_row.addWidget(QtWidgets.QLabel("ln(p/p°)"))
        fraction_layout.addLayout(numerator_row)
        fraction_layout.addWidget(fraction_line)
        fraction_layout.addLayout(denominator_row)

        outer_layout.addLayout(fraction_layout)
        outer_layout.addWidget(QtWidgets.QLabel(")^"))
        outer_layout.addWidget(exponent_spin)
        outer_layout.addStretch(1)
        return frame

    def _make_broekhoff_de_boer_formula_widget(self, *, context: str = "t_plot") -> QtWidgets.QWidget:
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        outer_layout = QtWidgets.QHBoxLayout(frame)
        outer_layout.setContentsMargins(8, 8, 8, 8)
        outer_layout.setSpacing(4)

        outer_layout.addWidget(QtWidgets.QLabel("log(p/p°) ="))
        inverse_spin = self._make_param_spin_for_method(
            "broekhoff_de_boer", "inverse_square", -1000000.0, 1000000.0, 4, width=116, context=context
        )
        factor_spin = self._make_param_spin_for_method(
            "broekhoff_de_boer", "exponential_factor", -1000000.0, 1000000.0, 4, width=112, context=context
        )
        rate_spin = self._make_param_spin_for_method(
            "broekhoff_de_boer", "exponential_rate", -1000000.0, 1000000.0, 4, width=112, context=context
        )

        fraction_layout = QtWidgets.QVBoxLayout()
        fraction_layout.setContentsMargins(0, 0, 0, 0)
        fraction_layout.setSpacing(3)
        top_row = QtWidgets.QHBoxLayout()
        top_row.setContentsMargins(0, 0, 0, 0)
        top_row.addWidget(inverse_spin)
        fraction_line = QtWidgets.QFrame()
        fraction_line.setFrameShape(QtWidgets.QFrame.HLine)
        fraction_line.setFrameShadow(QtWidgets.QFrame.Plain)
        fraction_line.setLineWidth(2)
        bottom_row = QtWidgets.QHBoxLayout()
        bottom_row.setContentsMargins(0, 0, 0, 0)
        bottom_row.addWidget(QtWidgets.QLabel("t²"))
        fraction_layout.addLayout(top_row)
        fraction_layout.addWidget(fraction_line)
        fraction_layout.addLayout(bottom_row)

        outer_layout.addLayout(fraction_layout)
        outer_layout.addWidget(QtWidgets.QLabel("+"))
        outer_layout.addWidget(factor_spin)
        outer_layout.addWidget(QtWidgets.QLabel("e^"))
        outer_layout.addWidget(rate_spin)
        outer_layout.addWidget(QtWidgets.QLabel("t"))
        outer_layout.addStretch(1)
        return frame

    def _make_carbon_black_stsa_formula_widget(self, *, context: str = "t_plot") -> QtWidgets.QWidget:
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        outer_layout = QtWidgets.QHBoxLayout(frame)
        outer_layout.setContentsMargins(8, 8, 8, 8)
        outer_layout.setSpacing(4)

        constant_spin = self._make_param_spin_for_method(
            "carbon_black_stsa", "constant", -1000000.0, 1000000.0, 4, width=108, context=context
        )
        linear_spin = self._make_param_spin_for_method(
            "carbon_black_stsa", "linear", -1000000.0, 1000000.0, 4, width=108, context=context
        )
        quadratic_spin = self._make_param_spin_for_method(
            "carbon_black_stsa", "quadratic", -1000000.0, 1000000.0, 4, width=108, context=context
        )
        outer_layout.addWidget(QtWidgets.QLabel("t ="))
        outer_layout.addWidget(constant_spin)
        outer_layout.addWidget(QtWidgets.QLabel("+"))
        outer_layout.addWidget(linear_spin)
        outer_layout.addWidget(QtWidgets.QLabel("(p/p°) +"))
        outer_layout.addWidget(quadratic_spin)
        outer_layout.addWidget(QtWidgets.QLabel("(p/p°)²"))
        outer_layout.addStretch(1)
        return frame

    def _make_pending_formula_widget(self) -> QtWidgets.QWidget:
        frame = QtWidgets.QFrame()
        frame.setFrameShape(QtWidgets.QFrame.StyledPanel)
        layout = QtWidgets.QVBoxLayout(frame)
        layout.setContentsMargins(8, 6, 8, 6)
        label = QtWidgets.QLabel("公式参数待补充")
        label.setStyleSheet("color: #6b7280;")
        layout.addWidget(label)
        return frame

    def _make_param_spin(
        self,
        value: float,
        minimum: float,
        maximum: float,
        decimals: int,
        *,
        width: int | None = None,
    ) -> QtWidgets.QDoubleSpinBox:
        spin = QtWidgets.QDoubleSpinBox()
        spin.setRange(minimum, maximum)
        spin.setDecimals(decimals)
        spin.setValue(float(value))
        spin.setSingleStep(10 ** -decimals)
        spin.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        if width is None:
            width = 112
        spin.setMinimumWidth(width)
        spin.setMaximumWidth(width)
        return spin

    def _toggle_t_plot_formula(self, key: str) -> None:
        widget = self.t_plot_formula_widgets.get(key)
        button = self.t_plot_formula_buttons.get(key)
        if widget is None or button is None:
            return
        is_visible = not self.t_plot_formula_expanded.get(key, False)
        widget.setVisible(is_visible)
        self.t_plot_formula_expanded[key] = is_visible
        button.setArrowType(QtCore.Qt.UpArrow if is_visible else QtCore.Qt.DownArrow)
        self._update_t_plot_options_panel_width()

    def _update_t_plot_options_panel_width(self, panel: QtWidgets.QWidget | None = None) -> None:
        panel = panel or getattr(self, "t_plot_options_panel", None)
        if panel is None:
            return
        expanded = any(self.t_plot_formula_expanded.values())
        panel.setFixedWidth(T_PLOT_PANEL_EXPANDED_WIDTH if expanded else T_PLOT_PANEL_COLLAPSED_WIDTH)

    def _toggle_bjh_formula(self, key: str) -> None:
        widget = self.bjh_formula_widgets.get(key)
        button = self.bjh_formula_buttons.get(key)
        if widget is None or button is None:
            return
        is_visible = not self.bjh_formula_expanded.get(key, False)
        widget.setVisible(is_visible)
        self.bjh_formula_expanded[key] = is_visible
        button.setArrowType(QtCore.Qt.UpArrow if is_visible else QtCore.Qt.DownArrow)
        self._update_bjh_options_panel_width()

    def _update_bjh_options_panel_width(self, panel: QtWidgets.QWidget | None = None) -> None:
        panel = panel or getattr(self, "bjh_options_panel", None)
        if panel is None:
            return
        expanded = any(self.bjh_formula_expanded.values())
        panel.setFixedWidth(BJH_PANEL_EXPANDED_WIDTH if expanded else BJH_PANEL_COLLAPSED_WIDTH)

    def _toggle_dh_formula(self, key: str) -> None:
        widget = self.dh_formula_widgets.get(key)
        button = self.dh_formula_buttons.get(key)
        if widget is None or button is None:
            return
        is_visible = not self.dh_formula_expanded.get(key, False)
        widget.setVisible(is_visible)
        self.dh_formula_expanded[key] = is_visible
        button.setArrowType(QtCore.Qt.UpArrow if is_visible else QtCore.Qt.DownArrow)
        self._update_dh_options_panel_width()

    def _update_dh_options_panel_width(self, panel: QtWidgets.QWidget | None = None) -> None:
        panel = panel or getattr(self, "dh_options_panel", None)
        if panel is None:
            return
        expanded = any(self.dh_formula_expanded.values())
        panel.setFixedWidth(BJH_PANEL_EXPANDED_WIDTH if expanded else BJH_PANEL_COLLAPSED_WIDTH)

    def _thickness_method_for_context(self, context: str) -> str:
        if context == "bjh":
            return self.bjh_thickness_method
        if context == "dh":
            return self.dh_thickness_method
        return self.t_plot_thickness_method

    def _on_t_plot_thickness_method_changed(self, method_key: str, checked: bool) -> None:
        if self._syncing_t_plot_controls:
            return
        if not checked:
            return
        self.t_plot_thickness_method = method_key
        self.t_plot_thickness_params = dict(
            self.t_plot_thickness_params_by_method.get(
                method_key,
                T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS),
            )
        )
        self._syncing_t_plot_controls = True
        try:
            self._set_t_plot_formula_spins_for_method(method_key, self.t_plot_thickness_params)
        finally:
            self._syncing_t_plot_controls = False
        self._save_t_plot_settings_for_active()
        self._refresh_t_plot_for_option_change(refresh_table=True)

    def _on_t_plot_thickness_param_changed(self, method_key: str | None = None) -> None:
        if self._syncing_t_plot_controls:
            return
        method_key = method_key or self.t_plot_thickness_method
        params = self._read_t_plot_thickness_params(method_key)
        self.t_plot_thickness_params_by_method[method_key] = dict(params)
        if method_key == self.t_plot_thickness_method:
            self.t_plot_thickness_params = dict(params)
        self._save_t_plot_settings_for_active()
        if method_key == self.t_plot_thickness_method:
            self._refresh_t_plot_for_option_change(refresh_table=True)
        else:
            self._refresh_sample_t_plot_cell(self.active_index)

    def _read_t_plot_thickness_params(self, method_key: str) -> dict[str, object]:
        if method_key == "reference":
            return self._sync_reference_params_from_table("t_plot")
        params = dict(T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS))
        for param_key, spin in self.t_plot_param_spins.get(method_key, {}).items():
            params[param_key] = float(spin.value())
        return params

    def _set_t_plot_formula_spins_for_method(self, method_key: str, params: dict[str, float]) -> None:
        for param_key, spin in self.t_plot_param_spins.get(method_key, {}).items():
            if param_key in params:
                spin.setValue(float(params[param_key]))

    def _set_all_t_plot_formula_spins(self) -> None:
        for method_key, params in self.t_plot_thickness_params_by_method.items():
            self._set_t_plot_formula_spins_for_method(method_key, params)

    def _on_bjh_thickness_method_changed(self, method_key: str, checked: bool) -> None:
        if self._syncing_bjh_controls or not checked:
            return
        self.bjh_thickness_method = method_key
        self.bjh_thickness_params = dict(
            self.bjh_thickness_params_by_method.get(
                method_key,
                T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS),
            )
        )
        self._syncing_bjh_controls = True
        try:
            self._set_bjh_formula_spins_for_method(method_key, self.bjh_thickness_params)
        finally:
            self._syncing_bjh_controls = False
        self._save_bjh_settings_for_active()
        self.refresh_bjh_plot()
        self._refresh_all_sample_bjh_pore_cells()

    def _on_bjh_thickness_param_changed(self, method_key: str | None = None) -> None:
        if self._syncing_bjh_controls:
            return
        method_key = method_key or self.bjh_thickness_method
        params = self._read_bjh_thickness_params(method_key)
        self.bjh_thickness_params_by_method[method_key] = dict(params)
        if method_key == self.bjh_thickness_method:
            self.bjh_thickness_params = dict(params)
        self._save_bjh_settings_for_active()
        if method_key == self.bjh_thickness_method:
            self.refresh_bjh_plot()
            self._refresh_all_sample_bjh_pore_cells()

    def _read_bjh_thickness_params(self, method_key: str) -> dict[str, object]:
        if method_key == "reference":
            return self._sync_reference_params_from_table("bjh")
        params = dict(T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS))
        for param_key, spin in self.bjh_param_spins.get(method_key, {}).items():
            params[param_key] = float(spin.value())
        return params

    def _set_bjh_formula_spins_for_method(self, method_key: str, params: dict[str, float]) -> None:
        for param_key, spin in self.bjh_param_spins.get(method_key, {}).items():
            if param_key in params:
                spin.setValue(float(params[param_key]))

    def _set_all_bjh_formula_spins(self) -> None:
        for method_key, params in self.bjh_thickness_params_by_method.items():
            self._set_bjh_formula_spins_for_method(method_key, params)

    def _on_dh_thickness_method_changed(self, method_key: str, checked: bool) -> None:
        if self._syncing_dh_controls or not checked:
            return
        self.dh_thickness_method = method_key
        self.dh_thickness_params = dict(
            self.dh_thickness_params_by_method.get(
                method_key,
                T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS),
            )
        )
        self._syncing_dh_controls = True
        try:
            self._set_dh_formula_spins_for_method(method_key, self.dh_thickness_params)
        finally:
            self._syncing_dh_controls = False
        self._save_dh_settings_for_active()
        self._dh_distribution_cache.clear()
        self.refresh_dh_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DH:
            self._refresh_all_sample_bjh_pore_cells()

    def _on_dh_thickness_param_changed(self, method_key: str | None = None) -> None:
        if self._syncing_dh_controls:
            return
        method_key = method_key or self.dh_thickness_method
        params = self._read_dh_thickness_params(method_key)
        self.dh_thickness_params_by_method[method_key] = dict(params)
        if method_key == self.dh_thickness_method:
            self.dh_thickness_params = dict(params)
            self._save_dh_settings_for_active()
            self._dh_distribution_cache.clear()
            self.refresh_dh_plot()
            if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DH:
                self._refresh_all_sample_bjh_pore_cells()

    def _read_dh_thickness_params(self, method_key: str) -> dict[str, object]:
        if method_key == "reference":
            return self._sync_reference_params_from_table("dh")
        params = dict(T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method_key, DEFAULT_T_PLOT_THICKNESS_PARAMS))
        for param_key, spin in self.dh_param_spins.get(method_key, {}).items():
            params[param_key] = float(spin.value())
        return params

    def _set_dh_formula_spins_for_method(self, method_key: str, params: dict[str, float]) -> None:
        for param_key, spin in self.dh_param_spins.get(method_key, {}).items():
            if param_key in params:
                spin.setValue(float(params[param_key]))

    def _set_all_dh_formula_spins(self) -> None:
        for method_key, params in self.dh_thickness_params_by_method.items():
            self._set_dh_formula_spins_for_method(method_key, params)

    def _on_bjh_option_changed(self, *_args) -> None:
        if self._syncing_bjh_controls:
            return
        if self.bjh_kjs_correction_radio.isChecked():
            self.bjh_correction = "kjs"
        elif self.bjh_faas_correction_radio.isChecked():
            self.bjh_correction = "faas"
        else:
            self.bjh_correction = "standard"
        self.bjh_open_pore_fraction = float(self.bjh_open_fraction_spin.value())
        self.bjh_smooth_derivative = self.bjh_smooth_checkbox.isChecked()
        self.bjh_show_adsorption = self.bjh_adsorption_checkbox.isChecked()
        self.bjh_show_desorption = self.bjh_desorption_checkbox.isChecked()
        self._save_bjh_settings_for_active()
        self.refresh_bjh_plot()
        self._refresh_all_sample_bjh_pore_cells()

    def _set_bjh_display_combo(self) -> None:
        combo = getattr(self, "bjh_display_combo", None)
        if combo is None:
            return
        target = _normalize_bjh_display_metric(self.bjh_display_metrics)
        for index in range(combo.count()):
            if combo.itemData(index) == target:
                combo.setCurrentIndex(index)
                return

    def _default_bjh_display_metrics_for_result(self, result) -> list[str]:
        if result is not None:
            vendor_mode = result.method_options.get("vendor_bjh_differential_mode")
            if vendor_mode is not None:
                return _normalize_bjh_display_metrics(vendor_mode)
        return list(DEFAULT_BJH_DISPLAY_METRICS)

    def _on_bjh_display_mode_changed(self, *_args) -> None:
        combo = getattr(self, "bjh_display_combo", None)
        metric = combo.currentData() if combo is not None else DEFAULT_BJH_DISPLAY_METRICS[0]
        self.bjh_display_metrics = _normalize_bjh_display_metrics(metric)
        self.bjh_differential_mode = self.bjh_display_metrics[0]
        if self._syncing_bjh_controls:
            return
        self.refresh_bjh_plot()

    def _current_bjh_settings_snapshot(self) -> dict[str, object]:
        return {
            "thickness_method": self.bjh_thickness_method,
            "thickness_params_by_method": {
                method_key: dict(params)
                for method_key, params in self.bjh_thickness_params_by_method.items()
            },
            "thickness_params": dict(self.bjh_thickness_params),
            "correction": self.bjh_correction,
            "open_pore_fraction": self.bjh_open_pore_fraction,
            "smooth_derivative": self.bjh_smooth_derivative,
            "show_adsorption": self.bjh_show_adsorption,
            "show_desorption": self.bjh_show_desorption,
        }

    def apply_bjh_settings_to_all(self) -> None:
        if not self.results:
            return
        settings = self._current_bjh_settings_snapshot()
        for result in self.results:
            self.custom_bjh_settings[id(result)] = copy.deepcopy(settings)
        self._bjh_distribution_cache.clear()
        self.refresh_bjh_plot()
        self._refresh_all_sample_bjh_pore_cells()
        self.statusBar().showMessage("已将当前 BJH 设置应用到所有样品", 3000)

    def reset_all_bjh_to_default(self) -> None:
        self.custom_bjh_settings.clear()
        self._bjh_distribution_cache.clear()
        self.reset_bjh_to_default(reset_region=True)
        self.statusBar().showMessage("所有样品已恢复各自默认 BJH 设置", 3000)

    def reset_bjh_to_default(self, *, reset_region: bool = False) -> None:
        active = self.active_result()
        if active is not None:
            self.custom_bjh_settings.pop(id(active), None)
        settings = self._default_bjh_settings() if active is None else self._bjh_default_settings_for_result(active)
        self._syncing_bjh_controls = True
        try:
            self.bjh_thickness_method = str(settings["thickness_method"])
            self.bjh_thickness_params_by_method = {
                method_key: dict(params)
                for method_key, params in dict(settings["thickness_params_by_method"]).items()
            }
            self.bjh_thickness_params = dict(settings["thickness_params"])
            self.bjh_correction = str(settings["correction"])
            self.bjh_open_pore_fraction = float(settings["open_pore_fraction"])
            self.bjh_smooth_derivative = bool(settings["smooth_derivative"])
            self.bjh_show_adsorption = bool(settings["show_adsorption"])
            self.bjh_show_desorption = bool(settings["show_desorption"])
            self.bjh_display_metrics = self._default_bjh_display_metrics_for_result(active)
            self.bjh_differential_mode = self.bjh_display_metrics[0]

            for key, radio in self.bjh_method_radios.items():
                radio.setChecked(key == self.bjh_thickness_method)
            self._set_all_bjh_formula_spins()
            self.bjh_standard_radio.setChecked(self.bjh_correction == "standard")
            self.bjh_kjs_correction_radio.setChecked(self.bjh_correction == "kjs")
            self.bjh_faas_correction_radio.setChecked(self.bjh_correction == "faas")
            self.bjh_open_fraction_spin.setValue(self.bjh_open_pore_fraction)
            self.bjh_smooth_checkbox.setChecked(self.bjh_smooth_derivative)
            self.bjh_adsorption_checkbox.setChecked(self.bjh_show_adsorption)
            self.bjh_desorption_checkbox.setChecked(self.bjh_show_desorption)
            self._set_bjh_display_combo()
            self._sync_reference_table_for_context("bjh")
        finally:
            self._syncing_bjh_controls = False
        if reset_region:
            self.bjh_pore_volume_range = DEFAULT_BJH_PORE_VOLUME_RANGE
            self._remove_bjh_region()
        self.refresh_bjh_plot()
        self._refresh_all_sample_bjh_pore_cells()

    def _on_dh_option_changed(self, *_args) -> None:
        if self._syncing_dh_controls:
            return
        self.dh_smooth_derivative = self.dh_smooth_checkbox.isChecked()
        self.dh_show_adsorption = self.dh_adsorption_checkbox.isChecked()
        self.dh_show_desorption = self.dh_desorption_checkbox.isChecked()
        self._save_dh_settings_for_active()
        self.refresh_dh_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DH:
            self._refresh_all_sample_bjh_pore_cells()

    def _set_dh_display_combo(self) -> None:
        combo = getattr(self, "dh_display_combo", None)
        if combo is None:
            return
        target = _normalize_bjh_display_metric(self.dh_display_metrics)
        for index in range(combo.count()):
            if combo.itemData(index) == target:
                combo.setCurrentIndex(index)
                return

    def _on_dh_display_mode_changed(self, *_args) -> None:
        combo = getattr(self, "dh_display_combo", None)
        metric = combo.currentData() if combo is not None else DEFAULT_DH_DISPLAY_METRICS[0]
        self.dh_display_metrics = _normalize_bjh_display_metrics(metric)
        self.dh_differential_mode = self.dh_display_metrics[0]
        if self._syncing_dh_controls:
            return
        self._save_dh_settings_for_active()
        self.refresh_dh_plot()

    def _current_dh_settings_snapshot(self) -> dict[str, object]:
        return {
            "thickness_method": self.dh_thickness_method,
            "thickness_params_by_method": {
                method_key: dict(params)
                for method_key, params in self.dh_thickness_params_by_method.items()
            },
            "thickness_params": dict(self.dh_thickness_params),
            "smooth_derivative": self.dh_smooth_derivative,
            "show_adsorption": self.dh_show_adsorption,
            "show_desorption": self.dh_show_desorption,
            "display_metrics": list(self.dh_display_metrics),
        }

    def apply_dh_settings_to_all(self) -> None:
        if not self.results:
            return
        settings = self._current_dh_settings_snapshot()
        for result in self.results:
            self.custom_dh_settings[id(result)] = copy.deepcopy(settings)
        self._dh_distribution_cache.clear()
        self.refresh_dh_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DH:
            self._refresh_all_sample_bjh_pore_cells()
        else:
            self._refresh_sample_bjh_pore_cell(self.active_index)
        self.statusBar().showMessage("已将当前 DH 设置应用到所有样品", 3000)

    def reset_all_dh_to_default(self) -> None:
        self.custom_dh_settings.clear()
        self._dh_distribution_cache.clear()
        self.reset_dh_to_default(reset_region=True)
        self.statusBar().showMessage("所有样品已恢复各自默认 DH 设置", 3000)

    def reset_dh_to_default(self, *, reset_region: bool = False) -> None:
        active = self.active_result()
        if active is not None:
            self.custom_dh_settings.pop(id(active), None)
        settings = self._default_dh_settings() if active is None else self._dh_default_settings_for_result(active)
        self._syncing_dh_controls = True
        try:
            self.dh_thickness_method = str(settings["thickness_method"])
            self.dh_thickness_params_by_method = {
                method_key: dict(params)
                for method_key, params in dict(settings["thickness_params_by_method"]).items()
            }
            self.dh_thickness_params = dict(settings["thickness_params"])
            self.dh_smooth_derivative = bool(settings["smooth_derivative"])
            self.dh_show_adsorption = bool(settings["show_adsorption"])
            self.dh_show_desorption = bool(settings["show_desorption"])
            self.dh_display_metrics = list(settings["display_metrics"])
            self.dh_differential_mode = self.dh_display_metrics[0]
            for key, radio in self.dh_method_radios.items():
                radio.setChecked(key == self.dh_thickness_method)
            self._set_all_dh_formula_spins()
            self.dh_smooth_checkbox.setChecked(self.dh_smooth_derivative)
            self.dh_adsorption_checkbox.setChecked(self.dh_show_adsorption)
            self.dh_desorption_checkbox.setChecked(self.dh_show_desorption)
            self._set_dh_display_combo()
            self._sync_reference_table_for_context("dh")
        finally:
            self._syncing_dh_controls = False
        if reset_region:
            self.bjh_pore_volume_range = DEFAULT_BJH_PORE_VOLUME_RANGE
            self._remove_dh_region()
        self._dh_distribution_cache.clear()
        self.refresh_dh_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DH:
            self._refresh_all_sample_bjh_pore_cells()

    def _read_hk_interaction_parameter(self) -> float:
        try:
            value = float(self.hk_interaction_edit.text())
        except (AttributeError, TypeError, ValueError):
            value = self.hk_interaction_parameter
        if not (np.isfinite(value) and value > 0.0):
            value = DEFAULT_HK_INTERACTION_PARAMETER
        return float(value)

    def _update_hk_interaction_controls(self) -> None:
        calculated = calculate_hk_interaction_parameter(
            self.hk_adsorbent_properties,
            self.hk_adsorptive_properties,
            adsorbent_key=self.hk_adsorbent_key,
            adsorptive_key=self.hk_adsorptive_key,
        )
        if hasattr(self, "hk_calculated_value_label"):
            self.hk_calculated_value_label.setText(f"{calculated:.3e} erg·cm^4")
        input_mode = self.hk_interaction_parameter_mode != "calculated"
        if hasattr(self, "hk_interaction_edit"):
            self.hk_interaction_edit.setEnabled(input_mode)

    def _on_hk_option_changed(self, *_args) -> None:
        if self._syncing_hk_controls:
            return
        for key, radio in getattr(self, "hk_geometry_radios", {}).items():
            if radio.isChecked():
                self.hk_geometry = key
                break
        self.hk_interaction_parameter_mode = "calculated" if self.hk_calculated_radio.isChecked() else "input"
        self.hk_interaction_parameter = self._read_hk_interaction_parameter()
        self.hk_cheng_yang_correction = self.hk_cheng_yang_checkbox.isChecked()
        self.hk_smooth_derivative = self.hk_smooth_checkbox.isChecked()
        self._update_hk_interaction_controls()
        self._save_hk_settings_for_active()
        self.refresh_hk_plot()

    def _set_hk_display_combo(self) -> None:
        combo = getattr(self, "hk_display_combo", None)
        if combo is None:
            return
        target = normalize_hk_display_metric(self.hk_display_metric)
        for index in range(combo.count()):
            if combo.itemData(index) == target:
                combo.setCurrentIndex(index)
                return

    def _on_hk_display_mode_changed(self, *_args) -> None:
        combo = getattr(self, "hk_display_combo", None)
        self.hk_display_metric = normalize_hk_display_metric(combo.currentData() if combo is not None else DEFAULT_HK_DISPLAY_METRIC)
        if self._syncing_hk_controls:
            return
        self._save_hk_settings_for_active()
        self.refresh_hk_plot()

    def _current_hk_settings_snapshot(self) -> dict[str, object]:
        if self.hk_interaction_parameter_mode == "calculated":
            interaction_parameter = calculate_hk_interaction_parameter(
                self.hk_adsorbent_properties,
                self.hk_adsorptive_properties,
                adsorbent_key=self.hk_adsorbent_key,
                adsorptive_key=self.hk_adsorptive_key,
            )
        else:
            interaction_parameter = self._read_hk_interaction_parameter()
        return {
            "geometry": self.hk_geometry,
            "adsorbent_key": self.hk_adsorbent_key,
            "adsorptive_key": self.hk_adsorptive_key,
            "adsorbent_properties": dict(self.hk_adsorbent_properties),
            "adsorptive_properties": dict(self.hk_adsorptive_properties),
            "interaction_parameter_mode": self.hk_interaction_parameter_mode,
            "interaction_parameter": interaction_parameter,
            "cheng_yang_correction": self.hk_cheng_yang_correction,
            "smooth_derivative": self.hk_smooth_derivative,
            "display_metric": self.hk_display_metric,
        }

    def apply_hk_settings_to_all(self) -> None:
        if not self.results:
            return
        settings = self._current_hk_settings_snapshot()
        self.hk_interaction_parameter = float(settings["interaction_parameter"])
        if hasattr(self, "hk_interaction_edit"):
            self.hk_interaction_edit.setText(f"{self.hk_interaction_parameter:.3e}")
        for result in self.results:
            self.custom_hk_settings[id(result)] = copy.deepcopy(settings)
        self.refresh_hk_plot()
        self.statusBar().showMessage("已将当前 HK 设置应用到所有样品", 3000)

    def reset_all_hk_to_default(self) -> None:
        self.custom_hk_settings.clear()
        self.reset_hk_to_default(reset_region=True)
        self.statusBar().showMessage("所有样品已恢复默认 HK 设置", 3000)

    def _open_hk_properties_dialog(self) -> None:
        dialog = HorvathKawazoePropertiesDialog(
            self,
            adsorbent_key=self.hk_adsorbent_key,
            adsorptive_key=self.hk_adsorptive_key,
            adsorbent_properties=self.hk_adsorbent_properties,
            adsorptive_properties=self.hk_adsorptive_properties,
        )
        if dialog.exec_() != QtWidgets.QDialog.Accepted:
            return
        self.hk_adsorbent_key = dialog.selected_adsorbent_key()
        self.hk_adsorptive_key = dialog.selected_adsorptive_key()
        self.hk_adsorbent_properties = dialog.adsorbent_properties()
        self.hk_adsorptive_properties = dialog.adsorptive_properties()
        self._update_hk_interaction_controls()
        if self.hk_interaction_parameter_mode == "calculated":
            self.hk_interaction_parameter = calculate_hk_interaction_parameter(
                self.hk_adsorbent_properties,
                self.hk_adsorptive_properties,
                adsorbent_key=self.hk_adsorbent_key,
                adsorptive_key=self.hk_adsorptive_key,
            )
        self._save_hk_settings_for_active()
        self.refresh_hk_plot()

    def reset_hk_to_default(self, *, reset_region: bool = False) -> None:
        active = self.active_result()
        if active is not None:
            self.custom_hk_settings.pop(id(active), None)
        settings = self._default_hk_settings()
        self._syncing_hk_controls = True
        try:
            self._apply_hk_settings(settings)
            self._sync_hk_controls_from_state()
        finally:
            self._syncing_hk_controls = False
        if reset_region:
            self.hk_pore_volume_range = DEFAULT_HK_PORE_VOLUME_RANGE
            self._remove_hk_region()
        self.refresh_hk_plot()

    def _on_t_plot_surface_area_mode_changed(self) -> None:
        if self._syncing_t_plot_controls:
            return
        if self.surface_area_bet_radio.isChecked():
            self.t_plot_surface_area_mode = "BET"
        elif self.surface_area_langmuir_radio.isChecked():
            self.t_plot_surface_area_mode = "Langmuir"
        elif self.surface_area_input_radio.isChecked():
            self.t_plot_surface_area_mode = "Input"
        self.surface_area_input_spin.setEnabled(self.t_plot_surface_area_mode == "Input")
        self._save_t_plot_settings_for_active()
        self._refresh_sample_t_plot_cell(self.active_index)
        self.refresh_metrics()

    def _on_t_plot_surface_area_input_changed(self) -> None:
        if self._syncing_t_plot_controls:
            return
        self.t_plot_surface_area_input = float(self.surface_area_input_spin.value())
        self._save_t_plot_settings_for_active()
        if self.t_plot_surface_area_mode == "Input":
            self._refresh_sample_t_plot_cell(self.active_index)
            self.refresh_metrics()

    def _on_t_plot_surface_area_correction_changed(self) -> None:
        if self._syncing_t_plot_controls:
            return
        self.t_plot_surface_area_correction = float(self.surface_area_correction_spin.value())
        self._save_t_plot_settings_for_active()
        self._refresh_sample_t_plot_cell(self.active_index)
        self.refresh_metrics()

    def _refresh_t_plot_for_option_change(self, *, refresh_table: bool = False) -> None:
        active = self.active_result()
        if active is None:
            return
        pressure_range = self._current_pressure_region()
        p_min, p_max = pressure_range if pressure_range else (None, None)
        self._refresh_t_plot_plot(active, p_min, p_max, reset_region=False)
        if refresh_table:
            self._refresh_sample_t_plot_cell(self.active_index)
        self.refresh_metrics()

    def _default_t_plot_settings(self) -> dict[str, object]:
        params_by_method = _default_t_plot_thickness_params_by_method()
        return {
            "thickness_method": DEFAULT_T_PLOT_THICKNESS_METHOD,
            "thickness_params_by_method": params_by_method,
            "thickness_params": dict(params_by_method[DEFAULT_T_PLOT_THICKNESS_METHOD]),
            "surface_area_mode": DEFAULT_T_PLOT_SURFACE_AREA_MODE,
            "surface_area_input": DEFAULT_T_PLOT_SURFACE_AREA_INPUT,
            "surface_area_correction": DEFAULT_T_PLOT_SURFACE_AREA_CORRECTION,
        }

    def _t_plot_settings_for_result(self, result) -> dict[str, object]:
        settings = self._default_t_plot_settings()
        custom = self.custom_t_plot_settings.get(id(result))
        if custom:
            settings.update(custom)
            params_by_method = _default_t_plot_thickness_params_by_method()
            if "thickness_params_by_method" in custom:
                for method_key, params in dict(custom["thickness_params_by_method"]).items():
                    if method_key in params_by_method:
                        params_by_method[method_key] = {
                            **params_by_method[method_key],
                            **dict(params),
                        }
            elif "thickness_params" in custom:
                method_key = str(settings["thickness_method"])
                if method_key in params_by_method:
                    params_by_method[method_key] = {
                        **params_by_method[method_key],
                        **dict(custom["thickness_params"]),
                    }
            settings["thickness_params_by_method"] = params_by_method
            method_key = str(settings["thickness_method"])
            settings["thickness_params"] = dict(
                params_by_method.get(method_key, params_by_method[DEFAULT_T_PLOT_THICKNESS_METHOD])
            )
        return settings

    def _save_t_plot_settings_for_active(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self.custom_t_plot_settings[id(active)] = {
            "thickness_method": self.t_plot_thickness_method,
            "thickness_params_by_method": {
                method_key: dict(params)
                for method_key, params in self.t_plot_thickness_params_by_method.items()
            },
            "thickness_params": dict(self.t_plot_thickness_params),
            "surface_area_mode": self.t_plot_surface_area_mode,
            "surface_area_input": self.t_plot_surface_area_input,
            "surface_area_correction": self.t_plot_surface_area_correction,
        }

    def _load_t_plot_settings_for_active(self) -> None:
        active = self.active_result()
        settings = self._default_t_plot_settings() if active is None else self._t_plot_settings_for_result(active)
        self._syncing_t_plot_controls = True
        try:
            self.t_plot_thickness_method = str(settings["thickness_method"])
            self.t_plot_thickness_params_by_method = {
                method_key: dict(params)
                for method_key, params in dict(settings["thickness_params_by_method"]).items()
            }
            self.t_plot_thickness_params = dict(settings["thickness_params"])
            self.t_plot_surface_area_mode = str(settings["surface_area_mode"])
            self.t_plot_surface_area_input = float(settings["surface_area_input"])
            self.t_plot_surface_area_correction = float(settings["surface_area_correction"])

            for key, radio in self.t_plot_method_radios.items():
                radio.setChecked(key == self.t_plot_thickness_method)
            self._set_all_t_plot_formula_spins()
            self.surface_area_bet_radio.setChecked(self.t_plot_surface_area_mode == "BET")
            self.surface_area_langmuir_radio.setChecked(self.t_plot_surface_area_mode == "Langmuir")
            self.surface_area_input_radio.setChecked(self.t_plot_surface_area_mode == "Input")
            self.surface_area_input_spin.setValue(self.t_plot_surface_area_input)
            self.surface_area_input_spin.setEnabled(self.t_plot_surface_area_mode == "Input")
            self.surface_area_correction_spin.setValue(self.t_plot_surface_area_correction)
            self._sync_reference_table_for_context("t_plot")
        finally:
            self._syncing_t_plot_controls = False

    def _default_bjh_settings(self) -> dict[str, object]:
        params_by_method = _default_bjh_thickness_params_by_method()
        return {
            "thickness_method": DEFAULT_BJH_THICKNESS_METHOD,
            "thickness_params_by_method": params_by_method,
            "thickness_params": dict(params_by_method[DEFAULT_BJH_THICKNESS_METHOD]),
            "correction": DEFAULT_BJH_CORRECTION,
            "open_pore_fraction": DEFAULT_BJH_OPEN_PORE_FRACTION,
            "smooth_derivative": DEFAULT_BJH_SMOOTH_DERIVATIVE,
            "show_adsorption": DEFAULT_BJH_SHOW_ADSORPTION,
            "show_desorption": DEFAULT_BJH_SHOW_DESORPTION,
        }

    def _bjh_default_settings_for_result(self, result) -> dict[str, object]:
        settings = self._default_bjh_settings()
        vendor_method = (
            result.method_options.get("vendor_bjh_thickness_method")
            or result.method_options.get("bsd_bjh_thickness_method")
        )
        if vendor_method:
            params_by_method = _default_bjh_thickness_params_by_method()
            method_key = str(vendor_method)
            if method_key in params_by_method:
                settings["thickness_method"] = method_key
                settings["thickness_params_by_method"] = params_by_method
                settings["thickness_params"] = dict(params_by_method[method_key])
        vendor_correction = result.method_options.get("vendor_bjh_correction")
        if vendor_correction in {"standard", "kjs", "faas"}:
            settings["correction"] = str(vendor_correction)
        vendor_smooth = result.method_options.get("vendor_bjh_smooth_derivative")
        if vendor_smooth is not None:
            settings["smooth_derivative"] = bool(vendor_smooth)
        return settings

    def _bjh_settings_for_result(self, result) -> dict[str, object]:
        settings = self._bjh_default_settings_for_result(result)
        custom = self.custom_bjh_settings.get(id(result))
        if custom:
            settings.update(custom)
            params_by_method = {
                method_key: dict(params)
                for method_key, params in dict(settings["thickness_params_by_method"]).items()
            }
            if "thickness_params_by_method" in custom:
                for method_key, params in dict(custom["thickness_params_by_method"]).items():
                    if method_key in params_by_method:
                        params_by_method[method_key] = {
                            **params_by_method[method_key],
                            **dict(params),
                        }
            elif "thickness_params" in custom:
                method_key = str(settings["thickness_method"])
                if method_key in params_by_method:
                    params_by_method[method_key] = {
                        **params_by_method[method_key],
                        **dict(custom["thickness_params"]),
                    }
            settings["thickness_params_by_method"] = params_by_method
            method_key = str(settings["thickness_method"])
            settings["thickness_params"] = dict(
                params_by_method.get(method_key, params_by_method[DEFAULT_BJH_THICKNESS_METHOD])
            )
        return settings

    def _save_bjh_settings_for_active(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self.custom_bjh_settings[id(active)] = {
            "thickness_method": self.bjh_thickness_method,
            "thickness_params_by_method": {
                method_key: dict(params)
                for method_key, params in self.bjh_thickness_params_by_method.items()
            },
            "thickness_params": dict(self.bjh_thickness_params),
            "correction": self.bjh_correction,
            "open_pore_fraction": self.bjh_open_pore_fraction,
            "smooth_derivative": self.bjh_smooth_derivative,
            "show_adsorption": self.bjh_show_adsorption,
            "show_desorption": self.bjh_show_desorption,
        }

    def _load_bjh_settings_for_active(self) -> None:
        active = self.active_result()
        settings = self._default_bjh_settings() if active is None else self._bjh_settings_for_result(active)
        self._syncing_bjh_controls = True
        try:
            self.bjh_thickness_method = str(settings["thickness_method"])
            self.bjh_thickness_params_by_method = {
                method_key: dict(params)
                for method_key, params in dict(settings["thickness_params_by_method"]).items()
            }
            self.bjh_thickness_params = dict(settings["thickness_params"])
            self.bjh_correction = str(settings["correction"])
            self.bjh_open_pore_fraction = float(settings["open_pore_fraction"])
            self.bjh_smooth_derivative = bool(settings["smooth_derivative"])
            self.bjh_show_adsorption = bool(settings["show_adsorption"])
            self.bjh_show_desorption = bool(settings["show_desorption"])

            for key, radio in self.bjh_method_radios.items():
                radio.setChecked(key == self.bjh_thickness_method)
            self._set_all_bjh_formula_spins()
            self.bjh_standard_radio.setChecked(self.bjh_correction == "standard")
            self.bjh_kjs_correction_radio.setChecked(self.bjh_correction == "kjs")
            self.bjh_faas_correction_radio.setChecked(self.bjh_correction == "faas")
            self.bjh_open_fraction_spin.setValue(self.bjh_open_pore_fraction)
            self.bjh_smooth_checkbox.setChecked(self.bjh_smooth_derivative)
            self.bjh_adsorption_checkbox.setChecked(self.bjh_show_adsorption)
            self.bjh_desorption_checkbox.setChecked(self.bjh_show_desorption)
            self._sync_reference_table_for_context("bjh")
        finally:
            self._syncing_bjh_controls = False

    def _default_dh_settings(self) -> dict[str, object]:
        params_by_method = _default_bjh_thickness_params_by_method()
        return {
            "thickness_method": DEFAULT_DH_THICKNESS_METHOD,
            "thickness_params_by_method": params_by_method,
            "thickness_params": dict(params_by_method[DEFAULT_DH_THICKNESS_METHOD]),
            "smooth_derivative": DEFAULT_DH_SMOOTH_DERIVATIVE,
            "show_adsorption": DEFAULT_DH_SHOW_ADSORPTION,
            "show_desorption": DEFAULT_DH_SHOW_DESORPTION,
            "display_metrics": list(DEFAULT_DH_DISPLAY_METRICS),
        }

    def _dh_default_settings_for_result(self, result) -> dict[str, object]:
        settings = self._default_dh_settings()
        if result is None:
            return settings
        params_by_method = dict(settings["thickness_params_by_method"])
        method = str(result.method_options.get("vendor_dh_thickness_method", settings["thickness_method"]))
        if method not in params_by_method:
            method = DEFAULT_DH_THICKNESS_METHOD
        settings["thickness_method"] = method
        settings["thickness_params"] = dict(params_by_method[method])
        if "vendor_dh_smooth_derivative" in result.method_options:
            settings["smooth_derivative"] = bool(result.method_options.get("vendor_dh_smooth_derivative"))
        if "vendor_dh_differential_mode" in result.method_options:
            settings["display_metrics"] = _normalize_bjh_display_metrics(
                result.method_options.get("vendor_dh_differential_mode")
            )
        return settings

    def _dh_settings_for_result(self, result) -> dict[str, object]:
        settings = self._dh_default_settings_for_result(result)
        custom = self.custom_dh_settings.get(id(result))
        if custom:
            settings.update(custom)
            params_by_method = _default_bjh_thickness_params_by_method()
            params_by_method.update(
                {
                    method_key: dict(params)
                    for method_key, params in dict(settings.get("thickness_params_by_method", {})).items()
                }
            )
            if "thickness_params" in custom:
                method_key = str(settings["thickness_method"])
                if method_key in params_by_method:
                    params_by_method[method_key] = {
                        **params_by_method[method_key],
                        **dict(custom["thickness_params"]),
                    }
            settings["thickness_params_by_method"] = params_by_method
            method_key = str(settings["thickness_method"])
            settings["thickness_params"] = dict(
                params_by_method.get(method_key, params_by_method[DEFAULT_DH_THICKNESS_METHOD])
            )
            settings["display_metrics"] = _normalize_bjh_display_metrics(
                settings.get("display_metrics", DEFAULT_DH_DISPLAY_METRICS)
            )
        return settings

    def _save_dh_settings_for_active(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self.custom_dh_settings[id(active)] = {
            "thickness_method": self.dh_thickness_method,
            "thickness_params_by_method": {
                method_key: dict(params)
                for method_key, params in self.dh_thickness_params_by_method.items()
            },
            "thickness_params": dict(self.dh_thickness_params),
            "smooth_derivative": self.dh_smooth_derivative,
            "show_adsorption": self.dh_show_adsorption,
            "show_desorption": self.dh_show_desorption,
            "display_metrics": list(self.dh_display_metrics),
        }

    def _load_dh_settings_for_active(self) -> None:
        active = self.active_result()
        settings = self._default_dh_settings() if active is None else self._dh_settings_for_result(active)
        self._syncing_dh_controls = True
        try:
            self.dh_thickness_method = str(settings["thickness_method"])
            self.dh_thickness_params_by_method = {
                method_key: dict(params)
                for method_key, params in dict(settings["thickness_params_by_method"]).items()
            }
            self.dh_thickness_params = dict(settings["thickness_params"])
            self.dh_smooth_derivative = bool(settings["smooth_derivative"])
            self.dh_show_adsorption = bool(settings["show_adsorption"])
            self.dh_show_desorption = bool(settings["show_desorption"])
            self.dh_display_metrics = _normalize_bjh_display_metrics(settings["display_metrics"])
            self.dh_differential_mode = self.dh_display_metrics[0]

            for key, radio in self.dh_method_radios.items():
                radio.setChecked(key == self.dh_thickness_method)
            self._set_all_dh_formula_spins()
            self.dh_smooth_checkbox.setChecked(self.dh_smooth_derivative)
            self.dh_adsorption_checkbox.setChecked(self.dh_show_adsorption)
            self.dh_desorption_checkbox.setChecked(self.dh_show_desorption)
            self._set_dh_display_combo()
            self._sync_reference_table_for_context("dh")
        finally:
            self._syncing_dh_controls = False

    def _default_hk_settings(self) -> dict[str, object]:
        return {
            "geometry": DEFAULT_HK_GEOMETRY,
            "adsorbent_key": DEFAULT_HK_ADSORBENT,
            "adsorptive_key": DEFAULT_HK_ADSORPTIVE,
            "adsorbent_properties": dict(HK_ADSORBENT_PRESETS[DEFAULT_HK_ADSORBENT]),
            "adsorptive_properties": dict(HK_ADSORPTIVE_PRESETS[DEFAULT_HK_ADSORPTIVE]),
            "interaction_parameter_mode": DEFAULT_HK_INTERACTION_PARAMETER_MODE,
            "interaction_parameter": DEFAULT_HK_INTERACTION_PARAMETER,
            "cheng_yang_correction": DEFAULT_HK_CHENG_YANG_CORRECTION,
            "smooth_derivative": DEFAULT_HK_SMOOTH_DERIVATIVE,
            "display_metric": DEFAULT_HK_DISPLAY_METRIC,
        }

    def _hk_settings_for_result(self, result) -> dict[str, object]:
        settings = self._default_hk_settings()
        custom = self.custom_hk_settings.get(id(result))
        if custom:
            settings.update(custom)
            settings["adsorbent_properties"] = {
                **dict(HK_ADSORBENT_PRESETS.get(str(settings["adsorbent_key"]), HK_ADSORBENT_PRESETS[DEFAULT_HK_ADSORBENT])),
                **dict(custom.get("adsorbent_properties", {})),
            }
            settings["adsorptive_properties"] = {
                **dict(HK_ADSORPTIVE_PRESETS.get(str(settings["adsorptive_key"]), HK_ADSORPTIVE_PRESETS[DEFAULT_HK_ADSORPTIVE])),
                **dict(custom.get("adsorptive_properties", {})),
            }
        return settings

    def _apply_hk_settings(self, settings: dict[str, object]) -> None:
        self.hk_geometry = str(settings.get("geometry", DEFAULT_HK_GEOMETRY))
        if self.hk_geometry not in {"slit", "cylinder"}:
            self.hk_geometry = DEFAULT_HK_GEOMETRY
        self.hk_adsorbent_key = str(settings.get("adsorbent_key", DEFAULT_HK_ADSORBENT))
        if self.hk_adsorbent_key not in HK_ADSORBENT_PRESETS:
            self.hk_adsorbent_key = DEFAULT_HK_ADSORBENT
        self.hk_adsorptive_key = str(settings.get("adsorptive_key", DEFAULT_HK_ADSORPTIVE))
        if self.hk_adsorptive_key not in HK_ADSORPTIVE_PRESETS:
            self.hk_adsorptive_key = DEFAULT_HK_ADSORPTIVE
        self.hk_adsorbent_properties = {
            **dict(HK_ADSORBENT_PRESETS[self.hk_adsorbent_key]),
            **dict(settings.get("adsorbent_properties", {})),
        }
        self.hk_adsorptive_properties = {
            **dict(HK_ADSORPTIVE_PRESETS[self.hk_adsorptive_key]),
            **dict(settings.get("adsorptive_properties", {})),
        }
        mode = str(settings.get("interaction_parameter_mode", DEFAULT_HK_INTERACTION_PARAMETER_MODE))
        self.hk_interaction_parameter_mode = "calculated" if mode == "calculated" else "input"
        try:
            self.hk_interaction_parameter = float(settings.get("interaction_parameter", DEFAULT_HK_INTERACTION_PARAMETER))
        except (TypeError, ValueError):
            self.hk_interaction_parameter = DEFAULT_HK_INTERACTION_PARAMETER
        if not (np.isfinite(self.hk_interaction_parameter) and self.hk_interaction_parameter > 0.0):
            self.hk_interaction_parameter = DEFAULT_HK_INTERACTION_PARAMETER
        self.hk_cheng_yang_correction = bool(settings.get("cheng_yang_correction", DEFAULT_HK_CHENG_YANG_CORRECTION))
        self.hk_smooth_derivative = bool(settings.get("smooth_derivative", DEFAULT_HK_SMOOTH_DERIVATIVE))
        self.hk_display_metric = normalize_hk_display_metric(settings.get("display_metric", DEFAULT_HK_DISPLAY_METRIC))

    def _save_hk_settings_for_active(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self.custom_hk_settings[id(active)] = {
            "geometry": self.hk_geometry,
            "adsorbent_key": self.hk_adsorbent_key,
            "adsorptive_key": self.hk_adsorptive_key,
            "adsorbent_properties": dict(self.hk_adsorbent_properties),
            "adsorptive_properties": dict(self.hk_adsorptive_properties),
            "interaction_parameter_mode": self.hk_interaction_parameter_mode,
            "interaction_parameter": self.hk_interaction_parameter,
            "cheng_yang_correction": self.hk_cheng_yang_correction,
            "smooth_derivative": self.hk_smooth_derivative,
            "display_metric": self.hk_display_metric,
        }

    def _sync_hk_controls_from_state(self) -> None:
        for key, radio in getattr(self, "hk_geometry_radios", {}).items():
            radio.setChecked(key == self.hk_geometry)
        if hasattr(self, "hk_calculated_radio"):
            self.hk_calculated_radio.setChecked(self.hk_interaction_parameter_mode == "calculated")
        if hasattr(self, "hk_input_radio"):
            self.hk_input_radio.setChecked(self.hk_interaction_parameter_mode != "calculated")
        if hasattr(self, "hk_interaction_edit"):
            self.hk_interaction_edit.setText(f"{self.hk_interaction_parameter:.3e}")
        if hasattr(self, "hk_cheng_yang_checkbox"):
            self.hk_cheng_yang_checkbox.setChecked(self.hk_cheng_yang_correction)
        if hasattr(self, "hk_smooth_checkbox"):
            self.hk_smooth_checkbox.setChecked(self.hk_smooth_derivative)
        self._set_hk_display_combo()
        self._update_hk_interaction_controls()

    def _load_hk_settings_for_active(self) -> None:
        active = self.active_result()
        settings = self._default_hk_settings() if active is None else self._hk_settings_for_result(active)
        self._syncing_hk_controls = True
        try:
            self._apply_hk_settings(settings)
            self._sync_hk_controls_from_state()
        finally:
            self._syncing_hk_controls = False

    def _default_dft_settings(self) -> dict[str, object]:
        return {
            "analysis_type": DEFAULT_DFT_ANALYSIS_TYPE,
            "geometry": DEFAULT_DFT_GEOMETRY,
            "model": DEFAULT_DFT_MODEL,
            "regularization": DEFAULT_DFT_REGULARIZATION,
        }

    def _dft_default_settings_for_result(self, result) -> dict[str, object]:
        settings = self._default_dft_settings()
        if result is None:
            return settings
        vendor_model = result.method_options.get("vendor_dft_model")
        if vendor_model:
            settings["model"] = str(vendor_model)
        vendor_regularization = result.method_options.get("vendor_dft_regularization")
        if vendor_regularization is not None:
            try:
                settings["regularization"] = float(vendor_regularization)
            except (TypeError, ValueError):
                pass
        return settings

    def _dft_settings_for_result(self, result) -> dict[str, object]:
        settings = self._dft_default_settings_for_result(result)
        custom = self.custom_dft_settings.get(id(result))
        if custom:
            settings.update(custom)
        return settings

    def _apply_dft_settings(self, settings: dict[str, object]) -> None:
        self.dft_analysis_type = str(settings.get("analysis_type", DEFAULT_DFT_ANALYSIS_TYPE))
        if self.dft_analysis_type not in {"dft_pore", "typical"}:
            self.dft_analysis_type = DEFAULT_DFT_ANALYSIS_TYPE
        self.dft_geometry = str(settings.get("geometry", DEFAULT_DFT_GEOMETRY))
        if self.dft_geometry not in {"slit", "cylinder"}:
            self.dft_geometry = DEFAULT_DFT_GEOMETRY
        self.dft_model = str(settings.get("model", DEFAULT_DFT_MODEL))
        try:
            regularization = float(settings.get("regularization", DEFAULT_DFT_REGULARIZATION))
        except (TypeError, ValueError):
            regularization = DEFAULT_DFT_REGULARIZATION
        if not np.isfinite(regularization):
            regularization = DEFAULT_DFT_REGULARIZATION
        self.dft_regularization = max(0.0, min(regularization, 10.0))

    def _sync_dft_controls_from_state(self) -> None:
        self._set_combo_data(getattr(self, "dft_type_combo", None), self.dft_analysis_type)
        self._set_combo_data(getattr(self, "dft_geometry_combo", None), self.dft_geometry)
        self._set_combo_data(getattr(self, "dft_model_combo", None), self.dft_model)
        slider = getattr(self, "dft_regularization_slider", None)
        if slider is not None:
            slider.setValue(self.dft_regularization, emit=False)

    @staticmethod
    def _set_combo_data(combo: QtWidgets.QComboBox | None, value: object) -> None:
        if combo is None:
            return
        target = str(value)
        for index in range(combo.count()):
            if str(combo.itemData(index)) == target:
                combo.setCurrentIndex(index)
                return

    def _save_dft_settings_for_active(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self._store_dft_settings_for_result(active, self._current_dft_settings_snapshot())

    def _store_dft_settings_for_result(self, result, settings: dict[str, object]) -> None:
        defaults = self._dft_default_settings_for_result(result)
        if _settings_mapping_equal(settings, defaults):
            self.custom_dft_settings.pop(id(result), None)
        else:
            self.custom_dft_settings[id(result)] = copy.deepcopy(settings)

    def _apply_dft_regularization_to_all(self, value: float) -> None:
        for result in self.results:
            settings = self._dft_settings_for_result(result)
            settings["regularization"] = float(value)
            self._store_dft_settings_for_result(result, settings)

    def _on_dft_regularization_apply_all_toggled(self, checked: bool) -> None:
        self.dft_regularization_apply_all = bool(checked)
        if not checked:
            self._pending_dft_regularization_apply_all_value = None
            self._dft_regularization_preview_active = False
            self._dft_regularization_refresh_timer.stop()

    def _sync_dft_diagnostic_line_to_regularization(self) -> None:
        line = getattr(self, "_dft_diagnostic_line", None)
        if line is None:
            return
        try:
            position = math.log10(max(float(self.dft_regularization), 1e-6))
        except (TypeError, ValueError):
            return
        try:
            line.blockSignals(True)
            line.setValue(position)
        finally:
            line.blockSignals(False)

    def _refresh_dft_regularization_dependents(self, *, preview: bool = False) -> None:
        self._sync_dft_diagnostic_line_to_regularization()
        self.refresh_dft_plot(refresh_diagnostics=not preview, refresh_pore_cells=not preview)
        if not preview and self._is_dft_tab_active():
            self.refresh_isotherm_plot()
        if not preview and self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT:
            self._refresh_all_sample_bjh_pore_cells()

    def _schedule_deferred_dft_regularization_refresh(self) -> None:
        if not self._dft_regularization_refresh_timer.isActive():
            self._dft_regularization_refresh_timer.start()

    def _preview_deferred_dft_regularization_refresh(self) -> None:
        if self._pending_dft_regularization_apply_all_value is None:
            return
        value = float(self._pending_dft_regularization_apply_all_value)
        self._pending_dft_regularization_apply_all_value = None
        self.dft_regularization = max(0.0, min(value, 10.0))
        self._sync_dft_diagnostic_line_to_regularization()
        self._apply_dft_regularization_to_all(value)
        self._dft_regularization_preview_active = True
        self._refresh_dft_regularization_dependents(preview=True)

    def _finish_deferred_dft_regularization_refresh(self, *_args) -> None:
        if not getattr(self, "dft_regularization_apply_all", False):
            return
        self._dft_regularization_refresh_timer.stop()
        if self._pending_dft_regularization_apply_all_value is None and not self._dft_regularization_preview_active:
            return
        if self._pending_dft_regularization_apply_all_value is not None:
            value = float(self._pending_dft_regularization_apply_all_value)
            self._pending_dft_regularization_apply_all_value = None
            self.dft_regularization = max(0.0, min(value, 10.0))
            self._sync_dft_diagnostic_line_to_regularization()
            self._apply_dft_regularization_to_all(value)
        self._dft_regularization_preview_active = False
        self._refresh_dft_regularization_dependents(preview=False)

    def _load_dft_settings_for_active(self) -> None:
        active = self.active_result()
        settings = self._default_dft_settings() if active is None else self._dft_settings_for_result(active)
        self._syncing_dft_controls = True
        try:
            self._apply_dft_settings(settings)
            self._sync_dft_controls_from_state()
        finally:
            self._syncing_dft_controls = False

    def _current_dft_settings_snapshot(self) -> dict[str, object]:
        return {
            "analysis_type": self.dft_analysis_type,
            "geometry": self.dft_geometry,
            "model": self.dft_model,
            "regularization": float(self.dft_regularization),
        }

    def _on_dft_option_changed(self, *_args) -> None:
        if self._syncing_dft_controls:
            return
        self.dft_analysis_type = str(self.dft_type_combo.currentData())
        self.dft_geometry = str(self.dft_geometry_combo.currentData())
        self.dft_model = str(self.dft_model_combo.currentData())
        self._save_dft_settings_for_active()
        self.refresh_dft_plot()
        if self._is_dft_tab_active():
            self.refresh_isotherm_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT:
            self._refresh_all_sample_bjh_pore_cells()

    def _on_dft_regularization_changed(self, value: float) -> None:
        if self._syncing_dft_controls:
            return
        self.dft_regularization = max(0.0, min(float(value), 10.0))
        self._sync_dft_diagnostic_line_to_regularization()
        if getattr(self, "dft_regularization_apply_all", False):
            slider = getattr(self, "dft_regularization_slider", None)
            if slider is not None and slider.isDragging():
                self._pending_dft_regularization_apply_all_value = self.dft_regularization
                self._schedule_deferred_dft_regularization_refresh()
                return
            self._pending_dft_regularization_apply_all_value = None
            self._dft_regularization_preview_active = False
            self._dft_regularization_refresh_timer.stop()
            self._apply_dft_regularization_to_all(self.dft_regularization)
        else:
            self._save_dft_settings_for_active()
        self._refresh_dft_regularization_dependents()

    def _on_dft_diagnostic_line_changed(self) -> None:
        if self._syncing_dft_controls or self._dft_diagnostic_line is None:
            return
        try:
            value = 10.0 ** float(self._dft_diagnostic_line.value())
        except (TypeError, ValueError):
            return
        if value <= 1e-6:
            value = 0.0
        self._syncing_dft_controls = True
        try:
            self.dft_regularization = max(0.0, min(value, 10.0))
            if hasattr(self, "dft_regularization_slider"):
                self.dft_regularization_slider.setValue(self.dft_regularization, emit=False)
        finally:
            self._syncing_dft_controls = False
        if getattr(self, "dft_regularization_apply_all", False):
            self._apply_dft_regularization_to_all(self.dft_regularization)
        else:
            self._save_dft_settings_for_active()
        self._refresh_dft_regularization_dependents()

    def apply_dft_settings_to_all(self) -> None:
        if not self.results:
            return
        settings = self._current_dft_settings_snapshot()
        for result in self.results:
            self._store_dft_settings_for_result(result, settings)
        self.refresh_dft_plot()
        if self._is_dft_tab_active():
            self.refresh_isotherm_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT:
            self._refresh_all_sample_bjh_pore_cells()
        self.statusBar().showMessage("已将当前 DFT 设置应用到所有样品", 3000)

    def reset_all_dft_to_default(self) -> None:
        self.custom_dft_settings.clear()
        self._dft_result_cache.clear()
        self.reset_dft_to_default(reset_region=True)
        self.statusBar().showMessage("所有样品已恢复默认 DFT 设置", 3000)

    def reset_dft_to_default(self, *, reset_region: bool = False) -> None:
        active = self.active_result()
        if active is not None:
            self.custom_dft_settings.pop(id(active), None)
        settings = self._default_dft_settings() if active is None else self._dft_default_settings_for_result(active)
        self._syncing_dft_controls = True
        try:
            self._apply_dft_settings(settings)
            self._sync_dft_controls_from_state()
        finally:
            self._syncing_dft_controls = False
        if reset_region:
            self.dft_pore_volume_range = DEFAULT_DFT_PORE_VOLUME_RANGE
            self._remove_dft_region()
        self.refresh_dft_plot()
        if self._is_dft_tab_active():
            self.refresh_isotherm_plot()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT:
            self._refresh_all_sample_bjh_pore_cells()

    def _read_directory_setting(self, key: str) -> Path:
        value = self.settings.value(key, "")
        if value:
            path = Path(str(value))
            if path.is_dir():
                return path
        return _default_user_directory()

    def _write_directory_setting(self, key: str, path: Path | str) -> None:
        directory = Path(path)
        if directory.is_file():
            directory = directory.parent
        if directory.is_dir():
            self.settings.setValue(key, str(directory))

    def open_files(self) -> None:
        existing_paths = [result.header.file_path for result in self.results]
        dialog = FileImportDialog(
            self,
            self.import_directory,
            existing_paths=existing_paths,
            available_sort=self._import_available_sort,
        )
        if dialog.exec_() != QtWidgets.QDialog.Accepted:
            self._import_available_sort = dialog.available_sort()
            return
        paths = dialog.selected_paths()
        if not paths:
            self._import_available_sort = dialog.available_sort()
            return
        self.import_directory = dialog.current_directory
        self._write_directory_setting("import_directory", self.import_directory)
        self._import_available_sort = dialog.available_sort()
        self.settings.setValue("import_available_sort_column", self._import_available_sort[0])
        self.settings.setValue("import_available_sort_order", int(self._import_available_sort[1]))
        self.sync_files(paths)

    def append_files(self, paths: list[str]) -> None:
        self.load_files(paths, replace=False)

    @staticmethod
    def _path_key(path: str) -> str:
        try:
            return str(Path(path).resolve()).lower()
        except OSError:
            return str(path).lower()

    def sync_files(self, paths: list[str]) -> None:
        """Make the sample list match ``paths``: reuse already-loaded results
        (keeping their per-sample settings), parse newly added files, and drop
        any sample no longer present."""
        existing_by_key = {self._path_key(result.header.file_path): result for result in self.results}
        new_results = []
        errors = []
        seen_keys: set[str] = set()
        for path in paths:
            key = self._path_key(path)
            if key in seen_keys:
                continue
            seen_keys.add(key)
            existing = existing_by_key.get(key)
            if existing is not None:
                new_results.append(existing)
                continue
            try:
                result = load_file(path)
            except (OSError, TriStarParseError, BELMasterParseError, ExcelParseError, ValueError) as exc:
                errors.append(f"{Path(path).name}: {exc}")
                continue
            if result.point_count <= 0:
                errors.append(f"{Path(path).name}: 没有解析到实际等温线，已跳过")
                continue
            new_results.append(result)

        if errors:
            QtWidgets.QMessageBox.warning(self, "部分文件未加载", "\n".join(errors))
        if not new_results:
            return

        kept_visibility = {id(result): visible for result, visible in zip(self.results, self.visible_results)}
        active_result = self.results[self.active_index] if 0 <= self.active_index < len(self.results) else None

        # Drop per-sample settings for removed samples (keyed by object id).
        retained_ids = {id(result) for result in new_results}
        for removed in self.results:
            if id(removed) not in retained_ids:
                self._discard_sample_settings(removed)

        self.results = new_results
        self.visible_results = [kept_visibility.get(id(result), True) for result in new_results]
        if active_result in new_results:
            self.active_index = new_results.index(active_result)
        else:
            self.active_index = 0
        self.refresh_all()

    def _discard_sample_settings(self, result) -> None:
        self.custom_bet_fit_ranges.pop(id(result), None)
        self.custom_langmuir_fit_ranges.pop(id(result), None)
        self.custom_t_plot_fit_ranges.pop(id(result), None)
        self.custom_t_plot_settings.pop(id(result), None)
        self.custom_bjh_settings.pop(id(result), None)
        self.custom_dh_settings.pop(id(result), None)
        self.custom_hk_settings.pop(id(result), None)
        self.custom_dft_settings.pop(id(result), None)
        self._discard_fit_analysis_cache_for_result(result)
        self._discard_bjh_distribution_cache_for_result(result)
        self._discard_dh_distribution_cache_for_result(result)
        self._discard_hk_distribution_cache_for_result(result)
        self._discard_dft_result_cache_for_result(result)

    def load_files(self, paths: Iterable[str], *, replace: bool) -> None:
        parsed = []
        errors = []
        for path in paths:
            try:
                result = load_file(path)
            except (OSError, TriStarParseError, BELMasterParseError, ExcelParseError, ValueError) as exc:
                errors.append(f"{Path(path).name}: {exc}")
                continue
            if result.point_count <= 0:
                errors.append(f"{Path(path).name}: 没有解析到实际等温线，已跳过")
                continue
            parsed.append(result)

        if errors:
            QtWidgets.QMessageBox.warning(self, "部分文件未加载", "\n".join(errors))
        if not parsed:
            return

        if replace:
            self.results = parsed
            self.visible_results = [True] * len(parsed)
            self.active_index = 0
            self.custom_bet_fit_ranges.clear()
            self.custom_langmuir_fit_ranges.clear()
            self.custom_t_plot_fit_ranges.clear()
            self.custom_t_plot_settings.clear()
            self.custom_bjh_settings.clear()
            self.custom_dh_settings.clear()
            self.custom_hk_settings.clear()
            self.custom_dft_settings.clear()
            self._fit_analysis_cache.clear()
            self._bjh_distribution_cache.clear()
            self._dh_distribution_cache.clear()
            self._hk_distribution_cache.clear()
            self._dft_result_cache.clear()
            self.bjh_pore_volume_range = DEFAULT_BJH_PORE_VOLUME_RANGE
            self.hk_pore_volume_range = DEFAULT_HK_PORE_VOLUME_RANGE
            self.dft_pore_volume_range = DEFAULT_DFT_PORE_VOLUME_RANGE
            self._remove_bjh_region()
            self._remove_dh_region()
            self._remove_hk_region()
            self._remove_dft_region()
            self._isotherm_region_custom = False
            self._last_isotherm_region_range = None
        else:
            self.results.extend(parsed)
            self.visible_results.extend([True] * len(parsed))
            if self.active_index < 0:
                self.active_index = 0
        self.refresh_all()

    def export_xlsx(self) -> None:
        selected = [result for result, visible in zip(self.results, self.visible_results) if visible]
        if not selected:
            QtWidgets.QMessageBox.information(self, "导出文件", "没有勾选可导出的样品。")
            return

        pore_range = self._selected_pore_volume_range()
        d_min, d_max = sorted((float(pore_range[0]), float(pore_range[1])))
        pore_volume_header = f"{self._pore_volume_method_label()} {_fmt_nm(d_min)}nm-{_fmt_nm(d_max)}nm孔容量(cm3/g)"
        pore_volumes = [self._selected_pore_volume_for_result(result) for result in selected]

        default_name = f"BET解析导出_{len(selected)}个样品.xlsx"
        path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "导出文件",
            str(self.export_directory / default_name),
            "Excel 工作簿 (*.xlsx)",
        )
        if not path:
            return
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"
        self.export_directory = Path(path).parent
        self._write_directory_setting("export_directory", self.export_directory)
        try:
            export_results_xlsx(
                selected,
                path,
                pore_volume_header=pore_volume_header,
                pore_volumes=pore_volumes,
            )
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "导出失败", str(exc))
            return
        self.statusBar().showMessage(f"已导出: {path}", 6000)

    def on_sample_item_changed(self, item: QtWidgets.QTableWidgetItem) -> None:
        if self._updating_table or self._updating_sample_checks or item.column() != VISIBLE_COLUMN:
            return
        row = item.row()
        if row >= len(self.visible_results):
            return
        self.visible_results[row] = _check_state_value(item.checkState()) == _check_state_value(QtCore.Qt.Checked)
        self._sync_select_all_state()
        self._refresh_visibility_dependent_ui()

    def on_active_cell_changed(self, current_row: int, current_column: int, previous_row: int, previous_column: int) -> None:
        if self._updating_table:
            return
        if current_row < 0 or current_row >= len(self.results):
            return
        if current_row == self.active_index:
            return
        self.active_index = current_row
        self._reset_all_fit_regions()
        self._load_t_plot_settings_for_active()
        self._load_bjh_settings_for_active()
        self._load_dh_settings_for_active()
        self._load_hk_settings_for_active()
        self._load_dft_settings_for_active()
        self.refresh_isotherm_plot()
        self.refresh_active_views()
        self.refresh_analysis_plots()

    def _select_sample_from_curve(self, row: int) -> None:
        if row < 0 or row >= len(self.results):
            return
        current_column = self.sample_table.currentColumn()
        if current_column < 0:
            current_column = FILE_COLUMN
        self.sample_table.setCurrentCell(int(row), int(current_column))
        self.sample_table.selectRow(int(row))
        try:
            self.sample_table.scrollToItem(self.sample_table.item(int(row), FILE_COLUMN), QtWidgets.QAbstractItemView.PositionAtCenter)
        except Exception:
            pass
        self.statusBar().showMessage(f"已切换到样品：{_display_file_name(self.results[int(row)])}", 2200)

    def on_sample_header_clicked(self, section: int) -> None:
        if len(self.results) < 2:
            return
        if section == TEST_TIME_COLUMN:
            self.test_time_sort_ascending = not self.test_time_sort_ascending
            self.sort_samples_by_test_time(self.test_time_sort_ascending)
        elif section == BET_COLUMN:
            self.bet_sort_ascending = not self.bet_sort_ascending
            self.sort_samples_by_bet(self.bet_sort_ascending)
        elif section == LANGMUIR_COLUMN:
            self.langmuir_sort_ascending = not self.langmuir_sort_ascending
            self.sort_samples_by_langmuir(self.langmuir_sort_ascending)
        elif section == T_PLOT_COLUMN:
            self.t_plot_sort_ascending = not self.t_plot_sort_ascending
            self.sort_samples_by_t_plot(self.t_plot_sort_ascending)
        elif section == BJH_PORE_VOLUME_COLUMN:
            self.bjh_pore_sort_ascending = not self.bjh_pore_sort_ascending
            self.sort_samples_by_bjh_pore_volume(self.bjh_pore_sort_ascending)

    def on_sample_header_resized(self, logical_index: int, old_size: int, new_size: int) -> None:
        self._position_header_controls()
        if self._updating_sample_column_widths or not self._sample_column_widths_initialized:
            return
        self.sample_column_widths[int(logical_index)] = int(new_size)

    def show_sample_context_menu(self, position: QtCore.QPoint) -> None:
        row = self.sample_table.rowAt(position.y())
        if row < 0 or row >= len(self.results):
            return
        self.sample_table.selectRow(row)
        menu = QtWidgets.QMenu(self.sample_table)
        delete_action = menu.addAction("删除")
        global_pos = self.sample_table.viewport().mapToGlobal(position)
        exec_menu = getattr(menu, "exec_", None) or getattr(menu, "exec")
        if exec_menu(global_pos) == delete_action:
            self.delete_sample_row(row)

    def delete_sample_row(self, row: int) -> None:
        if row < 0 or row >= len(self.results):
            return
        deleted = self.results.pop(row)
        self.visible_results.pop(row)
        self._discard_sample_settings(deleted)

        if not self.results:
            self.active_index = -1
            self._reset_all_fit_regions()
            self._isotherm_region_custom = False
            self.refresh_all()
            self.statusBar().showMessage("已删除样品", 3000)
            return

        if row < self.active_index:
            self.active_index -= 1
        elif row == self.active_index:
            self.active_index = min(row, len(self.results) - 1)
        self.refresh_all()
        self.statusBar().showMessage("已删除样品", 3000)

    def sort_samples_by_test_time(self, ascending: bool) -> None:
        active = self.active_result()
        rows = list(zip(self.results, self.visible_results))
        rows.sort(key=lambda row: self._test_time_sort_key(row[0]), reverse=not ascending)
        self.results = [row[0] for row in rows]
        self.visible_results = [row[1] for row in rows]
        self.active_index = 0
        if active is not None:
            for index, result in enumerate(self.results):
                if result is active:
                    self.active_index = index
                    break
        self.refresh_all()

    def sort_samples_by_bet(self, ascending: bool) -> None:
        active = self.active_result()
        rows = list(zip(self.results, self.visible_results))
        rows.sort(key=lambda row: self._bet_sort_key(row[0]), reverse=not ascending)
        self.results = [row[0] for row in rows]
        self.visible_results = [row[1] for row in rows]
        self.active_index = 0
        if active is not None:
            for index, result in enumerate(self.results):
                if result is active:
                    self.active_index = index
                    break
        self.refresh_all()

    def sort_samples_by_langmuir(self, ascending: bool) -> None:
        active = self.active_result()
        rows = list(zip(self.results, self.visible_results))
        rows.sort(key=lambda row: self._langmuir_sort_key(row[0]), reverse=not ascending)
        self.results = [row[0] for row in rows]
        self.visible_results = [row[1] for row in rows]
        self.active_index = 0
        if active is not None:
            for index, result in enumerate(self.results):
                if result is active:
                    self.active_index = index
                    break
        self.refresh_all()

    def sort_samples_by_t_plot(self, ascending: bool) -> None:
        active = self.active_result()
        rows = list(zip(self.results, self.visible_results))
        rows.sort(key=lambda row: self._t_plot_sort_key(row[0]), reverse=not ascending)
        self.results = [row[0] for row in rows]
        self.visible_results = [row[1] for row in rows]
        self.active_index = 0
        if active is not None:
            for index, result in enumerate(self.results):
                if result is active:
                    self.active_index = index
                    break
        self.refresh_all()

    def sort_samples_by_bjh_pore_volume(self, ascending: bool) -> None:
        active = self.active_result()
        rows = list(zip(self.results, self.visible_results))
        rows.sort(key=lambda row: self._bjh_pore_volume_sort_key(row[0]), reverse=not ascending)
        self.results = [row[0] for row in rows]
        self.visible_results = [row[1] for row in rows]
        self.active_index = 0
        if active is not None:
            for index, result in enumerate(self.results):
                if result is active:
                    self.active_index = index
                    break
        self.refresh_all()

    def move_sample_row(self, source_index: int, insert_index: int) -> None:
        if len(self.results) < 2 or not (0 <= source_index < len(self.results)):
            return

        insert_index = max(0, min(int(insert_index), len(self.results)))
        if insert_index in (source_index, source_index + 1):
            return

        active = self.active_result()
        moved_result = self.results.pop(source_index)
        moved_visible = self.visible_results.pop(source_index)

        if insert_index > source_index:
            insert_index -= 1

        self.results.insert(insert_index, moved_result)
        self.visible_results.insert(insert_index, moved_visible)

        self.active_index = insert_index
        if active is not None:
            for index, result in enumerate(self.results):
                if result is active:
                    self.active_index = index
                    break
        self.refresh_all()

    def refresh_all(self) -> None:
        self._load_t_plot_settings_for_active()
        self._load_bjh_settings_for_active()
        self._load_dh_settings_for_active()
        self._load_hk_settings_for_active()
        self._load_dft_settings_for_active()
        self.refresh_isotherm_plot()
        self.refresh_sample_table()
        self.refresh_active_views()
        self.refresh_analysis_plots()

    def refresh_sample_table(self) -> None:
        horizontal_scroll_bar = self.sample_table.horizontalScrollBar()
        horizontal_scroll_value = horizontal_scroll_bar.value()
        self._updating_table = True
        try:
            self.sample_table.setRowCount(len(self.results))
            self.sample_items = []
            for row, result in enumerate(self.results):
                bet = self._bet_analysis_for_result(result)
                langmuir = self._langmuir_analysis_for_result(result)
                t_plot = self._t_plot_analysis_for_result(result)
                visible_item = QtWidgets.QTableWidgetItem()
                visible_item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsUserCheckable)
                visible_item.setCheckState(QtCore.Qt.Checked if self.visible_results[row] else QtCore.Qt.Unchecked)
                visible_item.setTextAlignment(QtCore.Qt.AlignCenter)
                self.sample_table.setItem(row, VISIBLE_COLUMN, visible_item)
                self.sample_items.append(visible_item)

                file_item = self._table_item(_display_file_name(result), tooltip=result.header.file_path)
                self.sample_table.setItem(row, FILE_COLUMN, file_item)

                test_time_item = self._table_item(result.test_started_time)
                test_time_item.setToolTip("SMP 优先来自日志 Started 时间；DAT/QPS 来自测量日期")
                self.sample_table.setItem(row, TEST_TIME_COLUMN, test_time_item)

                bet_item = self._table_item(_fmt(bet.surface_area_m2_g), alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
                self._style_sample_bet_item(bet_item, result)
                self.sample_table.setItem(row, BET_COLUMN, bet_item)

                langmuir_item = self._table_item(
                    _fmt(langmuir.surface_area_m2_g),
                    alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter,
                )
                self._style_sample_langmuir_item(langmuir_item, result)
                self.sample_table.setItem(row, LANGMUIR_COLUMN, langmuir_item)

                t_plot_item = self._table_item(
                    _fmt(t_plot.external_surface_area_m2_g),
                    alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter,
                )
                self._style_sample_t_plot_item(t_plot_item, result)
                self.sample_table.setItem(row, T_PLOT_COLUMN, t_plot_item)

                bjh_volume_item = self._table_item(
                    _fmt(self._selected_pore_volume_for_result(result)),
                    alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter,
                )
                self._style_sample_bjh_pore_item(bjh_volume_item, result)
                self.sample_table.setItem(row, BJH_PORE_VOLUME_COLUMN, bjh_volume_item)
            if self.results and self.active_index >= 0:
                self.sample_table.selectRow(min(self.active_index, len(self.results) - 1))
            self._sync_select_all_state()
            self._resize_sample_columns()
            horizontal_scroll_bar.setValue(
                max(horizontal_scroll_bar.minimum(), min(horizontal_scroll_value, horizontal_scroll_bar.maximum()))
            )
            self._position_header_controls()
        finally:
            self._updating_table = False

    def _refresh_sample_bet_cell(self, row: int) -> None:
        if row < 0 or row >= len(self.results):
            return
        result = self.results[row]
        bet = self._bet_analysis_for_result(result)
        bet_item = self._table_item(_fmt(bet.surface_area_m2_g), alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter)
        self._style_sample_bet_item(bet_item, result)
        self.sample_table.setItem(row, BET_COLUMN, bet_item)

    def _refresh_sample_langmuir_cell(self, row: int) -> None:
        if row < 0 or row >= len(self.results):
            return
        result = self.results[row]
        langmuir = self._langmuir_analysis_for_result(result)
        langmuir_item = self._table_item(
            _fmt(langmuir.surface_area_m2_g),
            alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter,
        )
        self._style_sample_langmuir_item(langmuir_item, result)
        self.sample_table.setItem(row, LANGMUIR_COLUMN, langmuir_item)

    def _refresh_sample_t_plot_cell(self, row: int) -> None:
        if row < 0 or row >= len(self.results):
            return
        result = self.results[row]
        t_plot = self._t_plot_analysis_for_result(result)
        t_plot_item = self._table_item(
            _fmt(t_plot.external_surface_area_m2_g),
            alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter,
        )
        self._style_sample_t_plot_item(t_plot_item, result)
        self.sample_table.setItem(row, T_PLOT_COLUMN, t_plot_item)

    def _refresh_sample_bjh_pore_cell(self, row: int) -> None:
        if row < 0 or row >= len(self.results):
            return
        result = self.results[row]
        item = self._table_item(
            _fmt(self._selected_pore_volume_for_result(result)),
            alignment=QtCore.Qt.AlignRight | QtCore.Qt.AlignVCenter,
        )
        self._style_sample_bjh_pore_item(item, result)
        self.sample_table.setItem(row, BJH_PORE_VOLUME_COLUMN, item)

    def _refresh_all_sample_bjh_pore_cells(self) -> None:
        for row in range(len(self.results)):
            self._refresh_sample_bjh_pore_cell(row)

    def _style_sample_bet_item(self, item: QtWidgets.QTableWidgetItem, result) -> None:
        if not self._is_custom_bet_fit(result):
            return
        item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
        item.setToolTip("BET 拟合区间已人工调整")

    def _style_sample_langmuir_item(self, item: QtWidgets.QTableWidgetItem, result) -> None:
        if not self._is_custom_langmuir_fit(result):
            return
        item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
        item.setToolTip("Langmuir 拟合区间已人工调整")

    def _style_sample_t_plot_item(self, item: QtWidgets.QTableWidgetItem, result) -> None:
        if not self._is_custom_t_plot_fit(result):
            return
        item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
        item.setToolTip("t-Plot 厚度曲线、表面积参数或拟合厚度区间已人工调整")

    def _style_sample_bjh_pore_item(self, item: QtWidgets.QTableWidgetItem, result) -> None:
        label = self._pore_volume_method_label()
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT and self._has_custom_dft_settings(result):
            item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
            item.setToolTip("DFT 模型、结构或正则化参数已人工调整")
            return
        if self._is_hk_tab_active() and self._has_custom_hk_settings(result):
            item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
            item.setToolTip("HK 孔型、物性、作用参数或平滑选项已人工调整")
            return
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DH and self._has_custom_dh_settings(result):
            item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
            item.setToolTip("DH 厚度曲线、公式参数、平滑或显示分支已人工调整")
            return
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_BJH and self._has_custom_bjh_settings(result):
            item.setForeground(QtGui.QBrush(QtGui.QColor(CUSTOM_BET_COLOR)))
            item.setToolTip("BJH 厚度曲线、公式参数、校正或显示分支已人工调整")
        else:
            item.setToolTip(f"{label} 当前绿色选区孔容量")

    def refresh_active_views(self) -> None:
        self.refresh_metrics()
        self.refresh_condition_table()
        self.refresh_isotherm_table()
        self.refresh_target_table()
        self.refresh_report_module_table()
        self.refresh_log_table()

    def refresh_plots(self) -> None:
        self.refresh_isotherm_plot()
        self.refresh_analysis_plots()

    def on_plot_tab_changed(self, _index: int) -> None:
        self.refresh_isotherm_plot()
        self._update_dft_diagnostic_visibility()
        self._update_pore_volume_header()
        self._refresh_all_sample_bjh_pore_cells()
        self.refresh_analysis_plots()

    def _update_dft_diagnostic_visibility(self) -> None:
        plot = getattr(self, "dft_diagnostic_plot", None)
        if plot is not None:
            plot.setVisible(self._is_dft_tab_active())

    def refresh_isotherm_plot(self) -> None:
        raw_region = self._current_pressure_region() if self._isotherm_region_custom else None
        pressure = self._all_pressure_values()
        selected_range = None
        if pressure.size:
            selected_range = self._clamp_pressure_region(raw_region or self._default_isotherm_region(pressure), pressure)
        self._remove_region()
        self._remove_isotherm_selection()
        dft_active = self._is_dft_tab_active()
        dft_fit_rows = self._active_dft_fit_rows() if dft_active else None
        plot_isotherm_multi(
            self.isotherm_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            fade_inactive=dft_active,
            active_fit_rows=dft_fit_rows,
            x_log=dft_active,
            hollow_base_points=not dft_active,
        )
        if selected_range is not None and not dft_active:
            self._add_region(selected_range, pressure)
            self._refresh_isotherm_selection(selected_range)

    def refresh_analysis_plots(self) -> None:
        self._refresh_current_analysis_plot(reset_region=True)

    def _refresh_current_analysis_plot(self, *, reset_region: bool) -> None:
        active = self.active_result()
        if active is None:
            self._clear_analysis_plots()
            return
        pressure_range = self._current_pressure_region()
        p_min, p_max = pressure_range if pressure_range else (None, None)
        current_tab = self.plot_tabs.currentWidget() if getattr(self, "plot_tabs", None) is not None else None
        if current_tab is self.langmuir_tab:
            self._refresh_langmuir_plot(active, p_min, p_max, reset_region=reset_region)
        elif current_tab is self.t_plot_tab:
            self._refresh_t_plot_plot(active, p_min, p_max, reset_region=reset_region)
        elif current_tab is self.bjh_tab:
            self.refresh_bjh_plot()
        elif current_tab is self.dh_tab:
            self.refresh_dh_plot()
        elif current_tab is self.hk_tab:
            self.refresh_hk_plot()
        elif current_tab is self.dft_tab:
            self.refresh_dft_plot()
        else:
            self._refresh_bet_plot(active, p_min, p_max, reset_region=reset_region)

    def refresh_bjh_plot(self) -> None:
        target_diameter_range = self.bjh_pore_volume_range
        self._remove_bjh_region()
        self._remove_bjh_selection()
        self._bjh_distribution_rows_by_key = {}
        self._bjh_diameter_log_bounds = None
        if not self.results:
            plot_pore_distribution_placeholder(
                self.pore_plot,
                differential_mode=self.bjh_differential_mode,
                display_metrics=self.bjh_display_metrics,
            )
            return
        pressure_range = self._bjh_pressure_range()
        bjh_settings_by_index = {
            index: self._bjh_settings_for_result(result)
            for index, result in enumerate(self.results)
        }
        self._bjh_distribution_rows_by_key = plot_bjh_distribution_multi(
            self.pore_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            thickness_method=self.bjh_thickness_method,
            thickness_params=self.bjh_thickness_params,
            correction=self.bjh_correction,
            open_pore_fraction=self.bjh_open_pore_fraction,
            show_adsorption=self.bjh_show_adsorption,
            show_desorption=self.bjh_show_desorption,
            smooth=self.bjh_smooth_derivative,
            pressure_range=pressure_range,
            bjh_settings_by_index=bjh_settings_by_index,
            distribution_provider=self._cached_bjh_distribution_rows,
            differential_mode=self.bjh_differential_mode,
            display_metrics=self.bjh_display_metrics,
        )
        self._bjh_diameter_log_bounds = self._bjh_log_bounds_from_rows(self._bjh_distribution_rows_by_key)
        if self._is_bjh_default_region_active():
            if pressure_range is not None:
                self._set_default_isotherm_region(pressure_range)
        self._sync_bjh_region_to_diameter_range(target_diameter_range)
        current_range = self._current_bjh_diameter_range()
        if current_range is not None:
            self.bjh_pore_volume_range = tuple(sorted((float(target_diameter_range[0]), float(target_diameter_range[1]))))
            self._refresh_bjh_selection(self.bjh_pore_volume_range)
            self._refresh_all_sample_bjh_pore_cells()

    def refresh_dh_plot(self) -> None:
        target_diameter_range = self.bjh_pore_volume_range
        self._remove_dh_region()
        self._remove_dh_selection()
        self._dh_distribution_rows_by_key = {}
        self._dh_diameter_log_bounds = None
        if not self.results:
            plot_dh_distribution_placeholder(
                self.dh_plot,
                display_metrics=self.dh_display_metrics,
            )
            return
        pressure_range = self._dh_pressure_range()
        dh_settings_by_index = {
            index: self._dh_settings_for_result(result)
            for index, result in enumerate(self.results)
        }
        self._dh_distribution_rows_by_key = plot_dh_distribution_multi(
            self.dh_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            thickness_method=self.dh_thickness_method,
            thickness_params=self.dh_thickness_params,
            show_adsorption=self.dh_show_adsorption,
            show_desorption=self.dh_show_desorption,
            smooth=self.dh_smooth_derivative,
            pressure_range=pressure_range,
            dh_settings_by_index=dh_settings_by_index,
            distribution_provider=self._cached_dh_distribution_rows,
            display_metrics=self.dh_display_metrics,
        )
        self._dh_diameter_log_bounds = self._bjh_log_bounds_from_rows(self._dh_distribution_rows_by_key)
        if self._is_dh_default_region_active():
            if pressure_range is not None:
                self._set_default_isotherm_region(pressure_range)
        self._sync_dh_region_to_diameter_range(target_diameter_range)
        current_range = self._current_dh_diameter_range()
        if current_range is not None:
            self.bjh_pore_volume_range = tuple(sorted((float(target_diameter_range[0]), float(target_diameter_range[1]))))
            self._refresh_dh_selection(self.bjh_pore_volume_range)
            self._refresh_all_sample_bjh_pore_cells()

    def refresh_hk_plot(self) -> None:
        target_width_range = self.hk_pore_volume_range
        self._remove_hk_region()
        self._remove_hk_selection()
        self._hk_distribution_rows_by_key = {}
        self._hk_width_log_bounds = None
        if not self.results:
            plot_hk_distribution_placeholder(
                self.hk_plot,
                display_metric=self.hk_display_metric,
            )
            return
        hk_settings_by_index = {
            index: self._hk_settings_for_result(result)
            for index, result in enumerate(self.results)
        }
        pressure_range = self._hk_pressure_range()
        self._hk_distribution_rows_by_key = plot_hk_distribution_multi(
            self.hk_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            geometry=self.hk_geometry,
            adsorbent_key=self.hk_adsorbent_key,
            adsorptive_key=self.hk_adsorptive_key,
            adsorbent_properties=self.hk_adsorbent_properties,
            adsorptive_properties=self.hk_adsorptive_properties,
            interaction_parameter_erg_cm4=self.hk_interaction_parameter,
            interaction_parameter_mode=self.hk_interaction_parameter_mode,
            cheng_yang_correction=self.hk_cheng_yang_correction,
            smooth=self.hk_smooth_derivative,
            pressure_range=pressure_range,
            hk_settings_by_index=hk_settings_by_index,
            distribution_provider=self._cached_hk_distribution_rows,
            display_metric=self.hk_display_metric,
        )
        self._hk_width_log_bounds = self._hk_log_bounds_from_rows(self._hk_distribution_rows_by_key)
        if self._is_hk_default_region_active() and pressure_range is not None:
            self._set_default_isotherm_region(pressure_range)
        self._sync_hk_region_to_width_range(target_width_range)
        current_range = self._current_hk_width_range()
        if current_range is not None:
            self.hk_pore_volume_range = tuple(sorted((float(target_width_range[0]), float(target_width_range[1]))))
            self._refresh_hk_selection(self.hk_pore_volume_range)
        self._refresh_all_sample_bjh_pore_cells()

    def refresh_dft_plot(self, *, refresh_diagnostics: bool = True, refresh_pore_cells: bool = True) -> None:
        target_width_range = self.dft_pore_volume_range
        self._remove_dft_region()
        self._remove_dft_selection()
        self._dft_distribution_rows_by_index = {}
        self._dft_width_log_bounds = None
        if not self.results:
            self.dft_plot.clear()
            if hasattr(self, "dft_diagnostic_plot"):
                self.dft_diagnostic_plot.clear()
            return
        dft_settings_by_index = {
            index: self._dft_settings_for_result(result)
            for index, result in enumerate(self.results)
        }
        self._dft_distribution_rows_by_index = plot_dft_distribution_multi(
            self.dft_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            analysis_type=self.dft_analysis_type,
            geometry=self.dft_geometry,
            model=self.dft_model,
            regularization=self.dft_regularization,
            dft_settings_by_index=dft_settings_by_index,
            result_provider=self._cached_dft_result,
        )
        self._dft_width_log_bounds = self._dft_log_bounds_from_rows(self._dft_distribution_rows_by_index)
        self._sync_dft_region_to_width_range(target_width_range)
        current_range = self._current_dft_width_range()
        if current_range is not None:
            self.dft_pore_volume_range = tuple(sorted((float(target_width_range[0]), float(target_width_range[1]))))
            self._refresh_dft_selection(self.dft_pore_volume_range)

        if refresh_diagnostics:
            self._refresh_dft_diagnostics()
        if refresh_pore_cells and self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT:
            self._refresh_all_sample_bjh_pore_cells()

    def _refresh_dft_diagnostics(self) -> None:
        plot = getattr(self, "dft_diagnostic_plot", None)
        if plot is None:
            return
        active = self.active_result()
        self._dft_diagnostic_line = None
        if active is None:
            plot.clear()
            return
        settings = self._dft_settings_for_result(active)
        dft_result = self._cached_dft_result(
            active,
            analysis_type=str(settings["analysis_type"]),
            geometry=str(settings["geometry"]),
            model=str(settings["model"]),
            regularization=float(settings["regularization"]),
            include_diagnostics=True,
        )
        self._dft_diagnostic_line = plot_dft_diagnostics(
            plot,
            list(getattr(dft_result, "diagnostic_rows", [])),
            float(settings["regularization"]),
        )
        if self._dft_diagnostic_line is not None:
            self._dft_diagnostic_line.sigPositionChangeFinished.connect(self._on_dft_diagnostic_line_changed)

    def _bjh_pressure_range(self) -> tuple[float, float] | None:
        if not self._isotherm_region_custom:
            return self._default_bjh_pressure_range()
        return self._current_pressure_region()

    def _is_bjh_default_region_active(self) -> bool:
        return self._is_bjh_tab_active() and not self._isotherm_region_custom

    def _dh_pressure_range(self) -> tuple[float, float] | None:
        if not self._isotherm_region_custom:
            return self._default_dh_pressure_range()
        return self._current_pressure_region()

    def _is_dh_default_region_active(self) -> bool:
        return self._is_dh_tab_active() and not self._isotherm_region_custom

    def _hk_pressure_range(self) -> tuple[float, float] | None:
        if not self._isotherm_region_custom:
            pressure = self._all_pressure_values()
            if pressure.size == 0:
                return None
            return tuple(self._clamp_pressure_region(HK_DEFAULT_PRESSURE_RANGE, pressure))
        return self._current_pressure_region()

    def _is_hk_default_region_active(self) -> bool:
        return self._is_hk_tab_active() and not self._isotherm_region_custom

    def _default_bjh_pressure_range(self) -> tuple[float, float] | None:
        pressure = self._all_pressure_values()
        if pressure.size == 0:
            return None
        data_min = float(np.nanmin(pressure))
        data_max = float(np.nanmax(pressure))
        start_pressure = self._pressure_for_bjh_diameter(DEFAULT_BJH_PORE_VOLUME_RANGE[0])
        if start_pressure is None or not np.isfinite(start_pressure):
            return self._full_pressure_region(pressure)
        start_pressure = max(data_min, min(float(start_pressure), data_max))
        if start_pressure >= data_max:
            return self._full_pressure_region(pressure)
        return (start_pressure, data_max)

    def _default_dh_pressure_range(self) -> tuple[float, float] | None:
        pressure = self._all_pressure_values()
        if pressure.size == 0:
            return None
        data_min = float(np.nanmin(pressure))
        data_max = float(np.nanmax(pressure))
        start_pressure = self._pressure_for_dh_diameter(DEFAULT_BJH_PORE_VOLUME_RANGE[0])
        if start_pressure is None or not np.isfinite(start_pressure):
            return self._full_pressure_region(pressure)
        start_pressure = max(data_min, min(float(start_pressure), data_max))
        if start_pressure >= data_max:
            return self._full_pressure_region(pressure)
        return (start_pressure, data_max)

    def _update_analysis_plots_for_region(self) -> None:
        """等温线选区变化时调用：刷新三个分析图但保留各自的拟合选区。"""
        self._refresh_current_analysis_plot(reset_region=False)

    def _visible_analysis_indices(self) -> list[int]:
        draw_order = [i for i in range(len(self.results)) if i != self.active_index]
        if 0 <= self.active_index < len(self.results):
            draw_order.append(self.active_index)
        return [i for i in draw_order if i < len(self.visible_results) and self.visible_results[i]]

    def _analysis_sample_color(self, index: int) -> str:
        return self.sample_colors[index % len(self.sample_colors)] if self.sample_colors else DEFAULT_COLORS[0]

    def _refresh_bet_plot(self, active, p_min=None, p_max=None, *, reset_region: bool = False) -> None:
        raw_region = None
        if not reset_region and self.bet_region is not None:
            try:
                raw_region = list(self.bet_region.getRegion())
            except RuntimeError:
                pass
        self._remove_bet_region()
        self._remove_bet_selection()
        self._bet_fit_line = None
        self._bet_x_range = None
        data_p_min = p_min if p_min is not None else BET_PLOT_RANGE[0]
        data_p_max = p_max if p_max is not None else BET_PLOT_RANGE[1]
        self._bet_plot_p_range = (data_p_min, data_p_max)
        x_by_index = plot_bet_multi(
            self.bet_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            p_min=data_p_min,
            p_max=data_p_max,
            analysis_provider=self._cached_bet_analysis,
        )
        for index in self._visible_analysis_indices():
            result = self.results[index]
            x_values = x_by_index.get(index)
            if x_values is None or x_values.size < 2:
                continue
            x_min = float(np.nanmin(x_values))
            x_max = float(np.nanmax(x_values))
            if x_min >= x_max:
                continue
            is_active = result is active
            target_region = (
                raw_region
                if is_active and raw_region and not reset_region
                else self._bet_fit_range_for_result(result)
            )
            lo, hi = self._clamp_fit_region(target_region, x_min, x_max, False)
            item, _ = replace_bet_fit_line(
                self.bet_plot,
                None,
                result,
                lo,
                hi,
                line_x_min=x_min,
                line_x_max=x_max,
                color=self._analysis_sample_color(index),
                name="线性拟合" if is_active else None,
                width=ACTIVE_LINE_WIDTH if is_active else 1,
                analysis_provider=self._cached_bet_analysis,
            )
            if not is_active:
                continue
            self._bet_x_range = (x_min, x_max)
            self._bet_fit_line = item
            was_setting_bet_region = self._setting_bet_region
            self._setting_bet_region = True
            try:
                self._add_bet_region([lo, hi], [x_min, x_max])
            finally:
                self._setting_bet_region = was_setting_bet_region
            self._refresh_bet_selection(active, (lo, hi), data_p_min, data_p_max)

    def reset_bet_fit_to_default(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self._clear_custom_bet_fit_range(active)
        pressure_range = self._current_pressure_region()
        p_min, p_max = pressure_range if pressure_range else (None, None)
        self._setting_bet_region = True
        try:
            self._refresh_bet_plot(active, p_min, p_max, reset_region=True)
        finally:
            self._setting_bet_region = False
        self.refresh_sample_table()
        self.refresh_metrics()

    def _refresh_langmuir_plot(self, active, p_min, p_max, *, reset_region: bool = False) -> None:
        raw_region = None
        if not reset_region and self.langmuir_region is not None:
            try:
                raw_region = list(self.langmuir_region.getRegion())
            except RuntimeError:
                pass
        self._remove_langmuir_region()
        self._remove_langmuir_selection()
        self._langmuir_fit_line = None
        self._langmuir_x_range = None
        data_p_min = p_min if p_min is not None else LANGMUIR_PLOT_RANGE[0]
        data_p_max = p_max if p_max is not None else LANGMUIR_PLOT_RANGE[1]
        self._langmuir_plot_p_range = (data_p_min, data_p_max)
        x_by_index = plot_langmuir_points_multi(
            self.langmuir_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            p_min=data_p_min,
            p_max=data_p_max,
            analysis_provider=self._cached_langmuir_analysis,
        )
        for index in self._visible_analysis_indices():
            result = self.results[index]
            x_values = x_by_index.get(index)
            if x_values is None or x_values.size < 2:
                continue
            x_min = float(np.nanmin(x_values))
            x_max = float(np.nanmax(x_values))
            if x_min >= x_max:
                continue
            is_active = result is active
            target_region = (
                raw_region
                if is_active and raw_region and not reset_region
                else self._langmuir_fit_range_for_result(result)
            )
            lo, hi = self._clamp_fit_region(target_region, x_min, x_max, False)
            item, _ = replace_langmuir_fit_line(
                self.langmuir_plot,
                None,
                result,
                lo,
                hi,
                line_x_min=x_min,
                line_x_max=x_max,
                color=self._analysis_sample_color(index),
                name="线性拟合" if is_active else None,
                width=ACTIVE_LINE_WIDTH if is_active else 1,
                analysis_provider=self._cached_langmuir_analysis,
            )
            if not is_active:
                continue
            self._langmuir_x_range = (x_min, x_max)
            self._langmuir_fit_line = item
            was_setting_langmuir_region = self._setting_langmuir_region
            self._setting_langmuir_region = True
            try:
                self._add_langmuir_region([lo, hi], [x_min, x_max])
            finally:
                self._setting_langmuir_region = was_setting_langmuir_region
            self._refresh_langmuir_selection(active, (lo, hi), data_p_min, data_p_max)

    def reset_langmuir_fit_to_default(self) -> None:
        active = self.active_result()
        if active is None:
            return
        self._clear_custom_langmuir_fit_range(active)
        pressure_range = self._current_pressure_region()
        p_min, p_max = pressure_range if pressure_range else (None, None)
        self._setting_langmuir_region = True
        try:
            self._refresh_langmuir_plot(active, p_min, p_max, reset_region=True)
        finally:
            self._setting_langmuir_region = False
        self.refresh_sample_table()
        self.refresh_metrics()

    def _refresh_t_plot_plot(self, active, p_min, p_max, *, reset_region: bool = False) -> None:
        raw_region = None
        if not reset_region and self.t_plot_region is not None:
            try:
                raw_region = list(self.t_plot_region.getRegion())
            except RuntimeError:
                pass
        self._remove_t_plot_region()
        self._remove_t_plot_selection()
        self._t_plot_fit_line = None
        self._t_plot_x_range = None
        data_p_min = p_min if p_min is not None else T_PLOT_PLOT_RANGE[0]
        data_p_max = p_max if p_max is not None else T_PLOT_PLOT_RANGE[1]
        self._t_plot_p_range = (data_p_min, data_p_max)
        thickness_params_by_index = {
            index: dict(self._t_plot_settings_for_result(result)["thickness_params"])
            for index, result in enumerate(self.results)
        }
        thickness_method_by_index = {
            index: str(self._t_plot_settings_for_result(result)["thickness_method"])
            for index, result in enumerate(self.results)
        }
        x_by_index = plot_t_plot_points_multi(
            self.t_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            active_index=self.active_index,
            p_min=data_p_min,
            p_max=data_p_max,
            thickness_params_by_index=thickness_params_by_index,
            thickness_method_by_index=thickness_method_by_index,
            analysis_provider=self._cached_t_plot_pressure_analysis,
        )
        for index in self._visible_analysis_indices():
            result = self.results[index]
            x_values = x_by_index.get(index)
            if x_values is None or x_values.size < 2:
                continue
            x_min = float(np.nanmin(x_values))
            x_max = float(np.nanmax(x_values))
            if x_min >= x_max:
                continue
            is_active = result is active
            target_region = (
                raw_region
                if is_active and raw_region and not reset_region
                else self._t_plot_fit_range_for_result(result)
            )
            lo, hi = self._clamp_fit_region(target_region, x_min, x_max, False)
            thickness_params = thickness_params_by_index.get(index, self.t_plot_thickness_params)
            thickness_method = thickness_method_by_index.get(index, self.t_plot_thickness_method)
            item, _ = replace_t_plot_fit_line(
                self.t_plot,
                None,
                result,
                lo,
                hi,
                line_x_min=x_min,
                line_x_max=x_max,
                data_p_min=data_p_min,
                data_p_max=data_p_max,
                thickness_params=thickness_params,
                thickness_method=thickness_method,
                color=self._analysis_sample_color(index),
                name="线性拟合" if is_active else None,
                width=ACTIVE_LINE_WIDTH if is_active else 1,
                analysis_provider=self._cached_t_plot_thickness_analysis,
            )
            if not is_active:
                continue
            self._t_plot_x_range = (x_min, x_max)
            self._t_plot_fit_line = item
            was_setting_t_plot_region = self._setting_t_plot_region
            self._setting_t_plot_region = True
            try:
                self._add_t_plot_region([lo, hi], [x_min, x_max])
            finally:
                self._setting_t_plot_region = was_setting_t_plot_region
            self._refresh_t_plot_selection(active, (lo, hi), data_p_min, data_p_max)

    def reset_t_plot_fit_to_default(self) -> None:
        active = self.active_result()
        if active is None:
            self._load_t_plot_settings_for_active()
            self.refresh_metrics()
            return
        self.custom_t_plot_settings.pop(id(active), None)
        self._clear_custom_t_plot_fit_range(active)
        self._load_t_plot_settings_for_active()
        pressure_range = self._current_pressure_region()
        p_min, p_max = pressure_range if pressure_range else (None, None)
        self._setting_t_plot_region = True
        try:
            self._refresh_t_plot_plot(active, p_min, p_max, reset_region=True)
        finally:
            self._setting_t_plot_region = False
        self._refresh_sample_t_plot_cell(self.active_index)
        self.refresh_metrics()

    @staticmethod
    def _clamp_fit_region(raw_region, x_min: float, x_max: float, reset: bool) -> tuple[float, float]:
        if raw_region and not reset:
            lo = max(x_min, min(float(raw_region[0]), x_max))
            hi = max(x_min, min(float(raw_region[1]), x_max))
            if lo < hi - 1e-10:
                return lo, hi
        return x_min, x_max

    def _remove_bet_region(self) -> None:
        if self.bet_region is None:
            return
        try:
            self.bet_region.sigRegionChanged.disconnect(self.on_bet_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.bet_plot.removeItem(self.bet_region)
        except RuntimeError:
            pass
        self.bet_region = None

    def _add_bet_region(self, values: list, bounds: list) -> None:
        region = self._make_selection_region(values, bounds=bounds, movable=True)
        region.sigRegionChanged.connect(self.on_bet_region_changed)
        self.bet_plot.addItem(region, ignoreBounds=True)
        self.bet_region = region

    def on_bet_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_bet_region:
            return
        if not self._bet_region_pending:
            self._bet_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_bet_from_region)

    def _update_bet_from_region(self) -> None:
        self._bet_region_pending = False
        active = self.active_result()
        if active is None:
            return
        bet_fit_range = self._current_bet_fit_range()
        if bet_fit_range is None:
            return
        self._set_custom_bet_fit_range(active, bet_fit_range)
        lx_min = self._bet_x_range[0] if self._bet_x_range else None
        lx_max = self._bet_x_range[1] if self._bet_x_range else None
        self._bet_fit_line, _ = replace_bet_fit_line(
            self.bet_plot, self._bet_fit_line, active,
            bet_fit_range[0], bet_fit_range[1],
            line_x_min=lx_min, line_x_max=lx_max,
            color=self._analysis_sample_color(self.active_index),
            width=ACTIVE_LINE_WIDTH,
            analysis_provider=self._cached_bet_analysis,
        )
        data_p_min, data_p_max = self._bet_plot_p_range or BET_PLOT_RANGE
        self._refresh_bet_selection(active, bet_fit_range, data_p_min, data_p_max)
        self._refresh_sample_bet_cell(self.active_index)
        self.refresh_metrics()

    def _current_bet_fit_range(self) -> tuple[float, float] | None:
        if self.bet_region is None:
            return None
        try:
            lo, hi = self.bet_region.getRegion()
            return (min(float(lo), float(hi)), max(float(lo), float(hi)))
        except RuntimeError:
            return None

    # ── Langmuir region ──────────────────────────────────────────────────────

    def _remove_bet_selection(self) -> None:
        if self._bet_selection_item is None:
            return
        try:
            self.bet_plot.removeItem(self._bet_selection_item)
        except RuntimeError:
            pass
        self._bet_selection_item = None

    def _refresh_bet_selection(self, active, fit_range, data_p_min=None, data_p_max=None) -> None:
        self._remove_bet_selection()
        if fit_range is None:
            return
        self._bet_selection_item = plot_bet_selection(
            self.bet_plot,
            active,
            fit_range[0],
            fit_range[1],
            data_p_min=data_p_min,
            data_p_max=data_p_max,
            color=self._analysis_sample_color(self.active_index),
            analysis_provider=self._cached_bet_analysis,
        )

    def _remove_langmuir_selection(self) -> None:
        if self._langmuir_selection_item is None:
            return
        try:
            self.langmuir_plot.removeItem(self._langmuir_selection_item)
        except RuntimeError:
            pass
        self._langmuir_selection_item = None

    def _refresh_langmuir_selection(self, active, fit_range, data_p_min=None, data_p_max=None) -> None:
        self._remove_langmuir_selection()
        if fit_range is None:
            return
        self._langmuir_selection_item = plot_langmuir_selection(
            self.langmuir_plot,
            active,
            fit_range[0],
            fit_range[1],
            data_p_min=data_p_min,
            data_p_max=data_p_max,
            color=self._analysis_sample_color(self.active_index),
            analysis_provider=self._cached_langmuir_analysis,
        )

    def _remove_langmuir_region(self) -> None:
        if self.langmuir_region is None:
            return
        try:
            self.langmuir_region.sigRegionChanged.disconnect(self.on_langmuir_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.langmuir_plot.removeItem(self.langmuir_region)
        except RuntimeError:
            pass
        self.langmuir_region = None

    def _add_langmuir_region(self, values: list, bounds: list) -> None:
        region = self._make_selection_region(values, bounds=bounds, movable=True)
        region.sigRegionChanged.connect(self.on_langmuir_region_changed)
        self.langmuir_plot.addItem(region, ignoreBounds=True)
        self.langmuir_region = region

    def on_langmuir_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_langmuir_region:
            return
        if not self._langmuir_region_pending:
            self._langmuir_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_langmuir_from_region)

    def _update_langmuir_from_region(self) -> None:
        self._langmuir_region_pending = False
        active = self.active_result()
        if active is None:
            return
        fit_range = self._current_langmuir_fit_range()
        if fit_range is None:
            return
        self._set_custom_langmuir_fit_range(active, fit_range)
        lx_min = self._langmuir_x_range[0] if self._langmuir_x_range else None
        lx_max = self._langmuir_x_range[1] if self._langmuir_x_range else None
        self._langmuir_fit_line, _ = replace_langmuir_fit_line(
            self.langmuir_plot, self._langmuir_fit_line, active,
            fit_range[0], fit_range[1],
            line_x_min=lx_min, line_x_max=lx_max,
            color=self._analysis_sample_color(self.active_index),
            width=ACTIVE_LINE_WIDTH,
            analysis_provider=self._cached_langmuir_analysis,
        )
        data_p_min, data_p_max = self._langmuir_plot_p_range or LANGMUIR_PLOT_RANGE
        self._refresh_langmuir_selection(active, fit_range, data_p_min, data_p_max)
        self._refresh_sample_langmuir_cell(self.active_index)
        self.refresh_metrics()

    def _current_langmuir_fit_range(self) -> tuple[float, float] | None:
        if self.langmuir_region is None:
            return None
        try:
            lo, hi = self.langmuir_region.getRegion()
            return (min(float(lo), float(hi)), max(float(lo), float(hi)))
        except RuntimeError:
            return None

    # ── t-Plot region ─────────────────────────────────────────────────────────

    def _remove_t_plot_selection(self) -> None:
        if self._t_plot_selection_item is None:
            return
        try:
            self.t_plot.removeItem(self._t_plot_selection_item)
        except RuntimeError:
            pass
        self._t_plot_selection_item = None

    def _refresh_t_plot_selection(self, active, fit_range, data_p_min=None, data_p_max=None) -> None:
        self._remove_t_plot_selection()
        if fit_range is None:
            return
        self._t_plot_selection_item = plot_t_plot_selection(
            self.t_plot,
            active,
            fit_range[0],
            fit_range[1],
            data_p_min=data_p_min,
            data_p_max=data_p_max,
            thickness_params=self.t_plot_thickness_params,
            thickness_method=self.t_plot_thickness_method,
            color=self._analysis_sample_color(self.active_index),
            analysis_provider=self._cached_t_plot_thickness_analysis,
        )

    def _remove_t_plot_region(self) -> None:
        if self.t_plot_region is None:
            return
        try:
            self.t_plot_region.sigRegionChanged.disconnect(self.on_t_plot_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.t_plot.removeItem(self.t_plot_region)
        except RuntimeError:
            pass
        self.t_plot_region = None

    def _add_t_plot_region(self, values: list, bounds: list) -> None:
        region = self._make_selection_region(values, bounds=bounds, movable=True)
        region.sigRegionChanged.connect(self.on_t_plot_region_changed)
        self.t_plot.addItem(region, ignoreBounds=True)
        self.t_plot_region = region

    def on_t_plot_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_t_plot_region:
            return
        if not self._t_plot_region_pending:
            self._t_plot_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_t_plot_from_region)

    def _update_t_plot_from_region(self) -> None:
        self._t_plot_region_pending = False
        active = self.active_result()
        if active is None:
            return
        fit_range = self._current_t_plot_fit_range()
        if fit_range is None:
            return
        self._set_custom_t_plot_fit_range(active, fit_range)
        lx_min = self._t_plot_x_range[0] if self._t_plot_x_range else None
        lx_max = self._t_plot_x_range[1] if self._t_plot_x_range else None
        p_min, p_max = self._t_plot_p_range if self._t_plot_p_range else T_PLOT_PLOT_RANGE
        self._t_plot_fit_line, _ = replace_t_plot_fit_line(
            self.t_plot, self._t_plot_fit_line, active,
            fit_range[0], fit_range[1],
            line_x_min=lx_min, line_x_max=lx_max,
            data_p_min=p_min, data_p_max=p_max,
            thickness_params=self.t_plot_thickness_params,
            thickness_method=self.t_plot_thickness_method,
            color=self._analysis_sample_color(self.active_index),
            width=ACTIVE_LINE_WIDTH,
            analysis_provider=self._cached_t_plot_thickness_analysis,
        )
        self._refresh_t_plot_selection(active, fit_range, p_min, p_max)
        self._refresh_sample_t_plot_cell(self.active_index)
        self.refresh_metrics()

    def _current_t_plot_fit_range(self) -> tuple[float, float] | None:
        if self.t_plot_region is None:
            return None
        try:
            lo, hi = self.t_plot_region.getRegion()
            return (min(float(lo), float(hi)), max(float(lo), float(hi)))
        except RuntimeError:
            return None

    # ── reset all fit regions ─────────────────────────────────────────────────

    def _remove_bjh_selection(self) -> None:
        for item in self._bjh_selection_items:
            try:
                self.pore_plot.removeItem(item)
            except RuntimeError:
                pass
        self._bjh_selection_items = []

    def _refresh_bjh_selection(self, diameter_range=None) -> None:
        self._remove_bjh_selection()
        diameter_range = diameter_range or self._current_bjh_diameter_range()
        if diameter_range is None or not self._bjh_distribution_rows_by_key:
            return
        self._bjh_selection_items = plot_bjh_selection(
            self.pore_plot,
            self._bjh_distribution_rows_by_key,
            self.sample_colors,
            diameter_range,
            active_index=self.active_index,
            differential_mode=self.bjh_differential_mode,
            display_metrics=self.bjh_display_metrics,
        )

    def _remove_bjh_region(self) -> None:
        if self.bjh_region is None:
            return
        try:
            self.bjh_region.sigRegionChanged.disconnect(self.on_bjh_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.pore_plot.removeItem(self.bjh_region)
        except RuntimeError:
            pass
        self.bjh_region = None

    def _add_bjh_region(self, values: list[float], bounds: list[float]) -> None:
        region = self._make_selection_region(
            values,
            bounds=bounds,
            movable=True,
            line_color=BJH_REGION_LINE_COLOR,
            hover_line_color=BJH_REGION_LINE_HOVER_COLOR,
            fill_color=BJH_REGION_FILL_COLOR,
            hover_fill_color=BJH_REGION_FILL_HOVER_COLOR,
        )
        region.sigRegionChanged.connect(self.on_bjh_region_changed)
        self.pore_plot.addItem(region, ignoreBounds=True)
        self.bjh_region = region

    def on_bjh_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_bjh_region:
            return
        if not self._bjh_region_pending:
            self._bjh_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_bjh_pore_volume_from_region)

    def _update_bjh_pore_volume_from_region(self) -> None:
        self._bjh_region_pending = False
        diameter_range = self._current_bjh_diameter_range()
        if diameter_range is None:
            return
        self.bjh_pore_volume_range = diameter_range
        self._refresh_bjh_selection(diameter_range)
        self._refresh_all_sample_bjh_pore_cells()

    def _sync_bjh_region_to_diameter_range(self, diameter_range: tuple[float, float]) -> None:
        if self._setting_bjh_region or self._bjh_diameter_log_bounds is None:
            return
        values = self._diameter_range_to_log_region(diameter_range)
        if values is None:
            return
        bounds = list(self._bjh_diameter_log_bounds)
        values = [max(bounds[0], min(value, bounds[1])) for value in values]
        if values[0] >= values[1]:
            values = list(bounds)
        if values[0] >= values[1]:
            return
        self._setting_bjh_region = True
        try:
            if self.bjh_region is None:
                self._add_bjh_region(values, bounds)
            else:
                if hasattr(self.bjh_region, "setBounds"):
                    self.bjh_region.setBounds(bounds)
                self.bjh_region.setRegion(values)
        finally:
            self._setting_bjh_region = False

    def _diameter_range_to_log_region(self, diameter_range: tuple[float, float]) -> list[float] | None:
        lo, hi = sorted((float(diameter_range[0]), float(diameter_range[1])))
        if not (np.isfinite(lo) and np.isfinite(hi)) or lo <= 0.0 or hi <= 0.0:
            return None
        log_lo = math.log10(lo)
        log_hi = math.log10(hi)
        if log_lo == log_hi:
            log_lo -= 0.01
            log_hi += 0.01
        return [log_lo, log_hi]

    def _current_bjh_diameter_range(self) -> tuple[float, float] | None:
        if self.bjh_region is None:
            return None
        try:
            log_lo, log_hi = self.bjh_region.getRegion()
        except RuntimeError:
            return None
        lo, hi = sorted((float(log_lo), float(log_hi)))
        return (10.0**lo, 10.0**hi)

    def _pressure_for_bjh_diameter(self, diameter_nm: float) -> float | None:
        try:
            target = float(diameter_nm)
        except (TypeError, ValueError):
            return None
        if not np.isfinite(target) or target <= 0.0:
            return None
        rows = self._active_bjh_distribution_rows()
        candidates: list[tuple[float, float]] = []
        for row in rows:
            try:
                diameter = float(row["pore_diameter_nm"])
                p_low = float(row["relative_pressure_low"])
                p_high = float(row["relative_pressure_high"])
            except (KeyError, TypeError, ValueError):
                continue
            if not (np.isfinite(diameter) and diameter > 0.0 and np.isfinite(p_low) and np.isfinite(p_high)):
                continue
            distance = abs(math.log(diameter / target))
            candidates.append((distance, max(p_low, p_high)))
        if not candidates:
            return None
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    def _remove_dh_selection(self) -> None:
        for item in self._dh_selection_items:
            try:
                self.dh_plot.removeItem(item)
            except RuntimeError:
                pass
        self._dh_selection_items = []

    def _refresh_dh_selection(self, diameter_range=None) -> None:
        self._remove_dh_selection()
        diameter_range = diameter_range or self._current_dh_diameter_range()
        if diameter_range is None or not self._dh_distribution_rows_by_key:
            return
        self._dh_selection_items = plot_bjh_selection(
            self.dh_plot,
            self._dh_distribution_rows_by_key,
            self.sample_colors,
            diameter_range,
            active_index=self.active_index,
            differential_mode=self.dh_differential_mode,
            display_metrics=self.dh_display_metrics,
        )

    def _remove_dh_region(self) -> None:
        if self.dh_region is None:
            return
        try:
            self.dh_region.sigRegionChanged.disconnect(self.on_dh_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.dh_plot.removeItem(self.dh_region)
        except RuntimeError:
            pass
        self.dh_region = None

    def _add_dh_region(self, values: list[float], bounds: list[float]) -> None:
        region = self._make_selection_region(
            values,
            bounds=bounds,
            movable=True,
            line_color=BJH_REGION_LINE_COLOR,
            hover_line_color=BJH_REGION_LINE_HOVER_COLOR,
            fill_color=BJH_REGION_FILL_COLOR,
            hover_fill_color=BJH_REGION_FILL_HOVER_COLOR,
        )
        region.sigRegionChanged.connect(self.on_dh_region_changed)
        self.dh_plot.addItem(region, ignoreBounds=True)
        self.dh_region = region

    def on_dh_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_dh_region:
            return
        if not self._dh_region_pending:
            self._dh_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_dh_pore_volume_from_region)

    def _update_dh_pore_volume_from_region(self) -> None:
        self._dh_region_pending = False
        diameter_range = self._current_dh_diameter_range()
        if diameter_range is None:
            return
        self.bjh_pore_volume_range = diameter_range
        self._refresh_dh_selection(diameter_range)
        self._refresh_all_sample_bjh_pore_cells()

    def _sync_dh_region_to_diameter_range(self, diameter_range: tuple[float, float]) -> None:
        if self._setting_dh_region or self._dh_diameter_log_bounds is None:
            return
        values = self._diameter_range_to_log_region(diameter_range)
        if values is None:
            return
        bounds = list(self._dh_diameter_log_bounds)
        values = [max(bounds[0], min(value, bounds[1])) for value in values]
        if values[0] >= values[1]:
            values = list(bounds)
        if values[0] >= values[1]:
            return
        self._setting_dh_region = True
        try:
            if self.dh_region is None:
                self._add_dh_region(values, bounds)
            else:
                if hasattr(self.dh_region, "setBounds"):
                    self.dh_region.setBounds(bounds)
                self.dh_region.setRegion(values)
        finally:
            self._setting_dh_region = False

    def _current_dh_diameter_range(self) -> tuple[float, float] | None:
        if self.dh_region is None:
            return None
        try:
            log_lo, log_hi = self.dh_region.getRegion()
        except RuntimeError:
            return None
        lo, hi = sorted((float(log_lo), float(log_hi)))
        return (10.0**lo, 10.0**hi)

    def _remove_hk_selection(self) -> None:
        for item in self._hk_selection_items:
            try:
                self.hk_plot.removeItem(item)
            except RuntimeError:
                pass
        self._hk_selection_items = []

    def _refresh_hk_selection(self, width_range=None) -> None:
        self._remove_hk_selection()
        width_range = width_range or self._current_hk_width_range()
        if width_range is None or not self._hk_distribution_rows_by_key:
            return
        self._hk_selection_items = plot_hk_selection(
            self.hk_plot,
            self._hk_distribution_rows_by_key,
            self.sample_colors,
            width_range,
            active_index=self.active_index,
            display_metric=self.hk_display_metric,
        )

    def _remove_hk_region(self) -> None:
        if self.hk_region is None:
            return
        try:
            self.hk_region.sigRegionChanged.disconnect(self.on_hk_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.hk_plot.removeItem(self.hk_region)
        except RuntimeError:
            pass
        self.hk_region = None

    def _add_hk_region(self, values: list[float], bounds: list[float]) -> None:
        region = self._make_selection_region(
            values,
            bounds=bounds,
            movable=True,
            line_color=BJH_REGION_LINE_COLOR,
            hover_line_color=BJH_REGION_LINE_HOVER_COLOR,
            fill_color=BJH_REGION_FILL_COLOR,
            hover_fill_color=BJH_REGION_FILL_HOVER_COLOR,
        )
        region.sigRegionChanged.connect(self.on_hk_region_changed)
        self.hk_plot.addItem(region, ignoreBounds=True)
        self.hk_region = region

    def on_hk_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_hk_region:
            return
        if not self._hk_region_pending:
            self._hk_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_hk_pore_volume_from_region)

    def _update_hk_pore_volume_from_region(self) -> None:
        self._hk_region_pending = False
        width_range = self._current_hk_width_range()
        if width_range is None:
            return
        self.hk_pore_volume_range = width_range
        self._refresh_hk_selection(width_range)
        self._refresh_all_sample_bjh_pore_cells()

    def _sync_hk_region_to_width_range(self, width_range: tuple[float, float]) -> None:
        if self._setting_hk_region or self._hk_width_log_bounds is None:
            return
        values = self._diameter_range_to_log_region(width_range)
        if values is None:
            return
        bounds = list(self._hk_width_log_bounds)
        values = [max(bounds[0], min(value, bounds[1])) for value in values]
        if values[0] >= values[1]:
            values = list(bounds)
        if values[0] >= values[1]:
            return
        self._setting_hk_region = True
        try:
            if self.hk_region is None:
                self._add_hk_region(values, bounds)
            else:
                if hasattr(self.hk_region, "setBounds"):
                    self.hk_region.setBounds(bounds)
                self.hk_region.setRegion(values)
        finally:
            self._setting_hk_region = False

    def _current_hk_width_range(self) -> tuple[float, float] | None:
        if self.hk_region is None:
            return None
        try:
            log_lo, log_hi = self.hk_region.getRegion()
        except RuntimeError:
            return None
        lo, hi = sorted((float(log_lo), float(log_hi)))
        return (10.0**lo, 10.0**hi)

    def _remove_dft_selection(self) -> None:
        for item in self._dft_selection_items:
            try:
                self.dft_plot.removeItem(item)
            except RuntimeError:
                pass
        self._dft_selection_items = []

    def _refresh_dft_selection(self, width_range=None) -> None:
        self._remove_dft_selection()
        width_range = width_range or self._current_dft_width_range()
        if width_range is None or not self._dft_distribution_rows_by_index:
            return
        self._dft_selection_items = plot_dft_selection(
            self.dft_plot,
            self._dft_distribution_rows_by_index,
            self.sample_colors,
            width_range,
            active_index=self.active_index,
        )

    def _remove_dft_region(self) -> None:
        if self.dft_region is None:
            return
        try:
            self.dft_region.sigRegionChanged.disconnect(self.on_dft_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            self.dft_plot.removeItem(self.dft_region)
        except RuntimeError:
            pass
        self.dft_region = None

    def _add_dft_region(self, values: list[float], bounds: list[float]) -> None:
        region = self._make_selection_region(
            values,
            bounds=bounds,
            movable=True,
            line_color=BJH_REGION_LINE_COLOR,
            hover_line_color=BJH_REGION_LINE_HOVER_COLOR,
            fill_color=BJH_REGION_FILL_COLOR,
            hover_fill_color=BJH_REGION_FILL_HOVER_COLOR,
        )
        region.sigRegionChanged.connect(self.on_dft_region_changed)
        self.dft_plot.addItem(region, ignoreBounds=True)
        self.dft_region = region

    def on_dft_region_changed(self) -> None:
        if self._syncing_region_changes or self._setting_dft_region:
            return
        if not self._dft_region_pending:
            self._dft_region_pending = True
            QtCore.QTimer.singleShot(25, self._update_dft_pore_volume_from_region)

    def _update_dft_pore_volume_from_region(self) -> None:
        self._dft_region_pending = False
        width_range = self._current_dft_width_range()
        if width_range is None:
            return
        self.dft_pore_volume_range = width_range
        self._refresh_dft_selection(width_range)
        self._refresh_all_sample_bjh_pore_cells()

    def _sync_dft_region_to_width_range(self, width_range: tuple[float, float]) -> None:
        if self._setting_dft_region or self._dft_width_log_bounds is None:
            return
        values = self._diameter_range_to_log_region(width_range)
        if values is None:
            return
        bounds = list(self._dft_width_log_bounds)
        values = [max(bounds[0], min(value, bounds[1])) for value in values]
        if values[0] >= values[1]:
            values = list(bounds)
        if values[0] >= values[1]:
            return
        self._setting_dft_region = True
        try:
            if self.dft_region is None:
                self._add_dft_region(values, bounds)
            else:
                if hasattr(self.dft_region, "setBounds"):
                    self.dft_region.setBounds(bounds)
                self.dft_region.setRegion(values)
        finally:
            self._setting_dft_region = False

    def _current_dft_width_range(self) -> tuple[float, float] | None:
        if self.dft_region is None:
            return None
        try:
            log_lo, log_hi = self.dft_region.getRegion()
        except RuntimeError:
            return None
        lo, hi = sorted((float(log_lo), float(log_hi)))
        return (10.0**lo, 10.0**hi)

    def _pressure_for_dh_diameter(self, diameter_nm: float) -> float | None:
        try:
            target = float(diameter_nm)
        except (TypeError, ValueError):
            return None
        if not np.isfinite(target) or target <= 0.0:
            return None
        rows = self._active_dh_distribution_rows()
        candidates: list[tuple[float, float]] = []
        for row in rows:
            try:
                diameter = float(row["pore_diameter_nm"])
                p_low = float(row["relative_pressure_low"])
                p_high = float(row["relative_pressure_high"])
            except (KeyError, TypeError, ValueError):
                continue
            if not (np.isfinite(diameter) and diameter > 0.0 and np.isfinite(p_low) and np.isfinite(p_high)):
                continue
            distance = abs(math.log(diameter / target))
            candidates.append((distance, max(p_low, p_high)))
        if not candidates:
            return None
        candidates.sort(key=lambda item: item[0])
        return candidates[0][1]

    @staticmethod
    def _freeze_cache_value(value):
        if isinstance(value, dict):
            return tuple(
                sorted(
                    (str(key), MainWindow._freeze_cache_value(item))
                    for key, item in value.items()
                )
            )
        if isinstance(value, (list, tuple)):
            return tuple(MainWindow._freeze_cache_value(item) for item in value)
        if isinstance(value, np.ndarray):
            return MainWindow._freeze_cache_value(value.tolist())
        if isinstance(value, np.generic):
            return MainWindow._freeze_cache_value(value.item())
        if isinstance(value, float):
            if math.isnan(value):
                return ("float", "nan")
            if math.isinf(value):
                return ("float", "inf" if value > 0 else "-inf")
            return value
        if isinstance(value, Path):
            return str(value)
        try:
            hash(value)
        except TypeError:
            return repr(value)
        return value

    def _bjh_distribution_cache_key(
        self,
        result,
        *,
        phase: str,
        thickness_method: str,
        thickness_params: dict[str, object] | None,
        correction: str,
        open_pore_fraction: float,
        smooth: bool,
    ) -> tuple[object, ...]:
        return (
            id(result),
            str(getattr(getattr(result, "header", None), "file_path", "")),
            int(getattr(result, "point_count", 0)),
            "adsorption" if phase == "adsorption" else "desorption",
            str(thickness_method),
            self._freeze_cache_value(thickness_params or {}),
            str(correction),
            float(open_pore_fraction),
            bool(smooth),
        )

    def _discard_bjh_distribution_cache_for_result(self, result) -> None:
        result_id = id(result)
        self._bjh_distribution_cache = {
            key: rows
            for key, rows in self._bjh_distribution_cache.items()
            if key[0] != result_id
        }

    def _discard_dh_distribution_cache_for_result(self, result) -> None:
        result_id = id(result)
        self._dh_distribution_cache = {
            key: rows
            for key, rows in self._dh_distribution_cache.items()
            if key[0] != result_id
        }

    def _discard_hk_distribution_cache_for_result(self, result) -> None:
        result_id = id(result)
        self._hk_distribution_cache = {
            key: rows
            for key, rows in self._hk_distribution_cache.items()
            if key[0] != result_id
        }

    def _discard_dft_result_cache_for_result(self, result) -> None:
        result_id = id(result)
        self._dft_result_cache = {
            key: value
            for key, value in self._dft_result_cache.items()
            if key[0] != result_id
        }

    def _store_bjh_distribution_cache(self, cache_key: tuple[object, ...], rows: list[dict[str, float]]) -> None:
        self._bjh_distribution_cache[cache_key] = rows
        while len(self._bjh_distribution_cache) > BJH_DISTRIBUTION_CACHE_LIMIT:
            self._bjh_distribution_cache.pop(next(iter(self._bjh_distribution_cache)))

    def _cached_bjh_distribution_rows(
        self,
        result,
        *,
        phase: str,
        thickness_method: str,
        thickness_params: dict[str, object] | None,
        correction: str,
        open_pore_fraction: float,
        smooth: bool,
    ) -> list[dict[str, float]]:
        cache_key = self._bjh_distribution_cache_key(
            result,
            phase=phase,
            thickness_method=thickness_method,
            thickness_params=thickness_params,
            correction=correction,
            open_pore_fraction=open_pore_fraction,
            smooth=smooth,
        )
        cached = self._bjh_distribution_cache.get(cache_key)
        if cached is not None:
            return cached
        distribution = bjh_pore_distribution(
            result,
            phase=phase,
            thickness_method=thickness_method,
            thickness_params=thickness_params,
            correction=correction,
            open_pore_fraction=open_pore_fraction,
            smooth=smooth,
        )
        rows = list(distribution.rows)
        self._store_bjh_distribution_cache(cache_key, rows)
        return rows

    def _dh_distribution_cache_key(
        self,
        result,
        *,
        phase: str,
        thickness_method: str | None,
        thickness_params: dict[str, object] | None,
        smooth: bool,
    ) -> tuple[object, ...]:
        return (
            id(result),
            str(getattr(getattr(result, "header", None), "file_path", "")),
            int(getattr(result, "point_count", 0)),
            "adsorption" if phase == "adsorption" else "desorption",
            str(thickness_method or ""),
            self._freeze_cache_value(thickness_params or {}),
            bool(smooth),
        )

    def _store_dh_distribution_cache(self, cache_key: tuple[object, ...], rows: list[dict[str, float]]) -> None:
        self._dh_distribution_cache[cache_key] = rows
        while len(self._dh_distribution_cache) > BJH_DISTRIBUTION_CACHE_LIMIT:
            self._dh_distribution_cache.pop(next(iter(self._dh_distribution_cache)))

    def _cached_dh_distribution_rows(
        self,
        result,
        *,
        phase: str,
        thickness_method: str | None,
        thickness_params: dict[str, object] | None,
        smooth: bool,
    ) -> list[dict[str, float]]:
        cache_key = self._dh_distribution_cache_key(
            result,
            phase=phase,
            thickness_method=thickness_method,
            thickness_params=thickness_params,
            smooth=smooth,
        )
        cached = self._dh_distribution_cache.get(cache_key)
        if cached is not None:
            return cached
        distribution = dh_pore_distribution(
            result,
            phase=phase,
            thickness_method=thickness_method,
            thickness_params=thickness_params,
            smooth=smooth,
        )
        rows = list(distribution.rows)
        self._store_dh_distribution_cache(cache_key, rows)
        return rows

    def _hk_distribution_cache_key(
        self,
        result,
        *,
        geometry: str,
        adsorbent_key: str,
        adsorptive_key: str,
        adsorbent_properties: dict[str, object] | None,
        adsorptive_properties: dict[str, object] | None,
        interaction_parameter_erg_cm4: float,
        interaction_parameter_mode: str,
        cheng_yang_correction: bool,
        smooth: bool,
    ) -> tuple[object, ...]:
        return (
            id(result),
            str(getattr(getattr(result, "header", None), "file_path", "")),
            int(getattr(result, "point_count", 0)),
            str(geometry),
            str(adsorbent_key),
            str(adsorptive_key),
            self._freeze_cache_value(adsorbent_properties or {}),
            self._freeze_cache_value(adsorptive_properties or {}),
            float(interaction_parameter_erg_cm4),
            str(interaction_parameter_mode),
            bool(cheng_yang_correction),
            bool(smooth),
        )

    def _store_hk_distribution_cache(self, cache_key: tuple[object, ...], rows: list[dict[str, float]]) -> None:
        self._hk_distribution_cache[cache_key] = rows
        while len(self._hk_distribution_cache) > BJH_DISTRIBUTION_CACHE_LIMIT:
            self._hk_distribution_cache.pop(next(iter(self._hk_distribution_cache)))

    def _cached_hk_distribution_rows(
        self,
        result,
        *,
        geometry: str,
        adsorbent_key: str,
        adsorptive_key: str,
        adsorbent_properties: dict[str, object] | None,
        adsorptive_properties: dict[str, object] | None,
        interaction_parameter_erg_cm4: float,
        interaction_parameter_mode: str,
        cheng_yang_correction: bool,
        smooth: bool,
    ) -> list[dict[str, float]]:
        cache_key = self._hk_distribution_cache_key(
            result,
            geometry=geometry,
            adsorbent_key=adsorbent_key,
            adsorptive_key=adsorptive_key,
            adsorbent_properties=adsorbent_properties,
            adsorptive_properties=adsorptive_properties,
            interaction_parameter_erg_cm4=interaction_parameter_erg_cm4,
            interaction_parameter_mode=interaction_parameter_mode,
            cheng_yang_correction=cheng_yang_correction,
            smooth=smooth,
        )
        cached = self._hk_distribution_cache.get(cache_key)
        if cached is not None:
            return cached
        distribution = horvath_kawazoe_pore_distribution(
            result,
            geometry=geometry,
            adsorbent_key=adsorbent_key,
            adsorptive_key=adsorptive_key,
            adsorbent_properties=adsorbent_properties,
            adsorptive_properties=adsorptive_properties,
            interaction_parameter_erg_cm4=interaction_parameter_erg_cm4,
            interaction_parameter_mode=interaction_parameter_mode,
            cheng_yang_correction=cheng_yang_correction,
            smooth=smooth,
        )
        rows = list(distribution.rows)
        self._store_hk_distribution_cache(cache_key, rows)
        return rows

    def _dft_result_cache_key(
        self,
        result,
        *,
        analysis_type: str,
        geometry: str,
        model: str,
        regularization: float,
        include_diagnostics: bool,
    ) -> tuple[object, ...]:
        return (
            id(result),
            str(getattr(getattr(result, "header", None), "file_path", "")),
            int(getattr(result, "point_count", 0)),
            str(analysis_type),
            str(geometry),
            str(model),
            round(float(regularization), 10),
            bool(include_diagnostics),
        )

    def _store_dft_result_cache(self, cache_key: tuple[object, ...], result) -> None:
        self._dft_result_cache[cache_key] = result
        while len(self._dft_result_cache) > BJH_DISTRIBUTION_CACHE_LIMIT:
            self._dft_result_cache.pop(next(iter(self._dft_result_cache)))

    def _cached_dft_result(
        self,
        result,
        *,
        analysis_type: str,
        geometry: str,
        model: str,
        regularization: float,
        include_diagnostics: bool = False,
    ):
        cache_key = self._dft_result_cache_key(
            result,
            analysis_type=analysis_type,
            geometry=geometry,
            model=model,
            regularization=regularization,
            include_diagnostics=include_diagnostics,
        )
        cached = self._dft_result_cache.get(cache_key)
        if cached is not None:
            return cached
        distribution = dft_pore_distribution(
            result,
            analysis_type=analysis_type,
            geometry=geometry,
            model=model,
            regularization=regularization,
            include_diagnostics=include_diagnostics,
        )
        self._store_dft_result_cache(cache_key, distribution)
        return distribution

    def _active_bjh_distribution_rows(self) -> list[dict[str, float]]:
        active = self.active_result()
        if active is None:
            return []
        settings = self._bjh_settings_for_result(active)
        phase = self._bjh_pore_volume_phase(settings)
        if phase is None:
            return []
        return list(
            self._cached_bjh_distribution_rows(
                active,
                phase=phase,
                thickness_method=str(settings["thickness_method"]),
                thickness_params=dict(settings["thickness_params"]),
                correction=str(settings["correction"]),
                open_pore_fraction=float(settings["open_pore_fraction"]),
                smooth=bool(settings["smooth_derivative"]),
            )
        )

    def _active_dh_distribution_rows(self) -> list[dict[str, float]]:
        active = self.active_result()
        if active is None:
            return []
        settings = self._dh_settings_for_result(active)
        phase = self._dh_pore_volume_phase(settings)
        if phase is None:
            return []
        return list(
            self._cached_dh_distribution_rows(
                active,
                phase=phase,
                thickness_method=str(settings["thickness_method"]),
                thickness_params=dict(settings["thickness_params"]),
                smooth=bool(settings["smooth_derivative"]),
            )
        )

    def _active_dft_distribution_rows(self) -> list[dict[str, float]]:
        active = self.active_result()
        if active is None:
            return []
        settings = self._dft_settings_for_result(active)
        distribution = self._cached_dft_result(
            active,
            analysis_type=str(settings["analysis_type"]),
            geometry=str(settings["geometry"]),
            model=str(settings["model"]),
            regularization=float(settings["regularization"]),
        )
        return list(getattr(distribution, "rows", []))

    def _active_dft_fit_rows(self) -> list[dict[str, float]]:
        active = self.active_result()
        if active is None:
            return []
        settings = self._dft_settings_for_result(active)
        distribution = self._cached_dft_result(
            active,
            analysis_type=str(settings["analysis_type"]),
            geometry=str(settings["geometry"]),
            model=str(settings["model"]),
            regularization=float(settings["regularization"]),
        )
        return list(getattr(distribution, "fit_rows", []))

    @staticmethod
    def _bjh_log_bounds_from_rows(rows_by_key: dict[tuple[int, str], list[dict[str, float]]]) -> tuple[float, float] | None:
        diameters = []
        for rows in rows_by_key.values():
            for row in rows:
                try:
                    diameter = float(row["pore_diameter_nm"])
                except (KeyError, TypeError, ValueError):
                    continue
                if np.isfinite(diameter) and diameter > 0.0:
                    diameters.append(diameter)
        if not diameters:
            return None
        log_min = math.log10(min(diameters))
        log_max = math.log10(max(diameters))
        if log_min == log_max:
            log_min -= 0.01
            log_max += 0.01
        return (log_min, log_max)

    @staticmethod
    def _hk_log_bounds_from_rows(rows_by_key: dict[tuple[int, str], list[dict[str, float]]]) -> tuple[float, float] | None:
        widths = []
        for rows in rows_by_key.values():
            for row in rows:
                try:
                    width = float(row.get("pore_width_nm", row.get("pore_diameter_nm")))
                except (TypeError, ValueError):
                    continue
                if np.isfinite(width) and width > 0.0:
                    widths.append(width)
        if not widths:
            return None
        log_min = math.log10(min(widths))
        log_max = math.log10(max(widths))
        if log_min == log_max:
            log_min -= 0.01
            log_max += 0.01
        return (log_min, log_max)

    @staticmethod
    def _dft_log_bounds_from_rows(rows_by_index: dict[int, list[dict[str, float]]]) -> tuple[float, float] | None:
        widths = []
        for rows in rows_by_index.values():
            for row in rows:
                try:
                    width = float(row.get("pore_width_nm", row.get("pore_diameter_nm")))
                except (TypeError, ValueError):
                    continue
                if np.isfinite(width) and width > 0.0:
                    widths.append(width)
        if not widths:
            return None
        log_min = math.log10(min(widths))
        log_max = math.log10(max(widths))
        if log_min == log_max:
            log_min -= 0.01
            log_max += 0.01
        return (log_min, log_max)

    def _reset_all_fit_regions(self) -> None:
        self._remove_bet_region()
        self._remove_bet_selection()
        self._remove_langmuir_region()
        self._remove_langmuir_selection()
        self._remove_t_plot_region()
        self._remove_t_plot_selection()
        self._remove_bjh_region()
        self._remove_bjh_selection()
        self._remove_dh_region()
        self._remove_dh_selection()
        self._remove_hk_region()
        self._remove_hk_selection()
        self._remove_dft_region()
        self._remove_dft_selection()
        self._bet_fit_line = None
        self._bet_x_range = None
        self._bet_plot_p_range = None
        self._langmuir_fit_line = None
        self._langmuir_x_range = None
        self._langmuir_plot_p_range = None
        self._t_plot_fit_line = None
        self._t_plot_x_range = None
        self._t_plot_p_range = None
        self._bjh_distribution_rows_by_key = {}
        self._bjh_diameter_log_bounds = None
        self._dh_distribution_rows_by_key = {}
        self._dh_diameter_log_bounds = None
        self._hk_distribution_rows_by_key = {}
        self._hk_width_log_bounds = None
        self._dft_distribution_rows_by_index = {}
        self._dft_width_log_bounds = None

    def refresh_metrics(self) -> None:
        active = self.active_result()
        if active is None:
            self.metrics_table.setRowCount(0)
            return
        bet_fit_range = self._current_bet_fit_range() or self._bet_fit_range_for_result(active)
        langmuir_fit_range = self._current_langmuir_fit_range() or self._langmuir_fit_range_for_result(active)
        t_plot_fit_range = self._current_t_plot_fit_range() or self._t_plot_fit_range_for_result(active)
        pressure_range = self._current_pressure_region()
        bet = self._cached_bet_analysis(active, bet_fit_range[0], bet_fit_range[1])
        langmuir = self._cached_langmuir_analysis(active, langmuir_fit_range[0], langmuir_fit_range[1])
        if pressure_range is None:
            t_plot = self._cached_t_plot_thickness_analysis(
                active,
                t_plot_fit_range[0],
                t_plot_fit_range[1],
                thickness_params=self.t_plot_thickness_params,
                thickness_method=self.t_plot_thickness_method,
            )
        else:
            t_plot = self._cached_t_plot_thickness_analysis(
                active,
                t_plot_fit_range[0],
                t_plot_fit_range[1],
                pressure_range[0],
                pressure_range[1],
                self.t_plot_thickness_params,
                self.t_plot_thickness_method,
            )
        rows = active_metric_rows(
            active,
            pressure_range,
            bet_fit_range,
            langmuir_fit_range,
            t_plot_fit_range,
            t_plot_thickness_method=self.t_plot_thickness_method,
            t_plot_thickness_params=self.t_plot_thickness_params,
            t_plot_surface_area_mode=self.t_plot_surface_area_mode,
            t_plot_input_surface_area=self.t_plot_surface_area_input,
            t_plot_surface_area_correction=self.t_plot_surface_area_correction,
            bet_analysis_result=bet,
            langmuir_analysis_result=langmuir,
            t_plot_analysis_result=t_plot,
        )
        self._fill_two_column_table(self.metrics_table, rows)

    def refresh_condition_table(self) -> None:
        active = self.active_result()
        if active is None:
            self.condition_table.setRowCount(0)
            return
        rows = condition_rows(active)
        self._fill_two_column_table(self.condition_table, rows)

    def refresh_isotherm_table(self) -> None:
        active = self.active_result()
        if active is None:
            self.isotherm_table.setRowCount(0)
            return
        self.isotherm_table.setRowCount(len(active.isotherm))
        for row, point in enumerate(active.isotherm):
            values = [
                point.index,
                "吸附" if point.phase == "adsorption" else "脱附",
                _fmt(point.relative_pressure, 9),
                _fmt(point.absolute_pressure_mmHg, 6),
                _fmt(point.quantity_adsorbed_cm3_g_stp, 6),
                _fmt(point.quantity_adsorbed_mmol_g, 6),
                _fmt(point.saturation_pressure_mmHg, 6),
                point.elapsed_time,
            ]
            for column, value in enumerate(values):
                self._set_table_item(self.isotherm_table, row, column, str(value))

    def refresh_target_table(self) -> None:
        active = self.active_result()
        if active is None:
            self.target_table.setRowCount(0)
            return
        self.target_table.setRowCount(len(active.target_pressure_table))
        for row, item in enumerate(active.target_pressure_table):
            values = [
                item.row,
                "吸附" if item.branch == "adsorption" else "脱附",
                _fmt(item.starting_pressure_p_po, 9),
                _fmt(item.ending_pressure_p_po, 9),
                _fmt(item.pressure_increment_p_po, 9),
                item.ending_pressure_rel_offset,
            ]
            for column, value in enumerate(values):
                self._set_table_item(self.target_table, row, column, str(value))

    def refresh_log_table(self) -> None:
        active = self.active_result()
        if active is None:
            self.log_table.setRowCount(0)
            return
        rows = []
        rows.extend(("SUBSET705", item.rel_offset, item.text) for item in active.log_messages)
        rows.extend(("SUBSET1021", item.rel_offset, item.text) for item in active.sample_tube_strings)
        self.log_table.setRowCount(len(rows))
        for row, values in enumerate(rows):
            for column, value in enumerate(values):
                self._set_table_item(self.log_table, row, column, str(value))

    def refresh_report_module_table(self) -> None:
        active = self.active_result()
        if active is None:
            self.report_module_table.setRowCount(0)
            return
        report_subset_ids = (311, 312, 314, 315, 316, 331, 332)
        rows = []
        for subset_id in report_subset_ids:
            for item in active.raw_strings.get(subset_id, []):
                rows.append((f"SUBSET{subset_id}", item.rel_offset, item.text))
        self.report_module_table.setRowCount(len(rows))
        for row, values in enumerate(rows):
            for column, value in enumerate(values):
                self._set_table_item(self.report_module_table, row, column, str(value))

    def active_result(self):
        if self.active_index < 0 or self.active_index >= len(self.results):
            return None
        return self.results[self.active_index]

    def on_select_all_changed(self, state: int) -> None:
        if self._updating_sample_checks or not self.visible_results:
            return
        state_value = _check_state_value(state)
        if state_value == _check_state_value(QtCore.Qt.PartiallyChecked):
            return
        checked = state_value == _check_state_value(QtCore.Qt.Checked)
        self.visible_results = [checked] * len(self.visible_results)

        self.sample_table.blockSignals(True)
        for item in self.sample_items:
            item.setCheckState(QtCore.Qt.Checked if checked else QtCore.Qt.Unchecked)
        self.sample_table.blockSignals(False)

        self._refresh_visibility_dependent_ui()

    def _sync_select_all_state(self) -> None:
        if not hasattr(self, "select_all_check"):
            return
        if not self.visible_results:
            state = QtCore.Qt.Unchecked
        elif all(self.visible_results):
            state = QtCore.Qt.Checked
        elif any(self.visible_results):
            state = QtCore.Qt.PartiallyChecked
        else:
            state = QtCore.Qt.Unchecked

        self._updating_sample_checks = True
        self.select_all_check.setEnabled(bool(self.visible_results))
        self.select_all_check.setCheckState(state)
        self._updating_sample_checks = False

    def _pore_volume_method_label(self) -> str:
        method = self._active_pore_volume_method()
        if method == PORE_VOLUME_METHOD_DFT:
            return "DFT"
        if method == PORE_VOLUME_METHOD_HK:
            return "HK"
        return "DH" if method == PORE_VOLUME_METHOD_DH else "BJH"

    def _active_pore_volume_method(self) -> str:
        if self._is_dft_tab_active():
            return PORE_VOLUME_METHOD_DFT
        if self._is_hk_tab_active():
            return PORE_VOLUME_METHOD_HK
        return PORE_VOLUME_METHOD_DH if self._is_dh_tab_active() else PORE_VOLUME_METHOD_BJH

    def _update_pore_volume_header(self) -> None:
        item = self.sample_table.horizontalHeaderItem(BJH_PORE_VOLUME_COLUMN)
        label = self._pore_volume_method_label()
        if item is not None:
            item.setText("选区孔容量(cm3/g)")
            item.setToolTip(f"当前按 {label} 标签页绿色选区积分；切换 BJH/DH/HK 标签页会自动切换算法")
            item.setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignVCenter)
        self._position_header_controls()

    def _position_header_controls(self, *args) -> None:
        header = self.sample_table.frozen_header()
        if header.isVisible():
            size = self.select_all_check.sizeHint()
            x = header.sectionViewportPosition(VISIBLE_COLUMN) + (header.sectionSize(VISIBLE_COLUMN) - size.width()) // 2
            y = (header.height() - size.height()) // 2
            self.select_all_check.setVisible(x + size.width() > 0 and x < header.width())
            self.select_all_check.setGeometry(x, y, size.width(), size.height())

    def _refresh_visibility_dependent_ui(self) -> None:
        self.refresh_isotherm_plot()
        self.refresh_sample_table()
        self.refresh_metrics()
        self.refresh_analysis_plots()

    def _clear_analysis_plots(self) -> None:
        self._remove_bet_region()
        self._remove_bet_selection()
        self._remove_langmuir_region()
        self._remove_langmuir_selection()
        self._remove_t_plot_region()
        self._remove_t_plot_selection()
        self._remove_bjh_region()
        self._remove_bjh_selection()
        self._remove_dh_region()
        self._remove_dh_selection()
        self._remove_hk_region()
        self._remove_hk_selection()
        self._remove_dft_region()
        self._remove_dft_selection()
        self._bet_plot_p_range = None
        self._langmuir_plot_p_range = None
        self._t_plot_p_range = None
        self._bjh_distribution_rows_by_key = {}
        self._bjh_diameter_log_bounds = None
        self._dh_distribution_rows_by_key = {}
        self._dh_diameter_log_bounds = None
        self._hk_distribution_rows_by_key = {}
        self._hk_width_log_bounds = None
        self._dft_distribution_rows_by_index = {}
        self._dft_width_log_bounds = None
        for plot in (self.bet_plot, self.langmuir_plot, self.t_plot, self.pore_plot, self.dh_plot, self.hk_plot, self.dft_plot):
            plot.clear()
        if hasattr(self, "dft_diagnostic_plot"):
            self.dft_diagnostic_plot.clear()

    def _all_pressure_values(self) -> np.ndarray:
        values = []
        for result, visible in zip(self.results, self.visible_results):
            if not visible:
                continue
            for point in result.isotherm:
                try:
                    pressure = float(point.relative_pressure)
                except (TypeError, ValueError):
                    continue
                if np.isfinite(pressure):
                    values.append(pressure)
        return np.asarray(values, dtype=float)

    @staticmethod
    def _default_pressure_region(pressure: np.ndarray) -> list[float]:
        data_min = float(np.nanmin(pressure))
        data_max = float(np.nanmax(pressure))
        if data_min == data_max:
            return [data_min - 0.01, data_max + 0.01]
        bet_min = max(data_min, 0.05)
        bet_max = min(data_max, 0.30)
        if bet_min < bet_max:
            return [bet_min, bet_max]
        span = data_max - data_min
        return [data_min + span * 0.25, data_min + span * 0.55]

    def _default_isotherm_region(self, pressure: np.ndarray) -> list[float]:
        if self._is_bjh_tab_active():
            bjh_pressure_range = self._default_bjh_pressure_range()
            if bjh_pressure_range is not None:
                return self._clamp_pressure_region(bjh_pressure_range, pressure)
            return self._full_pressure_region(pressure)
        if self._is_dh_tab_active():
            dh_pressure_range = self._default_dh_pressure_range()
            if dh_pressure_range is not None:
                return self._clamp_pressure_region(dh_pressure_range, pressure)
            return self._full_pressure_region(pressure)
        if self._is_hk_tab_active():
            return self._clamp_pressure_region(HK_DEFAULT_PRESSURE_RANGE, pressure)
        if self._is_dft_tab_active():
            return self._full_pressure_region(pressure)
        return self._default_pressure_region(pressure)

    @staticmethod
    def _full_pressure_region(pressure: np.ndarray) -> list[float]:
        data_min = float(np.nanmin(pressure))
        data_max = float(np.nanmax(pressure))
        if data_min == data_max:
            return [data_min - 0.01, data_max + 0.01]
        return [data_min, data_max]

    def _is_bjh_tab_active(self) -> bool:
        return getattr(self, "plot_tabs", None) is not None and self.plot_tabs.currentWidget() is self.bjh_tab

    def _is_dh_tab_active(self) -> bool:
        return getattr(self, "plot_tabs", None) is not None and self.plot_tabs.currentWidget() is self.dh_tab

    def _is_hk_tab_active(self) -> bool:
        return getattr(self, "plot_tabs", None) is not None and self.plot_tabs.currentWidget() is self.hk_tab

    def _is_dft_tab_active(self) -> bool:
        return getattr(self, "plot_tabs", None) is not None and self.plot_tabs.currentWidget() is self.dft_tab

    def _clamp_pressure_region(self, raw_region: list[float] | tuple[float, float], pressure: np.ndarray) -> list[float]:
        data_min = float(np.nanmin(pressure))
        data_max = float(np.nanmax(pressure))
        region_min, region_max = sorted((float(raw_region[0]), float(raw_region[1])))
        region_min = max(data_min, min(region_min, data_max))
        region_max = max(data_min, min(region_max, data_max))
        if region_min < region_max:
            return [region_min, region_max]
        return self._default_pressure_region(pressure)

    def _make_selection_region(
        self,
        values,
        bounds=None,
        movable: bool = True,
        *,
        line_color: str = REGION_LINE_COLOR,
        hover_line_color: str = REGION_LINE_HOVER_COLOR,
        fill_color: tuple[int, int, int, int] = REGION_FILL_COLOR,
        hover_fill_color: tuple[int, int, int, int] = REGION_FILL_HOVER_COLOR,
    ):
        region = pg.LinearRegionItem(
            values,
            bounds=bounds,
            movable=movable,
            brush=pg.mkBrush(*fill_color),
            hoverBrush=pg.mkBrush(*hover_fill_color),
            pen=_region_pen(line_color),
            hoverPen=_region_pen(hover_line_color),
            swapMode="block",
        )
        for line in getattr(region, "lines", []):
            line.setPen(_region_pen(line_color))
            line.setHoverPen(_region_pen(hover_line_color))
            line.setCursor(QtCore.Qt.SizeHorCursor)
        return region

    def _remove_isotherm_selection(self) -> None:
        for item in self._isotherm_selection_items:
            try:
                self.isotherm_plot.removeItem(item)
            except RuntimeError:
                pass
        self._isotherm_selection_items = []

    def _refresh_isotherm_selection(self, pressure_range=None) -> None:
        self._remove_isotherm_selection()
        if pressure_range is None:
            pressure_range = self._current_pressure_region()
        if pressure_range is None:
            return
        self._isotherm_selection_items = plot_isotherm_selection(
            self.isotherm_plot,
            self.results,
            self.visible_results,
            self.sample_colors,
            pressure_range,
            active_index=self.active_index,
            fade_inactive=False,
        )

    def _add_region(self, raw_region: list[float] | tuple[float, float], pressure: np.ndarray) -> None:
        if pressure.size == 0:
            return
        region = self._clamp_pressure_region(raw_region, pressure)
        self._setting_isotherm_region = True
        try:
            self.region = self._make_selection_region(
                region,
                bounds=[float(np.nanmin(pressure)), float(np.nanmax(pressure))],
                movable=True,
            )
            self.isotherm_plot.addItem(self.region, ignoreBounds=True)
            self.region.sigRegionChanged.connect(self.on_region_changed)
            if hasattr(self.region, "sigRegionChangeFinished"):
                self.region.sigRegionChangeFinished.connect(self.on_region_change_finished)
        finally:
            self._setting_isotherm_region = False

    def _remove_region(self) -> None:
        if self.region is None:
            return
        try:
            self.region.sigRegionChanged.disconnect(self.on_region_changed)
        except (RuntimeError, TypeError):
            pass
        try:
            if hasattr(self.region, "sigRegionChangeFinished"):
                self.region.sigRegionChangeFinished.disconnect(self.on_region_change_finished)
        except (RuntimeError, TypeError):
            pass
        try:
            self.isotherm_plot.removeItem(self.region)
        except RuntimeError:
            pass
        self.region = None

    def _set_default_isotherm_region(self, pressure_range: tuple[float, float]) -> None:
        pressure = self._all_pressure_values()
        if pressure.size == 0:
            return
        region = self._clamp_pressure_region(pressure_range, pressure)
        self._last_isotherm_region_range = tuple(region)
        self._syncing_region_changes = True
        self._setting_isotherm_region = True
        try:
            if self.region is None:
                self._add_region(region, pressure)
            else:
                self.region.setRegion(region)
        finally:
            self._setting_isotherm_region = False
            self._syncing_region_changes = False
        self._isotherm_region_custom = False
        self._refresh_isotherm_selection(tuple(region))

    def on_region_changed(self) -> None:
        if self._syncing_region_changes:
            return
        self._mark_isotherm_region_custom()
        pressure_range = self._current_pressure_region()
        self._refresh_isotherm_selection(pressure_range)
        self.queue_metrics_update()

    def on_region_change_finished(self) -> None:
        if self._syncing_region_changes:
            return
        self._mark_isotherm_region_custom()
        pressure_range = self._current_pressure_region()
        self._refresh_isotherm_selection(pressure_range)
        self.queue_metrics_update()

    def _mark_isotherm_region_custom(self) -> None:
        if not self._setting_isotherm_region:
            self._isotherm_region_custom = True

    def queue_metrics_update(self) -> None:
        if self._metrics_pending:
            return
        self._metrics_pending = True
        QtCore.QTimer.singleShot(25, self.update_metrics_from_region)

    def update_metrics_from_region(self) -> None:
        self._metrics_pending = False
        if self._is_bjh_tab_active():
            self.refresh_bjh_plot()
            self._refresh_all_sample_bjh_pore_cells()
            self.refresh_metrics()
            return
        if self._is_dh_tab_active():
            self.refresh_dh_plot()
            self._refresh_all_sample_bjh_pore_cells()
            self.refresh_metrics()
            return
        if self._is_hk_tab_active():
            self.refresh_hk_plot()
            self.refresh_metrics()
            return
        if self._is_dft_tab_active():
            self.refresh_dft_plot()
            self.refresh_metrics()
            return
        self.refresh_sample_table()
        self._update_analysis_plots_for_region()
        self.refresh_metrics()

    def _current_pressure_region(self) -> tuple[float, float] | None:
        if self.region is None:
            return self._last_isotherm_region_range if self._isotherm_region_custom else None
        try:
            region_min, region_max = self.region.getRegion()
        except RuntimeError:
            return self._last_isotherm_region_range if self._isotherm_region_custom else None
        lo, hi = sorted((float(region_min), float(region_max)))
        self._last_isotherm_region_range = (lo, hi)
        return (lo, hi)

    def _analysis_bundle_for_range(self, result, pressure_range: tuple[float, float] | None):
        if pressure_range is None:
            return analysis_bundle(result)
        return analysis_bundle(result, pressure_range[0], pressure_range[1])

    def _analysis_cache_identity(self, result) -> tuple[object, ...]:
        return (
            id(result),
            str(getattr(getattr(result, "header", None), "file_path", "")),
            int(getattr(result, "point_count", 0)),
        )

    def _fit_analysis_cache_key(self, analysis_name: str, result, *parts) -> tuple[object, ...]:
        return (
            analysis_name,
            *self._analysis_cache_identity(result),
            self._freeze_cache_value(parts),
        )

    def _discard_fit_analysis_cache_for_result(self, result) -> None:
        result_id = id(result)
        self._fit_analysis_cache = {
            key: analysis
            for key, analysis in self._fit_analysis_cache.items()
            if len(key) < 2 or key[1] != result_id
        }

    def _store_fit_analysis_cache(self, cache_key: tuple[object, ...], analysis) -> None:
        self._fit_analysis_cache[cache_key] = analysis
        while len(self._fit_analysis_cache) > FIT_ANALYSIS_CACHE_LIMIT:
            self._fit_analysis_cache.pop(next(iter(self._fit_analysis_cache)))

    def _cached_bet_analysis(self, result, p_min: float, p_max: float):
        cache_key = self._fit_analysis_cache_key("bet", result, float(p_min), float(p_max))
        cached = self._fit_analysis_cache.get(cache_key)
        if cached is not None:
            return cached
        analysis = bet_analysis(result, p_min, p_max)
        self._store_fit_analysis_cache(cache_key, analysis)
        return analysis

    def _cached_langmuir_analysis(self, result, p_min: float, p_max: float):
        cache_key = self._fit_analysis_cache_key("langmuir", result, float(p_min), float(p_max))
        cached = self._fit_analysis_cache.get(cache_key)
        if cached is not None:
            return cached
        analysis = langmuir_analysis(result, p_min, p_max)
        self._store_fit_analysis_cache(cache_key, analysis)
        return analysis

    def _cached_t_plot_pressure_analysis(
        self,
        result,
        p_min: float,
        p_max: float,
        thickness_params: dict[str, object] | None = None,
        thickness_method: str = DEFAULT_T_PLOT_THICKNESS_METHOD,
    ):
        cache_key = self._fit_analysis_cache_key(
            "t_plot_pressure",
            result,
            float(p_min),
            float(p_max),
            str(thickness_method),
            thickness_params or {},
        )
        cached = self._fit_analysis_cache.get(cache_key)
        if cached is not None:
            return cached
        analysis = t_plot_analysis(result, p_min, p_max, thickness_params, thickness_method)
        self._store_fit_analysis_cache(cache_key, analysis)
        return analysis

    def _cached_t_plot_thickness_analysis(
        self,
        result,
        t_min: float,
        t_max: float,
        p_min: float | None = None,
        p_max: float | None = None,
        thickness_params: dict[str, object] | None = None,
        thickness_method: str = DEFAULT_T_PLOT_THICKNESS_METHOD,
    ):
        cache_key = self._fit_analysis_cache_key(
            "t_plot_thickness",
            result,
            float(t_min),
            float(t_max),
            None if p_min is None else float(p_min),
            None if p_max is None else float(p_max),
            str(thickness_method),
            thickness_params or {},
        )
        cached = self._fit_analysis_cache.get(cache_key)
        if cached is not None:
            return cached
        analysis = t_plot_analysis_by_thickness(
            result,
            t_min,
            t_max,
            p_min,
            p_max,
            thickness_params,
            thickness_method,
        )
        self._store_fit_analysis_cache(cache_key, analysis)
        return analysis

    def _bet_fit_range_for_result(self, result) -> tuple[float, float]:
        if id(result) in self.custom_bet_fit_ranges:
            return self.custom_bet_fit_ranges[id(result)]
        try:
            return automatic_bet_range(result)
        except Exception:
            return BET_DEFAULT_RANGE

    def _is_custom_bet_fit(self, result) -> bool:
        return id(result) in self.custom_bet_fit_ranges

    def _set_custom_bet_fit_range(self, result, fit_range: tuple[float, float]) -> None:
        lo, hi = sorted((float(fit_range[0]), float(fit_range[1])))
        self.custom_bet_fit_ranges[id(result)] = (lo, hi)

    def _clear_custom_bet_fit_range(self, result) -> None:
        self.custom_bet_fit_ranges.pop(id(result), None)

    def _bet_analysis_for_result(self, result):
        fit_range = self._bet_fit_range_for_result(result)
        return self._cached_bet_analysis(result, fit_range[0], fit_range[1])

    def _langmuir_fit_range_for_result(self, result) -> tuple[float, float]:
        return self.custom_langmuir_fit_ranges.get(id(result), automatic_langmuir_range(result))

    def _is_custom_langmuir_fit(self, result) -> bool:
        return id(result) in self.custom_langmuir_fit_ranges

    def _set_custom_langmuir_fit_range(self, result, fit_range: tuple[float, float]) -> None:
        lo, hi = sorted((float(fit_range[0]), float(fit_range[1])))
        self.custom_langmuir_fit_ranges[id(result)] = (lo, hi)

    def _clear_custom_langmuir_fit_range(self, result) -> None:
        self.custom_langmuir_fit_ranges.pop(id(result), None)

    def _langmuir_analysis_for_result(self, result):
        fit_range = self._langmuir_fit_range_for_result(result)
        return self._cached_langmuir_analysis(result, fit_range[0], fit_range[1])

    def _default_t_plot_fit_range(
        self,
        thickness_method: str | None = None,
        thickness_params: dict[str, float] | None = None,
        pressure_range: tuple[float, float] | None = None,
    ) -> tuple[float, float]:
        p_min, p_max = pressure_range or T_PLOT_DEFAULT_PRESSURE_RANGE
        thickness_method = thickness_method or self.t_plot_thickness_method
        thickness_params = thickness_params or self.t_plot_thickness_params
        t_values = [
            value
            for value in (
                thickness_nm(p_min, thickness_method, thickness_params),
                thickness_nm(p_max, thickness_method, thickness_params),
            )
            if value is not None
        ]
        if len(t_values) == 2:
            return (min(t_values), max(t_values))
        return T_PLOT_DEFAULT_PRESSURE_RANGE

    def _t_plot_fit_range_for_result(self, result) -> tuple[float, float]:
        settings = self._t_plot_settings_for_result(result)
        return self.custom_t_plot_fit_ranges.get(
            id(result),
            self._default_t_plot_fit_range(
                str(settings["thickness_method"]),
                dict(settings["thickness_params"]),
                automatic_t_plot_pressure_range(result),
            ),
        )

    def _is_custom_t_plot_fit(self, result) -> bool:
        return id(result) in self.custom_t_plot_fit_ranges or self._has_custom_t_plot_settings(result)

    def _has_custom_t_plot_settings(self, result) -> bool:
        if id(result) not in self.custom_t_plot_settings:
            return False
        settings = self._t_plot_settings_for_result(result)
        method = str(settings["thickness_method"])
        if method != DEFAULT_T_PLOT_THICKNESS_METHOD:
            return True

        default_params = T_PLOT_THICKNESS_PARAM_DEFAULTS.get(method, DEFAULT_T_PLOT_THICKNESS_PARAMS)
        params_by_method = dict(settings.get("thickness_params_by_method", {}))
        active_params = dict(settings.get("thickness_params") or params_by_method.get(method, {}))
        for key, default_value in default_params.items():
            if not _float_equal(active_params.get(key), default_value):
                return True

        if str(settings["surface_area_mode"]) != DEFAULT_T_PLOT_SURFACE_AREA_MODE:
            return True
        if not _float_equal(settings["surface_area_correction"], DEFAULT_T_PLOT_SURFACE_AREA_CORRECTION):
            return True
        return False

    def _set_custom_t_plot_fit_range(self, result, fit_range: tuple[float, float]) -> None:
        lo, hi = sorted((float(fit_range[0]), float(fit_range[1])))
        self.custom_t_plot_fit_ranges[id(result)] = (lo, hi)

    def _clear_custom_t_plot_fit_range(self, result) -> None:
        self.custom_t_plot_fit_ranges.pop(id(result), None)

    def _t_plot_analysis_for_result(self, result):
        settings = self._t_plot_settings_for_result(result)
        fit_range = self._t_plot_fit_range_for_result(result)
        return self._cached_t_plot_thickness_analysis(
            result,
            fit_range[0],
            fit_range[1],
            thickness_params=dict(settings["thickness_params"]),
            thickness_method=str(settings["thickness_method"]),
        )

    @staticmethod
    def _test_time_sort_key(result) -> tuple[int, str]:
        return (int(result.test_started_raw or 0), str(result.test_started_time or ""))

    def _bet_sort_key(self, result) -> float:
        try:
            value = self._bet_analysis_for_result(result).surface_area_m2_g
            return float(value) if value is not None else 0.0
        except Exception:
            return 0.0

    def _langmuir_sort_key(self, result) -> float:
        try:
            value = self._langmuir_analysis_for_result(result).surface_area_m2_g
            return float(value) if value is not None else 0.0
        except Exception:
            return 0.0

    def _t_plot_sort_key(self, result) -> float:
        try:
            value = self._t_plot_analysis_for_result(result).external_surface_area_m2_g
            return float(value) if value is not None else 0.0
        except Exception:
            return 0.0

    def _bjh_pore_volume_sort_key(self, result) -> float:
        try:
            value = self._selected_pore_volume_for_result(result)
            return float(value) if value is not None else 0.0
        except Exception:
            return 0.0

    def _selected_pore_volume_for_result(
        self,
        result,
        diameter_range: tuple[float, float] | None = None,
    ) -> float | None:
        method = self._active_pore_volume_method()
        if method == PORE_VOLUME_METHOD_DFT:
            return self._dft_pore_volume_for_result(result, width_range=diameter_range)
        if method == PORE_VOLUME_METHOD_HK:
            return self._hk_pore_volume_for_result(result, width_range=diameter_range)
        if method == PORE_VOLUME_METHOD_DH:
            return self._dh_pore_volume_for_result(result, diameter_range=diameter_range)
        return self._bjh_pore_volume_for_result(result, diameter_range=diameter_range)

    def _bjh_pore_volume_for_result(
        self,
        result,
        diameter_range: tuple[float, float] | None = None,
    ) -> float | None:
        settings = self._bjh_settings_for_result(result)
        phase = self._bjh_pore_volume_phase(settings)
        if phase is None:
            return None
        diameter_range = diameter_range or self._selected_bjh_pore_volume_range()
        d_min, d_max = sorted((float(diameter_range[0]), float(diameter_range[1])))
        rows = self._cached_bjh_distribution_rows(
            result,
            phase=phase,
            thickness_method=str(settings["thickness_method"]),
            thickness_params=dict(settings["thickness_params"]),
            correction=str(settings["correction"]),
            open_pore_fraction=float(settings["open_pore_fraction"]),
            smooth=bool(settings["smooth_derivative"]),
        )
        return self._bjh_pore_volume_from_rows(rows, (d_min, d_max))

    def _dh_pore_volume_for_result(
        self,
        result,
        diameter_range: tuple[float, float] | None = None,
    ) -> float | None:
        settings = self._dh_settings_for_result(result)
        phase = self._dh_pore_volume_phase(settings)
        if phase is None:
            return None
        diameter_range = diameter_range or self._selected_pore_volume_range()
        d_min, d_max = sorted((float(diameter_range[0]), float(diameter_range[1])))
        rows = self._cached_dh_distribution_rows(
            result,
            phase=phase,
            thickness_method=str(settings["thickness_method"]),
            thickness_params=dict(settings["thickness_params"]),
            smooth=bool(settings["smooth_derivative"]),
        )
        return self._bjh_pore_volume_from_rows(rows, (d_min, d_max))

    def _hk_pore_volume_for_result(
        self,
        result,
        width_range: tuple[float, float] | None = None,
    ) -> float | None:
        settings = self._hk_settings_for_result(result)
        width_range = width_range or self.hk_pore_volume_range
        w_min, w_max = sorted((float(width_range[0]), float(width_range[1])))
        rows = self._cached_hk_distribution_rows(
            result,
            geometry=str(settings["geometry"]),
            adsorbent_key=str(settings["adsorbent_key"]),
            adsorptive_key=str(settings["adsorptive_key"]),
            adsorbent_properties=dict(settings["adsorbent_properties"]),
            adsorptive_properties=dict(settings["adsorptive_properties"]),
            interaction_parameter_erg_cm4=float(settings["interaction_parameter"]),
            interaction_parameter_mode=str(settings["interaction_parameter_mode"]),
            cheng_yang_correction=bool(settings["cheng_yang_correction"]),
            smooth=bool(settings["smooth_derivative"]),
        )
        return self._hk_pore_volume_from_rows(rows, (w_min, w_max))

    def _dft_pore_volume_for_result(
        self,
        result,
        width_range: tuple[float, float] | None = None,
    ) -> float | None:
        settings = self._dft_settings_for_result(result)
        width_range = width_range or self.dft_pore_volume_range
        w_min, w_max = sorted((float(width_range[0]), float(width_range[1])))
        distribution = self._cached_dft_result(
            result,
            analysis_type=str(settings["analysis_type"]),
            geometry=str(settings["geometry"]),
            model=str(settings["model"]),
            regularization=float(settings["regularization"]),
        )
        return self._hk_pore_volume_from_rows(list(getattr(distribution, "rows", [])), (w_min, w_max))

    @staticmethod
    def _bjh_pore_volume_from_rows(
        rows: list[dict[str, float]],
        diameter_range: tuple[float, float],
    ) -> float | None:
        if not rows:
            return None
        d_min, d_max = sorted((float(diameter_range[0]), float(diameter_range[1])))
        volume = 0.0
        for row in rows:
            try:
                diameter = float(row["pore_diameter_nm"])
                increment = float(row["incremental_pore_volume_cm3_g"])
            except (KeyError, TypeError, ValueError):
                continue
            if np.isfinite(diameter) and np.isfinite(increment) and d_min <= diameter <= d_max:
                volume += increment
        return volume

    @staticmethod
    def _hk_pore_volume_from_rows(
        rows: list[dict[str, float]],
        width_range: tuple[float, float],
    ) -> float | None:
        if not rows:
            return None
        w_min, w_max = sorted((float(width_range[0]), float(width_range[1])))
        volume = 0.0
        for row in rows:
            try:
                width = float(row.get("pore_width_nm", row.get("pore_diameter_nm")))
                increment = float(row["incremental_pore_volume_cm3_g"])
            except (TypeError, ValueError, KeyError):
                continue
            if np.isfinite(width) and np.isfinite(increment) and w_min <= width <= w_max:
                volume += increment
        return volume

    def _selected_bjh_pore_volume_range(self) -> tuple[float, float]:
        return self.bjh_pore_volume_range

    def _selected_pore_volume_range(self) -> tuple[float, float]:
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_DFT:
            return self.dft_pore_volume_range
        if self._active_pore_volume_method() == PORE_VOLUME_METHOD_HK:
            return self.hk_pore_volume_range
        return self.bjh_pore_volume_range

    def _has_custom_bjh_settings(self, result) -> bool:
        if id(result) not in self.custom_bjh_settings:
            return False
        settings = self._bjh_settings_for_result(result)
        defaults = self._bjh_default_settings_for_result(result)
        method = str(settings["thickness_method"])
        default_method = str(defaults["thickness_method"])
        if method != default_method:
            return True
        default_params = dict(defaults["thickness_params"])
        active_params = dict(settings["thickness_params"])
        if not _thickness_params_equal(active_params, default_params):
            return True
        if str(settings["correction"]) != str(defaults["correction"]):
            return True
        if not _float_equal(settings["open_pore_fraction"], defaults["open_pore_fraction"]):
            return True
        if bool(settings["smooth_derivative"]) != bool(defaults["smooth_derivative"]):
            return True
        if bool(settings["show_adsorption"]) != bool(defaults["show_adsorption"]):
            return True
        if bool(settings["show_desorption"]) != bool(defaults["show_desorption"]):
            return True
        return False

    def _has_custom_dh_settings(self, result) -> bool:
        if id(result) not in self.custom_dh_settings:
            return False
        settings = self._dh_settings_for_result(result)
        defaults = self._dh_default_settings_for_result(result)
        if str(settings["thickness_method"]) != str(defaults["thickness_method"]):
            return True
        if not _thickness_params_equal(dict(settings["thickness_params"]), dict(defaults["thickness_params"])):
            return True
        if bool(settings["smooth_derivative"]) != bool(defaults["smooth_derivative"]):
            return True
        if bool(settings["show_adsorption"]) != bool(defaults["show_adsorption"]):
            return True
        if bool(settings["show_desorption"]) != bool(defaults["show_desorption"]):
            return True
        return False

    def _has_custom_hk_settings(self, result) -> bool:
        if id(result) not in self.custom_hk_settings:
            return False
        settings = self._hk_settings_for_result(result)
        defaults = self._default_hk_settings()
        for key in (
            "geometry",
            "adsorbent_key",
            "adsorptive_key",
            "interaction_parameter_mode",
        ):
            if str(settings[key]) != str(defaults[key]):
                return True
        if not _settings_mapping_equal(
            dict(settings["adsorbent_properties"]),
            dict(defaults["adsorbent_properties"]),
        ):
            return True
        if not _settings_mapping_equal(
            dict(settings["adsorptive_properties"]),
            dict(defaults["adsorptive_properties"]),
        ):
            return True
        if not _float_equal(settings["interaction_parameter"], defaults["interaction_parameter"]):
            return True
        if bool(settings["cheng_yang_correction"]) != bool(defaults["cheng_yang_correction"]):
            return True
        if bool(settings["smooth_derivative"]) != bool(defaults["smooth_derivative"]):
            return True
        return False

    def _has_custom_dft_settings(self, result) -> bool:
        if id(result) not in self.custom_dft_settings:
            return False
        settings = self._dft_settings_for_result(result)
        defaults = self._dft_default_settings_for_result(result)
        for key in ("analysis_type", "geometry", "model"):
            if str(settings.get(key)) != str(defaults.get(key)):
                return True
        if not _float_equal(settings.get("regularization"), defaults.get("regularization")):
            return True
        return False

    def _bjh_pore_volume_phase(self, settings: dict[str, object] | None = None) -> str | None:
        if settings is None:
            show_adsorption = self.bjh_show_adsorption
            show_desorption = self.bjh_show_desorption
        else:
            show_adsorption = bool(settings["show_adsorption"])
            show_desorption = bool(settings["show_desorption"])
        if show_adsorption:
            return "adsorption"
        if show_desorption:
            return "desorption"
        return None

    def _dh_pore_volume_phase(self, settings: dict[str, object] | None = None) -> str | None:
        if settings is None:
            show_adsorption = self.dh_show_adsorption
            show_desorption = self.dh_show_desorption
        else:
            show_adsorption = bool(settings["show_adsorption"])
            show_desorption = bool(settings["show_desorption"])
        if show_adsorption:
            return "adsorption"
        if show_desorption:
            return "desorption"
        return None

    def _resize_sample_columns(self) -> None:
        defaults = {
            VISIBLE_COLUMN: 30,
            FILE_COLUMN: 170,
            TEST_TIME_COLUMN: 250,
            BET_COLUMN: 120,
            LANGMUIR_COLUMN: 200,
            T_PLOT_COLUMN: 200,
            BJH_PORE_VOLUME_COLUMN: 190,
        }
        self._updating_sample_column_widths = True
        try:
            if not self._sample_column_widths_initialized:
                for column, width in defaults.items():
                    self.sample_table.setColumnWidth(column, width)
                self.sample_column_widths = {
                    column: self.sample_table.columnWidth(column)
                    for column in range(self.sample_table.columnCount())
                }
                self._sample_column_widths_initialized = True
            else:
                for column in range(self.sample_table.columnCount()):
                    width = self.sample_column_widths.get(column, defaults.get(column))
                    if width is not None:
                        self.sample_table.setColumnWidth(column, width)
        finally:
            self._updating_sample_column_widths = False
        self._position_header_controls()

    def _make_table(self, headers: list[str]) -> QtWidgets.QTableWidget:
        table = QtWidgets.QTableWidget(0, len(headers))
        table.setHorizontalHeaderLabels(headers)
        table.verticalHeader().setVisible(False)
        table.setAlternatingRowColors(True)
        table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        table.setVerticalScrollMode(QtWidgets.QAbstractItemView.ScrollPerPixel)
        table.horizontalHeader().setStretchLastSection(True)
        return table

    def _fill_two_column_table(self, table: QtWidgets.QTableWidget, rows: list[tuple[str, str]]) -> None:
        table.setRowCount(len(rows))
        for row, (name, value) in enumerate(rows):
            self._set_table_item(table, row, 0, name)
            self._set_table_item(table, row, 1, value)

    def _set_table_item(self, table: QtWidgets.QTableWidget, row: int, column: int, text: str) -> None:
        table.setItem(row, column, self._table_item(text))

    def _table_item(self, text: str, *, tooltip: str | None = None, alignment=None) -> QtWidgets.QTableWidgetItem:
        item = QtWidgets.QTableWidgetItem(str(text))
        item.setFlags(QtCore.Qt.ItemIsEnabled | QtCore.Qt.ItemIsSelectable)
        item.setForeground(QtGui.QBrush(QtGui.QColor("#111827")))
        if tooltip:
            item.setToolTip(tooltip)
        if alignment is not None:
            item.setTextAlignment(alignment)
        return item


def active_metric_rows(
    result,
    pressure_range: tuple[float, float] | None = None,
    bet_fit_range: tuple[float, float] | None = None,
    langmuir_fit_range: tuple[float, float] | None = None,
    t_plot_fit_range: tuple[float, float] | None = None,
    t_plot_thickness_method: str = DEFAULT_T_PLOT_THICKNESS_METHOD,
    t_plot_thickness_params: dict[str, float] | None = None,
    t_plot_surface_area_mode: str = "BET",
    t_plot_input_surface_area: float | None = None,
    t_plot_surface_area_correction: float = SURFACE_AREA_CORRECTION_FACTOR,
    bet_analysis_result=None,
    langmuir_analysis_result=None,
    t_plot_analysis_result=None,
) -> list[tuple[str, str]]:
    analyses = None
    if bet_analysis_result is None or langmuir_analysis_result is None or t_plot_analysis_result is None:
        analyses = analysis_bundle(result) if pressure_range is None else analysis_bundle(result, pressure_range[0], pressure_range[1])

    from tristar_bet.analysis import bet_analysis as _bet_analysis
    if bet_analysis_result is not None:
        bet = bet_analysis_result
    elif bet_fit_range is not None:
        bet = _bet_analysis(result, bet_fit_range[0], bet_fit_range[1])
    else:
        bet = analyses["BET"]
    if langmuir_analysis_result is not None:
        langmuir = langmuir_analysis_result
    elif langmuir_fit_range is not None:
        langmuir = langmuir_analysis(result, langmuir_fit_range[0], langmuir_fit_range[1])
    else:
        langmuir = analyses["Langmuir"]
    if t_plot_analysis_result is not None:
        t_plot = t_plot_analysis_result
    elif t_plot_fit_range is not None:
        if pressure_range is None:
            t_plot = t_plot_analysis_by_thickness(
                result,
                t_plot_fit_range[0],
                t_plot_fit_range[1],
                thickness_params=t_plot_thickness_params,
                thickness_method=t_plot_thickness_method,
            )
        else:
            t_plot = t_plot_analysis_by_thickness(
                result,
                t_plot_fit_range[0],
                t_plot_fit_range[1],
                pressure_range[0],
                pressure_range[1],
                t_plot_thickness_params,
                t_plot_thickness_method,
            )
    else:
        t_plot = analyses["t-Plot"]

    rows = [
        ("文件名", _display_file_name(result)),
        ("样品名称", result.sample_name),
        ("测试时间", result.test_started_time),
        ("测试结束时间", result.test_completed_time),
        ("测试耗时", result.test_duration_time),
        ("样品保存时间", result.sample_saved_time),
        ("设备厂家", _instrument_manufacturer(result)),
        ("设备型号", _instrument_model(result)),
        ("当前选区", _pressure_range_text(pressure_range)),
    ]
    if bet_fit_range is not None:
        rows.append(("BET 拟合区间", _pressure_range_text(bet_fit_range)))
    if langmuir_fit_range is not None:
        rows.append(("Langmuir 拟合区间", _pressure_range_text(langmuir_fit_range)))
    if t_plot_fit_range is not None:
        rows.append(("t-Plot 厚度区间", _thickness_range_text(t_plot_fit_range)))
    t_plot_regression = _t_plot_mmol_regression(t_plot)
    bet_correlation = _correlation_from_r_squared(bet.r_squared)
    t_plot_correlation = _correlation_from_r_squared(t_plot.r_squared)
    t_plot_total_surface_area = _t_plot_total_surface_area(
        t_plot_surface_area_mode,
        bet.surface_area_m2_g,
        langmuir.surface_area_m2_g,
        t_plot_input_surface_area,
    )
    micropore_area = _micropore_area_m2_g(
        t_plot_total_surface_area,
        t_plot.external_surface_area_m2_g,
        t_plot_surface_area_correction,
    )
    rows += [
        ("样品质量", f"{_fmt(result.sample.sample_mass_g)} g"),
        ("吸附质", _adsorptive_label(result)),
        ("数据点数", str(result.point_count)),
        ("BET 状态", status_text(bet.status)),
        ("BET 比表面积", _value_pm_text(bet.surface_area_m2_g, bet.surface_area_standard_error, "m2/g")),
        ("BET 斜率", _value_pm_text(bet.slope, bet.slope_standard_error, "g/cm3 STP")),
        ("BET Y 截距", _value_pm_text(bet.intercept, bet.intercept_standard_error, "g/cm3 STP")),
        ("BET 单层容量", f"{_fmt(bet.monolayer_capacity_cm3_g_stp)} cm3/g STP"),
        ("BET C 常数", _fmt(bet.c_constant)),
        ("BET 相关系数", _fmt(bet_correlation, 7)),
        ("BET R2", _fmt(bet.r_squared, 6)),
        ("Langmuir 状态", status_text(langmuir.status)),
        ("Langmuir 比表面积", f"{_fmt(langmuir.surface_area_m2_g)} m2/g"),
        ("Langmuir 单层容量", f"{_fmt(langmuir.monolayer_capacity_cm3_g_stp)} cm3/g STP"),
        ("Langmuir R2", _fmt(langmuir.r_squared, 6)),
        ("t-Plot 状态", status_text(t_plot.status)),
        ("t-Plot 微孔体积", f"{_fmt(t_plot.micropore_volume_cm3_g)} cm3/g"),
        ("t-Plot 微孔面积", f"{_fmt(micropore_area)} m2/g"),
        ("t-Plot 外比表面积", f"{_fmt(t_plot.external_surface_area_m2_g)} m2/g"),
        ("t-Plot 斜率", _value_pm_text(t_plot_regression["slope"], t_plot_regression["slope_se"], "mmol/g/nm")),
        ("t-Plot Y 截距", _value_pm_text(t_plot_regression["intercept"], t_plot_regression["intercept_se"], "mmol/g")),
        ("t-Plot 相关系数", _fmt(t_plot_correlation, 6)),
        ("t-Plot R2", _fmt(t_plot.r_squared, 6)),
        ("t-Plot 比表面积修正因子", _fmt(t_plot_surface_area_correction, 3)),
        ("t-Plot 密度转换因子", _fmt(density_conversion_factor(result), 9)),
        (f"t-Plot 总表面积({_t_plot_surface_area_label(t_plot_surface_area_mode)})", f"{_fmt(t_plot_total_surface_area)} m2/g"),
        ("t-Plot 厚度方程", _t_plot_thickness_label(t_plot_thickness_method)),
        ("自由空间来源", result.free_space.vfree_factor_source),
    ]
    return rows


def condition_rows(result) -> list[tuple[str, str]]:
    sample = result.sample
    run = result.run_conditions
    free = result.free_space
    props = result.adsorptive_properties
    rows = [
        ("文件路径", result.header.file_path),
        ("测试时间", result.test_started_time),
        ("测试结束时间", result.test_completed_time),
        ("测试耗时", result.test_duration_time),
        ("文件创建时间", result.header.created_time),
        ("文件修改时间", result.header.modified_time),
        ("样品保存时间", result.sample_saved_time),
        ("样品名称", result.sample_name),
        ("操作员", sample.operator),
        ("样品质量", f"{_fmt(sample.sample_mass_g)} g"),
        ("样品密度", f"{_fmt(sample.sample_density_g_cm3)} g/cm3"),
        ("吸附质助记符", run.adsorptive_short),
        ("吸附质名称", run.adsorptive_name),
        ("浴温", f"{_fmt(run.bath_temperature_K)} K"),
        ("Po 参考压力", f"{_fmt(run.po_reference_mmHg)} mmHg"),
        ("平衡间隔", f"{_fmt(run.equilibration_interval_s)} s"),
        ("自由空间平衡时间", f"{_fmt(run.free_space_equilibration_time_h)} h"),
        ("输入常温自由空间", f"{_fmt(run.ambient_free_space_entered_cm3)} cm3"),
        ("输入分析自由空间", f"{_fmt(run.analysis_free_space_entered_cm3)} cm3"),
        ("实际温自由空间", f"{_fmt(free.warm_free_space_cm3)} cm3"),
        ("冷自由空间", f"{_fmt(free.cold_free_space_cm3)} cm3"),
        ("Stem volume", f"{_fmt(free.stem_volume_cm3)} cm3"),
        ("Vbath", f"{_fmt(free.vbath_cm3)} cm3"),
        ("Vfree factor", f"{_fmt(free.vfree_factor_cm3)} cm3"),
        ("非理想因子", _fmt(free.nonideality_factor, 9)),
    ]
    if props is not None:
        rows.extend(
            [
                ("吸附质属性", props.adsorptive),
                ("最大歧管压力", f"{_fmt(props.max_manifold_pressure_kPa)} kPa"),
                ("分子截面积", f"{_fmt(props.molecular_cross_sectional_area_nm2)} nm2"),
                ("密度转换因子", _fmt(props.density_conversion_factor, 9)),
                ("Psat 表行数", str(len(props.psat_table))),
            ]
        )
    return rows


def status_text(status: str) -> str:
    return {
        "ok": "区间计算完成",
        "warning_negative_c": "区间计算完成；BET C<=0，需核对报告选点",
        "not_enough_points": "区间有效点不足",
        "not_enough_valid_points": "区间有效数值不足",
        "invalid_monolayer_capacity": "单层容量无效",
    }.get(status, status)


def _display_file_name(result) -> str:
    return Path(result.file_name).stem


def _instrument_manufacturer(result) -> str:
    value = result.method_options.get("instrument_manufacturer", "")
    return str(value) if value else "Micromeritics"


def _instrument_model(result) -> str:
    value = result.method_options.get("instrument_model", "")
    return str(value) if value else "TriStar II"


def _adsorptive_label(result) -> str:
    value = result.run_conditions.adsorptive_short or result.run_conditions.adsorptive_name
    if value:
        return value
    props = result.adsorptive_properties
    if props is not None:
        return props.mnemonic or props.adsorptive
    return ""


def _pressure_range_text(pressure_range: tuple[float, float] | None) -> str:
    if pressure_range is None:
        return "默认算法区间"
    return f"P/P0 {_fmt(pressure_range[0], 6)} - {_fmt(pressure_range[1], 6)}"


def _thickness_range_text(thickness_range: tuple[float, float] | None) -> str:
    if thickness_range is None:
        return "默认厚度区间"
    return f"{_fmt(thickness_range[0], 6)} - {_fmt(thickness_range[1], 6)} nm"


def _micropore_area_m2_g(total_surface_area, external_surface_area, correction_factor: float = 1.0):
    if total_surface_area is None or external_surface_area is None:
        return None
    try:
        return max(0.0, float(total_surface_area) - float(external_surface_area) * float(correction_factor))
    except (TypeError, ValueError):
        return None


def _t_plot_total_surface_area(mode: str, bet_area, langmuir_area, input_area):
    if mode == "Langmuir":
        return langmuir_area
    if mode == "Input":
        return input_area
    return bet_area


def _t_plot_surface_area_label(mode: str) -> str:
    if mode == "Langmuir":
        return "Langmuir"
    if mode == "Input":
        return "输入"
    return "BET"


def _correlation_from_r_squared(r_squared):
    if r_squared is None:
        return None
    try:
        return math.sqrt(max(0.0, float(r_squared)))
    except (TypeError, ValueError):
        return None


def _value_pm_text(value, error=None, unit: str = "") -> str:
    if value is None:
        return "n/a"
    text = _fmt(value, 6)
    if error is not None:
        text += f" ± {_fmt(error, 6)}"
    if unit:
        text += f" {unit}"
    return text


def _t_plot_mmol_regression(t_plot) -> dict[str, float | None]:
    result = {"slope": None, "slope_se": None, "intercept": None, "intercept_se": None}
    rows = getattr(t_plot, "rows", None) or []
    x_values = []
    y_values = []
    for row in rows:
        try:
            x = float(row["thickness_nm"])
            y = float(row["quantity_adsorbed_cm3_g_stp"]) / CM3_STP_PER_MMOL
        except (KeyError, TypeError, ValueError):
            continue
        if np.isfinite(x) and np.isfinite(y):
            x_values.append(x)
            y_values.append(y)
    if len(x_values) < 2:
        return result

    x = np.asarray(x_values, dtype=float)
    y = np.asarray(y_values, dtype=float)
    slope, intercept = np.polyfit(x, y, 1)
    result["slope"] = float(slope)
    result["intercept"] = float(intercept)

    if x.size <= 2:
        return result
    fitted = slope * x + intercept
    residual = y - fitted
    sxx = float(np.sum((x - np.mean(x)) ** 2))
    if sxx <= 0.0:
        return result
    residual_variance = float(np.sum(residual ** 2)) / float(x.size - 2)
    result["slope_se"] = math.sqrt(residual_variance / sxx)
    result["intercept_se"] = math.sqrt(residual_variance * (1.0 / x.size + float(np.mean(x)) ** 2 / sxx))
    return result


def export_results_xlsx(
    results,
    path: str | Path,
    *,
    pore_volume_header: str = "选区孔容量(cm3/g)",
    pore_volumes: list[float | None] | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "摘要"
    _write_rows(
        summary_sheet,
        [
            [
                "文件",
                "样品",
                "测试时间",
                "测试结束时间",
                "测试耗时",
                "质量(g)",
                "点数",
                "BET状态",
                "BET面积(m2/g)",
                "BET Vm(cm3/g)",
                "BET C",
                "Langmuir状态",
                "Langmuir面积(m2/g)",
                "t-Plot状态",
                "t-Plot外比表面积(m2/g)",
                "t-Plot微孔体积(cm3/g)",
                pore_volume_header,
            ]
        ],
        bold_first=True,
    )
    for index, result in enumerate(results):
        analyses = analysis_bundle(result)
        bet = analyses["BET"]
        langmuir = analyses["Langmuir"]
        t_plot = analyses["t-Plot"]
        pore_volume = pore_volumes[index] if pore_volumes is not None and index < len(pore_volumes) else None
        summary_sheet.append(
            [
                result.file_name,
                result.sample_name,
                result.test_started_time,
                result.test_completed_time,
                result.test_duration_time,
                result.sample.sample_mass_g,
                result.point_count,
                status_text(bet.status),
                bet.surface_area_m2_g,
                bet.monolayer_capacity_cm3_g_stp,
                bet.c_constant,
                status_text(langmuir.status),
                langmuir.surface_area_m2_g,
                status_text(t_plot.status),
                t_plot.external_surface_area_m2_g,
                t_plot.micropore_volume_cm3_g,
                pore_volume,
            ]
        )

    isotherm_sheet = workbook.create_sheet("实际等温线")
    isotherm_sheet.append(
        [
            "文件",
            "样品",
            "点",
            "阶段",
            "P/P0",
            "压力(mmHg)",
            "吸附量(cm3/g STP)",
            "吸附量(mmol/g)",
            "Po(mmHg)",
            "Elapsed(s)",
            "Elapsed",
        ]
    )
    for result in results:
        for point in result.isotherm:
            isotherm_sheet.append(
                [
                    result.file_name,
                    result.sample_name,
                    point.index,
                    "吸附" if point.phase == "adsorption" else "脱附",
                    point.relative_pressure,
                    point.absolute_pressure_mmHg,
                    point.quantity_adsorbed_cm3_g_stp,
                    point.quantity_adsorbed_mmol_g,
                    point.saturation_pressure_mmHg,
                    point.elapsed_seconds,
                    point.elapsed_time,
                ]
            )

    target_sheet = workbook.create_sheet("目标压力表")
    target_sheet.append(["文件", "样品", "行", "阶段", "起始P/P0", "终止P/P0", "步长P/P0", "偏移"])
    for result in results:
        for item in result.target_pressure_table:
            target_sheet.append(
                [
                    result.file_name,
                    result.sample_name,
                    item.row,
                    "吸附" if item.branch == "adsorption" else "脱附",
                    item.starting_pressure_p_po,
                    item.ending_pressure_p_po,
                    item.pressure_increment_p_po,
                    item.ending_pressure_rel_offset,
                ]
            )

    for name in ("BET", "Langmuir", "t-Plot"):
        sheet = workbook.create_sheet(name)
        sheet.append(["文件", "样品", "字段", "值"])
        for result in results:
            fit = analysis_bundle(result)[name]
            for key, value in fit.__dict__.items():
                if key == "rows":
                    continue
                sheet.append([result.file_name, result.sample_name, key, value])
            sheet.append([])
            if fit.rows:
                headers = list(fit.rows[0].keys())
                sheet.append(["文件", "样品", *headers])
                for row in fit.rows:
                    sheet.append([result.file_name, result.sample_name, *[row.get(header) for header in headers]])

    conditions_sheet = workbook.create_sheet("样品条件")
    conditions_sheet.append(["文件", "样品", "字段", "值"])
    for result in results:
        for name, value in condition_rows(result):
            conditions_sheet.append([result.file_name, result.sample_name, name, value])

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        for column_cells in sheet.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 10), 48)

    workbook.save(path)


def _write_rows(sheet, rows: list[list[object]], *, bold_first: bool = False) -> None:
    from openpyxl.styles import Font

    for row in rows:
        sheet.append(row)
    if bold_first and rows:
        for cell in sheet[1]:
            cell.font = Font(bold=True)


def _fmt(value, digits: int = 6) -> str:
    if value is None:
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if not math_isfinite(number):
        return ""
    return f"{number:.{digits}g}"


def _fmt_nm(value: float) -> str:
    """Format a diameter (nm) for column labels: trim trailing zeros (2.0 -> '2')."""
    text = f"{float(value):.2f}".rstrip("0").rstrip(".")
    return text or "0"


def math_isfinite(value: float) -> bool:
    return value == value and value not in (float("inf"), float("-inf"))


def run(argv: list[str] | None = None) -> int:
    QtCore.QCoreApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True)
    app = QtWidgets.QApplication(argv or sys.argv)
    app.setApplicationName("UnifiedBET")
    app.setApplicationDisplayName("")
    app.setStyle("windowsvista")
    if APP_ICON_PATH.exists():
        app.setWindowIcon(QtGui.QIcon(str(APP_ICON_PATH)))
    app.setFont(QtGui.QFont("Microsoft YaHei UI", 9))
    window = MainWindow()
    window.show()
    exec_func = getattr(app, "exec", app.exec_)
    return exec_func()
